# -*- coding: utf-8 -*-
"""
SQM Inventory Engine - CRUD Mixin
=================================

v2.9.91 - Extracted from inventory.py

CRUD operations and search functions
"""

import sqlite3
import logging
from datetime import datetime, date
from typing import Dict

logger = logging.getLogger(__name__)


class CRUDMixin:
    """
    CRUD and search mixin
    
    Methods for inventory CRUD operations and LOT search
    """
    
    def add_inventory(self, lot_no: str, sap_no: str = None, bl_no: str = None,
                      container_no: str = None, product: str = None,
                      product_code: str = None, mxbg_pallet: int = 20,
                      net_weight: float = 10000, warehouse: str = 'GY',
                      arrival_date=None, stock_date=None, **kwargs) -> Dict:
        """
        Add single LOT inventory (v3.8.7: 18열 전체 지원)
        
        Args:
            lot_no: LOT number (required)
            sap_no: SAP number
            bl_no: B/L number
            container_no: Container number
            product: Product name
            product_code: Product code
            mxbg_pallet: Number of tonbags (default 20)
            net_weight: Total weight in kg (default 10000)
            warehouse: Warehouse code (default 'GY')
            arrival_date: Arrival date
            stock_date: Stock date
            **kwargs: lot_sqm, gross_weight, salar_invoice_no, ship_date, free_time 등
        
        Returns:
            Result dict with success, lot_no, tonbags_created
        """
        try:
            # v3.8.7: kwargs에서 추가 필드 추출
            lot_sqm = kwargs.get('lot_sqm', '')
            gross_weight = kwargs.get('gross_weight', net_weight)
            salar_invoice_no = kwargs.get('salar_invoice_no', '') or kwargs.get('invoice_no', '')
            ship_date = kwargs.get('ship_date', '')
            free_time = kwargs.get('free_time', 0)
            
            # Date handling
            if arrival_date is None:
                arrival_date = date.today()
            if stock_date is None:
                stock_date = date.today()
            
            # Convert dates to strings
            if hasattr(arrival_date, 'isoformat'):
                arrival_date = arrival_date.isoformat()
            if hasattr(stock_date, 'isoformat'):
                stock_date = stock_date.isoformat()
            
            # Check duplicate
            existing = self.db.fetchone(
                "SELECT id FROM inventory WHERE lot_no = ?", (lot_no,)
            )
            
            if existing:
                return {
                    'success': False,
                    'lot_no': lot_no,
                    'tonbags_created': 0,
                    'message': f'LOT already exists: {lot_no}'
                }
            
            # Calculate weight per bag (v5.6.0 대원칙: 샘플 1kg 제외 후 균등 분배)
            # LOT 총무게 = (톤백수 × 단가) + 샘플 1kg
            # → 톤백 단가 = (총무게 - 1kg) / 톤백수
            SAMPLE_WEIGHT_KG = 1.0
            weight_per_bag = (net_weight - SAMPLE_WEIGHT_KG) / mxbg_pallet if mxbg_pallet > 0 else 500
            
            with self.db.transaction("IMMEDIATE"):
                now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                # Insert inventory (v3.8.7: 18열 전체)
                self.db.execute("""
                    INSERT INTO inventory (
                        lot_no, sap_no, bl_no, container_no, product, product_code,
                        lot_sqm, mxbg_pallet, net_weight, gross_weight,
                        current_weight, initial_weight, picked_weight,
                        salar_invoice_no, ship_date, arrival_date, free_time,
                        warehouse, stock_date, status, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?, ?, ?, ?, ?, 'AVAILABLE', ?)
                """, (lot_no, sap_no, bl_no, container_no, product, product_code,
                      lot_sqm, mxbg_pallet, net_weight, gross_weight,
                      net_weight, net_weight,
                      salar_invoice_no, ship_date, arrival_date, free_time,
                      warehouse, stock_date, now))
                
                # P3: DB 독립적 ID 조회 (SQLite: lastrowid, PG: RETURNING)
                if hasattr(self.db, 'insert_returning_id'):
                    inv_id = self.db.insert_returning_id("""
                        SELECT id FROM inventory WHERE lot_no = ?
                    """, (lot_no,))
                else:
                    inv_row = self.db.fetchone("SELECT id FROM inventory WHERE lot_no = ?", (lot_no,))
                    inv_id = inv_row['id'] if inv_row else None
                
                # Insert tonbags (일반: 1~mxbg_pallet)
                for sub in range(1, mxbg_pallet + 1):
                    self.db.execute("""
                        INSERT INTO inventory_tonbag (
                            inventory_id, lot_no, sub_lt, weight, status, 
                            is_sample, created_at
                        ) VALUES (?, ?, ?, ?, 'AVAILABLE', 0, ?)
                    """, (inv_id, lot_no, sub, weight_per_bag, now))
                
                # v3.9.1: 샘플 톤백 자동 생성 (sub_lt=0, 1kg, is_sample=1)
                sample_weight = 1.0  # 1kg (= 0.001 MT)
                self.db.execute("""
                    INSERT INTO inventory_tonbag (
                        inventory_id, lot_no, sub_lt, weight, status,
                        is_sample, created_at
                    ) VALUES (?, ?, 0, ?, 'AVAILABLE', 1, ?)
                """, (inv_id, lot_no, sample_weight, now))
                
                logger.info(f"[add_inventory] 샘플 톤백 생성: {lot_no}/0 (1kg)")
                
                # Movement history
                self.db.execute("""
                    INSERT INTO stock_movement (
                        movement_type, lot_no, qty_kg, created_at
                    ) VALUES ('INBOUND', ?, ?, ?)
                """, (lot_no, net_weight, now))
            
            logger.info(f"[add_inventory] Success: {lot_no}, {mxbg_pallet} tonbags + 1 sample")
            
            return {
                'success': True,
                'lot_no': lot_no,
                'tonbags_created': mxbg_pallet + 1,  # v3.9.1: 일반 + 샘플
                'sample_created': True,
                'message': 'OK'
            }
            
        except (sqlite3.OperationalError, sqlite3.IntegrityError, OSError) as e:
            logger.error(f"[add_inventory] Error: {e}")
            return {
                'success': False,
                'lot_no': lot_no,
                'tonbags_created': 0,
                'message': str(e)
            }
    
    def delete_inventory(self, lot_no: str, force: bool = False, 
                         confirmed: bool = False) -> Dict:
        """
        Delete inventory LOT
        
        Args:
            lot_no: LOT number
            force: Force delete even if not AVAILABLE (requires confirmed=True)
            confirmed: User confirmation for deletion (required for actual deletion)
            
        Returns:
            Result dict
            
        Note:
            데이터 보호 정책에 따라 confirmed=True 없이는 삭제되지 않습니다.
        """
        result = {'success': False, 'error': None}
        
        # 데이터 보호: confirmed 체크
        if not confirmed:
            result['error'] = "삭제 승인 필요: confirmed=True를 전달해주세요"
            result['requires_confirmation'] = True
            logger.warning(f"[delete_inventory] Blocked - no confirmation: {lot_no}")
            return result
        
        # force 사용 시에도 confirmed 필수
        if force and not confirmed:
            result['error'] = "force 옵션 사용 시 confirmed=True 필수"
            return result
        
        try:
            # Check if LOT exists
            lot = self.db.fetchone(
                "SELECT status FROM inventory WHERE lot_no = ?", (lot_no,)
            )
            
            if not lot:
                result['error'] = f"LOT not found: {lot_no}"
                return result
            
            # Check status unless forced
            if not force and lot['status'] != 'AVAILABLE':
                result['error'] = f"Cannot delete LOT with status: {lot['status']}"
                return result
            
            # Check if any tonbags are picked
            if not force:
                picked = self.db.fetchone("""
                    SELECT COUNT(*) as cnt FROM inventory_tonbag
                    WHERE lot_no = ? AND status != 'AVAILABLE'
                """, (lot_no,))
                
                if picked and picked['cnt'] > 0:
                    result['error'] = f"Cannot delete: {picked['cnt']} tonbags are not AVAILABLE"
                    return result
            
            with self.db.transaction("IMMEDIATE"):
                # Delete tonbags
                self.db.execute(
                    "DELETE FROM inventory_tonbag WHERE lot_no = ?", (lot_no,)
                )
                
                # Delete inventory
                self.db.execute(
                    "DELETE FROM inventory WHERE lot_no = ?", (lot_no,)
                )
                
                # Record movement
                from datetime import datetime
                now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                self.db.execute("""
                    INSERT INTO stock_movement (
                        movement_type, lot_no, qty_kg, created_at, remarks
                    ) VALUES ('DELETE', ?, 0, ?, 'Manual deletion')
                """, (lot_no, now))
            
            result['success'] = True
            logger.info(f"[delete_inventory] Deleted: {lot_no}")
            
        except (sqlite3.OperationalError, sqlite3.IntegrityError, OSError) as e:
            result['error'] = str(e)
            logger.error(f"[delete_inventory] Error: {e}")
        
        return result
    
    # NOTE: search_lots → QueryMixin으로 이관 완료 (v3.8.4 데드코드 정리)
    
    def update_inventory(self, lot_no: str, confirmed: bool = False, **updates) -> Dict:
        """
        Update inventory fields
        
        Args:
            lot_no: LOT number
            confirmed: User confirmation for critical field updates
            **updates: Fields to update (e.g., product='NICKEL', warehouse='GY2')
            
        Returns:
            Result dict
            
        Note:
            중요 필드(sap_no, bl_no, net_weight 등) 수정 시 confirmed=True 필요
        """
        result = {'success': False, 'error': None}
        
        # Allowed fields for update
        allowed_fields = {
            'sap_no', 'bl_no', 'container_no', 'product', 'product_code',
            'warehouse', 'remark', 'condition', 'sold_to', 'sale_ref'
        }
        
        # 중요 필드 (confirmed 필요)
        critical_fields = {'sap_no', 'bl_no', 'net_weight', 'initial_weight'}
        
        # Filter valid fields
        valid_updates = {k: v for k, v in updates.items() if k in allowed_fields}
        
        if not valid_updates:
            result['error'] = "No valid fields to update"
            return result
        
        # 중요 필드 수정 시 confirmed 체크
        updating_critical = any(f in valid_updates for f in critical_fields)
        if updating_critical and not confirmed:
            result['error'] = f"중요 필드 수정은 승인 필요: {critical_fields & set(valid_updates.keys())}"
            result['requires_confirmation'] = True
            return result
        
        try:
            # Build update query
            set_clauses = [f"{k} = ?" for k in valid_updates.keys()]
            values = list(valid_updates.values())
            values.append(lot_no)
            
            query = f"""
                UPDATE inventory 
                SET {', '.join(set_clauses)}, updated_at = CURRENT_TIMESTAMP
                WHERE lot_no = ?
            """
            
            self.db.execute(query, tuple(values))
            result['success'] = True
            logger.info(f"[update_inventory] Updated: {lot_no}")
            
        except (sqlite3.OperationalError, sqlite3.IntegrityError, OSError) as e:
            result['error'] = str(e)
            logger.error(f"[update_inventory] Error: {e}")
        
        return result
    
    # NOTE: get_lot_detail → QueryMixin으로 이관 완료 (v3.8.4 데드코드 정리)
    
    def export_lot_report(self, lot_no: str, filepath: str = None) -> Dict:
        """
        Export LOT detail report to Excel
        
        Args:
            lot_no: LOT number
            filepath: Output file path (optional)
            
        Returns:
            Result dict with filepath
        """
        import os
        from openpyxl import Workbook
        from openpyxl.styles import Font, Border, Side
        
        result = {'success': False, 'filepath': None, 'error': None}
        
        # Get LOT detail (QueryMixin 포맷 호환 — v3.8.4)
        detail = self.get_lot_detail(lot_no)
        if detail is None or detail.get('error'):
            result['error'] = detail.get('error', 'LOT not found') if detail else 'LOT not found'
            return result
        
        # QueryMixin 반환을 CRUDMixin export 포맷으로 래핑
        # QueryMixin: {lot_no, product, ..., tonbags: [...]}
        # export 기대: {success: True, inventory: {...}, tonbags: [...]}
        tonbags = detail.pop('tonbags', [])
        detail_wrapped = {
            'success': True,
            'inventory': detail,
            'tonbags': tonbags if isinstance(tonbags, list) else [],
        }
        detail = detail_wrapped
        
        # Generate filepath if not provided
        if not filepath:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filepath = f"output/reports/LOT_{lot_no}_{timestamp}.xlsx"
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = f"LOT {lot_no}"
            
            # Styles
            header_font = Font(bold=True)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # LOT Info section
            inv = detail['inventory']
            info_rows = [
                ('LOT NO', inv['lot_no']),
                ('SAP NO', inv.get('sap_no', '')),
                ('BL NO', inv.get('bl_no', '')),
                ('Product', inv.get('product', '')),
                ('Warehouse', inv.get('warehouse', '')),
                ('Net Weight', inv.get('net_weight', 0)),
                ('Status', inv.get('status', '')),
            ]
            
            for idx, (label, value) in enumerate(info_rows, 1):
                ws.cell(row=idx, column=1, value=label).font = header_font
                ws.cell(row=idx, column=2, value=value)
            
            # Tonbag section
            start_row = len(info_rows) + 2
            tonbag_headers = ['Sub LT', 'Weight', 'Status', 'Location', 'Inbound', 'Outbound', 'Customer']
            
            for col, h in enumerate(tonbag_headers, 1):
                cell = ws.cell(row=start_row, column=col, value=h)
                cell.font = header_font
                cell.border = thin_border
            
            for row_idx, tb in enumerate(detail['tonbags'], start_row + 1):
                data = [tb['sub_lt'], tb['weight'], tb['status'], tb.get('location', ''),
                        tb.get('inbound_date', ''), tb.get('outbound_date', ''), tb.get('picked_to', '')]
                for col, val in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=val or '')
                    cell.border = thin_border
            
            wb.save(filepath)
            result['success'] = True
            result['filepath'] = filepath
            
        except (OSError, ValueError) as e:
            result['error'] = str(e)
            logger.error(f"[export_lot_report] Error: {e}")
        
        return result
