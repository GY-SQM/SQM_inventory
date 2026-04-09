# -*- coding: utf-8 -*-
"""보고서 API — Detail of Outbound, Sales Order DN, Picking List PDF, Sales Order 업로드."""
import os
import logging
import tempfile
from typing import Optional
from fastapi import APIRouter, HTTPException, Query, UploadFile, File, BackgroundTasks
from fastapi.responses import FileResponse

from react_api.utils.db import get_db, get_engine, now_str

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/reports", tags=["reports"])


# ── 1. Detail of Outbound 미리보기 (JSON) ────────────────────────────────────
@router.get("/outbound-detail/preview")
def outbound_detail_preview(
    sale_ref:      Optional[str] = Query(None),
    outbound_date: Optional[str] = Query(None),
    lot_no:        Optional[str] = Query(None),
):
    """Detail of Outbound 데이터 미리보기 (JSON)."""
    try:
        with get_db() as db:
            conditions = ["s.status IN ('OUTBOUND','SOLD')"]
            params = []
            if sale_ref:
                conditions.append("s.sales_order_no LIKE ?")
                params.append(f"%{sale_ref}%")
            if outbound_date:
                conditions.append("s.delivery_date = ?")
                params.append(outbound_date)
            if lot_no:
                conditions.append("s.lot_no = ?")
                params.append(lot_no)
            where = " AND ".join(conditions)
            rows = db.fetchall(f"""
                SELECT s.lot_no, COALESCE(s.sap_no,'') AS sap_no,
                       COALESCE(s.bl_no,'') AS bl_no,
                       COALESCE(s.customer,'') AS customer,
                       COALESCE(s.sku,'') AS sku,
                       COALESCE(s.sales_order_no,'') AS sales_order_no,
                       COALESCE(s.picking_no,'') AS picking_no,
                       s.delivery_date,
                       COALESCE(s.sold_qty_mt, 0) AS sold_qty_mt,
                       COALESCE(s.gross_weight_kg, 0) AS gross_weight_kg,
                       COALESCE(s.ct_plt, 0) AS ct_plt,
                       COALESCE(s.is_sample, 0) AS is_sample,
                       COALESCE(i.product,'LITHIUM CARBONATE') AS description
                FROM sold_table s
                LEFT JOIN inventory i ON s.lot_no = i.lot_no
                WHERE {where}
                ORDER BY s.lot_no, s.is_sample
            """, tuple(params))
            data = [dict(r) for r in rows]
            total_nw = sum(float(r.get('sold_qty_mt') or 0) for r in data)
            total_gw = sum(float(r.get('gross_weight_kg') or 0) / 1000.0 for r in data)
            total_ct = sum(int(r.get('ct_plt') or 0) for r in data)
            return {
                'success': True, 'total': len(data),
                'total_nw_mt': round(total_nw, 3),
                'total_gw_mt': round(total_gw, 5),
                'total_ct': total_ct,
                'rows': data, 'generated_at': now_str(),
            }
    except Exception as exc:
        logger.error("outbound_detail_preview 실패: %s", exc, exc_info=True)
        raise HTTPException(500, f"미리보기 조회 실패: {exc}")


# ── 2. Detail of Outbound Excel 다운로드 ─────────────────────────────────────
@router.get("/outbound-detail/download")
def outbound_detail_download(
    sale_ref:      Optional[str] = Query(None),
    outbound_date: Optional[str] = Query(None),
    lot_no:        Optional[str] = Query(None),
):
    """Detail of Outbound Excel 생성 → 파일 다운로드."""
    try:
        with get_engine() as engine:
            tmp_dir  = tempfile.mkdtemp(prefix="sqm_rpt_")
            out_path = os.path.join(tmp_dir, "Detail_of_Outbound.xlsx")
            result   = engine._export_outbound_report(
                output_path=out_path,
                sale_ref=sale_ref,
                outbound_date=outbound_date,
                lot_no=lot_no,
            )
        if not result or not os.path.exists(result):
            raise HTTPException(404, "출고 데이터가 없습니다.")
        return FileResponse(
            path=result,
            filename=os.path.basename(result),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("outbound_detail_download 실패: %s", exc, exc_info=True)
        raise HTTPException(500, f"Detail of Outbound 생성 실패: {exc}")


# ── 3. Sales Order DN Excel 다운로드 ─────────────────────────────────────────
@router.get("/sales-order-dn/download")
def sales_order_dn_download(
    sale_ref:      Optional[str] = Query(None),
    outbound_date: Optional[str] = Query(None),
    lot_no:        Optional[str] = Query(None),
):
    """Sales Order DN Excel 생성 → 파일 다운로드."""
    try:
        with get_engine() as engine:
            tmp_dir  = tempfile.mkdtemp(prefix="sqm_dn_")
            out_path = os.path.join(tmp_dir, "Sales_Order_DN.xlsx")
            result   = engine._export_sales_order_dn_report(
                output_path=out_path,
                sale_ref=sale_ref,
                outbound_date=outbound_date,
                lot_no=lot_no,
            )
        if not result:
            raise HTTPException(404, "DN 데이터가 없습니다.")
        if isinstance(result, str) and result.startswith("INCOMPLETE:"):
            parts = result.split(":")[1].split("/")
            raise HTTPException(409, f"Sales Order 미완료 — {parts[0]}/{parts[1]} 출고 완료.")
        if not os.path.exists(result):
            raise HTTPException(404, "파일 생성 실패.")
        return FileResponse(
            path=result,
            filename=os.path.basename(result),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("sales_order_dn_download 실패: %s", exc, exc_info=True)
        raise HTTPException(500, f"Sales Order DN 생성 실패: {exc}")


# ── 4. Picking List PDF 파싱 ─────────────────────────────────────────────────
@router.post("/picking-list/parse")
async def picking_list_parse(file: UploadFile = File(...)):
    """Picking List PDF → 파싱 결과 반환 (DB 저장 없음)."""
    ext = os.path.splitext(file.filename or '')[1].lower()
    if ext != '.pdf':
        raise HTTPException(400, "PDF 파일만 지원합니다.")
    content = await file.read()
    if len(content) > 50 * 1024 * 1024:
        raise HTTPException(400, "파일 크기 50MB 초과.")
    tmp_dir  = tempfile.mkdtemp(prefix="sqm_pick_")
    tmp_path = os.path.join(tmp_dir, os.path.basename(file.filename or "picking.pdf"))
    try:
        with open(tmp_path, 'wb') as f:
            f.write(content)
        try:
            from features.parsers.picking_list_parser import PickingListParser
            result = PickingListParser().parse(tmp_path)
        except ImportError:
            raise HTTPException(500, "PickingListParser 모듈 없음.")
        return {
            'success':         result.get('parse_ok', False),
            'parse_ok':        result.get('parse_ok', False),
            'parse_method':    result.get('parse_method', ''),
            'total_lots':      result.get('total_lots', 0),
            'total_normal_mt': result.get('total_normal_mt', 0),
            'total_sample_kg': result.get('total_sample_kg', 0),
            'items':           result.get('items', []),
            'warnings':        result.get('warnings', []),
            'errors':          result.get('errors', []),
            'generated_at':    now_str(),
        }
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("picking_list_parse 실패: %s", exc, exc_info=True)
        raise HTTPException(500, f"Picking List 파싱 실패: {exc}")
    finally:
        try:
            if os.path.exists(tmp_path): os.remove(tmp_path)
            if os.path.exists(tmp_dir):  os.rmdir(tmp_dir)
        except OSError:
            pass


# ── 5. Sales Order Excel 업로드 + SOLD 처리 ──────────────────────────────────
@router.post("/sales-order/process")
async def sales_order_process(file: UploadFile = File(...)):
    """Sales Order Excel → picking_table 매칭 → SOLD 처리."""
    ext = os.path.splitext(file.filename or '')[1].lower()
    if ext not in ('.xlsx', '.xls'):
        raise HTTPException(400, "Excel 파일만 지원합니다.")
    content  = await file.read()
    tmp_dir  = tempfile.mkdtemp(prefix="sqm_so_")
    tmp_path = os.path.join(tmp_dir, os.path.basename(file.filename or "so.xlsx"))
    try:
        with open(tmp_path, 'wb') as f:
            f.write(content)
        try:
            from features.parsers.sales_order_engine import SalesOrderEngine
        except ImportError:
            raise HTTPException(500, "SalesOrderEngine 모듈 없음.")
        with get_engine() as engine:
            result = SalesOrderEngine(db=engine.db).process(
                file_path=tmp_path,
                sales_order_file=file.filename or '',
            )
        sold    = result.get('sold', 0)
        pending = result.get('pending', 0)
        return {
            'success':          result.get('success', False),
            'sales_order_no':   result.get('sales_order_no'),
            'sold':             sold,
            'pending':          pending,
            'remaining_picked': result.get('remaining_picked', 0),
            'warnings':         result.get('warnings', []),
            'skipped':          result.get('skipped', []),
            'elapsed_ms':       result.get('elapsed_ms', 0),
            'message': (
                f"✅ SOLD {sold}건 완료"
                if result.get('success') else
                f"⚠️ PENDING {pending}건 — picking_table 미매칭"
            ),
            'generated_at': now_str(),
        }
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("sales_order_process 실패: %s", exc, exc_info=True)
        raise HTTPException(500, f"Sales Order 처리 실패: {exc}")
    finally:
        try:
            if os.path.exists(tmp_path): os.remove(tmp_path)
            if os.path.exists(tmp_dir):  os.rmdir(tmp_dir)
        except OSError:
            pass


# ── 6. 거래명세서 Excel 생성 ─────────────────────────────────────────────────
@router.get("/invoice/download")
def invoice_download(
    background_tasks: BackgroundTasks,
    customer:      Optional[str] = Query(None),
    date_from:     Optional[str] = Query(None),
    date_to:       Optional[str] = Query(None),
    sale_ref:      Optional[str] = Query(None),
):
    """고객별 거래명세서 Excel 생성 → 다운로드."""
    try:
        with get_db() as db:
            conditions = ["s.status IN ('OUTBOUND','SOLD')"]
            params = []
            if customer:
                conditions.append("s.customer LIKE ?"); params.append(f"%{customer}%")
            if sale_ref:
                conditions.append("s.sales_order_no LIKE ?"); params.append(f"%{sale_ref}%")
            if date_from:
                conditions.append("s.delivery_date >= ?"); params.append(date_from)
            if date_to:
                conditions.append("s.delivery_date <= ?"); params.append(date_to)
            where = " AND ".join(conditions)
            rows = db.fetchall(f"""
                SELECT s.lot_no, s.sap_no, s.bl_no, s.customer,
                       s.sales_order_no, s.picking_no, s.delivery_date,
                       COALESCE(s.sold_qty_mt, 0) AS sold_qty_mt,
                       COALESCE(s.gross_weight_kg, 0) AS gross_weight_kg,
                       COALESCE(s.ct_plt, 0) AS ct_plt,
                       COALESCE(i.product, 'LITHIUM CARBONATE') AS product,
                       COALESCE(i.sap_no, s.sap_no, '') AS sap_no2
                FROM sold_table s
                LEFT JOIN inventory i ON i.lot_no = s.lot_no
                WHERE {where}
                ORDER BY s.delivery_date, s.lot_no
            """, tuple(params))

        if not rows:
            raise HTTPException(404, "해당 조건의 출고 데이터가 없습니다.")

        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        data = [dict(r) for r in rows]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "거래명세서"

        title_font = Font(bold=True, size=14)
        hdr_font   = Font(bold=True, size=10, color="FFFFFF")
        hdr_fill   = PatternFill("solid", fgColor="1F497D")
        thin       = Side(style="thin", color="000000")
        border     = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws['A1'] = "거래명세서"
        ws['A1'].font = title_font
        ws['A2'] = f"고객: {customer or '전체'}   기간: {date_from or '-'} ~ {date_to or '-'}"
        ws['A2'].font = Font(size=11)

        COLS = [
            ("LOT NO",16),("제품",22),("SAP NO",12),("BL NO",18),
            ("Sales Order",16),("Picking No",14),("출고일",12),
            ("NW(MT)",10),("GW(MT)",12),("CT/PLT",8),
        ]
        for ci, (h, w) in enumerate(COLS, 1):
            cell = ws.cell(row=4, column=ci, value=h)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            ws.column_dimensions[get_column_letter(ci)].width = w

        for ri, r in enumerate(data, 5):
            vals = [
                r.get('lot_no',''), r.get('product',''), r.get('sap_no',''),
                r.get('bl_no',''), r.get('sales_order_no',''), r.get('picking_no',''),
                str(r.get('delivery_date',''))[:10],
                float(r.get('sold_qty_mt') or 0),
                float(r.get('gross_weight_kg') or 0) / 1000.0,
                int(r.get('ct_plt') or 0),
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.border = border
                if ci == 8: cell.number_format = '#,##0.000'
                if ci == 9: cell.number_format = '#,##0.00000'

        total_nw = sum(float(r.get('sold_qty_mt') or 0) for r in data)
        total_gw = sum(float(r.get('gross_weight_kg') or 0) / 1000.0 for r in data)
        tr = len(data) + 5
        ws.cell(row=tr, column=7, value="합계").font = Font(bold=True)
        tc = ws.cell(row=tr, column=8, value=total_nw)
        tc.font = Font(bold=True); tc.number_format = '#,##0.000'
        tc = ws.cell(row=tr, column=9, value=total_gw)
        tc.font = Font(bold=True); tc.number_format = '#,##0.00000'

        fd, out_path = tempfile.mkstemp(suffix=".xlsx", prefix="sqm_invoice_")
        os.close(fd)
        wb.save(out_path)
        cust_str = (customer or "ALL").replace("/","_").replace(" ","_")
        background_tasks.add_task(lambda p=out_path: __import__('os').remove(p) if __import__('os').path.exists(p) else None)
        return FileResponse(path=out_path, filename=f"거래명세서_{cust_str}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("invoice_download 실패: %s", exc, exc_info=True)
        raise HTTPException(500, f"거래명세서 생성 실패: {exc}")


# ── 7. 입출고 내역 PDF ────────────────────────────────────────────────────────
@router.get("/transaction-pdf/download")
def transaction_pdf_download(
    background_tasks: BackgroundTasks,
    customer:  Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to:   Optional[str] = Query(None),
):
    """입출고 내역 PDF 생성 → 다운로드."""
    try:
        from gui_app_modular.utils.pdf_report_gen import generate_transaction_statement
        fd, out_path = tempfile.mkstemp(suffix=".pdf", prefix="sqm_trans_")
        os.close(fd)
        with get_engine() as engine:
            result = generate_transaction_statement(
                engine,
                customer=customer or '',
                date_from=date_from or '',
                date_to=date_to or '',
                output_path=out_path,
            )
        path = result if isinstance(result, str) and os.path.exists(result) else out_path
        if not os.path.exists(path):
            raise HTTPException(404, "PDF 생성 실패 — reportlab 미설치 또는 데이터 없음")
        background_tasks.add_task(lambda p=path: __import__('os').remove(p) if __import__('os').path.exists(p) else None)
        return FileResponse(path=path, filename="입출고내역.pdf", media_type="application/pdf")
    except HTTPException:
        raise
    except ImportError as exc:
        raise HTTPException(501, f"PDF 생성 모듈 없음: {exc}")
    except Exception as exc:
        logger.error("transaction_pdf_download 실패: %s", exc, exc_info=True)
        raise HTTPException(500, f"입출고 내역 PDF 실패: {exc}")


# ── 8. LOT 상세 PDF ───────────────────────────────────────────────────────────
@router.get("/lot-detail-pdf/download")
def lot_detail_pdf_download(background_tasks: BackgroundTasks, lot_no: str = Query(...)):
    """특정 LOT 상세 이력 PDF 생성 → 다운로드."""
    try:
        from gui_app_modular.utils.pdf_report_gen import generate_lot_outbound_history_pdf
        with get_db() as db:
            history = db.fetchall("""
                SELECT movement_type AS type, description,
                       COALESCE(qty_kg,0) AS qty_kg, created_at
                FROM stock_movement
                WHERE lot_no = ? ORDER BY created_at
            """, (lot_no,))
            history = [dict(r) for r in (history or [])]
        fd, out_path = tempfile.mkstemp(suffix=".pdf", prefix="sqm_lot_")
        os.close(fd)
        with get_engine() as engine:
            result = generate_lot_outbound_history_pdf(
                engine=engine, lot_no=lot_no,
                history=history, output_path=out_path,
            )
        path = result if isinstance(result, str) and os.path.exists(result) else out_path
        if not os.path.exists(path):
            raise HTTPException(404, "PDF 생성 실패")
        background_tasks.add_task(lambda p=path: __import__('os').remove(p) if __import__('os').path.exists(p) else None)
        return FileResponse(path=path, filename=f"LOT_{lot_no}_상세.pdf", media_type="application/pdf")
    except HTTPException:
        raise
    except ImportError as exc:
        raise HTTPException(501, f"PDF 생성 모듈 없음: {exc}")
    except Exception as exc:
        logger.error("lot_detail_pdf 실패: %s", exc, exc_info=True)
        raise HTTPException(500, f"LOT 상세 PDF 실패: {exc}")


# ── 9. 일일 재고 현황 PDF ─────────────────────────────────────────────────────
@router.post("/daily-pdf/download")
def daily_pdf_download(background_tasks: BackgroundTasks):
    """일일 재고 현황 PDF → 다운로드."""
    try:
        from gui_app_modular.utils.pdf_report_gen import generate_daily_inventory_report
        fd, out_path = tempfile.mkstemp(suffix=".pdf", prefix="sqm_daily_")
        os.close(fd); os.makedirs(os.path.dirname(out_path), exist_ok=True)
        with get_engine() as engine:
            result = generate_daily_inventory_report(engine, output_dir=os.path.dirname(out_path))
        path = result if result and os.path.exists(result) else out_path
        if not os.path.exists(path):
            raise HTTPException(404, "PDF 생성 실패 — reportlab 미설치 또는 데이터 없음")
        from datetime import datetime
        fname = f"SQM_재고현황_{datetime.now().strftime('%Y%m%d')}.pdf"
        background_tasks.add_task(lambda p=path: __import__('os').remove(p) if __import__('os').path.exists(p) else None)
        return FileResponse(path=path, filename=fname, media_type="application/pdf")
    except HTTPException:
        raise
    except ImportError as exc:
        raise HTTPException(501, f"PDF 생성 모듈 없음: {exc}")
    except Exception as exc:
        logger.error("daily_pdf 실패: %s", exc, exc_info=True)
        raise HTTPException(500, f"일일 PDF 실패: {exc}")


# ── 10. 월간 실적 PDF ─────────────────────────────────────────────────────────
@router.post("/monthly-pdf/download")
def monthly_pdf_download(background_tasks: BackgroundTasks, payload: dict = None):
    """월간 실적 PDF → 다운로드.  payload: { year, month }"""
    payload = payload or {}
    year  = int(payload.get('year')  or 0) or None
    month = int(payload.get('month') or 0) or None
    try:
        from gui_app_modular.utils.pdf_report_gen import generate_monthly_report
        fd, out_path = tempfile.mkstemp(suffix=".pdf", prefix="sqm_monthly_")
        os.close(fd)
        with get_engine() as engine:
            result = generate_monthly_report(engine, year=year, month=month,
                                             output_path=out_path)
        path = result if result and os.path.exists(result) else out_path
        if not os.path.exists(path):
            raise HTTPException(404, "PDF 생성 실패 — reportlab 미설치 또는 데이터 없음")
        from datetime import datetime
        now = datetime.now()
        fname = f"SQM_월간실적_{year or now.year}{month or now.month:02d}.pdf"
        return FileResponse(path=path, filename=fname, media_type="application/pdf")
    except HTTPException:
        raise
    except ImportError as exc:
        raise HTTPException(501, f"PDF 생성 모듈 없음: {exc}")
    except Exception as exc:
        logger.error("monthly_pdf 실패: %s", exc, exc_info=True)
        raise HTTPException(500, f"월간 PDF 실패: {exc}")


# ── 11. PDF → Excel 변환 ──────────────────────────────────────────────────────
@router.post("/pdf-to-excel")
async def pdf_to_excel(file: UploadFile = File(...)):
    """PDF → Excel 변환 (pdfplumber 기반)."""
    ext = os.path.splitext(file.filename or '')[1].lower()
    if ext != '.pdf':
        raise HTTPException(400, "PDF 파일만 지원합니다.")
    content  = await file.read()
    tmp_dir  = tempfile.mkdtemp(prefix="sqm_conv_")
    tmp_path = os.path.join(tmp_dir, os.path.basename(file.filename or "input.pdf"))
    try:
        with open(tmp_path, 'wb') as f:
            f.write(content)
        try:
            import pdfplumber, openpyxl
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            with pdfplumber.open(tmp_path) as pdf:
                for pi, page in enumerate(pdf.pages, 1):
                    ws = wb.create_sheet(title=f"Page{pi}")
                    tables = page.extract_tables()
                    if tables:
                        for table in tables:
                            for row in table:
                                ws.append([c or '' for c in row])
                    else:
                        text = page.extract_text() or ''
                        for line in text.splitlines():
                            ws.append([line])
            fd, out_path = tempfile.mkstemp(suffix=".xlsx", prefix="sqm_excel_")
            os.close(fd)
            wb.save(out_path)
            base = os.path.splitext(file.filename or 'converted')[0]
            return FileResponse(path=out_path, filename=f"{base}.xlsx",
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except ImportError as exc:
            raise HTTPException(501, f"pdfplumber 또는 openpyxl 미설치: {exc}")
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("pdf_to_excel 실패: %s", exc, exc_info=True)
        raise HTTPException(500, f"PDF→Excel 변환 실패: {exc}")
    finally:
        try:
            if os.path.exists(tmp_path): os.remove(tmp_path)
            if os.path.exists(tmp_dir):  os.rmdir(tmp_dir)
        except OSError:
            pass
