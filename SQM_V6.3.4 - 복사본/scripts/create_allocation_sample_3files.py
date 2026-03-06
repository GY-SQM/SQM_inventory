"""
Allocation 형식 샘플 파일 3개 생성 (DB 불필요)

이미지 양식: Product, SAP NO, ETA BUSAN, Date in stock, QTY (MT), Lot No, WH, Customs, GW, SALE REF
- 1행: 타이틀 (Allocation - PT LBM - September / CN Semarang - xxx MT of MIC9000)
- 2행: 빈 행 (E2에 SUM 수식)
- 3행: 헤더
- 4행~: 데이터

사용: python scripts/create_allocation_sample_3files.py
출력: generated_allocation/Allocation_샘플_1.xlsx, _2.xlsx, _3.xlsx
"""

import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
OUTPUT_DIR = PROJECT_ROOT / "generated_allocation"

# 샘플 데이터: (행 수, 총 MT 설명) × 3
SAMPLE_SPECS = [
    (20, 100.0),   # 샘플 1: 20행 × 5MT = 100MT
    (15, 75.0),    # 샘플 2: 15행 × 5MT = 75MT
    (10, 50.0),    # 샘플 3: 10행 × 5MT = 50MT
]

# 가상 LOT 번호 (10자리)
LOT_BASE = [
    "1126011509", "1126010151", "1126010636", "1126010642", "1126010706", "1126011511",
    "1126011209", "1126010379", "1126011136", "1126010801", "1126010902", "1126011003",
    "1126011104", "1126011305", "1126011406", "1126011607", "1126011708", "1126011809",
    "1126011910", "1126012011",
]
# SAP NO (8자리)
SAP_BASE = ["2200034276", "2200034275", "2200034274", "2200034273", "2200034272"]
# Date in stock (YYYY-MM-DD)
DATES = ["2026-02-20", "2026-03-24", "2026-03-10", "2026-02-15", "2026-04-01"]


def _create_one_excel(filepath: Path, num_rows: int, total_mt: float) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Allocation Table"

    title_font = Font(bold=True, size=14, color="2C3E50")
    header_font = Font(bold=True, color="2C3E50", size=10)
    data_font = Font(size=10)
    header_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    header_yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    data_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    data_grey = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    # Row 1: 타이틀
    ws.merge_cells('A1:J1')
    ws['A1'] = f"Allocation - PT LBM - September / CN Semarang - {total_mt:.1f}MT of MIC9000"
    ws['A1'].font = title_font
    ws.row_dimensions[1].height = 30

    # Row 2: 빈 행 (E2에 SUM)
    ws.row_dimensions[2].height = 6

    # Row 3: 헤더
    headers = [
        ('Product', 16), ('SAP NO', 14), ('ETA BUSAN', 14), ('Date in stock', 14),
        ('QTY (MT)', 12), ('Lot No', 14), ('WH', 8), ('Customs', 12),
        ('GW', 12), ('SALE REF', 12),
    ]
    yellow_cols = (3, 8, 9)
    for col, (text, width) in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=text)
        cell.font = header_font
        cell.fill = header_yellow if col in yellow_cols else header_fill
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width

    # Row 4~: 데이터 (5.000 MT, 5.130 GW, 광양, Cleared, SALE REF 1955)
    qty_mt = 5.000
    gw = 5.130
    for i in range(num_rows):
        row_num = 4 + i
        lot_no = LOT_BASE[i % len(LOT_BASE)]
        sap_no = SAP_BASE[i % len(SAP_BASE)]
        date_stock = DATES[i % len(DATES)]
        vals = [
            "LITHIUM CARBONATE",
            sap_no,
            "",  # ETA BUSAN 비움
            date_stock,
            qty_mt,
            lot_no,
            "광양",
            "Cleared",
            gw,
            "1955",
        ]
        fill = data_grey if i % 2 == 1 else data_white
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = data_font
            cell.fill = fill
            cell.border = thin_border
            if col in (5, 9):
                cell.number_format = '#,##0.000'
                cell.alignment = right_align
            else:
                cell.alignment = center

    last_row = 4 + num_rows - 1
    ws['E2'].value = f"=SUM(E4:E{last_row})"
    ws['E2'].number_format = '#,##0.0000'
    ws['E2'].font = Font(bold=True)

    wb.save(filepath)


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    for i, (num_rows, total_mt) in enumerate(SAMPLE_SPECS, 1):
        fpath = OUTPUT_DIR / f"Allocation_샘플_{i}.xlsx"
        _create_one_excel(fpath, num_rows, total_mt)
        print(f"생성: {fpath} ({num_rows}행, {total_mt} MT)")
    print(f"총 3개 파일 생성 완료: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
    sys.exit(0)
