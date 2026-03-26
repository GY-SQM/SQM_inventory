# -*- coding: utf-8 -*-
"""
SQM v7.9.2 — Allocation Excel 자동 생성 스크립트
================================================
generate_allocation_from_tonbag.py

기능:
    DB의 AVAILABLE LOT을 조회하여 N개 파일로 균등 분배한 Allocation Excel을 생성합니다.

개선 사항 (v7.7.1):
    - 이미 RESERVED/STAGED 상태인 LOT 자동 제외 (파일 간 중복 원천 방지)
    - 파일 수(N) 자동 계산 또는 수동 지정
    - 샘플 행(0.001MT) 각 LOT마다 자동 포함
    - 파일간 LOT 중복 Zero 보장

사용법 (CLI):
    python generate_allocation_from_tonbag.py
    python generate_allocation_from_tonbag.py --files 3 --customer "PT LBM" --sale-ref 1955

작성자: Ruby (남기동)
버전: v7.9.2
"""

import argparse
import logging
import os
import sqlite3
import sys
from datetime import date
from pathlib import Path

logger = logging.getLogger(__name__)


# ── 경로 설정 ────────────────────────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parent.parent
DB_CANDIDATES = [
    PROJECT_ROOT / "sqm_inventory.db",
    PROJECT_ROOT / "data" / "sqm_inventory.db",
    PROJECT_ROOT / "db" / "sqm_inventory.db",
]
OUTPUT_DIR = PROJECT_ROOT / "generated_allocation"


def _find_db() -> Path:
    for p in DB_CANDIDATES:
        if p.exists():
            return p
    # 재귀 탐색 (최대 2단계)
    for p in PROJECT_ROOT.glob("**/*.db"):
        if "sqm" in p.name.lower() or "inventory" in p.name.lower():
            return p
    raise FileNotFoundError(
        f"SQM DB 파일을 찾을 수 없습니다. 확인된 경로: {[str(p) for p in DB_CANDIDATES]}"
    )


# ── DB 조회 ──────────────────────────────────────────────────────────────────
def fetch_available_lots(db_path: Path, sale_ref: str = None) -> list:
    """
    AVAILABLE 상태 LOT 조회.
    sale_ref가 주어지면 해당 sale_ref로 이미 RESERVED/STAGED된 LOT 제외.
    """
    con = sqlite3.connect(str(db_path))
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    # 이미 예약된 LOT 목록
    reserved_lots = set()
    if sale_ref:
        cur.execute(
            "SELECT DISTINCT lot_no FROM allocation_plan "
            "WHERE sale_ref=? AND status IN ('RESERVED','STAGED','PENDING_APPROVAL')",
            (str(sale_ref).strip(),)
        )
        reserved_lots = {str(r[0]).strip() for r in cur.fetchall()}
        logger.info(f"[GEN] sale_ref={sale_ref} 기존 예약 LOT {len(reserved_lots)}개 제외 예정")

    # AVAILABLE LOT 조회 (샘플 없는 일반 톤백 기준)
    cur.execute("""
        SELECT
            i.lot_no,
            i.sap_no,
            i.product,
            i.product_code,
            i.net_weight,
            i.gross_weight,
            i.mxbg_pallet,
            i.warehouse,
            i.arrival_date,
            i.inbound_date,
            i.con_return
        FROM inventory i
        WHERE i.status = 'AVAILABLE'
          AND EXISTS (
              SELECT 1 FROM inventory_tonbag t
              WHERE t.lot_no = i.lot_no
                AND t.status = 'AVAILABLE'
                AND COALESCE(t.is_sample, 0) = 0
          )
        ORDER BY COALESCE(i.arrival_date, i.inbound_date, i.created_at) ASC, i.lot_no
    """)
    rows = [dict(r) for r in cur.fetchall()]
    con.close()

    # 이미 예약된 LOT 제외
    if reserved_lots:
        before = len(rows)
        rows = [r for r in rows if str(r['lot_no']).strip() not in reserved_lots]
        logger.info(f"[GEN] 예약 제외 후: {before}개 → {len(rows)}개 LOT")

    return rows


# ── Excel 생성 ────────────────────────────────────────────────────────────────
def _make_xlsx(file_no: int, lots: list, customer: str,
               sale_ref: str, out_dir: Path, total_files: int) -> Path:
    """LOT 목록으로 Allocation Excel 1개 생성."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl 미설치: pip install openpyxl")
        sys.exit(1)

    total_mt = sum(5.0 for _ in lots)  # LOT당 5MT 기준
    title = (
        f"Allocation - GY - {customer} "
        f"({total_files}분할 {file_no}/{total_files}) — {total_mt:.2f} MT"
    )

    TITLE_FILL = PatternFill('solid', start_color='1F3864')
    HDR_FILL   = PatternFill('solid', start_color='2E75B6')
    EVEN_FILL  = PatternFill('solid', start_color='DCE6F1')
    ODD_FILL   = PatternFill('solid', start_color='FFFFFF')
    SAMP_FILL  = PatternFill('solid', start_color='FFF2CC')  # 샘플행 노란색
    TITLE_FONT = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    HDR_FONT   = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    DATA_FONT  = Font(name='Arial', size=10)
    SAMP_FONT  = Font(name='Arial', size=10, italic=True, color='7F6000')
    CENTER = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='B8CCE4')
    BDR  = Border(left=thin, right=thin, top=thin, bottom=thin)

    HEADERS = ['Product', 'SAP NO', 'ETA BUSAN', 'Date in stock',
               'QTY (MT)', 'Lot No', 'WH', 'Customs', 'GW', 'SALE REF']
    COL_W   = [22, 14, 14, 14, 10, 14, 8, 10, 8, 10]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Allocation'

    # 행1: 타이틀
    ws.merge_cells('A1:J1')
    c = ws['A1']
    c.value = title
    c.font  = TITLE_FONT
    c.fill  = TITLE_FILL
    c.alignment = CENTER
    ws.row_dimensions[1].height = 22

    # 행2: 합계 공식 (본품+샘플 합산)
    n_total_rows = len(lots) * 2  # 본품 + 샘플 각 1행
    ws.merge_cells('A2:J2')
    ws['A2'].value = f'=SUM(E4:E{3 + n_total_rows})'
    ws['A2'].font  = Font(name='Arial', bold=True, size=10)
    ws['A2'].alignment = CENTER
    ws.row_dimensions[2].height = 18

    # 행3: 헤더
    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(3, ci)
        c.value     = h
        c.font      = HDR_FONT
        c.fill      = HDR_FILL
        c.alignment = CENTER
        c.border    = BDR
    ws.row_dimensions[3].height = 20

    today_str = date.today().strftime('%Y-%m-%d')

    # 행4~: 본품 + 샘플 교대 삽입
    excel_row = 4
    for idx, lot in enumerate(lots):
        lot_no   = str(lot['lot_no']).split('.')[0]
        sap_no   = str(lot.get('sap_no') or '').split('.')[0]
        product  = str(lot.get('product') or 'LITHIUM CARBONATE')
        gw       = float(lot.get('gross_weight') or 0)
        gw_mt    = round(gw / 1000, 3) if gw > 0 else 5.13
        wh       = str(lot.get('warehouse') or '광양')
        eta      = str(lot.get('arrival_date') or lot.get('inbound_date') or '')
        fill     = EVEN_FILL if idx % 2 == 0 else ODD_FILL

        # 본품 행 (5.0 MT)
        main_row = [product, sap_no, eta, today_str,
                    5.0, lot_no, wh, '', gw_mt, sale_ref or '']
        for ci, val in enumerate(main_row, 1):
            c = ws.cell(excel_row, ci)
            c.value     = val
            c.font      = DATA_FONT
            c.fill      = fill
            c.alignment = CENTER
            c.border    = BDR
        ws.row_dimensions[excel_row].height = 18
        excel_row += 1

        # 샘플 행 (0.001 MT = 1 kg)
        samp_row = [product, sap_no, eta, today_str,
                    0.001, lot_no, wh, '', 0.001, sale_ref or '']
        for ci, val in enumerate(samp_row, 1):
            c = ws.cell(excel_row, ci)
            c.value     = val
            c.font      = SAMP_FONT
            c.fill      = SAMP_FILL
            c.alignment = CENTER
            c.border    = BDR
        ws.row_dimensions[excel_row].height = 16
        excel_row += 1

    # 열 너비
    for ci, w in enumerate(COL_W, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # 파일 저장
    fname = f"Allocation_GY_{customer.replace(' ', '_')}_{file_no}of{total_files}.xlsx"
    out_path = out_dir / fname
    wb.save(str(out_path))
    return out_path


# ── 메인 ────────────────────────────────────────────────────────────────────
def generate(
    n_files: int = 3,
    customer: str = "PT LBM",
    sale_ref: str = "",
    lots_per_file: int = None,
    db_path: Path = None,
    out_dir: Path = None,
) -> list:
    """
    Allocation Excel N개 생성.

    Args:
        n_files:       분할 파일 수 (기본 3)
        customer:      고객사명 (기본 PT LBM)
        sale_ref:      판매참조번호 — 이미 예약된 LOT 제외 기준
        lots_per_file: 파일당 LOT 수 (지정 시 n_files 자동 계산)
        db_path:       DB 경로 (None이면 자동 탐색)
        out_dir:       출력 디렉터리 (None이면 generated_allocation/)

    Returns:
        생성된 파일 경로 리스트
    """
    logging.basicConfig(level=logging.INFO,
                        format='%(asctime)s [%(levelname)s] %(message)s')

    # DB 경로 결정
    if db_path is None:
        db_path = _find_db()
    logger.info(f"[GEN] DB: {db_path}")

    # 출력 디렉터리
    if out_dir is None:
        out_dir = OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)

    # AVAILABLE LOT 조회 (이미 예약된 것 제외)
    lots = fetch_available_lots(db_path, sale_ref=sale_ref)
    if not lots:
        logger.warning("[GEN] AVAILABLE LOT이 없습니다.")
        return []
    logger.info(f"[GEN] 가용 LOT: {len(lots)}개")

    # 파일 수 결정
    if lots_per_file and lots_per_file > 0:
        import math
        n_files = math.ceil(len(lots) / lots_per_file)
    n_files = max(1, n_files)

    # LOT 균등 분배 (파일 간 중복 없음)
    chunks = [[] for _ in range(n_files)]
    for i, lot in enumerate(lots):
        chunks[i % n_files].append(lot)

    # Excel 생성
    generated = []
    for fi, chunk in enumerate(chunks, 1):
        if not chunk:
            continue
        out_path = _make_xlsx(fi, chunk, customer, sale_ref, out_dir, n_files)
        generated.append(out_path)
        logger.info(f"[GEN] 파일{fi}: {out_path.name} ({len(chunk)}개 LOT)")

    # 검증: 전체 LOT 중복 없음 확인
    all_lots_out = []
    for chunk in chunks:
        all_lots_out.extend(str(r['lot_no']).split('.')[0] for r in chunk)
    from collections import Counter
    dup_check = {l: c for l, c in Counter(all_lots_out).items() if c > 1}
    if dup_check:
        logger.error(f"[GEN] ❌ 중복 LOT 감지: {dup_check}")
    else:
        logger.info(f"[GEN] ✅ 파일 간 LOT 중복 없음 ({len(all_lots_out)}개)")

    logger.info(f"[GEN] 완료: {len(generated)}개 파일 → {out_dir}")
    return generated


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="SQM Allocation Excel 자동 생성 v7.9.2")
    parser.add_argument("--files",    type=int,   default=3,        help="분할 파일 수 (기본 3)")
    parser.add_argument("--customer", type=str,   default="PT LBM", help="고객사명")
    parser.add_argument("--sale-ref", type=str,   default="",       help="SALE REF (기존 예약 제외용)")
    parser.add_argument("--lots-per-file", type=int, default=None,  help="파일당 LOT 수 (지정 시 파일수 자동)")
    parser.add_argument("--db",       type=str,   default=None,     help="DB 파일 경로")
    parser.add_argument("--out",      type=str,   default=None,     help="출력 디렉터리")
    args = parser.parse_args()

    db   = Path(args.db)  if args.db  else None
    out  = Path(args.out) if args.out else None

    files = generate(
        n_files       = args.files,
        customer      = args.customer,
        sale_ref      = args.sale_ref,
        lots_per_file = args.lots_per_file,
        db_path       = db,
        out_dir       = out,
    )
    for f in files:
        print(f"생성: {f}")
