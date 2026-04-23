"""
SQM v864.3 — Menubar API (62 엔드포인트)
자동 생성: Ruby, Stage 2 BACKEND, 2026-04-21
기능 수: 62
"""
from fastapi import APIRouter, HTTPException
from backend.common.errors import wrap_engine_call, NotReadyError, ok_response

router = APIRouter(prefix="/api/menu", tags=["menubar"])

# ── F001 | menubar | 파일 > 입고 | 📄  PDF 스캔 입고 ──
# tkinter_callback: _on_pdf_inbound
# source: gui_app_modular/handlers/inbound_processor.py
@router.post("/-on-pdf-inbound", summary="📄  PDF 스캔 입고")
async def ononpdfinbound(payload: dict | None = None):
    """Feature F001: 📄  PDF 스캔 입고"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F001 📄  PDF 스캔 입고 — GUI 재구현 필요 (Phase 4-B)")
# ── F002 | menubar | 파일 > 입고 | 📊  엑셀 파일 수동 입고 ──
# tkinter_callback: _bulk_import_inventory_simple
# source: gui_app_modular/handlers/import_handlers.py
@router.post("/-bulk-import-inventory-simple", summary="📊  엑셀 파일 수동 입고")
async def onbulkimportinventorysimple(payload: dict | None = None):
    """Feature F002: 📊  엑셀 파일 수동 입고"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F002 📊  엑셀 파일 수동 입고 — GUI 재구현 필요 (Phase 4-B)")
# ── F003 | menubar | 파일 > 입고 | 📋  D/O 후속 연결 ──
# tkinter_callback: _on_do_update
# source: gui_app_modular/handlers/inbound_processor.py
@router.post("/-on-do-update", summary="📋  D/O 후속 연결")
async def onondoupdate(payload: dict | None = None):
    """Feature F003: 📋  D/O 후속 연결"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F003 📋  D/O 후속 연결 — GUI 재구현 필요 (Phase 4-B)")
# ── F004 | menubar | 파일 > 입고 | 📍  톤백 위치 매핑 ──
# tkinter_callback: _on_tonbag_location_upload
# source: unknown
@router.post("/-on-tonbag-location-upload", summary="📍  톤백 위치 매핑")
async def onontonbaglocationupload(payload: dict | None = None):
    """Feature F004: 📍  톤백 위치 매핑"""
    raise NotReadyError("F004 📍  톤백 위치 매핑")

# ── F005 | menubar | 파일 > 입고 | ✅  대량 이동 승인 ──
# tkinter_callback: _on_move_approval_queue
# source: gui_app_modular/mixins/advanced_dialogs_mixin.py
@router.post("/-on-move-approval-queue", summary="✅  대량 이동 승인")
async def ononmoveapprovalqueue(payload: dict | None = None):
    """Feature F005: ✅  대량 이동 승인"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F005 ✅  대량 이동 승인 — GUI 재구현 필요 (Phase 4-B)")
# ── F006 | menubar | 파일 > 입고 | 🔄  반품 (재입고) ──
# tkinter_callback: _show_return_dialog
# source: gui_app_modular/mixins/advanced_dialogs_mixin.py
@router.post("/-show-return-dialog", summary="🔄  반품 (재입고)")
async def onshowreturndialog(payload: dict | None = None):
    """Feature F006: 🔄  반품 (재입고)"""
    try:
        from gui_app_modular.mixins.advanced_dialogs_mixin import _show_return_dialog  # type: ignore
    except ImportError:
        raise NotReadyError("F006 🔄  반품 (재입고)")
    # v864.3 방안 A: GUI 재구현 필요 → 명시적 NOT_READY
    raise NotReadyError("F006 🔄  반품 (재입고) — GUI 재구현 필요 (Phase 4-B)")

# ── F007 | menubar | 파일 > 입고 | 📂  반품 입고 (Excel) ──
# tkinter_callback: _on_return_inbound_upload
# source: gui_app_modular/mixins/advanced_dialogs_mixin.py
@router.post("/-on-return-inbound-upload", summary="📂  반품 입고 (Excel)")
async def ononreturninboundupload(payload: dict | None = None):
    """Feature F007: 📂  반품 입고 (Excel)"""
    try:
        from gui_app_modular.mixins.advanced_dialogs_mixin import _on_return_inbound_upload  # type: ignore
    except ImportError:
        raise NotReadyError("F007 📂  반품 입고 (Excel)")
    # v864.3 방안 A: GUI 재구현 필요 → 명시적 NOT_READY
    raise NotReadyError("F007 📂  반품 입고 (Excel) — GUI 재구현 필요 (Phase 4-B)")

# ── F008 | menubar | 파일 > 입고 | 📊  반품 사유 통계 ──
# tkinter_callback: _show_return_statistics
# source: gui_app_modular/mixins/advanced_dialogs_mixin.py
@router.post("/-show-return-statistics", summary="📊  반품 사유 통계")
async def onshowreturnstatistics(payload: dict | None = None):
    """Feature F008: 📊  반품 사유 통계"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F008 📊  반품 사유 통계 — GUI 재구현 필요 (Phase 4-B)")
# ── F009 | menubar | 파일 > 입고 | 📋  입고 현황 조회 ──
# tkinter_callback: _bulk_import_inventory
# source: gui_app_modular/mixins/bulk_import_mixin.py
@router.post("/-bulk-import-inventory", summary="📋  입고 현황 조회")
async def onbulkimportinventory(payload: dict | None = None):
    """Feature F009: 📋  입고 현황 조회"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F009 📋  입고 현황 조회 — GUI 재구현 필요 (Phase 4-B)")
# ── F010 | menubar | 파일 > 입고 | 📝  입고 파싱 템플릿 관리 ──
# tkinter_callback: _on_inbound_template_manage
# source: gui_app_modular/mixins/advanced_dialogs_mixin.py
@router.post("/-on-inbound-template-manage", summary="📝  입고 파싱 템플릿 관리")
async def ononinboundtemplatemanage(payload: dict | None = None):
    """Feature F010: 📝  입고 파싱 템플릿 관리"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F010 📝  입고 파싱 템플릿 관리 — GUI 재구현 필요 (Phase 4-B)")
# ── F011 | menubar | 파일 > 입고 | 📦  제품 마스터 관리 ──
# tkinter_callback: _show_product_master
# source: unknown
@router.post("/-show-product-master", summary="📦  제품 마스터 관리")
async def onshowproductmaster(payload: dict | None = None):
    """Feature F011: 📦  제품 마스터 관리"""
    raise NotReadyError("F011 📦  제품 마스터 관리")

# ── F012 | menubar | 파일 > 입고 | ⚙️  이메일 설정 ──
# tkinter_callback: _show_email_config
# source: gui_app_modular/mixins/advanced_dialogs_mixin.py
@router.post("/-show-email-config", summary="⚙️  이메일 설정")
async def onshowemailconfig(payload: dict | None = None):
    """Feature F012: ⚙️  이메일 설정"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F012 ⚙️  이메일 설정 — GUI 재구현 필요 (Phase 4-B)")
# ── F013 | menubar | 파일 > 입고 | 🔍  정합성 검증 (시각화) ──
# tkinter_callback: _on_integrity_report_v760
# source: gui_app_modular/mixins/advanced_dialogs_mixin.py
@router.post("/-on-integrity-report-v760", summary="🔍  정합성 검증 (시각화)")
async def ononintegrityreportv760(payload: dict | None = None):
    """Feature F013: 🔍  정합성 검증 (시각화)"""
    try:
        from gui_app_modular.mixins.advanced_dialogs_mixin import _on_integrity_report_v760  # type: ignore
    except ImportError:
        raise NotReadyError("F013 🔍  정합성 검증 (시각화)")
    # v864.3 방안 A: GUI 재구현 필요 → 명시적 NOT_READY
    raise NotReadyError("F013 🔍  정합성 검증 (시각화) — GUI 재구현 필요 (Phase 4-B)")

# ── F014 | menubar | 파일 > 입고 | 🛠️  LOT 상태 정합성 복구 ──
# tkinter_callback: _on_fix_lot_status_integrity
# source: gui_app_modular/mixins/toolbar_mixin.py
@router.post("/-on-fix-lot-status-integrity", summary="🛠️  LOT 상태 정합성 복구")
async def ononfixlotstatusintegrity(payload: dict | None = None):
    """Feature F014: 🛠️  LOT 상태 정합성 복구"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F014 🛠️  LOT 상태 정합성 복구 — GUI 재구현 필요 (Phase 4-B)")
# ── F015 | menubar | 파일 > 출고 | 🚀  즉시 출고 (원스톱) ──
# tkinter_callback: _on_s1_onestop_outbound
# source: gui_app_modular/handlers/outbound_handlers.py
@router.post("/-on-s1-onestop-outbound", summary="🚀  즉시 출고 (원스톱)")
async def onons1onestopoutbound(payload: dict | None = None):
    """Feature F015: 🚀  즉시 출고 (원스톱)"""
    try:
        from gui_app_modular.handlers.outbound_handlers import _on_s1_onestop_outbound  # type: ignore
    except ImportError:
        raise NotReadyError("F015 🚀  즉시 출고 (원스톱)")
    # v864.3 방안 A: GUI 재구현 필요 → 명시적 NOT_READY
    raise NotReadyError("F015 🚀  즉시 출고 (원스톱) — GUI 재구현 필요 (Phase 4-B)")

# ── F016 | menubar | 파일 > 출고 | 📤  빠른 출고 (붙여넣기) ──
# tkinter_callback: _on_quick_outbound_paste
# source: gui_app_modular/handlers/outbound_handlers.py
@router.post("/-on-quick-outbound-paste", summary="📤  빠른 출고 (붙여넣기)")
async def ononquickoutboundpaste(payload: dict | None = None):
    """Feature F016: 📤  빠른 출고 (붙여넣기)"""
    try:
        from gui_app_modular.handlers.outbound_handlers import _on_quick_outbound_paste  # type: ignore
    except ImportError:
        raise NotReadyError("F016 📤  빠른 출고 (붙여넣기)")
    # v864.3 방안 A: GUI 재구현 필요 → 명시적 NOT_READY
    raise NotReadyError("F016 📤  빠른 출고 (붙여넣기) — GUI 재구현 필요 (Phase 4-B)")

# ── F017 | menubar | 파일 > 출고 | 📋  Picking List 업로드 (PDF) ──
# tkinter_callback: _on_picking_list_upload
# source: gui_app_modular/handlers/outbound_handlers.py
@router.post("/-on-picking-list-upload", summary="📋  Picking List 업로드 (PDF)")
async def ononpickinglistupload(payload: dict | None = None):
    """Feature F017: 📋  Picking List 업로드 (PDF)"""
    try:
        from gui_app_modular.handlers.outbound_handlers import _on_picking_list_upload  # type: ignore
    except ImportError:
        raise NotReadyError("F017 📋  Picking List 업로드 (PDF)")
    # v864.3 방안 A: GUI 재구현 필요 → 명시적 NOT_READY
    raise NotReadyError("F017 📋  Picking List 업로드 (PDF) — GUI 재구현 필요 (Phase 4-B)")

# ── F018 | menubar | 파일 > 출고 | 📊  바코드 스캔 업로드 ──
# tkinter_callback: _on_barcode_scan_upload
# source: gui_app_modular/handlers/outbound_handlers.py
@router.post("/-on-barcode-scan-upload", summary="📊  바코드 스캔 업로드")
async def ononbarcodescanupload(payload: dict | None = None):
    """Feature F018: 📊  바코드 스캔 업로드"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F018 📊  바코드 스캔 업로드 — GUI 재구현 필요 (Phase 4-B)")
# ── F019 | menubar | 파일 > 출고 | 📷  스캔 탭으로 이동 ──
# tkinter_callback: _on_go_scan_tab
# source: gui_app_modular/handlers/outbound_handlers.py
@router.post("/-on-go-scan-tab", summary="📷  스캔 탭으로 이동")
async def onongoscantab(payload: dict | None = None):
    """Feature F019: 📷  스캔 탭으로 이동"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F019 📷  스캔 탭으로 이동 — GUI 재구현 필요 (Phase 4-B)")
# ── F020 | menubar | 파일 > 출고 | 📋  Allocation 입력 ──
# tkinter_callback: _on_allocation_input_unified
# source: gui_app_modular/handlers/outbound_handlers.py
@router.post("/-on-allocation-input-unified", summary="📋  Allocation 입력")
async def ononallocationinputunified(payload: dict | None = None):
    """Feature F020: 📋  Allocation 입력"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F020 📋  Allocation 입력 — GUI 재구현 필요 (Phase 4-B)")
# ── F021 | menubar | 파일 > 출고 | ✅  승인 대기 ──
# tkinter_callback: _show_allocation_approval_queue
# source: gui_app_modular/handlers/outbound_handlers.py
@router.post("/-show-allocation-approval-queue", summary="✅  승인 대기")
async def onshowallocationapprovalqueue(payload: dict | None = None):
    """Feature F021: ✅  승인 대기"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F021 ✅  승인 대기 — GUI 재구현 필요 (Phase 4-B)")
# ── F022 | menubar | 파일 > 출고 | 📌  예약 반영 (승인분) ──
# tkinter_callback: _apply_approved_allocation
# source: gui_app_modular/handlers/outbound_handlers.py
@router.post("/-apply-approved-allocation", summary="📌  예약 반영 (승인분)")
async def onapplyapprovedallocation(payload: dict | None = None):
    """Feature F022: 📌  예약 반영 (승인분)"""
    try:
        from gui_app_modular.handlers.outbound_handlers import _apply_approved_allocation  # type: ignore
    except ImportError:
        raise NotReadyError("F022 📌  예약 반영 (승인분)")
    # v864.3 방안 A: GUI 재구현 필요 → 명시적 NOT_READY
    raise NotReadyError("F022 📌  예약 반영 (승인분) — GUI 재구현 필요 (Phase 4-B)")

# ── F023 | menubar | 파일 > 출고 | 📜  승인 이력 조회 ──
# tkinter_callback: _show_allocation_approval_history
# source: gui_app_modular/handlers/outbound_handlers.py
@router.post("/-show-allocation-approval-history", summary="📜  승인 이력 조회")
async def onshowallocationapprovalhistory(payload: dict | None = None):
    """Feature F023: 📜  승인 이력 조회"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F023 📜  승인 이력 조회 — GUI 재구현 필요 (Phase 4-B)")
# ── F024 | menubar | 파일 > 출고 | 📋  판매 배정 탭으로 이동 ──
# tkinter_callback: _on_go_allocation_tab
# source: gui_app_modular/handlers/outbound_handlers.py
@router.post("/-on-go-allocation-tab", summary="📋  판매 배정 탭으로 이동")
async def onongoallocationtab(payload: dict | None = None):
    """Feature F024: 📋  판매 배정 탭으로 이동"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F024 📋  판매 배정 탭으로 이동 — GUI 재구현 필요 (Phase 4-B)")
# ── F025 | menubar | 파일 > 출고 | 📋  출고 현황 조회 ──
# tkinter_callback: _show_outbound_history
# source: gui_app_modular/mixins/advanced_dialogs_mixin.py
@router.post("/-show-outbound-history", summary="📋  출고 현황 조회")
async def onshowoutboundhistory(payload: dict | None = None):
    """Feature F025: 📋  출고 현황 조회"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F025 📋  출고 현황 조회 — GUI 재구현 필요 (Phase 4-B)")
# ── F026 | menubar | 파일 > 출고 | 📊  Sales Order 업로드 ──
# tkinter_callback: _on_sales_order_upload
# source: gui_app_modular/mixins/toolbar_mixin.py
@router.post("/-on-sales-order-upload", summary="📊  Sales Order 업로드")
async def ononsalesorderupload(payload: dict | None = None):
    """Feature F026: 📊  Sales Order 업로드"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F026 📊  Sales Order 업로드 — GUI 재구현 필요 (Phase 4-B)")
# ── F027 | menubar | 파일 > 출고 | 🔁  Swap 리포트 ──
# tkinter_callback: _show_swap_report_dialog
# source: gui_app_modular/handlers/outbound_handlers.py
@router.post("/-show-swap-report-dialog", summary="🔁  Swap 리포트")
async def onshowswapreportdialog(payload: dict | None = None):
    """Feature F027: 🔁  Swap 리포트"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F027 🔁  Swap 리포트 — GUI 재구현 필요 (Phase 4-B)")
# ── F028 | menubar | 파일 > 출고 | 📦  출고 피킹 템플릿 관리 ──
# tkinter_callback: _on_picking_template_manage
# source: gui_app_modular/mixins/advanced_dialogs_mixin.py
@router.post("/-on-picking-template-manage", summary="📦  출고 피킹 템플릿 관리")
async def ononpickingtemplatemanage(payload: dict | None = None):
    """Feature F028: 📦  출고 피킹 템플릿 관리"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F028 📦  출고 피킹 템플릿 관리 — GUI 재구현 필요 (Phase 4-B)")
# ── F029 | menubar | 파일 > 백업 | 💾 백업 생성 ──
# tkinter_callback: _on_backup_click
# source: gui_app_modular/handlers/backup_handlers.py
@router.post("/-on-backup-click", summary="💾 백업 생성")
async def ononbackupclick(payload: dict | None = None):
    """Feature F029: 💾 백업 생성"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F029 💾 백업 생성 — GUI 재구현 필요 (Phase 4-B)")
# ── F030 | menubar | 파일 > 백업 | 🔄 복원 ──
# tkinter_callback: _on_restore_click
# source: gui_app_modular/handlers/backup_handlers.py
@router.post("/-on-restore-click", summary="🔄 복원")
async def ononrestoreclick(payload: dict | None = None):
    """Feature F030: 🔄 복원"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F030 🔄 복원 — GUI 재구현 필요 (Phase 4-B)")
# ── F031 | menubar | 파일 > 백업 | 📋 백업 목록 ──
# tkinter_callback: _show_backup_list
# source: gui_app_modular/handlers/backup_handlers.py
@router.post("/-show-backup-list", summary="📋 백업 목록")
async def onshowbackuplist(payload: dict | None = None):
    """Feature F031: 📋 백업 목록"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F031 📋 백업 목록 — GUI 재구현 필요 (Phase 4-B)")
# ── F032 | menubar | 파일 > AI 도구 | 🚢 선사 BL 등록 도구 ──
# tkinter_callback: _on_bl_carrier_register
# source: unknown
@router.post("/-on-bl-carrier-register", summary="🚢 선사 BL 등록 도구")
async def ononblcarrierregister(payload: dict | None = None):
    """Feature F032: 🚢 선사 BL 등록 도구"""
    raise NotReadyError("F032 🚢 선사 BL 등록 도구")

# ── F033 | menubar | 파일 > AI 도구 | 🔬 선사 패턴 분석 ──
# tkinter_callback: _on_bl_carrier_analyze
# source: unknown
@router.post("/-on-bl-carrier-analyze", summary="🔬 선사 패턴 분석")
async def ononblcarrieranalyze(payload: dict | None = None):
    """Feature F033: 🔬 선사 패턴 분석"""
    raise NotReadyError("F033 🔬 선사 패턴 분석")

# ── F034 | menubar | 도구 | 📋 감사 로그 조회 / Export ──
# tkinter_callback: _s1_open_audit_viewer
# source: gui_app_modular/handlers/outbound_handlers.py
@router.post("/-s1-open-audit-viewer", summary="📋 감사 로그 조회 / Export")
async def ons1openauditviewer(payload: dict | None = None):
    """Feature F034: 📋 감사 로그 조회 / Export"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F034 📋 감사 로그 조회 / Export — GUI 재구현 필요 (Phase 4-B)")
# ── F035 | menubar | 재고 | 📊 LOT 리스트 Excel ──
# tkinter_callback: _on_export_click
# source: gui_app_modular/handlers/export_handlers.py
@router.post("/-on-export-click", summary="📊 LOT 리스트 Excel")
async def ononexportclick(payload: dict | None = None):
    """Feature F035: 📊 LOT 리스트 Excel"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F035 📊 LOT 리스트 Excel — GUI 재구현 필요 (Phase 4-B)")
# ── F036 | menubar | 재고 | 🎒 톤백리스트 Excel ──
# tkinter_callback: _on_export_click
# source: gui_app_modular/handlers/export_handlers.py
@router.post("/-on-export-click", summary="🎒 톤백리스트 Excel")
async def ononexportclick_f036(payload: dict | None = None):
    """Feature F036: 🎒 톤백리스트 Excel"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F036 🎒 톤백리스트 Excel — GUI 재구현 필요 (Phase 4-B)")
# ── F037 | menubar | 재고 | 📋 출고 현황 조회 ──
# tkinter_callback: _show_outbound_history
# source: gui_app_modular/mixins/advanced_dialogs_mixin.py
@router.post("/-show-outbound-history", summary="📋 출고 현황 조회")
async def onshowoutboundhistory_f037(payload: dict | None = None):
    """Feature F037: 📋 출고 현황 조회"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F037 📋 출고 현황 조회 — GUI 재구현 필요 (Phase 4-B)")
# ── F038 | menubar | 재고 | 📊 재고 추이 차트 ──
# tkinter_callback: _show_snapshot_chart
# source: gui_app_modular/mixins/advanced_dialogs_mixin.py
@router.post("/-show-snapshot-chart", summary="📊 재고 추이 차트")
async def onshowsnapshotchart(payload: dict | None = None):
    """Feature F038: 📊 재고 추이 차트"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F038 📊 재고 추이 차트 — GUI 재구현 필요 (Phase 4-B)")
# ── F039 | menubar | 보고서 | 📄 거래명세서 생성 (headless) ──
# Phase 2 Step 3 (2026-04-21): Tkinter dialog -> headless payload 방식으로 재구현
# source: gui_app_modular/mixins/advanced_dialogs_mixin.py:1329 (참조만)
@router.post("/-generate-outbound-invoice", summary="📄 거래명세서 생성")
async def ongenerateoutboundinvoice(payload: dict | None = None):
    """
    Feature F039: 거래명세서 Excel 생성 (headless).

    payload: {customer, date_from (YYYY-MM-DD), date_to (YYYY-MM-DD)}
    payload 누락 시 required 안내 반환 (soft-fail).
    """
    from backend.api import engine as _engine, ENGINE_AVAILABLE as _ok
    if not _ok or _engine is None:
        raise NotReadyError("F039 거래명세서 - engine not loaded")

    p = payload or {}
    customer = str(p.get("customer", "")).strip()
    date_from = str(p.get("date_from", "")).strip()
    date_to = str(p.get("date_to", "")).strip()

    if not customer or not date_from or not date_to:
        return ok_response(
            data={"required": ["customer", "date_from", "date_to"]},
            message="payload 필요: {customer, date_from, date_to (YYYY-MM-DD)}",
        )

    def _gen():
        import os, openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        try:
            movements = _engine.db.fetchall(
                """SELECT lot_no, qty_kg, customer,
                      COALESCE(movement_date, created_at) AS movement_date
                   FROM stock_movement
                   WHERE customer = ? AND movement_type = 'OUTBOUND'
                     AND COALESCE(movement_date, created_at) >= ?
                     AND COALESCE(movement_date, created_at) <= ?
                   ORDER BY created_at""",
                (customer, date_from, date_to + ' 23:59:59'),
            )
        except Exception:
            movements = []

        if not movements:
            return {
                "ok": False, "filepath": None, "count": 0,
                "message": f"{customer} {date_from}~{date_to} 출고 이력 없음",
            }

        out_dir = "features/reports"
        os.makedirs(out_dir, exist_ok=True)
        save_path = os.path.join(
            out_dir, f"거래명세서_{customer}_{date_from}_{date_to}.xlsx",
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "거래명세서"
        ws.merge_cells('A1:F1')
        ws['A1'] = f"거래명세서 - {customer}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A2'] = f"기간: {date_from} ~ {date_to}"
        ws['A2'].font = Font(size=10, color='666666')

        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        headers = ['No', 'LOT NO', '수량(kg)', '수량(MT)', '출고일', '비고']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=4, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            c.border = border
            c.alignment = Alignment(horizontal='center')

        total_kg = 0
        for i, mv in enumerate(movements, 1):
            qty = mv['qty_kg'] or 0
            total_kg += qty
            ws.cell(row=4+i, column=1, value=i).border = border
            ws.cell(row=4+i, column=2, value=mv['lot_no']).border = border
            c3 = ws.cell(row=4+i, column=3, value=f"{qty:,.0f}")
            c3.border = border; c3.alignment = Alignment(horizontal='right')
            c4 = ws.cell(row=4+i, column=4, value=f"{qty/1000:.3f}")
            c4.border = border; c4.alignment = Alignment(horizontal='right')
            ws.cell(row=4+i, column=5, value=str(mv['movement_date'] or '')[:10]).border = border
            ws.cell(row=4+i, column=6, value='').border = border

        summary_row = 5 + len(movements)
        ws.cell(row=summary_row, column=1, value=f"합계: {total_kg:,.0f} kg ({total_kg/1000:.3f} MT)").font = Font(bold=True)
        wb.save(save_path)

        return {
            "ok": True, "filepath": save_path,
            "count": len(movements), "total_kg": total_kg,
            "message": f"{len(movements)}건 거래명세서 Excel 생성 완료",
        }

    return wrap_engine_call(_gen)

# ── F040 | menubar | 보고서 | 📦 Detail of Outbound ──
# tkinter_callback: _on_detail_of_outbound_report
# source: gui_app_modular/mixins/advanced_dialogs_mixin.py
@router.post("/-on-detail-of-outbound-report", summary="📦 Detail of Outbound")
async def onondetailofoutboundreport(payload: dict | None = None):
    """Feature F040: 📦 Detail of Outbound"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F040 📦 Detail of Outbound — GUI 재구현 필요 (Phase 4-B)")
# ── F041 | menubar | 보고서 | 📋 Sales Order DN ──
# tkinter_callback: _on_sales_order_dn_report
# source: gui_app_modular/mixins/advanced_dialogs_mixin.py
@router.post("/-on-sales-order-dn-report", summary="📋 Sales Order DN")
async def ononsalesorderdnreport(payload: dict | None = None):
    """Feature F041: 📋 Sales Order DN"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F041 📋 Sales Order DN — GUI 재구현 필요 (Phase 4-B)")
# ── F042 | menubar | 보고서 | 🔍 DN 교차검증 ──
# tkinter_callback: _on_dn_cross_check
# source: gui_app_modular/mixins/toolbar_mixin.py
@router.post("/-on-dn-cross-check", summary="🔍 DN 교차검증")
async def onondncrosscheck(payload: dict | None = None):
    """Feature F042: 🔍 DN 교차검증"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F042 🔍 DN 교차검증 — GUI 재구현 필요 (Phase 4-B)")
# ── F043 | menubar | 보고서 | 📝 고객 보고서 생성 ──
# tkinter_callback: _generate_customer_report
# source: unknown
@router.post("/-generate-customer-report", summary="📝 고객 보고서 생성")
async def ongeneratecustomerreport(payload: dict | None = None):
    """Feature F043: 📝 고객 보고서 생성"""
    raise NotReadyError("F043 📝 고객 보고서 생성")

# ── F044 | menubar | 보고서 | 📂 보고서 양식 관리 ──
# tkinter_callback: _manage_report_templates
# source: unknown
@router.post("/-manage-report-templates", summary="📂 보고서 양식 관리")
async def onmanagereporttemplates(payload: dict | None = None):
    """Feature F044: 📂 보고서 양식 관리"""
    raise NotReadyError("F044 📂 보고서 양식 관리")

# ── F045 | menubar | 보고서 | 📋 보고서 이력 조회 ──
# tkinter_callback: _show_report_history
# source: unknown
@router.post("/-show-report-history", summary="📋 보고서 이력 조회")
async def onshowreporthistory(payload: dict | None = None):
    """Feature F045: 📋 보고서 이력 조회"""
    raise NotReadyError("F045 📋 보고서 이력 조회")

# ── F046 | menubar | 보고서 | 📦 재고 현황 보고서 ──
# tkinter_callback: _generate_inventory_pdf_report
# source: gui_app_modular/handlers/pdf_report_handler.py
@router.post("/-generate-inventory-pdf-report", summary="📦 재고 현황 보고서")
async def ongenerateinventorypdfreport(payload: dict | None = None):
    """Feature F046: 📦 재고 현황 보고서"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F046 📦 재고 현황 보고서 — GUI 재구현 필요 (Phase 4-B)")
# ── F047 | menubar | 보고서 | 📈 입출고 내역 ──
# tkinter_callback: _generate_transaction_pdf
# source: gui_app_modular/handlers/pdf_handlers.py
@router.post("/-generate-transaction-pdf", summary="📈 입출고 내역")
async def ongeneratetransactionpdf(payload: dict | None = None):
    """Feature F047: 📈 입출고 내역"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F047 📈 입출고 내역 — GUI 재구현 필요 (Phase 4-B)")
# ── F048 | menubar | 보고서 | 📅 월간 실적 PDF ──
# tkinter_callback: _generate_monthly_pdf_v398
# source: gui_app_modular/handlers/pdf_handlers.py
@router.post("/-generate-monthly-pdf-v398", summary="📅 월간 실적 PDF")
async def ongeneratemonthlypdfv398(payload: dict | None = None):
    """Feature F048: 📅 월간 실적 PDF"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F048 📅 월간 실적 PDF — GUI 재구현 필요 (Phase 4-B)")
# ── F049 | menubar | 보고서 | 📊 일일 현황 PDF ──
# tkinter_callback: _generate_daily_pdf_v398
# source: gui_app_modular/handlers/pdf_handlers.py
@router.post("/-generate-daily-pdf-v398", summary="📊 일일 현황 PDF")
async def ongeneratedailypdfv398(payload: dict | None = None):
    """Feature F049: 📊 일일 현황 PDF"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F049 📊 일일 현황 PDF — GUI 재구현 필요 (Phase 4-B)")
# ── F050 | menubar | 보고서 | 🔖 LOT 상세 (headless) ──
# Phase 2 Step 3 (2026-04-21): Tkinter selection -> payload.lot_no 방식으로 재구현
# source: gui_app_modular/handlers/pdf_handlers.py:455 (참조만)
@router.post("/-generate-lot-detail-pdf", summary="🔖 LOT 상세")
async def ongeneratelotdetailpdf(payload: dict | None = None):
    """
    Feature F050: LOT 상세 보고서 PDF (headless).

    payload: {lot_no: str}
    payload 누락 시 required 안내 반환 (soft-fail).
    """
    from backend.api import engine as _engine, ENGINE_AVAILABLE as _ok
    if not _ok or _engine is None:
        raise NotReadyError("F050 LOT 상세 - engine not loaded")

    lot_no = str((payload or {}).get("lot_no", "")).strip()
    if not lot_no:
        return ok_response(
            data={"required": "lot_no"},
            message='payload 필요: {"lot_no": "ABC123"}',
        )

    def _gen():
        import os
        try:
            from gui_app_modular.utils.pdf_report_gen import generate_outbound_confirmation
        except ImportError as e:
            return {"ok": False, "filepath": None,
                    "message": f"pdf_report_gen 모듈 없음 ({e})"}

        filepath = generate_outbound_confirmation(_engine, lot_no=lot_no, customer='')
        if not filepath or not os.path.exists(filepath):
            return {"ok": False, "filepath": None,
                    "message": f"LOT {lot_no} 출고 데이터 없음 또는 reportlab 미설치"}
        return {"ok": True, "filepath": filepath,
                "message": f"LOT {lot_no} 상세 PDF 생성 완료"}

    return wrap_engine_call(_gen)

# ── F051 | menubar | 설정/도구 | 🔄 새로고침 (F5) ──
# tkinter_callback: _refresh_all_data
# source: gui_app_modular/mixins/toolbar_mixin.py
@router.post("/-refresh-all-data", summary="🔄 새로고침 (F5)")
async def onrefreshalldata(payload: dict | None = None):
    """Feature F051: 🔄 새로고침 (F5)"""
    try:
        from gui_app_modular.mixins.toolbar_mixin import _refresh_all_data  # type: ignore
    except ImportError:
        raise NotReadyError("F051 🔄 새로고침 (F5)")
    # v864.3 방안 A: GUI 재구현 필요 → 명시적 NOT_READY
    raise NotReadyError("F051 🔄 새로고침 (F5) — GUI 재구현 필요 (Phase 4-B)")

# ── F052 | menubar | 설정/도구 | 💾 현재 창 크기 저장 ──
# tkinter_callback: _on_save_window_size
# source: gui_app_modular/mixins/toolbar_mixin.py
@router.post("/-on-save-window-size", summary="💾 현재 창 크기 저장")
async def ononsavewindowsize(payload: dict | None = None):
    """Feature F052: 💾 현재 창 크기 저장"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F052 💾 현재 창 크기 저장 — GUI 재구현 필요 (Phase 4-B)")
# ── F053 | menubar | 설정/도구 | ↩️ 기본 창 크기 초기화 ──
# tkinter_callback: _on_reset_window_size
# source: gui_app_modular/mixins/toolbar_mixin.py
@router.post("/-on-reset-window-size", summary="↩️ 기본 창 크기 초기화")
async def ononresetwindowsize(payload: dict | None = None):
    """Feature F053: ↩️ 기본 창 크기 초기화"""
    # v864.3 방안 A: GUI(tkinter filedialog/messagebox) 필요 → 명시적 NOT_READY
    raise NotReadyError("F053 ↩️ 기본 창 크기 초기화 — GUI 재구현 필요 (Phase 4-B)")
# ── F054 | menubar | 설정/도구 | 📦 제품 마스터 관리 ──
# tkinter_callback: _show_product_master
# source: unknown
@router.post("/-show-product-master", summary="📦 제품 마스터 관리")
async def onshowproductmaster_f054(payload: dict | None = None):
    """Feature F054: 📦 제품 마스터 관리"""
    raise NotReadyError("F054 📦 제품 마스터 관리")

# ── F055 | menubar | 설정/도구 | 📊 제품별 재고 현황 ──
# tkinter_callback: _show_product_inventory_report
# source: unknown
@router.post("/-show-product-inventory-report", summary="📊 제품별 재고 현황")
async def onshowproductinventoryreport(payload: dict | None = None):
    """Feature F055: 📊 제품별 재고 현황"""
    raise NotReadyError("F055 📊 제품별 재고 현황")

# ── F056 | menubar | 설정/도구 | 📋 D/O 후속 연결 ──
# tkinter_callback: _on_do_update
# source: gui_app_modu

@router.post("/-on-settings", summary="⚙️ 환경 설정")
def on_settings(payload: dict = None):
    """환경 설정 다이얼로그 (현재: 설정 정보 반환)"""
    import sqlite3, os
    base = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    db_path = os.path.join(base, "data", "db", "sqm_inventory.db")
    return {
        "success": True,
        "title": "SQM v864.3 환경 설정",
        "settings": {
            "db_path": db_path,
            "db_exists": os.path.exists(db_path),
            "version": "864.3",
            "note": "고급 설정은 다음 버전에서 지원 예정"
        }
    }
