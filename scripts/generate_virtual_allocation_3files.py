"""
가상 Allocation Table Excel 3개 생성
- 입고 리스트에 있는 컬럼만 사용: Product, SAP NO, Lot No, WH, Customs, Date in stock, QTY (MT)
- ETA BUSAN, GW, SALE REF는 빈칸 (입고 리스트에 없음)
- 파일당 200 일반 톤백 + 200 샘플 = 400행
"""
import random
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl 필요: pip install openpyxl")
    raise

# 프로젝트 루트 기준 출력 폴더 (스크립트 상위 = 프로젝트 루트)
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
OUTPUT_DIR = PROJECT_ROOT / "generated_allocation"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

PRODUCTS = ['MIC9000', 'MIC7100', 'LCA', 'NSH']
SAP_NOS = ['2200032552', '2200032833', '2200032900', '2200033010']
CUSTOMS = ['Cleared', 'Cleared', 'Cleared', 'Uncleared']
DATE_STOCKS = ['2025-07-29', '2025-08-15', '2025-09-01', '2025-09-18', '2025-10-05']


def _generate_lot_no(used: set) -> str:
    """10자리 LOT 번호 생성 (1125xxxxxx)"""
    while True:
        lot = '1125' + str(random.randint(100000, 999999))
        if lot not in used:
            used.add(lot)
            return lot


def _create_allocation_file(filepath: Path, file_index: int, seed: int) -> None:
    random.seed(seed)
    used_lots = set()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Allocation Table"

    title_font = Font(bold=True, size=14, color="2C3E50")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    data_font = Font(size=10)
    sample_font = Font(size=10, color="0066CC")
    header_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    data_fill = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")
    sample_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    # Row 1: 타이틀
    ws.merge_cells('A1:J1')
    ws['A1'] = f"Allocation - PT LBM - File{file_index} / CIF Semarang - 2000MT of MIC9000"
    ws['A1'].font = title_font
    ws.row_dimensions[1].height = 30

    # Row 2: 합계 (나중에 수식)
    ws.row_dimensions[2].height = 20

    # Row 3: 헤더 (화주 양식 — 입고 리스트에 있는 것 + QTY)
    headers = [
        ('Product', 16), ('SAP NO', 14), ('ETA BUSAN', 14), ('Date in stock', 14),
        ('QTY (MT)', 12), ('Lot No', 14), ('WH', 8), ('Customs', 12),
        ('GW', 12), ('SALE REF', 12),
    ]
    for col, (text, width) in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width

    row_num = 4
    total_qty = 0.0

    # 200 일반 톤백
    for _ in range(200):
        lot = _generate_lot_no(used_lots)
        product = random.choice(PRODUCTS)
        sap = random.choice(SAP_NOS)
        date_stock = random.choice(DATE_STOCKS)
        qty = 5.0
        customs = random.choice(CUSTOMS)
        vals = [product, sap, '', date_stock, qty, lot, 'GY', customs, '', '']
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = data_font
            cell.fill = data_fill
            cell.border = thin_border
            if col == 5:
                cell.number_format = '#,##0.000'
                cell.alignment = right_align
            else:
                cell.alignment = center
        row_num += 1
        total_qty += qty

    # 200 샘플 톤백
    for _ in range(200):
        lot = _generate_lot_no(used_lots)
        product = random.choice(PRODUCTS) + ' Sample'
        sap = random.choice(SAP_NOS)
        date_stock = random.choice(DATE_STOCKS)
        qty = 0.001
        customs = random.choice(CUSTOMS)
        vals = [product, sap, '', date_stock, qty, lot, 'GY', customs, '', '']
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = sample_font
            cell.fill = sample_fill
            cell.border = thin_border
            if col == 5:
                cell.number_format = '#,##0.000'
                cell.alignment = right_align
            else:
                cell.alignment = center
        row_num += 1
        total_qty += qty

    last_row = row_num - 1
    ws['E2'].value = f"=SUM(E4:E{last_row})"
    ws['E2'].number_format = '#,##0.0000'
    ws['E2'].font = Font(bold=True)

    wb.save(filepath)
    print(f"  생성: {filepath.name} ({last_row - 3}행, 합계 QTY 약 {total_qty:.3f} MT)")


def main():
    print("가상 Allocation Table Excel 3개 생성 중...")
    print(f"출력 폴더: {OUTPUT_DIR}")
    for i in range(1, 4):
        fpath = OUTPUT_DIR / f"Allocation_가상_{i}.xlsx"
        _create_allocation_file(fpath, i, seed=100 + i)
    print("완료.")


if __name__ == '__main__':
    main()
