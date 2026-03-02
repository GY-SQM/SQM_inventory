"""
업로드한 톤백 테이블(DB)에서 Allocation Table 샘플 Excel 3개 생성

규칙:
- 88 LOT, 880 톤백, 88 샘플 중에서 200 톤백 + 20 샘플 allocation
- LOT당 0~5개 랜덤 선택, 총 200개까지 (A에서 5개, B에서 6개... 방식)
- 샘플: LOT당 1개씩 20개
- 원칙: 한 번 allocation된 톤백은 재사용 불가, 중복 불가, 재고 0 미만 불가
- 3개 파일 생성 시 톤백 풀 공유·중복 없음 (파일1 사용분 → 파일2에서 제외)
"""
import os
import random
import sys
from pathlib import Path

# 프로젝트 루트를 path에 추가
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
sys.path.insert(0, str(PROJECT_ROOT))

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl 필요: pip install openpyxl")
    raise

OUTPUT_DIR = PROJECT_ROOT / "generated_allocation"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def _tonbag_key(tb: dict) -> tuple:
    """톤백 고유 키 (lot_no, tonbag_no/sub_lt) — 중복·재사용 검사용"""
    lot = str(tb.get('lot_no') or '').strip()
    tn = tb.get('tonbag_no') or tb.get('sub_lt')
    if tn is None:
        tn = tb.get('tonbag_uid') or ''
    return (lot, str(tn) if tn is not None else '')


def _get_tonbags_from_db(db_path: str = None):
    """DB에서 톤백+재고 JOIN 데이터 조회"""
    if db_path is None:
        try:
            from core.config import DB_PATH
            db_path = str(DB_PATH)
        except ImportError:
            db_path = str(PROJECT_ROOT / "data" / "db" / "sqm_inventory.db")

    if not os.path.exists(db_path):
        return None, f"DB 파일 없음: {db_path}"

    try:
        from engine_modules.inventory_modular.engine import SQMInventoryEngineV3
        engine = SQMInventoryEngineV3(db_path=db_path)
        tonbags = engine.get_tonbags_with_inventory()
        return tonbags, None
    except Exception as e:
        return None, str(e)


def _to_allocation_row(tb: dict, is_sample: bool) -> list:
    """톤백 dict를 Allocation 행 데이터로 변환 (화주 양식)"""
    product = (tb.get('product') or '').strip() or 'MIC9000'
    if is_sample:
        product = product if 'sample' in product.lower() else f"{product} sample"
    sap_no = str(tb.get('sap_no') or '').strip() or '2200032552'
    lot_no = str(tb.get('lot_no') or '').strip()
    warehouse = str(tb.get('warehouse') or 'GY').strip()
    customs = str(tb.get('customs') or 'Cleared').strip()
    date_stock = tb.get('arrival_date') or tb.get('ship_date') or '2025-07-29'
    if date_stock:
        date_stock = str(date_stock)[:10]
    # 1 행 = 1 톤백. 일반 5MT, 샘플 0.001MT (이미지 양식 기준)
    qty_mt = 5.0 if not is_sample else 0.001
    gw = 5.13 if not is_sample else 0.00125  # 일반 5.13톤, 샘플 1.25kg
    sale_ref = str(tb.get('sale_ref') or '1955').strip()
    return [product, sap_no, '', date_stock, qty_mt, lot_no, warehouse, customs, gw, sale_ref]


def _create_allocation_excel(filepath: Path, rows_data: list, file_index: int) -> float:
    """Allocation Excel 파일 생성 (화주 양식: PT LBM / CN Semarang)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Allocation Table"

    # 스타일 — 이미지 기준
    title_font = Font(bold=True, size=14, color="2C3E50")
    header_font = Font(bold=True, color="2C3E50", size=10)
    data_font = Font(size=10)
    sample_font = Font(size=10, color="0066CC")
    # 헤더: 연한 초록 배경
    header_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    # ETA BUSAN, Customs, GW — 노란 강조
    header_yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    # 데이터: 흰색 / 연한 회색 교대
    data_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    data_grey = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    sample_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    total_qty = sum(float(vals[4]) for vals, _ in rows_data)

    # Row 1: 타이틀 (화주 양식)
    ws.merge_cells('A1:J1')
    ws['A1'] = f"Allocation - PT LBM - September / CN Semarang - {total_qty:.1f}MT of MIC9000"
    ws['A1'].font = title_font
    ws.row_dimensions[1].height = 30

    # Row 2: 합계
    ws.row_dimensions[2].height = 6

    # Row 3: 헤더 (연한 초록, ETA BUSAN·Customs·GW는 노란)
    headers = [
        ('Product', 16), ('SAP NO', 14), ('ETA BUSAN', 14), ('Date in stock', 14),
        ('QTY (MT)', 12), ('Lot No', 14), ('WH', 8), ('Customs', 12),
        ('GW', 12), ('SALE REF', 12),
    ]
    yellow_cols = (3, 8, 9)  # ETA BUSAN, Customs, GW (1-based)
    for col, (text, width) in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=text)
        cell.font = header_font
        cell.fill = header_yellow if col in yellow_cols else header_fill
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width

    # Row 4~: 데이터 (교대 색상, 샘플은 연두)
    row_num = 4
    for i, (vals, is_sample) in enumerate(rows_data):
        fill = sample_fill if is_sample else (data_grey if i % 2 == 1 else data_white)
        font = sample_font if is_sample else data_font
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = font
            cell.fill = fill
            cell.border = thin_border
            if col in (5, 9):  # QTY (MT), GW
                cell.number_format = '#,##0.000'
                cell.alignment = right_align
            else:
                cell.alignment = center
        row_num += 1

    last_row = row_num - 1
    ws['E2'].value = f"=SUM(E4:E{last_row})"
    ws['E2'].number_format = '#,##0.0000'
    ws['E2'].font = Font(bold=True)

    wb.save(filepath)
    return total_qty


def _pick_regular_tonbags(
    lots_dict: dict,
    used_keys: set,
    target_count: int = 200,
    seed: int = 0,
) -> tuple:
    """
    LOT당 0~5개 랜덤 allocation, 총 target_count개. used_keys 제외, 재고 0 미만 방지.
    Returns: (picked_list, used_keys_updated). picked에 중복 있으면 에러.
    """
    random.seed(seed)
    lots_with_regular = [
        (lot_no, [t for t in info['regular'] if _tonbag_key(t) not in used_keys])
        for lot_no, info in lots_dict.items()
        if info['regular']
    ]
    lots_with_regular = [(ln, lst) for ln, lst in lots_with_regular if lst]
    random.shuffle(lots_with_regular)
    picked = []
    seen = set()
    for lot_no, available in lots_with_regular:
        if len(picked) >= target_count:
            break
        n = random.randint(0, min(5, len(available)))
        if n == 0:
            continue
        candidates = [t for t in available if _tonbag_key(t) not in seen]
        take = min(n, target_count - len(picked), len(candidates))
        if take <= 0:
            continue
        chosen = random.sample(candidates, take)
        for t in chosen:
            k = _tonbag_key(t)
            if k in seen:
                raise ValueError(f"중복 톤백: {k}")
            seen.add(k)
            used_keys.add(k)
            picked.append(t)
    return picked, used_keys


def _pick_sample_tonbags(
    lots_dict: dict,
    used_keys: set,
    target_count: int = 20,
    seed: int = 0,
) -> tuple:
    """
    샘플 있는 LOT 20개 선택, LOT당 1개. used_keys 제외.
    """
    random.seed(seed)
    lots_with_sample = [
        (lot_no, info['sample'])
        for lot_no, info in lots_dict.items()
        if info['sample'] and _tonbag_key(info['sample']) not in used_keys
    ]
    chosen = random.sample(lots_with_sample, min(target_count, len(lots_with_sample)))
    picked = []
    for lot_no, tb in chosen:
        k = _tonbag_key(tb)
        if k in used_keys:
            raise ValueError(f"중복 샘플 톤백: {k}")
        used_keys.add(k)
        picked.append(tb)
    return picked, used_keys


def _create_dummy_rows(count_regular: int = 60, count_sample: int = 2, seed: int = 0) -> list:
    """DB 없이 화주 양식 더미 데이터 생성 (이미지 형식)"""
    random.seed(seed)
    base_lots = ['1125052654', '1125052707', '1125052708', '1125052709', '1125052710',
                 '1125052711', '1125052712', '1125052713', '1125052714', '1125052715',
                 '1125052716', '1125052717', '1125052718', '1125052719', '1125052720']
    rows_data = []
    for _ in range(count_regular):
        lot = random.choice(base_lots) + str(random.randint(0, 99)).zfill(2)
        tb = {'product': 'MIC9000', 'sap_no': '2200032552', 'lot_no': lot[:10],
              'warehouse': 'GY', 'customs': 'Cleared', 'arrival_date': '2025-07-29',
              'sale_ref': '1955'}
        rows_data.append((_to_allocation_row(tb, False), False))
    for _ in range(count_sample):
        lot = random.choice(base_lots)
        tb = {'product': 'MIC9000', 'sap_no': '2200032552', 'lot_no': lot,
              'warehouse': 'GY', 'customs': 'Cleared', 'arrival_date': '2025-07-29',
              'sale_ref': '1955', 'is_sample': True}
        rows_data.append((_to_allocation_row(tb, True), True))
    return rows_data


def main():
    print("Allocation Table 샘플 3개 생성 (화주 양식: PT LBM / CN Semarang)...")
    tonbags, err = _get_tonbags_from_db()
    use_db = tonbags and not err

    if use_db:
        available = [t for t in tonbags if str(t.get('tonbag_status') or '').upper() in ('AVAILABLE', 'RESERVED', '')]
        if not available:
            available = tonbags
        lots_dict = {}
        for t in available:
            lot_no = str(t.get('lot_no') or '').strip()
            if not lot_no:
                continue
            if lot_no not in lots_dict:
                lots_dict[lot_no] = {'regular': [], 'sample': None}
            if t.get('is_sample'):
                lots_dict[lot_no]['sample'] = t
            else:
                lots_dict[lot_no]['regular'].append(t)
        total_regular = sum(len(v['regular']) for v in lots_dict.values())
        lots_with_sample = sum(1 for v in lots_dict.values() if v['sample'])
        use_db = total_regular >= 200 and lots_with_sample >= 20

    if use_db:
        print("  DB 기반 생성 (일반 200+샘플 20)")
        used_tonbag_keys = set()
        for i in range(1, 4):
            try:
                regular_tbs, used_tonbag_keys = _pick_regular_tonbags(
                    lots_dict, used_tonbag_keys, 200, seed=100 + i
                )
                sample_tbs, used_tonbag_keys = _pick_sample_tonbags(
                    lots_dict, used_tonbag_keys, 20, seed=200 + i
                )
            except ValueError as e:
                print(f"  에러 (파일{i}): {e}")
                use_db = False
                break
            rows_data = []
            for t in regular_tbs:
                rows_data.append((_to_allocation_row(t, False), False))
            for t in sample_tbs:
                rows_data.append((_to_allocation_row(t, True), True))
            fpath = OUTPUT_DIR / f"Allocation_샘플_{i}.xlsx"
            total = _create_allocation_excel(fpath, rows_data, i)
            print(f"  생성: {fpath.name} ({len(rows_data)}행, QTY {total:.3f} MT)")

    if not use_db:
        print("  DB 부족 또는 없음 → 더미 데이터로 생성 (60+2행)")
        for i in range(1, 4):
            rows_data = _create_dummy_rows(60, 2, seed=500 + i)
            fpath = OUTPUT_DIR / f"Allocation_샘플_{i}.xlsx"
            total = _create_allocation_excel(fpath, rows_data, i)
            print(f"  생성: {fpath.name} ({len(rows_data)}행, QTY {total:.3f} MT)")

    print(f"\n출력 폴더: {OUTPUT_DIR}")
    print("완료.")


if __name__ == '__main__':
    main()
