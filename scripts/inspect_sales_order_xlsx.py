# -*- coding: utf-8 -*-
"""Sales Order 엑셀 구조 확인 (헤더·샘플 행)."""
import sys
from pathlib import Path

root = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(root))

xlsx = root.parent / "Sales order No (26.02.09)-3266.xlsx"
if not xlsx.exists():
    print("파일 없음:", xlsx)
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("openpyxl 필요: pip install openpyxl")
    sys.exit(1)

wb = openpyxl.load_workbook(str(xlsx), read_only=True, data_only=True)
for sn in wb.sheetnames[:5]:
    ws = wb[sn]
    print("=== Sheet:", sn, "===")
    for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True)):
        if row and any(c is not None for c in row):
            print(i + 1, list(row))
    print()
wb.close()
