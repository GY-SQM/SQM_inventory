# -*- coding: utf-8 -*-
"""Allocation 파일 재생성 스크립트 — 본품+샘플 쌍 구조"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 공통 스타일
hdr_fill = PatternFill('solid', fgColor='1F3864')
hdr_font = Font(bold=True, color='FFFFFF', size=10, name='맑은 고딕')
hdr_align = Alignment(horizontal='center', vertical='center')
data_font = Font(size=10, name='맑은 고딕')
data_align = Alignment(horizontal='center', vertical='center')
sample_font = Font(size=10, name='맑은 고딕', color='996600')
sample_fill = PatternFill('solid', fgColor='FFF2CC')
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)

SONG_HEADERS = ['Product', 'SAP NO', 'Date in stock', 'QTY (MT)', 'Lot No',
                'WH', 'Customs', 'SOLD TO', 'SALE REF', 'GW']
SONG_WIDTHS = [22, 14, 14, 10, 14, 6, 12, 28, 26, 10]

WOO_HEADERS = ['Product', 'SAP NO', 'Date in stock', 'QTY (MT)', 'Lot No',
               'WH', 'Customs', 'Export', 'SOLD TO', 'SALE REF', 'Balance', 'GW', 'Remark']
WOO_WIDTHS = [18, 14, 14, 10, 14, 6, 12, 8, 20, 26, 10, 10, 12]


def write_header_row(ws, row, headers, widths):
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row, c, h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(c)].width = w


def write_data_row(ws, row, vals, is_sample=False):
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row, c, v)
        cell.font = sample_font if is_sample else data_font
        cell.alignment = data_align
        cell.border = thin_border
        if is_sample:
            cell.fill = sample_fill


# ============================================================
# 1. Song 양식 재생성
# ============================================================
print("1. Song 양식 재생성...")
wb_src = openpyxl.load_workbook('Allocation_Song_202607.xlsx', data_only=True)
ws_src = wb_src.active

song_rows = []
seen = set()
for r in range(3, ws_src.max_row + 1):
    product = str(ws_src.cell(r, 1).value or '')
    lot_raw = ws_src.cell(r, 5).value
    if not lot_raw or 'sample' in product.lower():
        continue
    lot = str(lot_raw).split('.')[0]
    if lot in seen:
        continue
    seen.add(lot)
    song_rows.append({
        'product': ws_src.cell(r, 1).value or 'LITHIUM CARBONATE',
        'sap': ws_src.cell(r, 2).value,
        'date': ws_src.cell(r, 3).value,
        'qty': ws_src.cell(r, 4).value,
        'lot': lot,
        'wh': ws_src.cell(r, 6).value or 'GY',
        'customs': ws_src.cell(r, 7).value or 'uncleared',
        'sold_to': ws_src.cell(r, 8).value or 'CATL KOREA CO., LTD',
        'sale_ref': ws_src.cell(r, 9).value or 'JAKARTA-2026-07-SONG',
        'gw': ws_src.cell(r, 10).value,
    })

wb_new = openpyxl.Workbook()
ws = wb_new.active
ws.title = 'Allocation'

title_val = ws_src.cell(1, 1).value or 'Allocation - CATL KOREA CO., LTD - July 2026'
ws.cell(1, 1, title_val)
ws.cell(1, 1).font = Font(bold=True, size=12, name='맑은 고딕')
ws.merge_cells('A1:J1')

write_header_row(ws, 2, SONG_HEADERS, SONG_WIDTHS)

row_num = 3
for d in song_rows:
    qty = float(d['qty'] or 5)
    gw = float(d['gw'] or qty * 1.026)

    # 본품 행
    write_data_row(ws, row_num, [
        d['product'], d['sap'], d['date'], qty, d['lot'],
        d['wh'], d['customs'], d['sold_to'], d['sale_ref'], round(gw, 3)
    ])
    row_num += 1

    # 샘플 행
    sp = str(d['product'] or 'LITHIUM CARBONATE')
    if 'sample' not in sp.lower():
        sp += ' sample'
    write_data_row(ws, row_num, [
        sp, d['sap'], d['date'], 0.001, d['lot'],
        d['wh'], d['customs'], d['sold_to'], d['sale_ref'], 0.00125
    ], is_sample=True)
    row_num += 1

ws.freeze_panes = 'A3'
wb_new.save('Allocation_Song_202607.xlsx')
print(f"   Song 완료: {len(song_rows)} LOT x 2 = {len(song_rows)*2} 행")


# ============================================================
# 2. Woo 양식 재생성
# ============================================================
print("2. Woo 양식 재생성...")
wb_src2 = openpyxl.load_workbook('Allocation_Woo_202606.xlsx', data_only=True)
ws_src2 = wb_src2.active

woo_rows = []
seen2 = set()
for r in range(7, ws_src2.max_row + 1):
    product = str(ws_src2.cell(r, 1).value or '')
    lot_raw = ws_src2.cell(r, 5).value
    if not lot_raw or 'sample' in product.lower():
        continue
    lot = str(lot_raw).split('.')[0]
    if lot in seen2:
        continue
    seen2.add(lot)
    woo_rows.append({
        'product': ws_src2.cell(r, 1).value or 'MIC9000',
        'sap': ws_src2.cell(r, 2).value,
        'date': ws_src2.cell(r, 3).value,
        'qty': ws_src2.cell(r, 4).value,
        'lot': lot,
        'wh': ws_src2.cell(r, 6).value or 'GY',
        'customs': ws_src2.cell(r, 7).value or 'Uncleared',
        'export': ws_src2.cell(r, 8).value or '반송',
        'sold_to': ws_src2.cell(r, 9).value or 'PT LBM JAKARTA',
        'sale_ref': ws_src2.cell(r, 10).value or 'JAKARTA-2026-06-WOO',
        'balance': ws_src2.cell(r, 11).value,
        'gw': ws_src2.cell(r, 12).value,
        'remark': ws_src2.cell(r, 13).value,
    })

wb_new2 = openpyxl.Workbook()
ws2 = wb_new2.active
ws2.title = 'Sheet1'

title2_val = ws_src2.cell(1, 1).value or 'Allocation - PT LBM JAKARTA - June 2026'
ws2.cell(1, 1, title2_val)
ws2.cell(1, 1).font = Font(bold=True, size=12, name='맑은 고딕')
ws2.merge_cells('A1:M1')

write_header_row(ws2, 6, WOO_HEADERS, WOO_WIDTHS)

row_num2 = 7
for d in woo_rows:
    qty = float(d['qty'] or 5)
    gw = float(d['gw'] or qty * 1.026)
    balance = float(d['balance'] or qty)

    # 본품 행
    write_data_row(ws2, row_num2, [
        d['product'], d['sap'], d['date'], qty, d['lot'],
        d['wh'], d['customs'], d['export'], d['sold_to'], d['sale_ref'],
        balance, round(gw, 3), d['remark'] or ''
    ])
    row_num2 += 1

    # 샘플 행
    sp = str(d['product'] or 'MIC9000')
    if 'sample' not in sp.lower():
        sp += ' Sample'
    write_data_row(ws2, row_num2, [
        sp, d['sap'], d['date'], 0.001, d['lot'],
        d['wh'], d['customs'], d['export'], d['sold_to'], d['sale_ref'],
        0.001, 0.00125, ''
    ], is_sample=True)
    row_num2 += 1

ws2.freeze_panes = 'A7'
wb_new2.save('Allocation_Woo_202606.xlsx')
print(f"   Woo 완료: {len(woo_rows)} LOT x 2 = {len(woo_rows)*2} 행")


# ============================================================
# 3. Song 템플릿
# ============================================================
print("3. Song 템플릿 생성...")
wb_t1 = openpyxl.Workbook()
ws_t1 = wb_t1.active
ws_t1.title = 'Allocation'
ws_t1.cell(1, 1, 'Allocation - [고객명] - [월] [연도]')
ws_t1.cell(1, 1).font = Font(bold=True, size=12, name='맑은 고딕')
ws_t1.merge_cells('A1:J1')

write_header_row(ws_t1, 2, SONG_HEADERS, SONG_WIDTHS)

tmpl_data = [
    ('MIC9000', '2200032713', '2025-09-05', 5, '1125062056', 'GY', 'uncleared', 'LBM AP - January 550MT', '2903', 5.13),
    ('MIC9000', '2200032713', '2025-09-05', 5, '1125062057', 'GY', 'uncleared', 'LBM AP - January 550MT', '2903', 5.13),
    ('MIC9000', '2200032713', '2025-09-05', 5, '1125062058', 'GY', 'uncleared', 'LBM AP - January 550MT', '2903', 5.13),
]
r = 3
for sd in tmpl_data:
    write_data_row(ws_t1, r, list(sd))
    r += 1
    sv = list(sd)
    sv[0] = 'MIC9000 sample'
    sv[3] = 0.001
    sv[9] = 0.00125
    write_data_row(ws_t1, r, sv, is_sample=True)
    r += 1

ws_t1.freeze_panes = 'A3'
wb_t1.save('Template_Song_Allocation.xlsx')
print("   Template_Song 완료")


# ============================================================
# 4. Woo 템플릿
# ============================================================
print("4. Woo 템플릿 생성...")
wb_t2 = openpyxl.Workbook()
ws_t2 = wb_t2.active
ws_t2.title = 'Sheet1'
ws_t2.cell(1, 1, 'Allocation - [고객명] - [월] [연도]')
ws_t2.cell(1, 1).font = Font(bold=True, size=12, name='맑은 고딕')
ws_t2.merge_cells('A1:M1')

write_header_row(ws_t2, 6, WOO_HEADERS, WOO_WIDTHS)

tmpl_data2 = [
    ('MIC9000', '2200032902', '2025-10-13', 5, '1125080535', 'GY', 'Uncleared', '반송', 'LBM AP Q1 2026', '3184', 5, 5.13, ''),
    ('MIC9000', '2200032902', '2025-10-13', 5, '1125080536', 'GY', 'Uncleared', '반송', 'LBM AP Q1 2026', '3184', 5, 5.13, ''),
    ('MIC9000', '2200032902', '2025-10-13', 5, '1125080537', 'GY', 'Uncleared', '반송', 'LBM AP Q1 2026', '3184', 5, 5.13, ''),
]
r = 7
for sd in tmpl_data2:
    write_data_row(ws_t2, r, list(sd))
    r += 1
    sv = list(sd)
    sv[0] = 'MIC9000 Sample'
    sv[3] = 0.001
    sv[10] = 0.001
    sv[11] = 0.00125
    write_data_row(ws_t2, r, sv, is_sample=True)
    r += 1

ws_t2.freeze_panes = 'A7'
wb_t2.save('Template_Woo_Allocation.xlsx')
print("   Template_Woo 완료")

print("\n=== 모든 파일 생성 완료 ===")
print("  Allocation_Song_202607.xlsx  (88 LOT x 2 = 176행)")
print("  Allocation_Woo_202606.xlsx   (88 LOT x 2 = 176행)")
print("  Template_Song_Allocation.xlsx (3 LOT 샘플)")
print("  Template_Woo_Allocation.xlsx  (3 LOT 샘플)")
