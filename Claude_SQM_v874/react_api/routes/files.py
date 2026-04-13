# -*- coding: utf-8 -*-
"""POST /api/files/upload — 파일 업로드 + 파싱 엔드포인트.

file_type 파라미터:
  - 'allocation' : AllocationParser (6개 고객 양식 자동 인식)
  - 'inbound'    : 입고 Excel 파싱
  - 미지정/기타   : 확장자 자동 감지
"""
import os
import tempfile
import logging
from typing import Optional
from fastapi import APIRouter, UploadFile, File, Form

from react_api.schemas.write_models import WriteResponse
from react_api.utils.db import now_str

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/files", tags=["files"])

ALLOWED_EXTENSIONS = {'.pdf', '.xlsx', '.xls', '.csv'}
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB


# ── 메인 업로드 엔드포인트 ──────────────────────────────────────────────────────
@router.post("/upload", response_model=WriteResponse)
async def file_upload(
    file:      UploadFile     = File(...),
    file_type: Optional[str] = Form(default=None),
) -> WriteResponse:
    """파일 업로드 → file_type에 따라 파서 분기.

    file_type:
      'allocation' → AllocationParser (6개 양식 자동 인식)
      'inbound'    → 입고 Excel 파싱
      미지정       → 확장자 자동 판별
    """
    _, ext = os.path.splitext(file.filename or '')
    ext = ext.lower()
    if ext not in ALLOWED_EXTENSIONS:
        return WriteResponse(
            success=False,
            message=f"지원하지 않는 파일 형식: {ext} (허용: {', '.join(ALLOWED_EXTENSIONS)})",
            data={},
        )

    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        return WriteResponse(
            success=False,
            message=f"파일 크기 초과: {len(content) // (1024*1024)}MB (최대: 50MB)",
            data={},
        )

    tmp_dir  = tempfile.mkdtemp(prefix="sqm_upload_")
    tmp_path = os.path.join(tmp_dir, os.path.basename(file.filename or f"upload{ext}"))
    try:
        with open(tmp_path, 'wb') as f:
            f.write(content)

        # ── file_type 분기 ──────────────────────────────────────────────────
        ft = (file_type or '').strip().lower()
        if ft == 'allocation':
            result = _parse_allocation(tmp_path, file.filename or '')
        elif ft == 'inbound':
            result = _parse_inbound_excel(tmp_path)
        elif ext == '.pdf':
            result = _parse_pdf(tmp_path)
        elif ext in ('.xlsx', '.xls'):
            result = _parse_excel(tmp_path)
        elif ext == '.csv':
            result = _parse_csv(tmp_path)
        else:
            result = {'success': False, 'message': f'파서 없음: {ext}', 'data': {}}

        return WriteResponse(**result)

    except Exception as e:
        logger.exception("파일 처리 실패")
        return WriteResponse(
            success=False,
            message=f"파일 처리 실패: {str(e)}",
            data={},
        )
    finally:
        try:
            if os.path.exists(tmp_path): os.remove(tmp_path)
            if os.path.exists(tmp_dir):  os.rmdir(tmp_dir)
        except OSError:
            pass


# ── Allocation 파싱 (핵심 추가) ───────────────────────────────────────────────
def _parse_allocation(file_path: str, filename: str) -> dict:
    """
    AllocationParser v2.7.1 호출 — 6개 고객 양식 자동 인식.
    Song / Woo / Standard / Shipper Original / Easpring / Jakarta
    """
    try:
        from parsers.allocation_parser import AllocationParser
        parser = AllocationParser()
        data   = parser.parse(file_path)

        if data is None:
            return {
                'success':  False,
                'message':  'Allocation 파싱 실패 — 지원하지 않는 양식이거나 헤더를 찾을 수 없습니다.',
                'errors':   [],
                'warnings': [],
                'data':     {'rows': [], 'file_type': 'ALLOCATION'},
            }

        if not data.success:
            return {
                'success':  False,
                'message':  f'Allocation 파싱 실패: {"; ".join(data.errors)}',
                'errors':   data.errors,
                'warnings': [],
                'data':     {'rows': [], 'file_type': 'ALLOCATION'},
            }

        # AllocationRow → 프론트 전달용 dict 변환
        rows = []
        for r in data.rows:
            outbound_date = ''
            if r.outbound_date:
                try:
                    outbound_date = r.outbound_date.strftime('%Y-%m-%d')
                except Exception:
                    outbound_date = str(r.outbound_date)

            rows.append({
                'lot_no':        r.lot_no        or '',
                'sap_no':        r.sap_no        or '',
                'product':       r.product       or '',
                'customer':      r.sold_to       or '',
                'sale_ref':      r.sale_ref      or '',
                'qty_mt':        round(float(r.qty_mt or 0), 3),
                'outbound_date': outbound_date,
                'warehouse':     r.warehouse     or '',
                'sub_lt':        int(r.sub_lt    or 0),
                'is_sample':     bool(r.is_sample),
                'export_type':   r.export_type   or '',
                'gross_weight':  round(float(r.gross_weight or 0), 3),
            })

        # 헤더 정보
        header = data.header
        customer_name = getattr(header, 'customer', '') if header else ''
        sale_ref_hdr  = getattr(header, 'sale_ref',  '') if header else ''

        total_qty = round(float(data.total_qty or 0), 3)

        logger.info(
            "[AllocationParser] %s — %d행, 합계 %.3f MT, 고객: %s",
            filename, len(rows), total_qty, customer_name
        )

        return {
            'success':  True,
            'message':  f'Allocation 파싱 완료: {len(rows)}행 / 합계 {total_qty} MT',
            'errors':   [],
            'warnings': [],
            'data': {
                'file_type':    'ALLOCATION',
                'rows':         rows,
                'total_rows':   len(rows),
                'total_qty_mt': total_qty,
                'customer':     customer_name,
                'sale_ref':     sale_ref_hdr,
                'source_file':  filename,
            },
        }

    except ImportError as exc:
        logger.error("AllocationParser import 실패: %s", exc)
        return {
            'success':  False,
            'message':  f'AllocationParser 모듈 없음: {exc}',
            'errors':   [str(exc)],
            'warnings': [],
            'data':     {'rows': [], 'file_type': 'ALLOCATION'},
        }
    except Exception as exc:
        logger.error("Allocation 파싱 오류: %s", exc, exc_info=True)
        return {
            'success':  False,
            'message':  f'Allocation 파싱 오류: {exc}',
            'errors':   [str(exc)],
            'warnings': [],
            'data':     {'rows': [], 'file_type': 'ALLOCATION'},
        }


# ── 입고 Excel 파싱 ───────────────────────────────────────────────────────────
def _parse_inbound_excel(file_path: str) -> dict:
    """입고용 Excel — pandas로 읽어서 컬럼 자동 매핑."""
    try:
        import pandas as pd
        df = pd.read_excel(file_path, dtype=str).fillna('')
        headers = list(df.columns)
        rows    = df.head(20).to_dict('records')
        return {
            'success': True,
            'message': f'입고 Excel 파싱 완료 ({len(df)}행)',
            'data': {
                'file_type': 'INBOUND_EXCEL',
                'headers':   headers,
                'rows':      rows,
                'total_rows': len(df),
            },
        }
    except Exception as exc:
        return {'success': False, 'message': f'입고 Excel 파싱 실패: {exc}', 'data': {}}


# ── PDF 파싱 ──────────────────────────────────────────────────────────────────
def _parse_pdf(file_path: str) -> dict:
    """PDF → DocumentParserV3 (BL/PL/FA/DO 자동 감지)."""
    try:
        from parsers.document_parser_modular.parser import DocumentParserV3
        parser = DocumentParserV3()
        result = parser.parse_document(file_path)
        return {
            'success': True,
            'message': 'PDF 파싱 완료',
            'data': {
                'file_type': 'PDF',
                'parsed': result if isinstance(result, dict) else {'raw': str(result)},
            },
        }
    except ImportError:
        try:
            from parsers.pdf_parser import parse_pdf
            result = parse_pdf(file_path)
            return {
                'success': True,
                'message': 'PDF 파싱 완료 (fallback)',
                'data': {'file_type': 'PDF', 'parsed': result if isinstance(result, dict) else {'raw': str(result)}},
            }
        except Exception as exc:
            return {'success': False, 'message': f'PDF 파싱 실패: {exc}', 'data': {}}
    except Exception as exc:
        return {'success': False, 'message': f'PDF 파싱 실패: {exc}', 'data': {}}


# ── 일반 Excel 파싱 ───────────────────────────────────────────────────────────
def _parse_excel(file_path: str) -> dict:
    """일반 Excel → 시트별 미리보기."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheets_data = {}
        for sheet_name in wb.sheetnames:
            ws   = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(cell) if cell is not None else '' for cell in row])
            sheets_data[sheet_name] = {
                'row_count': len(rows),
                'headers':   rows[0] if rows else [],
                'preview':   rows[:10],
            }
        wb.close()
        return {
            'success': True,
            'message': f'Excel 파싱 완료 ({len(sheets_data)} 시트)',
            'data': {'file_type': 'EXCEL', 'sheets': sheets_data},
        }
    except Exception as exc:
        return {'success': False, 'message': f'Excel 파싱 실패: {exc}', 'data': {}}


# ── CSV 파싱 ──────────────────────────────────────────────────────────────────
def _parse_csv(file_path: str) -> dict:
    """CSV → 미리보기."""
    try:
        import csv
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            rows = list(csv.reader(f))
        return {
            'success': True,
            'message': f'CSV 파싱 완료 ({len(rows)} 행)',
            'data': {
                'file_type': 'CSV',
                'row_count': len(rows),
                'headers':   rows[0] if rows else [],
                'preview':   rows[:10],
            },
        }
    except Exception as exc:
        return {'success': False, 'message': f'CSV 파싱 실패: {exc}', 'data': {}}
