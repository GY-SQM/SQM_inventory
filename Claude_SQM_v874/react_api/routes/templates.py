# -*- coding: utf-8 -*-
"""4개 미구현 기능 API.

1. /api/templates/inbound  — 입고 파싱 템플릿 CRUD
2. /api/templates/picking  — 피킹 템플릿 CRUD
3. /api/move-approval      — 대량 이동 승인
4. /api/reports/swap       — Swap 리포트
"""
import logging
from typing import Optional
from fastapi import APIRouter, HTTPException, Query, BackgroundTasks
from fastapi.responses import FileResponse

from react_api.utils.db import get_db, get_engine, now_str

logger = logging.getLogger(__name__)

# 라우터 3개 (reports는 기존 파일에 추가 대신 여기서 별도 prefix)
inbound_tpl_router = APIRouter(prefix="/api/templates/inbound", tags=["templates"])
picking_tpl_router = APIRouter(prefix="/api/templates/picking",  tags=["templates"])
move_appr_router   = APIRouter(prefix="/api/move-approval",      tags=["move-approval"])
swap_router        = APIRouter(prefix="/api/swap",               tags=["swap"])


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 1. 입고 파싱 템플릿 CRUD
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

INBOUND_COLS = [
    'template_id', 'template_name', 'carrier_id', 'bag_weight_kg',
    'product_hint', 'weight_format', 'bl_format',
    'gemini_hint_packing', 'gemini_hint_invoice', 'gemini_hint_bl',
    'note', 'is_active',
]


@inbound_tpl_router.get("/list")
def inbound_tpl_list(active_only: bool = Query(True)):
    """입고 파싱 템플릿 목록 조회."""
    try:
        with get_db() as db:
            # 테이블 없으면 빈 목록 반환
            tbl = db.fetchone(
                "SELECT name FROM sqlite_master WHERE type='table' AND name='inbound_template'"
            )
            if not tbl:
                return {"success": True, "rows": [], "total": 0, "generated_at": now_str()}
            sql = (
                f"SELECT {','.join(INBOUND_COLS)} FROM inbound_template "
                + ("WHERE is_active=1 " if active_only else "")
                + "ORDER BY carrier_id, bag_weight_kg"
            )
            rows = db.fetchall(sql)
            data = []
            for r in (rows or []):
                data.append(dict(r))
            return {"success": True, "rows": data, "total": len(data), "generated_at": now_str()}
    except Exception as exc:
        logger.error("inbound_tpl_list 실패: %s", exc, exc_info=True)
        raise HTTPException(500, f"템플릿 조회 실패: {exc}")


@inbound_tpl_router.post("/save")
def inbound_tpl_save(payload: dict):
    """입고 파싱 템플릿 저장 (신규/수정).
    payload: { template_id, template_name, carrier_id, bag_weight_kg, ... }
    """
    try:
        with get_engine() as engine:
            from gui_app_modular.dialogs.inbound_template_dialog import save_template
            ok = save_template(engine, payload)
        if not ok:
            raise HTTPException(500, "템플릿 저장 실패")
        return {"success": True, "message": f"템플릿 '{payload.get('template_name','')}' 저장 완료"}
    except HTTPException:
        raise
    except ImportError:
        # 직접 DB 저장 fallback
        try:
            with get_db() as db:
                db.execute("""
                    INSERT OR REPLACE INTO inbound_template (
                        template_id, template_name, carrier_id, bag_weight_kg,
                        product_hint, weight_format, bl_format,
                        gemini_hint_packing, gemini_hint_invoice, gemini_hint_bl,
                        note, is_active
                    ) VALUES (
                        :template_id, :template_name, :carrier_id, :bag_weight_kg,
                        :product_hint, :weight_format, :bl_format,
                        :gemini_hint_packing, :gemini_hint_invoice, :gemini_hint_bl,
                        :note, :is_active
                    )
                """, payload)
            return {"success": True, "message": "템플릿 저장 완료"}
        except Exception as exc:
            raise HTTPException(500, f"저장 실패: {exc}")
    except Exception as exc:
        logger.error("inbound_tpl_save 실패: %s", exc, exc_info=True)
        raise HTTPException(500, f"저장 실패: {exc}")


@inbound_tpl_router.delete("/{template_id}")
def inbound_tpl_delete(template_id: str):
    """입고 파싱 템플릿 삭제."""
    try:
        with get_db() as db:
            db.execute(
                "UPDATE inbound_template SET is_active=0 WHERE template_id=?",
                (template_id,)
            )
        return {"success": True, "message": f"템플릿 '{template_id}' 비활성화 완료"}
    except Exception as exc:
        raise HTTPException(500, f"삭제 실패: {exc}")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 2. 피킹 템플릿 CRUD
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

PICKING_COLS = [
    'template_id', 'template_name', 'customer', 'customer_code',
    'port_loading', 'port_discharge', 'delivery_terms',
    'contact_person', 'contact_email',
    'bag_weight_kg', 'storage_location', 'note', 'is_active',
]


@picking_tpl_router.get("/list")
def picking_tpl_list(active_only: bool = Query(True)):
    """피킹 템플릿 목록 조회."""
    try:
        with get_db() as db:
            tbl = db.fetchone(
                "SELECT name FROM sqlite_master WHERE type='table' AND name='picking_template'"
            )
            if not tbl:
                return {"success": True, "rows": [], "total": 0, "generated_at": now_str()}
            sql = (
                f"SELECT {','.join(PICKING_COLS)} FROM picking_template "
                + ("WHERE is_active=1 " if active_only else "")
                + "ORDER BY customer"
            )
            rows = db.fetchall(sql)
            data = [
                dict(r)
                for r in (rows or [])
            ]
            return {"success": True, "rows": data, "total": len(data), "generated_at": now_str()}
    except Exception as exc:
        logger.error("picking_tpl_list 실패: %s", exc, exc_info=True)
        raise HTTPException(500, f"피킹 템플릿 조회 실패: {exc}")


@picking_tpl_router.post("/save")
def picking_tpl_save(payload: dict):
    """피킹 템플릿 저장."""
    try:
        with get_engine() as engine:
            from gui_app_modular.dialogs.picking_template_dialog import save_template
            ok = save_template(engine, payload)
        if not ok:
            raise HTTPException(500, "피킹 템플릿 저장 실패")
        return {"success": True, "message": f"템플릿 '{payload.get('template_name','')}' 저장 완료"}
    except HTTPException:
        raise
    except ImportError:
        try:
            with get_db() as db:
                cols_no_active = [c for c in PICKING_COLS if c != 'is_active']
                placeholders   = ', '.join(f':{c}' for c in cols_no_active)
                db.execute(
                    f"INSERT OR REPLACE INTO picking_template ({','.join(cols_no_active)}, is_active) "
                    f"VALUES ({placeholders}, :is_active)",
                    payload
                )
            return {"success": True, "message": "피킹 템플릿 저장 완료"}
        except Exception as exc:
            raise HTTPException(500, f"저장 실패: {exc}")
    except Exception as exc:
        logger.error("picking_tpl_save 실패: %s", exc, exc_info=True)
        raise HTTPException(500, f"저장 실패: {exc}")


@picking_tpl_router.delete("/{template_id}")
def picking_tpl_delete(template_id: str):
    """피킹 템플릿 삭제."""
    try:
        if template_id == 'UNKNOWN_CUSTOMER':
            raise HTTPException(400, "기본 템플릿은 삭제할 수 없습니다.")
        with get_db() as db:
            db.execute(
                "DELETE FROM picking_template WHERE template_id=?",
                (template_id,)
            )
        return {"success": True, "message": f"피킹 템플릿 '{template_id}' 삭제 완료"}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(500, f"삭제 실패: {exc}")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 3. 대량 이동 승인
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@move_appr_router.get("/pending")
def move_approval_pending():
    """PENDING 대량 이동 요청 목록."""
    try:
        with get_engine() as engine:
            rows = engine.get_pending_batch_moves()
        data = []
        for r in (rows or []):
            row = dict(r)
            data.append(row)
        return {"success": True, "rows": data, "total": len(data), "generated_at": now_str()}
    except Exception as exc:
        logger.error("move_approval_pending 실패: %s", exc, exc_info=True)
        raise HTTPException(500, f"대기 목록 조회 실패: {exc}")


@move_appr_router.post("/approve")
def move_approval_approve(payload: dict):
    """대량 이동 승인.  payload: { batch_id: str, approver?: str }"""
    batch_id = str(payload.get("batch_id", "")).strip()
    approver = str(payload.get("approver", "react_ui")).strip()
    if not batch_id:
        raise HTTPException(400, "batch_id 필수")
    try:
        with get_engine() as engine:
            # move_batch 테이블 존재 확인
            tbl = engine.db.fetchone(
                "SELECT name FROM sqlite_master WHERE type='table' AND name='move_batch'"
            )
            if not tbl:
                raise HTTPException(404, "move_batch 테이블 없음 — 대량이동 기능 미사용 중")
            result = engine.approve_batch_move(batch_id=batch_id, approver=approver)
        return {
            "success":  result.get("success", False),
            "applied":  result.get("applied", 0),
            "skipped":  result.get("skipped", 0),
            "message":  f"승인 완료 — {result.get('applied', 0)}건 반영",
            "errors":   result.get("errors", []),
            "generated_at": now_str(),
        }
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("move_approval_approve 실패: %s", exc, exc_info=True)
        raise HTTPException(500, f"승인 처리 실패: {exc}")


@move_appr_router.post("/reject")
def move_approval_reject(payload: dict):
    """대량 이동 반려.  payload: { batch_id: str, reason?: str }"""
    batch_id = str(payload.get("batch_id", "")).strip()
    reason   = str(payload.get("reason", "반려")).strip()
    if not batch_id:
        raise HTTPException(400, "batch_id 필수")
    try:
        with get_engine() as engine:
            tbl = engine.db.fetchone(
                "SELECT name FROM sqlite_master WHERE type='table' AND name='move_batch'"
            )
            if not tbl:
                raise HTTPException(404, "move_batch 테이블 없음 — 대량이동 기능 미사용 중")
            result = engine.reject_batch_move(
                batch_id=batch_id, rejector="react_ui", reason=reason
            )
        return {
            "success": result.get("success", False),
            "message": f"반려 완료: {batch_id}",
            "generated_at": now_str(),
        }
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("move_approval_reject 실패: %s", exc, exc_info=True)
        raise HTTPException(500, f"반려 처리 실패: {exc}")


@move_appr_router.get("/history")
def move_approval_history(
    page:      int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
):
    """대량 이동 처리 이력 (APPROVED / REJECTED)."""
    try:
        with get_db() as db:
            tbl = db.fetchone(
                "SELECT name FROM sqlite_master WHERE type='table' AND name='move_batch'"
            )
            if not tbl:
                return {"success": True, "rows": [], "total": 0, "generated_at": now_str()}
            offset = (page - 1) * page_size
            total  = int((db.fetchone(
                "SELECT COUNT(*) AS c FROM move_batch WHERE status IN ('APPROVED','REJECTED')"
            ) or {}).get("c") or 0)
            rows = db.fetchall(
                "SELECT batch_id, total_count, reason_code, submitted_by, submitted_at, "
                "status, note FROM move_batch "
                "WHERE status IN ('APPROVED','REJECTED') "
                "ORDER BY submitted_at DESC LIMIT ? OFFSET ?",
                (page_size, offset)
            )
            data = [dict(r) for r in (rows or [])]
            return {"success": True, "rows": data, "total": total, "generated_at": now_str()}
    except Exception as exc:
        raise HTTPException(500, f"이력 조회 실패: {exc}")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 4. Swap 리포트
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@swap_router.get("/list")
def swap_list(
    lot_no:    Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to:   Optional[str] = Query(None),
    page:      int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
):
    """Swap(UID 교체) 이력 조회."""
    try:
        with get_db() as db:
            # uid_swap_report 또는 stock_movement에서 SWAP 유형 조회
            tbl = db.fetchone(
                "SELECT name FROM sqlite_master WHERE type='table' AND name='uid_swap_report'"
            )
            offset = (page - 1) * page_size

            if tbl:
                conditions = ["1=1"]
                params: list = []
                if lot_no:
                    conditions.append("(old_lot_no LIKE ? OR new_lot_no LIKE ?)")
                    params.extend([f"%{lot_no}%", f"%{lot_no}%"])
                if date_from:
                    conditions.append("swap_date >= ?"); params.append(date_from)
                if date_to:
                    conditions.append("swap_date <= ?"); params.append(date_to)
                where = " AND ".join(conditions)
                total = int((db.fetchone(
                    f"SELECT COUNT(*) AS c FROM uid_swap_report WHERE {where}",
                    tuple(params)
                ) or {}).get("c") or 0)
                rows = db.fetchall(
                    f"SELECT * FROM uid_swap_report WHERE {where} "
                    f"ORDER BY swap_date DESC LIMIT ? OFFSET ?",
                    tuple(params + [page_size, offset])
                )
            else:
                # stock_movement에서 SWAP 유형으로 대체
                conditions = ["movement_type LIKE '%SWAP%'"]
                params = []
                if lot_no:
                    conditions.append("lot_no LIKE ?"); params.append(f"%{lot_no}%")
                if date_from:
                    conditions.append("created_at >= ?"); params.append(date_from)
                if date_to:
                    conditions.append("created_at <= ?"); params.append(date_to + " 23:59:59")
                where = " AND ".join(conditions)
                total = int((db.fetchone(
                    f"SELECT COUNT(*) AS c FROM stock_movement WHERE {where}",
                    tuple(params)
                ) or {}).get("c") or 0)
                rows = db.fetchall(
                    f"SELECT lot_no, description, qty_kg, created_at FROM stock_movement "
                    f"WHERE {where} ORDER BY created_at DESC LIMIT ? OFFSET ?",
                    tuple(params + [page_size, offset])
                )

            data = [dict(r) for r in (rows or [])]
            return {"success": True, "rows": data, "total": total,
                    "page": page, "generated_at": now_str()}
    except Exception as exc:
        logger.error("swap_list 실패: %s", exc, exc_info=True)
        raise HTTPException(500, f"Swap 목록 조회 실패: {exc}")


@swap_router.get("/download")
def swap_download(background_tasks: BackgroundTasks,
    lot_no:    Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to:   Optional[str] = Query(None),
):
    """Swap 리포트 Excel 다운로드."""
    import tempfile, os
    try:
        with get_db() as db:
            tbl = db.fetchone(
                "SELECT name FROM sqlite_master WHERE type='table' AND name='uid_swap_report'"
            )
            if tbl:
                rows = db.fetchall("SELECT * FROM uid_swap_report ORDER BY swap_date DESC") or []
                headers = ['old_uid', 'new_uid', 'old_lot_no', 'new_lot_no',
                           'reason', 'operator', 'swap_date']
            else:
                rows = db.fetchall(
                    "SELECT lot_no, description, qty_kg, created_at "
                    "FROM stock_movement WHERE movement_type LIKE '%SWAP%' "
                    "ORDER BY created_at DESC"
                ) or []
                headers = ['lot_no', 'description', 'qty_kg', 'created_at']

        if not rows:
            raise HTTPException(404, "Swap 이력이 없습니다.")

        import openpyxl
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Swap 리포트"
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F497D")

        for r in rows:
            row = dict(r)
            if row:
                ws.append([row.get(h, '') for h in headers])
            else:
                ws.append(row)

        fd, out_path = tempfile.mkstemp(suffix=".xlsx", prefix="sqm_swap_")
        os.close(fd)
        wb.save(out_path)

        from fastapi.responses import FileResponse
        return FileResponse(
            path=out_path,
            filename="SQM_Swap_리포트.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("swap_download 실패: %s", exc, exc_info=True)
        raise HTTPException(500, f"Swap 리포트 다운로드 실패: {exc}")
