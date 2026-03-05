"""
원본: D:\\프로그램\\Sqm 재고관리\\SQM-Inventory-2026_02_21.xlsx (또는 동일 이름의 재고리스트 엑셀)
→ Allocation - GY - PT LBM 300MT (2)-1.xlsx 형식으로 3개 생성

규칙:
- 데이터 소스: SQM-Inventory-2026_02_21.xlsx (재고리스트 시트에서 LOT NO, SAP NO, PRODUCT, WH, CUSTOMS 읽음)
- 한 파일당: 톤백 200개 (500kg = 0.5 MT), 샘플 20행 (각 QTY (MT) = 0.001)
- LOT 20개에 톤백 분배, 롯트당 2~10개 (행당 최대 5 MT = 톤백 10개만 허용)
- QTY (MT): 5 = 500kg 톤백 10개. 톤백 N개 → N*0.5 MT. 절대 행당 10 MT(20개) 금지.
- 출력: generated_allocation/ 또는 인자로 지정한 폴더
"""
import random
import sys
from pathlib import Path

import pandas as pd

# 원본: SQM-Inventory-2026_02_21.xlsx (D:\프로그램\Sqm 재고관리\ 아래 또는 인자로 경로 지정)
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
DEFAULT_SOURCE = PROJECT_ROOT.parent / "SQM-Inventory-2026_02_21.xlsx"
OUTPUT_DIR_DEFAULT = PROJECT_ROOT / "generated_allocation"
OUTPUT_NAME_TEMPLATE = "Allocation - GY - PT LBM 300MT (2)-{}.xlsx"

# 파일당: 20 LOT, 총 200 톤백 (500kg each). 롯트당 2~10개만 허용 → 행당 최대 5 MT
TOTAL_TONBAGS = 200
NUM_LOTS = 20
MIN_PER_LOT = 2
MAX_PER_LOT = 10   # 10개 = 5 MT (행당 5 초과 금지)
MAX_TONBAGS_PER_ROW = 10  # 500kg 톤백 10개 = 5 MT
NUM_SAMPLES = 20   # 샘플 20행 (각 0.001 MT)
MT_PER_TONBAG = 0.5   # 500kg = 0.5 MT
SAMPLE_MT = 0.001
GW_PER_TONBAG_MT = 0.513  # 5.13 / 10 (10 bags = 5 MT = 5.13 GW)


def _read_lots_from_inventory(excel_path: Path) -> list:
    """SQM-Inventory 엑셀에서 LOT 목록 읽기 (LOT NO, SAP NO, PRODUCT, WH, CUSTOMS 등)."""
    xl = pd.ExcelFile(excel_path)
    sheet = "재고리스트" if "재고리스트" in xl.sheet_names else xl.sheet_names[0]
    df_raw = pd.read_excel(excel_path, sheet_name=sheet, header=None)

    def _norm(s):
        return str(s).strip().upper().replace(" ", "") if pd.notna(s) else ""

    header_row = 0
    for i in range(min(5, len(df_raw))):
        row = df_raw.iloc[i]
        if any("LOT" in _norm(v) for v in row):
            header_row = i
            break

    df = pd.read_excel(excel_path, sheet_name=sheet, header=header_row)
    df = df.dropna(how="all")
    df.columns = [str(c).strip() for c in df.columns]

    lot_col = next((c for c in df.columns if "LOT" in c.upper() and "NO" in c.upper()), None)
    if lot_col is None:
        lot_col = next((c for c in df.columns if c.upper().startswith("LOT")), None)
    if lot_col is None:
        raise ValueError("LOT NO 컬럼을 찾을 수 없습니다.")

    sap_col = next((c for c in df.columns if "SAP" in c.upper()), None)
    product_col = next((c for c in df.columns if "PRODUCT" in c.upper()), None)
    wh_col = next((c for c in df.columns if c.strip().upper() == "WH"), None)
    customs_col = next((c for c in df.columns if "CUSTOMS" in c.upper()), None)

    lots = []
    seen = set()
    for _, r in df.iterrows():
        lot_no = str(r.get(lot_col, "")).strip()
        if not lot_no or lot_no in seen:
            continue
        try:
            if str(lot_no).upper().startswith("LOT") and len(lot_no) < 15:
                continue
        except (ValueError, TypeError):
            pass  # LOT 번호 형식 필터링 — 의도적 skip
        seen.add(lot_no)
        lots.append({
            "lot_no": lot_no,
            "sap_no": str(r.get(sap_col, "")).strip() if sap_col else "",
            "product": str(r.get(product_col, "LITHIUM CARBONATE")).strip() or "LITHIUM CARBONATE",
            "wh": str(r.get(wh_col, "광양")).strip() or "광양",
            "customs": str(r.get(customs_col, "Cleared")).strip() or "Cleared",
        })
    return lots


def _distribute_tonbags(total: int, num_lots: int, min_per: int, max_per: int, cap_per_lot: int = 10) -> list:
    """총 total개 톤백을 num_lots개 LOT에 분배. 롯트당 cap_per_lot(기본 10) 초과 금지 → QTY(MT) 행당 최대 5."""
    base = [random.randint(min_per, min(max_per, cap_per_lot)) for _ in range(num_lots)]
    s = sum(base)
    if s < total:
        need = total - s
        idx = num_lots - 1
        while need > 0 and idx >= 0:
            add = min(need, cap_per_lot - base[idx])
            base[idx] += add
            need -= add
            idx -= 1
        if need > 0:
            base[0] += min(need, cap_per_lot - base[0])
    elif s > total:
        need = s - total
        for i in range(num_lots):
            sub = min(need, base[i] - min_per)
            base[i] -= sub
            need -= sub
            if need <= 0:
                break
        if need > 0:
            base[-1] -= min(need, base[-1] - min_per)
    return base


def _create_allocation_excel(
    output_path: Path,
    lots: list,
    tonbags_per_lot: list,
    num_samples: int,
    file_index: int,
) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Allocation Table"

    title_font = Font(bold=True, size=14, color="2C3E50")
    header_font = Font(bold=True, color="2C3E50", size=10)
    data_font = Font(size=10)
    header_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    header_yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    total_mt = sum(n * MT_PER_TONBAG for n in tonbags_per_lot) + num_samples * SAMPLE_MT
    ws.merge_cells("A1:J1")
    ws["A1"] = f"Allocation - GY - PT LBM 300MT (2)-{file_index} — {total_mt:.2f} MT"
    ws["A1"].font = title_font
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 6

    headers = [
        ("Product", 16), ("SAP NO", 14), ("ETA BUSAN", 14), ("Date in stock", 14),
        ("QTY (MT)", 12), ("Lot No", 14), ("WH", 8), ("Customs", 12),
        ("GW", 12), ("SALE REF", 12),
    ]
    yellow_cols = (3, 8, 9)
    for col, (text, width) in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=text)
        cell.font = header_font
        cell.fill = header_yellow if col in yellow_cols else header_fill
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width

    dates = ["2026-02-20", "2026-03-24", "2026-03-10", "2026-02-15", "2026-04-01"]
    sale_ref = "1955"
    row_num = 4

    for i, (lot_info, num_tb) in enumerate(zip(lots, tonbags_per_lot)):
        qty_mt = round(num_tb * MT_PER_TONBAG, 3)
        gw = round(num_tb * GW_PER_TONBAG_MT, 3)
        date_stock = dates[i % len(dates)]
        vals = [
            lot_info["product"],
            lot_info["sap_no"] or "",
            "",
            date_stock,
            qty_mt,
            lot_info["lot_no"],
            lot_info["wh"],
            lot_info["customs"],
            gw,
            sale_ref,
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = data_font
            cell.border = thin_border
            if col in (5, 9):
                cell.number_format = "#,##0.000"
                cell.alignment = right_align
            else:
                cell.alignment = center
        row_num += 1

    # 샘플 20행: QTY (MT) = 0.001
    sample_lots = lots[:num_samples] if len(lots) >= num_samples else (lots * ((num_samples // len(lots)) + 1))[:num_samples]
    for i, lot_info in enumerate(sample_lots):
        vals = [
            lot_info["product"],
            lot_info["sap_no"] or "",
            "",
            dates[i % len(dates)],
            SAMPLE_MT,
            lot_info["lot_no"],
            lot_info["wh"],
            lot_info["customs"],
            round(0.001 * (5.13 / 5), 4),
            sale_ref,
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = data_font
            cell.border = thin_border
            if col in (5, 9):
                cell.number_format = "#,##0.000"
                cell.alignment = right_align
            else:
                cell.alignment = center
        row_num += 1

    last_row = row_num - 1
    ws["E2"].value = f"=SUM(E4:E{last_row})"
    ws["E2"].number_format = "#,##0.0000"
    ws["E2"].font = Font(bold=True)

    wb.save(output_path)


def main():
    source = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SOURCE
    out_dir = Path(sys.argv[2]) if len(sys.argv) > 2 else OUTPUT_DIR_DEFAULT
    if not source.exists():
        print(f"원본 파일 없음: {source}")
        sys.exit(1)
    out_dir.mkdir(parents=True, exist_ok=True)

    lots = _read_lots_from_inventory(source)
    if len(lots) < NUM_LOTS:
        print(f"LOT 수 부족: {len(lots)}개 (필요: {NUM_LOTS}개). LOT 수를 {len(lots)}개로 생성합니다.")
        num_lots_use = len(lots)
    else:
        num_lots_use = NUM_LOTS

    random.seed(42)
    for file_idx in range(1, 4):
        selected = random.sample(lots, num_lots_use) if len(lots) >= num_lots_use else lots
        tonbags_per_lot = _distribute_tonbags(
            TOTAL_TONBAGS, num_lots_use, MIN_PER_LOT, MAX_PER_LOT,
            cap_per_lot=MAX_TONBAGS_PER_ROW,
        )
        out_path = out_dir / OUTPUT_NAME_TEMPLATE.format(file_idx)
        _create_allocation_excel(out_path, selected, tonbags_per_lot, NUM_SAMPLES, file_idx)
        total_tb = sum(tonbags_per_lot)
        max_mt = max(n * MT_PER_TONBAG for n in tonbags_per_lot)
        print(f"Create: {out_path.name} | source={source.name} | LOTs {num_lots_use}, tonbags {total_tb}, samples {NUM_SAMPLES} | max QTY(MT)/row={max_mt}")

    print(f"Done. 3 files in: {out_dir}")


if __name__ == "__main__":
    main()
