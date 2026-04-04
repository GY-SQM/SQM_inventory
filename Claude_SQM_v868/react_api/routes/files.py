# -*- coding: utf-8 -*-
"""POST /api/files/upload — 파일 업로드 + 파싱 엔드포인트."""
import os
import tempfile
import logging
from fastapi import APIRouter, UploadFile, File

from react_api.schemas.write_models import WriteResponse

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/files", tags=["files"])

ALLOWED_EXTENSIONS = {'.pdf', '.xlsx', '.xls', '.csv'}
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB


@router.post("/upload", response_model=WriteResponse)
async def file_upload(file: UploadFile = File(...)) -> WriteResponse:
    """파일 업로드 → 확장자 검증 → 임시 저장 → 파서 호출."""
    # 확장자 검증
    _, ext = os.path.splitext(file.filename or '')
    ext = ext.lower()
    if ext not in ALLOWED_EXTENSIONS:
        return WriteResponse(
            success=False,
            message=f"지원하지 않는 파일 형식: {ext} (허용: {', '.join(ALLOWED_EXTENSIONS)})",
            data={},
        )

    # 파일 크기 검증
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        return WriteResponse(
            success=False,
            message=f"파일 크기 초과: {len(content) // (1024*1024)}MB (최대: 50MB)",
            data={},
        )

    # 임시 파일 저장
    tmp_dir = tempfile.mkdtemp(prefix="sqm_upload_")
    tmp_path = os.path.join(tmp_dir, file.filename or f"upload{ext}")
    try:
        with open(tmp_path, 'wb') as f:
            f.write(content)

        # 파서 호출
        parse_result = _parse_file(tmp_path, ext)
        return WriteResponse(**parse_result)

    except Exception as e:
        logger.exception("파일 처리 실패")
        return WriteResponse(
            success=False,
            message=f"파일 처리 실패: {str(e)}",
            data={},
        )
    finally:
        # 임시 파일 정리
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
            if os.path.exists(tmp_dir):
                os.rmdir(tmp_dir)
        except OSError:
            pass


def _parse_file(file_path: str, ext: str) -> dict:
    """확장자별 파서 호출."""
    if ext == '.pdf':
        return _parse_pdf(file_path)
    elif ext in ('.xlsx', '.xls'):
        return _parse_excel(file_path)
    elif ext == '.csv':
        return _parse_csv(file_path)
    return {'success': False, 'message': f'파서 없음: {ext}', 'data': {}}


def _parse_pdf(file_path: str) -> dict:
    """PDF 파싱 — parsers 모듈 사용."""
    try:
        from parsers.document_parser_modular.parser import DocumentParser
        parser = DocumentParser()
        result = parser.parse_document(file_path)
        return {
            'success': True,
            'message': 'PDF 파싱 완료',
            'data': {
                'file_type': 'PDF',
                'parsed': result if isinstance(result, dict) else {'raw': str(result)},
            }
        }
    except ImportError:
        logger.warning("DocumentParser를 가져올 수 없음, 기본 PDF 파서 시도")
        try:
            from parsers.pdf_parser import parse_pdf
            result = parse_pdf(file_path)
            return {
                'success': True,
                'message': 'PDF 파싱 완료 (fallback)',
                'data': {
                    'file_type': 'PDF',
                    'parsed': result if isinstance(result, dict) else {'raw': str(result)},
                }
            }
        except Exception as e:
            return {'success': False, 'message': f'PDF 파싱 실패: {str(e)}', 'data': {}}


def _parse_excel(file_path: str) -> dict:
    """Excel 파싱."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheets_data = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(cell) if cell is not None else '' for cell in row])
            sheets_data[sheet_name] = {
                'row_count': len(rows),
                'headers': rows[0] if rows else [],
                'preview': rows[:10],
            }
        wb.close()
        return {
            'success': True,
            'message': f'Excel 파싱 완료 ({len(sheets_data)} 시트)',
            'data': {
                'file_type': 'EXCEL',
                'sheets': sheets_data,
            }
        }
    except Exception as e:
        return {'success': False, 'message': f'Excel 파싱 실패: {str(e)}', 'data': {}}


def _parse_csv(file_path: str) -> dict:
    """CSV 파싱."""
    try:
        import csv
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            rows = list(reader)
        return {
            'success': True,
            'message': f'CSV 파싱 완료 ({len(rows)} 행)',
            'data': {
                'file_type': 'CSV',
                'row_count': len(rows),
                'headers': rows[0] if rows else [],
                'preview': rows[:10],
            }
        }
    except Exception as e:
        return {'success': False, 'message': f'CSV 파싱 실패: {str(e)}', 'data': {}}
