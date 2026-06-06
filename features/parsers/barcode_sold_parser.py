# -*- coding: utf-8 -*-
"""
SQM v8.7.0 — 출고 바코드 스캔 파서 (Barcode → SOLD)
========================================================
현장 직원이 출고 시점에 톤백 바코드를 스캔한 결과 Excel 파일을 파싱.
파일이 업로드되면 비로소 **PICKED → SOLD 확정** 단계로 들어간다 (사용자 정책).

입력 파일 형식 (예: sample_out/bar_code.xlsx):
    Sheet: 바코드리스트
    11 컬럼 (행 1 = 헤더):
        1. SAP NO
        2. BL NO
        3. Container
        4. 품목명
        5. 톤백 UID          ← 매칭 키 (예: 1126021635-001)
        6. Sub LT            ← 보조 매칭 키
        7. 톤백 번호
        8. 중량(kg)
        9. 상태              ← PICKED 여야 정상
       10. 실제 위치         ← 직원이 현장에서 스캔한 실제 위치 (이걸 DB에 반영)
       11. 셀 위치 식별     ← 시스템 추정 위치 (참고용)

출력 doc:
    {
        'parse_ok': bool,
        'parse_method': 'excel',
        'source_file': str,
        'total_rows': int,
        'items': [
            {
                'tonbag_uid': '1126021635-001',
                'lot_no': '1126021635',            # UID에서 파생
                'sub_lt': 1,                       # int
                'tonbag_no': 1,                    # int
                'weight_kg': 500.0,
                'sap_no': '2200034566',
                'bl_no': 'MEDUW9030661',
                'container_no': 'MEDU9454898',
                'product': 'LITHIUM CARBONATE',
                'src_status': 'PICKED',
                'actual_location': 'G6-05-01-01',   # ← 신뢰 위치
                'cell_location': 'G6-05-01-01',
            },
            ...
        ],
        'warnings': [str, ...],   # 헤더 누락, 빈 행 등
    }
"""
from __future__ import annotations
import logging
import os
import re
from typing import Any

logger = logging.getLogger(__name__)

# 헤더 라벨 → 표준 키 매핑 (소문자 + 공백/특수문자 제거 후 비교)
_HEADER_ALIASES: dict[str, str] = {
    'sapno':           'sap_no',
    'sap':             'sap_no',
    'blno':            'bl_no',
    'bl':              'bl_no',
    'container':       'container_no',
    'containerno':     'container_no',
    # ── 제품명 (실제 헤더 표기: 제품명 / 품목명 / Product) ──
    '제품명':           'product',
    '품목명':           'product',
    '품목':             'product',
    '제품':             'product',
    'product':         'product',
    'productname':     'product',
    # ── 톤백 UID ──
    '톤백uid':          'tonbag_uid',
    'tonbaguid':       'tonbag_uid',
    'uid':             'tonbag_uid',
    '톤백id':           'tonbag_uid',
    # ── Sub LT ──
    'sublt':           'sub_lt',
    'sub_lt':          'sub_lt',
    '서브lt':           'sub_lt',
    # ── 톤백 번호 ──
    '톤백번호':         'tonbag_no',
    'tonbagno':        'tonbag_no',
    # ── 중량 ──
    '중량(kg)':         'weight_kg',
    '중량':             'weight_kg',
    'weight':          'weight_kg',
    'weightkg':        'weight_kg',
    # ── 상태 ──
    '상태':             'src_status',
    'status':          'src_status',
    # ── 실제 위치 ──
    '실제위치':         'actual_location',
    'actuallocation':  'actual_location',
    'realposition':    'actual_location',
    'reallocation':    'actual_location',
    # ── 셀/랙 위치 후보 (실제 헤더: 랙 위치 후보) ──
    '랙위치후보':       'cell_location',
    '셀위치후보':       'cell_location',
    '셀위치식별':       'cell_location',
    '셀위치':           'cell_location',
    '랙위치':           'cell_location',
    'celllocation':    'cell_location',
    'racklocation':    'cell_location',
}

# 부분 매칭 폴백 — 정확한 키 매칭 실패 시 contains 검사
# (헤더 표기가 미세하게 다를 때 안전망)
_PARTIAL_MATCH: list[tuple[tuple[str, ...], str]] = [
    (('uid',),                              'tonbag_uid'),
    (('sub', 'lt'),                         'sub_lt'),
    (('톤백', '번호'),                       'tonbag_no'),
    (('실제', '위치'),                       'actual_location'),
    (('랙', '위치'),                         'cell_location'),
    (('셀', '위치'),                         'cell_location'),
    (('제품',),                              'product'),
    (('품목',),                              'product'),
    (('중량',),                              'weight_kg'),
    (('weight',),                           'weight_kg'),
    (('상태',),                              'src_status'),
    (('container',),                        'container_no'),
]


def _norm_header(s: Any) -> str:
    """헤더 라벨 정규화: 유니코드 NFC + 소문자 + 공백/특수문자 제거."""
    if s is None:
        return ''
    import unicodedata
    t = unicodedata.normalize('NFC', str(s)).strip().lower()
    t = re.sub(r'[\s\-_().\[\]/\\]+', '', t)
    return t


def _resolve_header(raw: Any) -> str | None:
    """헤더 → 표준 키. 정확 매칭 실패 시 부분 매칭 폴백."""
    norm = _norm_header(raw)
    if not norm:
        return None
    if norm in _HEADER_ALIASES:
        return _HEADER_ALIASES[norm]
    # 부분 매칭 (모든 키워드가 포함돼야 매칭)
    for keywords, std_key in _PARTIAL_MATCH:
        if all(kw in norm for kw in keywords):
            return std_key
    return None


def _parse_uid(uid: str) -> tuple[str, int | None]:
    """톤백 UID(예: '1126021635-001')에서 lot_no, sub_lt 분리."""
    if not uid:
        return ('', None)
    s = str(uid).strip()
    m = re.match(r'^(.+?)[-_/](\d+)$', s)
    if m:
        try:
            return (m.group(1).strip(), int(m.group(2)))
        except ValueError:
            return (m.group(1).strip(), None)
    return (s, None)


def _safe_num(v: Any) -> float:
    try:
        if v is None or v == '':
            return 0.0
        return float(str(v).replace(',', '').strip())
    except (ValueError, TypeError):
        return 0.0


def _safe_int(v: Any) -> int | None:
    try:
        if v is None or v == '':
            return None
        return int(float(str(v).replace(',', '').strip()))
    except (ValueError, TypeError):
        return None


def parse_barcode_sold_excel(path: str) -> dict:
    """현장 바코드 스캔 Excel → SOLD 확정용 doc 변환.

    Args:
        path: .xlsx 파일 경로

    Returns:
        doc dict (위 docstring 참조)
    """
    doc: dict[str, Any] = {
        'parse_ok': False,
        'parse_method': 'excel',
        'source_file': os.path.basename(path),
        'total_rows': 0,
        'items': [],
        'warnings': [],
    }

    try:
        import openpyxl
    except ImportError:
        doc['warnings'].append('openpyxl 미설치')
        return doc

    if not os.path.exists(path):
        doc['warnings'].append(f'파일 없음: {path}')
        return doc

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        doc['warnings'].append(f'Excel 열기 실패: {e}')
        return doc

    # 시트 선택: 첫 시트 (이름 무관)
    ws = wb[wb.sheetnames[0]]

    # 헤더 찾기 (행 1~5 안에 'UID' 또는 '실제 위치' 컬럼이 있는 행)
    header_row = None
    col_map: dict[int, str] = {}  # {column_index_1based: 표준 키}
    for r, values in enumerate(ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), max_col=min(ws.max_column, 30), values_only=True), start=1):
        candidate = {}
        for c, cell_val in enumerate(values, start=1):
            key = _resolve_header(cell_val)
            if key:
                candidate[c] = key
        # 필수: tonbag_uid + actual_location 둘 다 있어야 헤더 행으로 판정
        if 'tonbag_uid' in candidate.values() and 'actual_location' in candidate.values():
            header_row = r
            col_map = candidate
            break

    if header_row is None:
        doc['warnings'].append('헤더 행을 찾지 못함 (톤백 UID + 실제 위치 컬럼 필요)')
        try:
            wb.close()
        except Exception:
            pass
        return doc

    # 데이터 행 파싱
    items: list[dict] = []
    skipped_empty = 0
    max_needed_col = max(col_map) if col_map else 1
    blank_streak = 0
    for r, values in enumerate(
        ws.iter_rows(
            min_row=header_row + 1,
            max_row=ws.max_row,
            max_col=max_needed_col,
            values_only=True,
        ),
        start=header_row + 1,
    ):
        row_dict: dict[str, Any] = {}
        any_value = False
        for col_idx, key in col_map.items():
            v = values[col_idx - 1] if col_idx <= len(values) else None
            if v is not None and str(v).strip() != '':
                any_value = True
            row_dict[key] = v

        if not any_value:
            skipped_empty += 1
            blank_streak += 1
            if items and blank_streak >= 50:
                break
            continue
        blank_streak = 0

        uid = str(row_dict.get('tonbag_uid') or '').strip()
        if not uid:
            doc['warnings'].append(f'R{r}: 톤백 UID 없음 — 스킵')
            continue

        lot_no_from_uid, sub_lt_from_uid = _parse_uid(uid)
        sub_lt = _safe_int(row_dict.get('sub_lt')) or sub_lt_from_uid

        item = {
            'tonbag_uid':       uid,
            'lot_no':           lot_no_from_uid,
            'sub_lt':           sub_lt,
            'tonbag_no':        _safe_int(row_dict.get('tonbag_no')),
            'weight_kg':        _safe_num(row_dict.get('weight_kg')),
            'sap_no':           str(row_dict.get('sap_no') or '').strip(),
            'bl_no':            str(row_dict.get('bl_no') or '').strip(),
            'container_no':     str(row_dict.get('container_no') or '').strip(),
            'product':          str(row_dict.get('product') or '').strip(),
            'src_status':       str(row_dict.get('src_status') or '').strip().upper(),
            'actual_location':  str(row_dict.get('actual_location') or '').strip(),
            'cell_location':    str(row_dict.get('cell_location') or '').strip(),
        }
        if not item['actual_location']:
            doc['warnings'].append(f'R{r} ({uid}): 실제 위치 비어있음')
        items.append(item)

    try:
        wb.close()
    except Exception:
        pass

    doc['total_rows'] = len(items)
    doc['items'] = items
    doc['parse_ok'] = len(items) > 0
    if skipped_empty > 0:
        doc['warnings'].append(f'빈 행 {skipped_empty}건 스킵')

    logger.info(f'[barcode_sold_parser] {path} → {len(items)}건 파싱 (warnings={len(doc["warnings"])})')
    return doc


if __name__ == '__main__':
    # 빠른 단독 테스트
    import json, sys
    p = sys.argv[1] if len(sys.argv) > 1 else r'D:\program\SQM_inventory\sample_out\bar_code.xlsx'
    d = parse_barcode_sold_excel(p)
    print(f'parse_ok    : {d["parse_ok"]}')
    print(f'total_rows  : {d["total_rows"]}')
    print(f'warnings    : {d["warnings"][:5]}')
    print(f'first 3 items:')
    for it in d['items'][:3]:
        print(f'  {json.dumps(it, ensure_ascii=False)}')
