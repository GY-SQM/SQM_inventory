"""
Gemini AI API — v8.6.6 신규 (1차 작업)
=========================================
v8.6.6의 검증된 features/ai/gemini_utils.py를 thin wrapper로 노출.

엔드포인트:
- GET  /api/ai/settings    → 현재 키/모델/사용여부 조회 (키는 마스킹)
- POST /api/ai/settings    → API 키 저장 (keyring 우선, ini fallback)
- POST /api/ai/toggle      → 사용 ON/OFF
- GET  /api/ai/test        → API 연결 테스트 (간단한 핑)
"""
import logging
import os
import configparser
from fastapi import APIRouter, HTTPException
from pydantic import BaseModel

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/ai", tags=["AI"])


# ── Schemas ────────────────────────────────────────────────────
class ApiKeyPayload(BaseModel):
    api_key: str
    model: str = ""  # 선택


class PinUnlockPayload(BaseModel):
    pin: str

class PinChangePayload(BaseModel):
    current_pin: str
    new_pin: str

class TogglePayload(BaseModel):
    enabled: bool


class ChatPayload(BaseModel):
    message: str
    write_session_token: str = ""


# ── Helpers ────────────────────────────────────────────────────
def _mask_key(key: str) -> str:
    if not key:
        return ""
    if len(key) <= 8:
        return "****"
    return key[:4] + "****" + key[-4:]


def _fresh_api_key():
    """
    매번 호출 시 환경변수 → keyring → settings.ini 순으로 직접 읽기.
    Python 모듈 캐시 우회 (백엔드 재시작 없이 새 키 인식).
    Returns: (key, source, model)
    """
    # 1. 환경변수
    key = (os.environ.get("GEMINI_API_KEY", "") or "").strip()
    model_env = (os.environ.get("GEMINI_MODEL", "") or "").strip()
    if key and not key.startswith("your-"):
        return key, "ENV", model_env or "gemini-2.5-flash"

    # 2. keyring
    try:
        import keyring
        kr_key = keyring.get_password("SQM_Inventory", "GEMINI_API_KEY")
        if kr_key and kr_key.strip():
            return kr_key.strip(), "KEYRING", model_env or "gemini-2.5-flash"
    except Exception as _e:
        logger.debug(f"keyring 조회 suppressed: {_e}")

    # 3. settings.ini 직접 읽기
    try:
        from config import SETTINGS_FILE
        if SETTINGS_FILE.exists():
            cfg = configparser.ConfigParser()
            cfg.read(SETTINGS_FILE, encoding="utf-8")
            ini_key = (cfg.get("Gemini", "api_key", fallback="") or "").strip()
            ini_model = (cfg.get("Gemini", "model", fallback="") or "").strip()
            if ini_key and not ini_key.startswith("your-") and not ini_key.startswith("# "):
                return ini_key, "INI", ini_model or model_env or "gemini-2.5-flash"
    except Exception as _e:
        logger.debug(f"ini 조회 suppressed: {_e}")

    return "", "NONE", model_env or "gemini-2.5-flash"


# ── 설정 조회 ──────────────────────────────────────────────────
@router.get("/settings", summary="🔐 Gemini API 설정 조회")
def get_settings():
    try:
        # 매번 fresh 읽기 (백엔드 재시작 없이도 변경 반영)
        key, source, model = _fresh_api_key()
        try:
            from config import USE_GEMINI_DEFAULT
            enabled = bool(USE_GEMINI_DEFAULT)
        except Exception:
            enabled = True
        return {
            "success": True,
            "api_key_masked": _mask_key(key),
            "has_key": bool(key),
            "model": model,
            "enabled": enabled,
            "key_source": source,  # ENV / KEYRING / INI / NONE
        }
    except Exception as e:
        logger.exception("get_settings error")
        raise HTTPException(500, f"설정 조회 실패: {e}")


# ── API 키 저장 ────────────────────────────────────────────────
@router.post("/settings", summary="🔐 Gemini API 키 저장")
def save_settings(payload: ApiKeyPayload):
    if not payload.api_key or not payload.api_key.strip():
        raise HTTPException(400, "API 키가 비어있습니다")
    try:
        from config import save_api_key_secure, save_gemini_model
        source = save_api_key_secure(payload.api_key.strip())
        if source == "FAILED":
            raise HTTPException(
                500,
                "API 키 저장 실패: keyring(OS 자격증명 관리자) 사용이 필요합니다. "
                "환경변수 GEMINI_API_KEY 또는 keyring 설정을 확인하세요."
            )
        if payload.model and payload.model.strip():
            save_gemini_model(payload.model.strip())
        # Gemini Client 싱글턴 리셋 (다음 호출 시 새 키로 재생성)
        try:
            from features.ai.gemini_utils import reset_gemini_client
            reset_gemini_client()
        except Exception as _e:
            logger.debug(f"reset_gemini_client suppressed: {_e}")
        return {
            "success": True,
            "message": f"API 키가 {source}에 저장되었습니다. 다음 실행부터 적용됩니다.",
            "source": source,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("save_settings error")
        raise HTTPException(500, f"저장 실패: {e}")


# ── 사용 토글 ──────────────────────────────────────────────────
@router.post("/toggle", summary="🔀 Gemini AI 사용 ON/OFF")
def toggle_ai(payload: TogglePayload):
    try:
        import configparser
        from config import SETTINGS_FILE
        config = configparser.ConfigParser()
        if SETTINGS_FILE.exists():
            config.read(SETTINGS_FILE, encoding="utf-8")
        if not config.has_section("Gemini"):
            config.add_section("Gemini")
        config.set("Gemini", "use_gemini", "true" if payload.enabled else "false")
        with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
            config.write(f)
        return {
            "success": True,
            "enabled": payload.enabled,
            "message": f"Gemini AI 사용을 {'켰습니다' if payload.enabled else '껐습니다'}. 다음 실행부터 적용됩니다.",
        }
    except Exception as e:
        logger.exception("toggle_ai error")
        raise HTTPException(500, f"토글 실패: {e}")


# ── API 연결 테스트 ────────────────────────────────────────────
@router.get("/test", summary="🧪 Gemini API 연결 테스트")
def test_connection():
    """
    짧은 핑 프롬프트를 보내 키/모델/네트워크가 정상인지 검증.
    매번 fresh 키 읽기 + Client 강제 재생성 (캐시 무시).
    """
    # 1. fresh 키 읽기 (백엔드 캐시 우회)
    key, source, model = _fresh_api_key()
    if not key:
        return {
            "success": False,
            "ok": False,
            "message": "❌ API 키를 찾을 수 없음 (환경변수 GEMINI_API_KEY, keyring, settings.ini 모두 비어있음)",
            "source": "NONE",
        }

    # 2. google-genai 패키지 확인
    try:
        from google import genai
    except ImportError:
        return {
            "success": False,
            "ok": False,
            "message": "❌ google-genai 패키지 미설치. pip install google-genai",
            "source": source,
        }

    # 3. Client 강제 재생성 (캐시 무시)
    try:
        from features.ai.gemini_utils import reset_gemini_client, call_gemini_safe
        reset_gemini_client()
        client = genai.Client(api_key=key)
    except Exception as e:
        return {
            "success": False,
            "ok": False,
            "message": f"❌ Client 생성 실패: {e}",
            "source": source,
        }

    # 4. 핑 호출
    try:
        resp = call_gemini_safe(
            client=client,
            model_name=model,
            contents="ping",
            timeout=15,
            temperature=0.0,
            max_output_tokens=16,
        )
        text = ""
        try:
            text = (resp.text or "").strip() if resp else ""
        except Exception:
            text = ""
        return {
            "success": True,
            "ok": True,
            "model": model,
            "source": source,
            "key_masked": _mask_key(key),
            "reply": text or "(응답 비어있음)",
            "message": f"✅ {model} 연결 성공 (키 출처: {source})",
        }
    except TimeoutError:
        return {"success": False, "ok": False, "source": source, "message": "⏱ 타임아웃 (15초 초과)"}
    except PermissionError:
        return {"success": False, "ok": False, "source": source, "message": f"🔐 API 키 인증 실패 (401) — 키 출처: {source}, 마스킹: {_mask_key(key)}"}
    except RuntimeError as e:
        return {"success": False, "ok": False, "source": source, "message": f"⚠️ 할당량/런타임 에러: {e}"}
    except Exception as e:
        logger.exception("test_connection error")
        return {"success": False, "ok": False, "source": source, "message": f"❌ 연결 실패: {e}"}


# ── AI 채팅 싱글턴 캐시 (대화 히스토리 유지) ──────────────────
_chat_singleton = None
_chat_singleton_db = None
_chat_singleton_key = None


def _get_chat_singleton(db_path: str, api_key: str):
    """동일 db+key 조합이면 기존 인스턴스 재사용해 히스토리 유지."""
    global _chat_singleton, _chat_singleton_db, _chat_singleton_key
    import sys
    here = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.normpath(os.path.join(here, "..", ".."))
    if project_root not in sys.path:
        sys.path.insert(0, project_root)
    from features.ai.gemini_chat_query import GeminiChatQuery
    if (
        _chat_singleton is None
        or _chat_singleton_db != db_path
        or _chat_singleton_key != api_key
    ):
        _chat_singleton = GeminiChatQuery(db_path=db_path, api_key=api_key or "")
        _chat_singleton_db = db_path
        _chat_singleton_key = api_key
    return _chat_singleton


# ── AI 채팅 ────────────────────────────────────────────────────
@router.post("/chat", summary="💬 AI 재고 채팅")
def chat_message(payload: ChatPayload):
    if not payload.message or not payload.message.strip():
        raise HTTPException(400, "메시지가 비어있습니다")

    key, _source, _model = _fresh_api_key()

    here = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.normpath(os.path.join(here, "..", ".."))
    db_path = os.path.join(project_root, "data", "db", "sqm_inventory.db")

    try:
        from backend.api.ai_write_session import validate_session
        write_mode = validate_session(payload.write_session_token or "")
        chat = _get_chat_singleton(db_path, key or "")
        result = chat.ask(payload.message.strip(), write_mode=write_mode)
        return {
            "success": result["success"],
            "answer": result["answer"],
            "data": result.get("data", []),
            "columns": result.get("columns", []),
            "row_count": result.get("row_count", 0),
            "query_type": result.get("query_type", ""),
            "elapsed_ms": result.get("elapsed_ms", 0),
        }
    except Exception as e:
        logger.exception("chat_message error")
        raise HTTPException(500, f"채팅 오류: {type(e).__name__}: {e}")


@router.post("/chat/export-excel", summary="📥 AI 쿼리 결과 Excel 내보내기 (Phase 4-1)")
def chat_export_excel(payload: dict):
    """
    AI 채팅 결과 데이터를 Excel 파일로 반환.
    요청: { "columns": [...], "data": [[...], ...], "query_type": "..." }
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from fastapi.responses import StreamingResponse
        import io, datetime

        columns = payload.get("columns") or []
        data    = payload.get("data")    or []
        q_type  = (payload.get("query_type") or "AI_조회")[:30]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = q_type[:31]

        # 헤더
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        header_font = Font(bold=True, color="FFFFFF")
        for ci, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # 데이터
        for ri, row in enumerate(data, 2):
            if isinstance(row, dict):
                row = [row.get(c) for c in columns]
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=val)

        # 열 너비 자동
        for ci, col in enumerate(columns, 1):
            max_len = max((len(str(col)),), *(
                (len(str(r[ci-1])) if isinstance(r, list) and ci-1 < len(r) else 0,)
                for r in data
            ), default=10)
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(ci)
            ].width = min(max_len + 2, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"AI_{q_type}_{ts}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
        )
    except Exception as e:
        logger.exception("chat_export_excel error")
        raise HTTPException(500, f"Excel 생성 실패: {e}")


@router.get("/autocomplete", summary="💡 AI 채팅 자동완성 (Phase 4-2)")
def chat_autocomplete(q: str = ""):
    """입력 텍스트와 매칭되는 제품명/LOT번호 반환 (최대 10개)."""
    if not q or len(q.strip()) < 2:
        return {"items": []}
    try:
        here = os.path.dirname(os.path.abspath(__file__))
        project_root = os.path.normpath(os.path.join(here, "..", ".."))
        db_path = os.path.join(project_root, "data", "db", "sqm_inventory.db")

        import sqlite3 as _sq
        con = _sq.connect(db_path, timeout=3)
        pattern = f"%{q.strip()}%"
        rows = con.execute("""
            SELECT DISTINCT lot_no FROM inventory
            WHERE lot_no LIKE ? LIMIT 5
            UNION
            SELECT DISTINCT product FROM inventory
            WHERE product LIKE ? LIMIT 5
        """, (pattern, pattern)).fetchall()
        con.close()
        items = [r[0] for r in rows if r[0]]
        return {"items": items[:10]}
    except Exception as e:
        logger.warning(f"autocomplete 오류: {e}")
        return {"items": []}


@router.post("/chat/clear", summary="💬 AI 채팅 히스토리 초기화")
def chat_clear():
    global _chat_singleton
    if _chat_singleton is not None:
        try:
            _chat_singleton.clear_history()
        except Exception:
            pass
        _chat_singleton = None
    return {"success": True, "message": "대화 히스토리가 초기화되었습니다."}


# ── AI 수정 모드 PIN 관리 ───────────────────────────────────────
@router.post("/write-unlock", summary="🔐 수정 모드 잠금 해제")
def write_unlock(payload: PinUnlockPayload):
    from backend.api.ai_write_session import check_pin, create_session, ensure_edit_log_table
    ensure_edit_log_table()
    success, msg = check_pin(payload.pin)
    if not success:
        raise HTTPException(401, msg)
    token = create_session()
    return {"success": True, "token": token, "message": "수정 모드가 활성화됐습니다."}


@router.post("/write-lock", summary="🔒 수정 모드 잠금")
def write_lock():
    from backend.api.ai_write_session import _sessions
    _sessions.clear()
    return {"success": True, "message": "수정 모드가 잠겼습니다."}


@router.get("/write-status", summary="🔓 수정 모드 상태 조회")
def write_status(token: str = ""):
    from backend.api.ai_write_session import validate_session, get_session_remaining
    valid = validate_session(token)
    remaining = get_session_remaining(token) if valid else 0
    return {"active": valid, "remaining_seconds": remaining}


@router.post("/pin/change", summary="🔑 PIN 변경")
def change_pin_endpoint(payload: PinChangePayload):
    from backend.api.ai_write_session import change_pin
    success, msg = change_pin(payload.current_pin, payload.new_pin)
    if not success:
        raise HTTPException(400, msg)
    return {"success": True, "message": msg}
