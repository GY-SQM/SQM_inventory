"""
SQM v8.6.6 — Phase 4-C SQL 조회 엔드포인트
queries3.py: sales-order-dn, dn-cross-check, do-status,
             invoice-list, settings-info

모든 응답: ok_response(data=...) 표준 포맷
"""
import os
import re
import sqlite3
import logging
import tempfile
import json
import shutil
from datetime import date, datetime
from pathlib import Path
from fastapi import APIRouter, Body, File, HTTPException, Query as QP, UploadFile
from fastapi.responses import FileResponse
from backend.common.errors import ok_response, err_response

router = APIRouter(prefix="/api/q3", tags=["queries3"])
logger = logging.getLogger(__name__)


# ── DB 경로 헬퍼 ─────────────────────────────────────────────────
def _db() -> sqlite3.Connection:
    here = os.path.dirname(os.path.abspath(__file__))
    root = os.path.dirname(os.path.dirname(here))
    db_path = os.path.join(root, "data", "db", "sqm_inventory.db")
    con = sqlite3.connect(db_path, timeout=5, check_same_thread=False)
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA journal_mode=WAL")
    con.execute("PRAGMA busy_timeout=3000")
    return con


def _rows(rows) -> list:
    return [dict(r) for r in rows]


def _project_root() -> Path:
    here = Path(__file__).resolve()
    return here.parents[2]


def _sales_order_dn_template_path() -> Path:
    root = _project_root()
    candidates = [
        root / "data" / "templates" / "sales_order_dn_template.xlsx",
        root.parent / "sqm_2_merge_upload" / "data" / "templates" / "sales_order_dn_template.xlsx",
        Path(r"D:\program\SQM_inventory\sqm_2_merge_upload\data\templates\sales_order_dn_template.xlsx"),
    ]
    for p in candidates:
        if p.exists():
            return p
    raise HTTPException(
        status_code=404,
        detail="sales_order_dn_template.xlsx 템플릿을 찾을 수 없습니다.",
    )


_REPORT_TYPES = {
    "outbound_report": "Outbound Report",
    "export_work_report": "수출 작업 리포트",
    "sales_order_dn": "Sales Order DN",
    "storage_confirmation": "Storage Confirmation",
    "sold_inventory_report": "SOLD Inventory Report",
}

_REPORT_FIELDS = {
    "outbound_report": [
        {"field": "destination", "label": "Destination / 고객사"},
        {"field": "delivery_date", "label": "Delivery Date"},
        {"field": "lot_no", "label": "LOT NO"},
        {"field": "sap_no", "label": "SAP NO"},
        {"field": "bl_no", "label": "BL NO"},
        {"field": "container_no", "label": "Container No"},
        {"field": "sales_order_no", "label": "Sales Order No"},
        {"field": "picking_no", "label": "Picking No"},
        {"field": "sku", "label": "SKU"},
        {"field": "description", "label": "Description"},
        {"field": "nw_mt", "label": "NW(MT)"},
        {"field": "gw_mt", "label": "GW(MT)"},
        {"field": "qty", "label": "CT/PLT"},
        {"field": "is_sample", "label": "Sample Flag"},
    ],
    "export_work_report": [
        {"field": "fixed_1", "label": "Bag"},
        {"field": "description", "label": "Description of goods"},
        {"field": "lot_no", "label": "Lot No."},
        {"field": "qty", "label": "Q'ty"},
        {"field": "nw_mt", "label": "Net Weight"},
        {"field": "gw_mt", "label": "Gross Weight"},
        {"field": "container_no", "label": "Container No"},
        {"field": "seal_no", "label": "Seal No"},
        {"field": "size_type", "label": "Container Size"},
    ],
    "sales_order_dn": [
        {"field": "destination", "label": "Destination / 고객사"},
        {"field": "delivery_date", "label": "Delivery Date"},
        {"field": "lot_no", "label": "LOT NO"},
        {"field": "sap_no", "label": "SAP NO"},
        {"field": "bl_no", "label": "BL NO"},
        {"field": "sales_order_no", "label": "Sales Order No"},
        {"field": "picking_no", "label": "Picking No"},
        {"field": "sku", "label": "SKU"},
        {"field": "description", "label": "Description"},
        {"field": "nw_mt", "label": "NW(MT)"},
        {"field": "gw_mt", "label": "GW(MT)"},
        {"field": "qty", "label": "CT/PLT"},
    ],
    "storage_confirmation": [
        {"field": "no", "label": "NO"},
        {"field": "part_no", "label": "PART NO"},
        {"field": "part_description", "label": "PART DESCRIPTION"},
        {"field": "sap_no", "label": "SAP NO"},
        {"field": "lot_no", "label": "LOT NO"},
        {"field": "in_date", "label": "IN DATE"},
        {"field": "invoice_net_weight", "label": "INVOICE NET WEIGHT"},
        {"field": "inspection_net_weight", "label": "INSPCECTION NET WEIGHT"},
        {"field": "balance", "label": "BALANCE (+/-)"},
        {"field": "damage_weight", "label": "DAMAGE WEIGHT"},
        {"field": "damage_reason", "label": "DAMAGE 사유"},
        {"field": "container_no", "label": "CONT NO"},
        {"field": "bl_no", "label": "BL NO"},
    ],
    "sold_inventory_report": [
        {"field": "product", "label": "Product"},
        {"field": "sap_no", "label": "SAP NO"},
        {"field": "eta_busan", "label": "ETA BUSAN"},
        {"field": "date_in_stock", "label": "Date in stock"},
        {"field": "sc_rcvd", "label": "SC RCVD"},
        {"field": "days", "label": "Days"},
        {"field": "qty_mt", "label": "QTY (MT)"},
        {"field": "lot_no", "label": "Lot No"},
        {"field": "wh", "label": "WH"},
        {"field": "salar_invoice_no", "label": "Salar Invoice no."},
        {"field": "sold_to", "label": "SOLD TO"},
        {"field": "sale_ref", "label": "SALE REF"},
        {"field": "invoice_date", "label": "Invoice date"},
        {"field": "picked_up_qty_mt", "label": "Picked up Qty (MT)"},
        {"field": "balance", "label": "Balance"},
        {"field": "gw", "label": "GW"},
        {"field": "actual_pick_up", "label": "Actual pick up"},
        {"field": "old", "label": "Old"},
        {"field": "condition", "label": "Condition"},
        {"field": "remark", "label": "Remark"},
    ],
}

_DEFAULT_COLUMNS = {
    "outbound_report": [
        {"header": "Destination", "field": "destination", "enabled": True},
        {"header": "Delivery Date", "field": "delivery_date", "enabled": True},
        {"header": "LOT NO", "field": "lot_no", "enabled": True},
        {"header": "SAP NO", "field": "sap_no", "enabled": True},
        {"header": "BL NO", "field": "bl_no", "enabled": True},
        {"header": "Sales order No", "field": "sales_order_no", "enabled": True},
        {"header": "Picking No", "field": "picking_no", "enabled": True},
        {"header": "SKU", "field": "sku", "enabled": True},
        {"header": "Description", "field": "description", "enabled": True},
        {"header": "NW(MT)", "field": "nw_mt", "enabled": True},
        {"header": "GW(MT)", "field": "gw_mt", "enabled": True},
        {"header": "CT/PLT", "field": "qty", "enabled": True},
    ],
    "export_work_report": [
        {"header": "Bag", "field": "fixed_1", "enabled": True},
        {"header": "Description of goods", "field": "description", "enabled": True},
        {"header": "Lot No.", "field": "lot_no", "enabled": True},
        {"header": "Q'ty", "field": "qty", "enabled": True},
        {"header": "Net Weight", "field": "nw_mt", "enabled": True},
        {"header": "Gross Weight", "field": "gw_mt", "enabled": True},
    ],
    "sales_order_dn": [
        {"header": "Destination", "field": "destination", "enabled": True},
        {"header": "Delivery Date", "field": "delivery_date", "enabled": True},
        {"header": "LOT NO", "field": "lot_no", "enabled": True},
        {"header": "SAP NO", "field": "sap_no", "enabled": True},
        {"header": "BL NO", "field": "bl_no", "enabled": True},
        {"header": "Sales order No", "field": "sales_order_no", "enabled": True},
        {"header": "Picking No", "field": "picking_no", "enabled": True},
        {"header": "SKU", "field": "sku", "enabled": True},
        {"header": "Description", "field": "description", "enabled": True},
        {"header": "NW(MT)", "field": "nw_mt", "enabled": True},
        {"header": "GW(MT)", "field": "gw_mt", "enabled": True},
        {"header": "CT/PLT", "field": "qty", "enabled": True},
    ],
    "storage_confirmation": [
        {"header": "NO", "field": "no", "enabled": True},
        {"header": "PART NO", "field": "part_no", "enabled": True},
        {"header": "PART DESCRIPTION", "field": "part_description", "enabled": True},
        {"header": "SAP NO", "field": "sap_no", "enabled": True},
        {"header": "LOT NO", "field": "lot_no", "enabled": True},
        {"header": "IN DATE", "field": "in_date", "enabled": True},
        {"header": "INVOICE NET WEIGHT", "field": "invoice_net_weight", "enabled": True},
        {"header": "INSPCECTION NET WEIGHT", "field": "inspection_net_weight", "enabled": True},
        {"header": "BALANCE (+/-)", "field": "balance", "enabled": True},
        {"header": "DAMAGE WEIGHT", "field": "damage_weight", "enabled": True},
        {"header": "DAMAGE 사유", "field": "damage_reason", "enabled": True},
        {"header": "CONT NO", "field": "container_no", "enabled": True},
        {"header": "BL NO", "field": "bl_no", "enabled": True},
    ],
    "sold_inventory_report": [
        {"header": "Product", "field": "product", "enabled": True},
        {"header": "SAP NO", "field": "sap_no", "enabled": True},
        {"header": "ETA BUSAN", "field": "eta_busan", "enabled": True},
        {"header": "Date in stock", "field": "date_in_stock", "enabled": True},
        {"header": "SC RCVD", "field": "sc_rcvd", "enabled": True},
        {"header": "Days", "field": "days", "enabled": True},
        {"header": "QTY (MT)", "field": "qty_mt", "enabled": True},
        {"header": "Lot No", "field": "lot_no", "enabled": True},
        {"header": "WH", "field": "wh", "enabled": True},
        {"header": "Salar Invoice no.", "field": "salar_invoice_no", "enabled": True},
        {"header": "SOLD TO", "field": "sold_to", "enabled": True},
        {"header": "SALE REF", "field": "sale_ref", "enabled": True},
        {"header": "Invoice date", "field": "invoice_date", "enabled": True},
        {"header": "Picked up Qty (MT)", "field": "picked_up_qty_mt", "enabled": True},
        {"header": "Balance", "field": "balance", "enabled": True},
        {"header": "GW", "field": "gw", "enabled": True},
        {"header": "Actual pick up", "field": "actual_pick_up", "enabled": True},
        {"header": "Old", "field": "old", "enabled": True},
        {"header": "Condition", "field": "condition", "enabled": True},
        {"header": "Remark", "field": "remark", "enabled": True},
    ],
}


def _report_template_dir(report_type: str) -> Path:
    key = str(report_type or "").strip()
    if key not in _REPORT_TYPES:
        raise HTTPException(status_code=400, detail="알 수 없는 보고서 유형입니다.")
    d = _project_root() / "data" / "report_templates" / key
    d.mkdir(parents=True, exist_ok=True)
    return d


def _ensure_report_template_default(report_type: str) -> None:
    if report_type not in {"sales_order_dn", "storage_confirmation", "sold_inventory_report"}:
        return
    dest = _report_template_dir(report_type) / "template_1.xlsx"
    if dest.exists():
        return
    if report_type == "sales_order_dn":
        src = _sales_order_dn_template_path()
        shutil.copyfile(src, dest)
        return
    if report_type == "sold_inventory_report":
        src = _project_root().parent / "sample_out" / "new_SQM 재고관리파일(2026.04.30 SOLD).xlsx"
        if not src.exists():
            src = Path(r"D:\program\SQM_inventory\sample_out\new_SQM 재고관리파일(2026.04.30 SOLD).xlsx")
        if not src.exists():
            raise HTTPException(status_code=404, detail="SOLD Inventory Report 기본 양식을 찾을 수 없습니다.")
        shutil.copyfile(src, dest)
        return
    if report_type == "storage_confirmation":
        src = _project_root().parent / "sample_out" / "new_Storage confirmation(2200034857) - 2026.05.15.xls"
        if not src.exists():
            src = Path(r"D:\program\SQM_inventory\sample_out\new_Storage confirmation(2200034857) - 2026.05.15.xls")
        if not src.exists():
            raise HTTPException(status_code=404, detail="Storage Confirmation 기본 양식을 찾을 수 없습니다.")
        try:
            import pandas as pd
        except ImportError:
            raise HTTPException(status_code=500, detail="pandas 미설치: xls 기본 양식 변환 불가")
        sheets = pd.read_excel(src, sheet_name=None, header=None)
        with pd.ExcelWriter(dest, engine="openpyxl") as writer:
            for sheet_name, df in sheets.items():
                df.to_excel(writer, sheet_name=str(sheet_name)[:31], header=False, index=False)


def _report_template_path(report_type: str, template: str = "template_1") -> Path:
    safe = re.sub(r"[^A-Za-z0-9_-]+", "", str(template or "template_1")) or "template_1"
    _ensure_report_template_default(report_type)
    path = _report_template_dir(report_type) / f"{safe}.xlsx"
    if not path.exists():
        raise HTTPException(status_code=404, detail=f"{_REPORT_TYPES[report_type]} 템플릿 없음: {safe}")
    return path


def _report_template_meta_path(report_type: str) -> Path:
    return _report_template_dir(report_type) / "templates.json"


def _report_template_columns_path(report_type: str) -> Path:
    return _report_template_dir(report_type) / "columns.json"


def _load_report_template_meta(report_type: str) -> dict:
    path = _report_template_meta_path(report_type)
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _save_report_template_meta(report_type: str, meta: dict) -> None:
    _report_template_meta_path(report_type).write_text(
        json.dumps(meta, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _load_report_template_columns(report_type: str) -> dict:
    path = _report_template_columns_path(report_type)
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _save_report_template_columns(report_type: str, columns: dict) -> None:
    _report_template_columns_path(report_type).write_text(
        json.dumps(columns, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _get_template_columns(report_type: str, template: str) -> list[dict]:
    all_cols = _load_report_template_columns(report_type)
    cols = all_cols.get(template)
    if isinstance(cols, list) and cols:
        return cols
    return [dict(c) for c in _DEFAULT_COLUMNS.get(report_type, [])]


def _set_template_columns(report_type: str, template: str, columns: list[dict]) -> None:
    allowed = {f["field"] for f in _REPORT_FIELDS.get(report_type, [])}
    clean = []
    for col in columns or []:
        field = str(col.get("field") or "").strip()
        if field not in allowed:
            continue
        header = str(col.get("header") or next((f["label"] for f in _REPORT_FIELDS[report_type] if f["field"] == field), field)).strip()
        clean.append({"header": header or field, "field": field, "enabled": bool(col.get("enabled", True))})
    if not clean:
        clean = _get_template_columns(report_type, template)
    all_cols = _load_report_template_columns(report_type)
    all_cols[template] = clean
    _save_report_template_columns(report_type, all_cols)


def _template_display_name(report_type: str, template: str) -> str:
    meta = _load_report_template_meta(report_type)
    return (meta.get(template) or {}).get("display_name") or template


def _set_template_display_name(report_type: str, template: str, display_name: str, original_filename: str = "") -> None:
    meta = _load_report_template_meta(report_type)
    meta[template] = {
        "display_name": str(display_name or template).strip() or template,
        "original_filename": original_filename,
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    _save_report_template_meta(report_type, meta)


def _next_template_path(report_type: str) -> Path:
    d = _report_template_dir(report_type)
    nums = []
    for p in d.glob("template_*.xlsx"):
        m = re.match(r"template_(\d+)\.xlsx$", p.name)
        if m:
            nums.append(int(m.group(1)))
    return d / f"template_{(max(nums) if nums else 0) + 1}.xlsx"


def _ensure_default_template_names() -> None:
    defaults = {
        "outbound_report": "Outbound Report 기본 양식",
        "export_work_report": "수출 작업 리포트 기본 양식",
        "sales_order_dn": "Sales Order DN 기본 양식",
        "storage_confirmation": "Storage Confirmation 기본 양식",
        "sold_inventory_report": "SOLD Inventory Report 기본 양식",
    }
    for report_type, name in defaults.items():
        _ensure_report_template_default(report_type)
        p = _report_template_dir(report_type) / "template_1.xlsx"
        meta = _load_report_template_meta(report_type)
        if p.exists() and "template_1" not in meta:
            _set_template_display_name(report_type, "template_1", name, p.name)


def _safe_filename_part(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", str(value or "").strip())[:80] or "SalesOrder"


def _copy_cell_style(src, dst) -> None:
    from copy import copy
    if src.has_style:
        dst._style = copy(src._style)
    if src.number_format:
        dst.number_format = src.number_format
    if src.font:
        dst.font = copy(src.font)
    if src.fill:
        dst.fill = copy(src.fill)
    if src.border:
        dst.border = copy(src.border)
    if src.alignment:
        dst.alignment = copy(src.alignment)
    if src.protection:
        dst.protection = copy(src.protection)


def _copy_row_style(ws, source_row: int, target_row: int, start_col: int, end_col: int) -> None:
    if source_row <= 0 or target_row <= 0:
        return
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(start_col, end_col + 1):
        _copy_cell_style(ws.cell(source_row, col), ws.cell(target_row, col))


def _copy_rows_style(ws, source_row: int, start_row: int, row_count: int, start_col: int, end_col: int) -> None:
    for offset in range(max(0, row_count)):
        _copy_row_style(ws, source_row, start_row + offset, start_col, end_col)


def _fmt_date(value) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    return text[:10]


def _fmt_mt(value) -> float:
    try:
        return round(float(value or 0) / 1000.0, 4)
    except Exception:
        return 0.0


def _sales_order_dn_rows(con: sqlite3.Connection, sales_order_no: str) -> list:
    return _rows(con.execute(
        """
        WITH sold_group AS (
            SELECT
                lot_no,
                COALESCE(NULLIF(picking_no, ''), '') AS picking_no,
                COALESCE(NULLIF(sap_no, ''), '') AS sap_no,
                COALESCE(NULLIF(bl_no, ''), '') AS bl_no,
                COALESCE(NULLIF(customer, ''), '') AS customer,
                COALESCE(NULLIF(sku, ''), '') AS sku,
                COALESCE(NULLIF(delivery_date, ''), '') AS delivery_date,
                COALESCE(is_sample, 0) AS is_sample,
                SUM(COALESCE(sold_qty_kg, 0)) AS net_kg,
                SUM(COALESCE(gross_weight_kg, 0)) AS sold_gross_kg,
                COUNT(*) AS ct_plt
            FROM sold_table
            WHERE sales_order_no = ?
              AND COALESCE(status, '') IN ('SOLD', 'CONFIRMED')
            GROUP BY lot_no, picking_no, sap_no, bl_no, customer, sku, delivery_date, COALESCE(is_sample, 0)
        )
        SELECT
            sg.*,
            i.product,
            i.product_code,
            i.initial_weight AS inv_net_kg,
            i.gross_weight AS inv_gross_kg,
            d.final_destination,
            d.port_of_discharge,
            d.warehouse_name,
            d.gross_weight_kg AS do_gross_kg
        FROM sold_group sg
        LEFT JOIN inventory i ON i.lot_no = sg.lot_no
        LEFT JOIN document_do d ON d.lot_no = sg.lot_no
        ORDER BY sg.delivery_date, sg.lot_no, sg.is_sample, sg.picking_no
        """,
        (sales_order_no,),
    ).fetchall())


def _sales_order_dn_report_rows(con: sqlite3.Connection) -> list[dict]:
    rows = _rows(con.execute(
        """
        WITH sold_group AS (
            SELECT
                lot_no,
                COALESCE(NULLIF(sales_order_no, ''), '') AS sales_order_no,
                COALESCE(NULLIF(picking_no, ''), '') AS picking_no,
                COALESCE(NULLIF(sap_no, ''), '') AS sap_no,
                COALESCE(NULLIF(bl_no, ''), '') AS bl_no,
                COALESCE(NULLIF(customer, ''), '') AS customer,
                COALESCE(NULLIF(sku, ''), '') AS sku,
                COALESCE(NULLIF(delivery_date, ''), '') AS delivery_date,
                SUM(COALESCE(sold_qty_kg, 0)) AS net_kg,
                SUM(COALESCE(gross_weight_kg, 0)) AS sold_gross_kg,
                COUNT(*) AS ct_plt
            FROM sold_table
            WHERE COALESCE(status, '') IN ('SOLD', 'CONFIRMED')
              AND TRIM(COALESCE(sales_order_no, '')) != ''
            GROUP BY lot_no, sales_order_no, picking_no, sap_no, bl_no, customer, sku, delivery_date
        )
        SELECT
            sg.*,
            i.product,
            i.product_code,
            i.initial_weight AS inv_net_kg,
            i.gross_weight AS inv_gross_kg,
            i.container_no,
            d.final_destination,
            d.port_of_discharge,
            d.warehouse_name
        FROM sold_group sg
        LEFT JOIN inventory i ON i.lot_no = sg.lot_no
        LEFT JOIN document_do d ON d.lot_no = sg.lot_no
        ORDER BY sg.delivery_date, sg.sales_order_no, sg.lot_no, sg.picking_no
        """
    ).fetchall())
    for r in rows:
        r["destination"] = (
            r.get("customer")
            or r.get("final_destination")
            or r.get("port_of_discharge")
            or r.get("warehouse_name")
            or ""
        )
        r["sku"] = r.get("sku") or r.get("product_code") or ""
        r["description"] = r.get("product") or r.get("sku") or ""
        r["nw_mt"] = _fmt_mt(r.get("net_kg"))
        r["gw_mt"] = _fmt_mt(_estimate_gross_kg(r))
        r["qty"] = int(r.get("ct_plt") or 0)
    return rows


def _estimate_gross_kg(row: dict) -> float:
    net = float(row.get("net_kg") or 0)
    sold_gross = float(row.get("sold_gross_kg") or 0)
    if sold_gross > 0:
        return sold_gross
    inv_net = float(row.get("inv_net_kg") or 0)
    inv_gross = float(row.get("inv_gross_kg") or 0)
    if net > 0 and inv_net > 0 and inv_gross > 0:
        return net * (inv_gross / inv_net)
    return net


def _estimate_gross_mt_for_lot(row: dict, is_sample: bool) -> float:
    net_mt = float(row.get("nw_mt") or 0)
    inv_gross_kg = float(row.get("inv_gross_kg") or 0)
    if is_sample:
        return 0.00125 if net_mt > 0 else 0
    if inv_gross_kg > 0:
        sample_gross_mt = 0.00125 if int(row.get("sample_qty") or 0) > 0 else 0
        return max(0, round((inv_gross_kg / 1000.0) - sample_gross_mt, 6))
    return net_mt


def _outbound_report_rows(con: sqlite3.Connection) -> list[dict]:
    rows = _rows(con.execute(
        """
        WITH sold_group AS (
            SELECT
                s.lot_no,
                COALESCE(s.is_sample, 0) AS is_sample,
                MAX(COALESCE(NULLIF(s.customer, ''), '')) AS destination,
                MAX(COALESCE(NULLIF(s.delivery_date, ''), NULLIF(s.sold_date, ''), '')) AS delivery_date,
                MAX(COALESCE(NULLIF(s.sap_no, ''), '')) AS sap_no,
                MAX(COALESCE(NULLIF(s.bl_no, ''), '')) AS bl_no,
                MAX(COALESCE(NULLIF(s.sales_order_no, ''), '')) AS sales_order_no,
                MAX(COALESCE(NULLIF(s.picking_no, ''), '')) AS picking_no,
                MAX(COALESCE(NULLIF(s.sku, ''), '')) AS sold_sku,
                SUM(COALESCE(s.sold_qty_kg, 0)) / 1000.0 AS nw_mt,
                COUNT(*) AS qty
            FROM sold_table s
            WHERE COALESCE(s.status, '') IN ('SOLD', 'CONFIRMED')
            GROUP BY s.lot_no, COALESCE(s.is_sample, 0)
        ),
        sample_qty AS (
            SELECT lot_no, COUNT(*) AS sample_qty
            FROM sold_table
            WHERE COALESCE(status, '') IN ('SOLD', 'CONFIRMED') AND COALESCE(is_sample, 0) = 1
            GROUP BY lot_no
        )
        SELECT
            sg.*,
            i.product,
            i.product_code,
            i.container_no,
            i.gross_weight AS inv_gross_kg,
            COALESCE(sq.sample_qty, 0) AS sample_qty
        FROM sold_group sg
        LEFT JOIN inventory i ON i.lot_no = sg.lot_no
        LEFT JOIN sample_qty sq ON sq.lot_no = sg.lot_no
        ORDER BY sg.delivery_date, sg.is_sample, sg.lot_no
        """
    ).fetchall())
    for r in rows:
        is_sample = int(r.get("is_sample") or 0) == 1
        r["sku"] = r.get("sold_sku") or r.get("product_code") or ""
        base_desc = r.get("product") or ""
        r["description"] = (base_desc + " SAMPLE").strip() if is_sample and "SAMPLE" not in base_desc.upper() else base_desc
        r["gw_mt"] = _estimate_gross_mt_for_lot(r, is_sample)
    return rows


def _match_report_filter(row: dict, filter_type: str, filter_value: str, start_date: str = "", end_date: str = "") -> bool:
    ft = str(filter_type or "all").strip()
    fv = str(filter_value or "").strip().lower()
    if start_date:
        d = _fmt_date(row.get("delivery_date") or row.get("sold_date"))
        if d < start_date:
            return False
    if end_date:
        d = _fmt_date(row.get("delivery_date") or row.get("sold_date"))
        if d > end_date:
            return False
    if ft == "all" or not fv:
        return True
    field_map = {
        "sales_order_no": ["sales_order_no"],
        "customer": ["destination", "customer"],
        "bl_no": ["bl_no"],
        "container_no": ["container_no"],
        "lot_no": ["lot_no"],
        "picking_no": ["picking_no"],
    }
    values = []
    for key in field_map.get(ft, [ft]):
        values.append(str(row.get(key) or "").lower())
    return any(fv in v for v in values)


def _split_filter_values(filter_values: str = "", filter_value: str = "") -> list[str]:
    raw = str(filter_values or filter_value or "").strip()
    if not raw:
        return []
    return [v.strip() for v in raw.split(",") if v.strip()]


def _order_rows_by_filter_values(rows: list[dict], key: str, values: list[str]) -> list[dict]:
    if not values:
        return rows
    rank = {v.lower(): i for i, v in enumerate(values)}
    return sorted(rows, key=lambda r: (rank.get(str(r.get(key) or "").lower(), len(rank)), str(r.get("lot_no") or "")))


def _filter_outbound_rows(rows: list[dict], filter_type: str, filter_value: str, start_date: str = "", end_date: str = "", filter_values: str = "") -> list[dict]:
    vals = _split_filter_values(filter_values, filter_value)
    if filter_type == "container_no" and vals:
        allowed = {v.lower() for v in vals}
        rows = [r for r in rows if str(r.get("container_no") or "").lower() in allowed]
        return _order_rows_by_filter_values(rows, "container_no", vals)
    return [r for r in rows if _match_report_filter(r, filter_type, filter_value, start_date, end_date)]


def _filter_sales_order_dn_rows(rows: list[dict], filter_type: str, filter_value: str, start_date: str = "", end_date: str = "", filter_values: str = "") -> list[dict]:
    vals = _split_filter_values(filter_values, filter_value)
    if filter_type == "container_no" and vals:
        allowed = {v.lower() for v in vals}
        rows = [r for r in rows if str(r.get("container_no") or "").lower() in allowed]
        return _order_rows_by_filter_values(rows, "container_no", vals)
    return [r for r in rows if _match_report_filter(r, filter_type, filter_value, start_date, end_date)]


def _export_work_blocks(con: sqlite3.Connection) -> list[dict]:
    rows = _rows(con.execute(
        """
        SELECT
            i.container_no,
            i.lot_no,
            i.product,
            i.product_code,
            i.net_weight AS inv_net_kg,
            i.gross_weight AS inv_gross_kg,
            MAX(COALESCE(NULLIF(ci.seal_no, ''), '')) AS seal_no,
            MAX(COALESCE(NULLIF(ci.size_type, ''), '')) AS size_type,
            SUM(CASE WHEN COALESCE(s.is_sample, 0) = 0 THEN 1 ELSE 0 END) AS normal_qty,
            SUM(CASE WHEN COALESCE(s.is_sample, 0) = 0 THEN COALESCE(s.sold_qty_kg, 0) ELSE 0 END) / 1000.0 AS normal_nw_mt,
            SUM(CASE WHEN COALESCE(s.is_sample, 0) = 1 THEN 1 ELSE 0 END) AS sample_qty,
            SUM(CASE WHEN COALESCE(s.is_sample, 0) = 1 THEN COALESCE(s.sold_qty_kg, 0) ELSE 0 END) / 1000.0 AS sample_nw_mt
        FROM inventory i
        LEFT JOIN sold_table s ON s.lot_no = i.lot_no AND COALESCE(s.status, '') IN ('SOLD', 'CONFIRMED')
        LEFT JOIN container_info ci ON ci.lot_no = i.lot_no AND ci.container_no = i.container_no
        WHERE TRIM(COALESCE(i.container_no, '')) != ''
        GROUP BY i.container_no, i.lot_no, i.product, i.product_code, i.net_weight, i.gross_weight
        ORDER BY i.container_no, i.lot_no
        """
    ).fetchall())
    blocks: dict[str, dict] = {}
    for r in rows:
        cno = r.get("container_no") or ""
        b = blocks.setdefault(cno, {
            "container_no": cno,
            "seal_no": r.get("seal_no") or "",
            "size_type": r.get("size_type") or "20FT",
            "lots": [],
        })
        if not b["seal_no"] and r.get("seal_no"):
            b["seal_no"] = r.get("seal_no")
        if r.get("size_type"):
            b["size_type"] = r.get("size_type")
        normal_qty = int(r.get("normal_qty") or 0)
        sample_qty = int(r.get("sample_qty") or 0)
        inv_gross_kg = float(r.get("inv_gross_kg") or 0)
        sample_gw_mt = 0.00125 if sample_qty > 0 else 0
        normal_gw_mt = max(0, round((inv_gross_kg / 1000.0) - sample_gw_mt, 6)) if inv_gross_kg else float(r.get("normal_nw_mt") or 0)
        b["lots"].append({
            "lot_no": r.get("lot_no") or "",
            "description": r.get("product") or "",
            "normal_qty": normal_qty,
            "normal_nw_mt": float(r.get("normal_nw_mt") or 0),
            "normal_gw_mt": normal_gw_mt,
            "sample_qty": sample_qty,
            "sample_nw_mt": float(r.get("sample_nw_mt") or 0),
            "sample_gw_mt": sample_gw_mt,
        })
    return list(blocks.values())


def _filter_export_blocks(blocks: list[dict], filter_type: str, filter_value: str, filter_values: str = "") -> list[dict]:
    ft = str(filter_type or "all").strip()
    fv = str(filter_value or "").strip().lower()
    vals = _split_filter_values(filter_values, filter_value)
    if ft == "container_no" and vals:
        by_no = {str(b.get("container_no") or "").lower(): b for b in blocks}
        return [by_no[v.lower()] for v in vals if v.lower() in by_no]
    if ft == "all" or not fv:
        return blocks
    out = []
    for block in blocks:
        block_text = " ".join([
            str(block.get("container_no") or ""),
            str(block.get("seal_no") or ""),
            str(block.get("size_type") or ""),
        ]).lower()
        lots = block.get("lots") or []
        lot_match = any(fv in str(l.get("lot_no") or "").lower() for l in lots)
        if ft == "container_no" and fv in str(block.get("container_no") or "").lower():
            out.append(block)
        elif ft == "lot_no" and lot_match:
            out.append(block)
        elif ft in {"all", "seal_no", "size_type"} and fv in block_text:
            out.append(block)
    return out


def _summarize_outbound_rows(rows: list[dict]) -> dict:
    return {
        "rows": len(rows),
        "lots": len({r.get("lot_no") for r in rows if r.get("lot_no")}),
        "normal_rows": sum(1 for r in rows if int(r.get("is_sample") or 0) == 0),
        "sample_rows": sum(1 for r in rows if int(r.get("is_sample") or 0) == 1),
        "nw_mt": round(sum(float(r.get("nw_mt") or 0) for r in rows), 6),
        "gw_mt": round(sum(float(r.get("gw_mt") or 0) for r in rows), 6),
    }


def _summarize_export_blocks(blocks: list[dict]) -> dict:
    lots = [lot for b in blocks for lot in (b.get("lots") or [])]
    return {
        "containers": len(blocks),
        "lots": len({l.get("lot_no") for l in lots if l.get("lot_no")}),
        "normal_qty": sum(int(l.get("normal_qty") or 0) for l in lots),
        "sample_qty": sum(int(l.get("sample_qty") or 0) for l in lots),
        "nw_mt": round(sum(float(l.get("normal_nw_mt") or 0) + float(l.get("sample_nw_mt") or 0) for l in lots), 6),
        "gw_mt": round(sum(float(l.get("normal_gw_mt") or 0) + float(l.get("sample_gw_mt") or 0) for l in lots), 6),
    }


def _outbound_value(row: dict, field: str):
    if field == "fixed_1":
        return 1
    return row.get(field)


def _storage_confirmation_rows(con: sqlite3.Connection) -> list[dict]:
    rows = _rows(con.execute(
        """
        SELECT
            i.lot_no,
            i.sap_no,
            COALESCE(i.product_code, i.product, '') AS part_no,
            COALESCE(i.product, '') AS part_description,
            COALESCE(i.stock_date, i.inbound_date, '') AS in_date,
            COALESCE(di.net_weight_kg, i.initial_weight, i.net_weight, 0) AS invoice_net_weight,
            COALESCE(i.net_weight, i.current_weight, i.initial_weight, 0) AS inspection_net_weight,
            COALESCE(i.container_no, ci.container_no, '') AS container_no,
            COALESCE(i.bl_no, db.bl_no, dd.bl_no, '') AS bl_no,
            COALESCE(i.remarks, '') AS damage_reason
        FROM inventory i
        LEFT JOIN document_invoice di ON di.lot_no = i.lot_no
        LEFT JOIN document_bl db ON db.lot_no = i.lot_no
        LEFT JOIN document_do dd ON dd.lot_no = i.lot_no
        LEFT JOIN container_info ci ON ci.lot_no = i.lot_no
        GROUP BY i.lot_no
        ORDER BY i.sap_no, i.container_no, i.lot_no
        """
    ).fetchall())
    for idx, r in enumerate(rows, start=1):
        invoice = float(r.get("invoice_net_weight") or 0)
        inspection = float(r.get("inspection_net_weight") or 0)
        r["no"] = idx
        r["balance"] = round(inspection - invoice, 4)
        r["damage_weight"] = 0
    return rows


def _sold_inventory_rows(con: sqlite3.Connection) -> list[dict]:
    rows = _rows(con.execute(
        """
        WITH sold AS (
            SELECT
                lot_no,
                MAX(COALESCE(NULLIF(customer, ''), '')) AS sold_to,
                MAX(COALESCE(NULLIF(sales_order_no, ''), '')) AS sale_ref,
                MAX(COALESCE(NULLIF(delivery_date, ''), NULLIF(sold_date, ''), '')) AS actual_pick_up,
                MAX(COALESCE(NULLIF(remark, ''), '')) AS sold_remark,
                SUM(COALESCE(sold_qty_kg, 0)) / 1000.0 AS picked_up_qty_mt,
                SUM(COALESCE(gross_weight_kg, 0)) / 1000.0 AS sold_gw_mt
            FROM sold_table
            WHERE COALESCE(status, '') IN ('SOLD', 'CONFIRMED')
            GROUP BY lot_no
        ),
        ap AS (
            SELECT
                lot_no,
                MAX(COALESCE(NULLIF(customer, ''), '')) AS ap_sold_to,
                MAX(COALESCE(NULLIF(sale_ref, ''), '')) AS ap_sale_ref,
                MAX(COALESCE(NULLIF(sc_rcvd, ''), '')) AS sc_rcvd
            FROM allocation_plan
            GROUP BY lot_no
        )
        SELECT
            i.lot_no,
            COALESCE(i.product, '') AS product,
            COALESCE(i.sap_no, '') AS sap_no,
            COALESCE(i.arrival_date, dd.arrival_date, '') AS eta_busan,
            COALESCE(i.stock_date, i.inbound_date, dd.stock_date, '') AS date_in_stock,
            COALESCE(ap.sc_rcvd, '') AS sc_rcvd,
            COALESCE(i.initial_weight, i.net_weight, 0) / 1000.0 AS qty_mt,
            COALESCE(i.warehouse, i.location, dd.warehouse_name, '') AS wh,
            COALESCE(i.salar_invoice_no, di.salar_invoice_no, '') AS salar_invoice_no,
            COALESCE(s.sold_to, ap.ap_sold_to, i.sold_to, '') AS sold_to,
            COALESCE(s.sale_ref, ap.ap_sale_ref, i.sale_ref, '') AS sale_ref,
            COALESCE(i.invoice_date, di.invoice_date, '') AS invoice_date,
            COALESCE(s.picked_up_qty_mt, 0) AS picked_up_qty_mt,
            COALESCE(i.current_weight, 0) / 1000.0 AS balance,
            COALESCE(s.sold_gw_mt, i.gross_weight / 1000.0, 0) AS gw,
            COALESCE(s.actual_pick_up, '') AS actual_pick_up,
            COALESCE(i.remarks, s.sold_remark, '') AS remark
        FROM inventory i
        LEFT JOIN sold s ON s.lot_no = i.lot_no
        LEFT JOIN ap ON ap.lot_no = i.lot_no
        LEFT JOIN document_invoice di ON di.lot_no = i.lot_no
        LEFT JOIN document_do dd ON dd.lot_no = i.lot_no
        ORDER BY actual_pick_up, i.sap_no, i.lot_no
        """
    ).fetchall())
    today = date.today()
    for r in rows:
        r["days"] = ""
        r["old"] = ""
        r["condition"] = ""
        stock = _fmt_date(r.get("date_in_stock"))
        if stock:
            try:
                stock_dt = date.fromisoformat(stock)
                r["days"] = (today - stock_dt).days
                r["old"] = r["days"]
            except Exception:
                pass
    return rows


def _filter_generic_rows(rows: list[dict], filter_type: str, filter_value: str, start_date: str = "", end_date: str = "", filter_values: str = "") -> list[dict]:
    ft = str(filter_type or "all").strip()
    fv = str(filter_value or "").strip().lower()
    date_keys = ["actual_pick_up", "in_date", "date_in_stock", "invoice_date", "eta_busan"]
    if start_date or end_date:
        out = []
        for row in rows:
            d = ""
            for key in date_keys:
                d = _fmt_date(row.get(key))
                if d:
                    break
            if start_date and d and d < start_date:
                continue
            if end_date and d and d > end_date:
                continue
            out.append(row)
        rows = out
    if ft == "all" or not fv:
        return rows
    vals = _split_filter_values(filter_values, filter_value)
    if ft == "container_no" and vals:
        allowed = {v.lower() for v in vals}
        rows = [r for r in rows if str(r.get("container_no") or "").lower() in allowed]
        return _order_rows_by_filter_values(rows, "container_no", vals)
    return [r for r in rows if fv in str(r.get(ft) or "").lower()]


def _summarize_generic_rows(rows: list[dict]) -> dict:
    qty_total = 0.0
    for r in rows:
        if "qty_mt" in r:
            qty_total += float(r.get("qty_mt") or 0)
        elif "invoice_net_weight" in r:
            qty_total += float(r.get("invoice_net_weight") or 0) / 1000.0
    return {
        "rows": len(rows),
        "lots": len({r.get("lot_no") for r in rows if r.get("lot_no")}),
        "qty_mt": round(qty_total, 4),
        "picked_up_qty_mt": round(sum(float(r.get("picked_up_qty_mt") or 0) for r in rows), 4),
    }


def _export_lot_value(block: dict, lot: dict, field: str, sample: bool):
    if field == "fixed_1":
        return 1
    if field == "qty":
        return int(lot.get("sample_qty" if sample else "normal_qty") or 0)
    if field == "nw_mt":
        return round(float(lot.get("sample_nw_mt" if sample else "normal_nw_mt") or 0), 6)
    if field == "gw_mt":
        return round(float(lot.get("sample_gw_mt" if sample else "normal_gw_mt") or 0), 6)
    if field == "description":
        desc = lot.get("description") or ""
        return (desc + " SAMPLE").strip() if sample and "SAMPLE" not in desc.upper() else desc
    if field in {"container_no", "seal_no", "size_type"}:
        return block.get(field) or ("20FT" if field == "size_type" else "")
    return lot.get(field)


# ── Sales Order / DN 보고서 ───────────────────────────────────────
@router.get("/sales-order-dn", summary="📋 Sales Order DN 보고서 (F054)")
def get_sales_order_dn():
    """
    allocation_plan + document_do JOIN → Sales Order DN 현황
    allocation_plan이 비어있으면 document_do 단독 조회
    """
    try:
        con = _db()

        # allocation_plan 존재 여부 확인
        ap_count = con.execute("SELECT COUNT(*) FROM allocation_plan").fetchone()[0]

        if ap_count > 0:
            rows = con.execute("""
                SELECT ap.lot_no, ap.sub_lt, ap.customer, ap.sale_ref,
                       ap.qty_mt, ap.outbound_date, ap.status AS ap_status,
                       ap.picking_no, ap.bl_no,
                       d.do_no, d.vessel, d.voyage, d.port_of_discharge,
                       d.arrival_date, d.stock_date, d.free_time
                FROM allocation_plan ap
                LEFT JOIN document_do d ON d.lot_no = ap.lot_no
                ORDER BY ap.outbound_date DESC, ap.lot_no
                LIMIT 200
            """).fetchall()
        else:
            # allocation_plan 비어있음 → document_do 단독
            rows = con.execute("""
                SELECT lot_no, do_no, bl_no, sap_no,
                       vessel, voyage, carrier_id,
                       port_of_loading, port_of_discharge,
                       arrival_date, stock_date, free_time,
                       total_containers, total_packages,
                       gross_weight_kg, warehouse_name,
                       created_at
                FROM document_do
                ORDER BY arrival_date DESC, lot_no
                LIMIT 200
            """).fetchall()

        con.close()
        return ok_response(data={
            "items": _rows(rows),
            "total": len(rows),
            "source": "allocation_plan+document_do" if ap_count > 0 else "document_do",
            "note": "Sales Order DN 현황 보고서",
        })
    except Exception as e:
        logger.error("sales-order-dn error: %s", e)
        return err_response(str(e))


@router.get("/sales-order-nos", summary="📋 Sales Order No 목록")
def get_sales_order_nos(limit: int = QP(100, ge=1, le=500)):
    """sold_table 기준 Sales Order No 선택 목록."""
    try:
        con = _db()
        rows = con.execute(
            """
            SELECT
                sales_order_no,
                COUNT(*) AS row_count,
                ROUND(SUM(COALESCE(sold_qty_kg, 0)) / 1000.0, 4) AS total_mt,
                MAX(COALESCE(sold_date, created_at, '')) AS last_date
            FROM sold_table
            WHERE sales_order_no IS NOT NULL
              AND TRIM(sales_order_no) != ''
              AND COALESCE(status, '') IN ('SOLD', 'CONFIRMED')
            GROUP BY sales_order_no
            ORDER BY last_date DESC, sales_order_no DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
        con.close()
        return ok_response(data={"items": _rows(rows), "total": len(rows)})
    except Exception as e:
        logger.error("sales-order-nos error: %s", e)
        return err_response(str(e))


@router.get("/sales-order-dn-template", summary="📋 Sales Order DN 템플릿 Excel 생성")
def export_sales_order_dn_template(sales_order_no: str = QP(..., min_length=1)):
    """
    Sales Order No 선택 → sold_table 조회 → sales_order_dn_template.xlsx 복사/채움 → 다운로드.
    템플릿은 5행 헤더, 6행부터 데이터 영역으로 사용한다.
    """
    so_no = str(sales_order_no or "").strip()
    if not so_no:
        raise HTTPException(status_code=400, detail="Sales Order No가 필요합니다.")

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 미설치")

    con = _db()
    try:
        rows = _sales_order_dn_rows(con, so_no)
    finally:
        con.close()

    if not rows:
        raise HTTPException(status_code=404, detail=f"Sales Order No '{so_no}' 출고 데이터가 없습니다.")

    template = _sales_order_dn_template_path()
    wb = load_workbook(template)
    ws = wb["DN"] if "DN" in wb.sheetnames else wb.active

    ws["B3"] = "Sales order No :"
    ws["C3"] = so_no

    start_row = 6
    for r in range(start_row, max(ws.max_row, start_row) + 1):
        for c in range(2, 14):
            ws.cell(r, c).value = None

    for offset, row in enumerate(rows):
        excel_row = start_row + offset
        net_kg = float(row.get("net_kg") or 0)
        gross_kg = _estimate_gross_kg(row)
        destination = (
            row.get("customer")
            or row.get("final_destination")
            or row.get("port_of_discharge")
            or row.get("warehouse_name")
            or ""
        )
        sku = row.get("sku") or row.get("product_code") or ""
        description = row.get("product") or sku

        ws.cell(excel_row, 2).value = destination
        ws.cell(excel_row, 3).value = _fmt_date(row.get("delivery_date"))
        ws.cell(excel_row, 4).value = row.get("lot_no") or ""
        ws.cell(excel_row, 5).value = row.get("sap_no") or ""
        ws.cell(excel_row, 6).value = row.get("bl_no") or ""
        ws.cell(excel_row, 7).value = so_no
        ws.cell(excel_row, 8).value = row.get("picking_no") or ""
        ws.cell(excel_row, 9).value = sku
        ws.cell(excel_row, 10).value = description
        ws.cell(excel_row, 11).value = _fmt_mt(net_kg)
        ws.cell(excel_row, 12).value = _fmt_mt(gross_kg)
        ws.cell(excel_row, 13).value = int(row.get("ct_plt") or 0)

    out_dir = Path(tempfile.gettempdir()) / "sqm_exports"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_name = f"Sales_order_DN_{_safe_filename_part(so_no)}.xlsx"
    out_path = out_dir / out_name
    wb.save(out_path)

    return FileResponse(
        path=str(out_path),
        filename=out_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/report-templates", summary="보고서별 템플릿 목록")
def list_report_templates(report_type: str = QP(...)):
    _ensure_default_template_names()
    d = _report_template_dir(report_type)
    items = []
    for p in sorted(d.glob("template_*.xlsx")):
        st = p.stat()
        items.append({
            "name": p.stem,
            "display_name": _template_display_name(report_type, p.stem),
            "filename": p.name,
            "size_bytes": st.st_size,
            "modified_at": datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
        })
    return ok_response(data={"report_type": report_type, "label": _REPORT_TYPES[report_type], "items": items})


@router.post("/report-template-upload", summary="보고서별 템플릿 추가")
async def upload_report_template(
    report_type: str = QP(...),
    display_name: str = QP("", max_length=120),
    mode: str = QP("add"),
    template: str = QP(""),
    file: UploadFile = File(...),
):
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="xlsx 파일만 템플릿으로 저장할 수 있습니다.")
    mode_key = str(mode or "add").strip()
    if mode_key == "update_current":
        if not template:
            raise HTTPException(status_code=400, detail="업데이트할 템플릿을 선택해야 합니다.")
        dest = _report_template_path(report_type, template)
    elif mode_key == "update_default":
        dest = _report_template_dir(report_type) / "template_1.xlsx"
    else:
        dest = _next_template_path(report_type)
    content = await file.read()
    if len(content) > 50 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="파일 크기 50MB 초과")
    dest.write_bytes(content)
    if mode_key == "update_current" and not display_name:
        display = _template_display_name(report_type, dest.stem)
    elif mode_key == "update_default" and not display_name:
        display = f"{_REPORT_TYPES[report_type]} 기본 양식"
    else:
        display = display_name or dest.stem
    _set_template_display_name(report_type, dest.stem, display, file.filename or "")
    action_label = "업데이트" if mode_key in {"update_current", "update_default"} else "저장"
    return ok_response(data={
        "report_type": report_type,
        "template": dest.stem,
        "display_name": _template_display_name(report_type, dest.stem),
        "filename": dest.name,
        "size_bytes": len(content),
        "mode": mode_key,
    }, message=f"{_REPORT_TYPES[report_type]} {dest.stem} {action_label} 완료")


@router.delete("/report-template", summary="보고서 템플릿 삭제")
def delete_report_template(report_type: str = QP(...), template: str = QP(...)):
    path = _report_template_path(report_type, template)
    path.unlink()
    meta = _load_report_template_meta(report_type)
    meta.pop(template, None)
    _save_report_template_meta(report_type, meta)
    cols = _load_report_template_columns(report_type)
    cols.pop(template, None)
    _save_report_template_columns(report_type, cols)
    return ok_response(message=f"{_REPORT_TYPES[report_type]} {template} 삭제 완료")


@router.post("/report-template-columns", summary="보고서 템플릿 컬럼 구성 저장")
async def save_report_template_columns(report_type: str = QP(...), template: str = QP(...), payload: dict = Body(default_factory=dict)):
    _report_template_path(report_type, template)
    columns = (payload or {}).get("columns") or []
    _set_template_columns(report_type, template, columns)
    return ok_response(data={
        "report_type": report_type,
        "template": template,
        "columns": _get_template_columns(report_type, template),
    }, message="컬럼 구성을 저장했습니다.")


@router.get("/report-filter-values", summary="보고서 범위 기준값 목록")
def report_filter_values(report_type: str = QP(...), filter_type: str = QP("all"), limit: int = QP(300, ge=1, le=1000)):
    ft = str(filter_type or "all").strip()
    if ft == "all":
        return ok_response(data={"items": [], "values": []})
    con = _db()
    try:
        if report_type == "outbound_report":
            source = _outbound_report_rows(con)
            key_map = {
                "sales_order_no": "sales_order_no",
                "customer": "destination",
                "bl_no": "bl_no",
                "container_no": "container_no",
                "lot_no": "lot_no",
                "picking_no": "picking_no",
            }
            key = key_map.get(ft)
            vals = sorted({str(r.get(key) or "").strip() for r in source if key and str(r.get(key) or "").strip()})
            vals = vals[:limit]
            return ok_response(data={"items": vals, "values": vals})
        if report_type == "sales_order_dn":
            source = _sales_order_dn_report_rows(con)
            key_map = {
                "sales_order_no": "sales_order_no",
                "customer": "destination",
                "bl_no": "bl_no",
                "container_no": "container_no",
                "lot_no": "lot_no",
                "picking_no": "picking_no",
            }
            key = key_map.get(ft)
            vals = sorted({str(r.get(key) or "").strip() for r in source if key and str(r.get(key) or "").strip()})
            vals = vals[:limit]
            return ok_response(data={"items": vals, "values": vals})
        if report_type in {"storage_confirmation", "sold_inventory_report"}:
            source = _storage_confirmation_rows(con) if report_type == "storage_confirmation" else _sold_inventory_rows(con)
            vals = sorted({str(r.get(ft) or "").strip() for r in source if str(r.get(ft) or "").strip()})
            vals = vals[:limit]
            return ok_response(data={"items": vals, "values": vals})
        if report_type == "export_work_report":
            blocks = _export_work_blocks(con)
            vals = set()
            for b in blocks:
                if ft == "container_no" and b.get("container_no"):
                    vals.add(str(b.get("container_no")))
                elif ft == "seal_no" and b.get("seal_no"):
                    vals.add(str(b.get("seal_no")))
                elif ft == "size_type" and b.get("size_type"):
                    vals.add(str(b.get("size_type")))
                elif ft == "lot_no":
                    vals.update(str(l.get("lot_no")) for l in b.get("lots", []) if l.get("lot_no"))
            vals = sorted(vals)[:limit]
            return ok_response(data={"items": vals, "values": vals})
        raise HTTPException(status_code=400, detail="알 수 없는 보고서 유형입니다.")
    finally:
        con.close()


@router.get("/report-template-analyze", summary="보고서 템플릿 헤더 분석")
def analyze_report_template(report_type: str = QP(...), template: str = QP("template_1")):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 미설치")

    wb = load_workbook(_report_template_path(report_type, template), data_only=False)
    ws = wb["Outbound"] if report_type == "outbound_report" and "Outbound" in wb.sheetnames else wb.active
    if report_type in {"outbound_report", "sales_order_dn", "storage_confirmation", "sold_inventory_report"}:
        header_row = 3 if report_type == "storage_confirmation" else (2 if report_type == "sold_inventory_report" else 5)
        data_start = header_row + 1
        headers = []
        defaults = {
            "Destination": "customer",
            "Delivery Date": "delivery_date",
            "LOT NO": "lot_no",
            "SAP NO": "sap_no",
            "BL NO": "bl_no",
            "Sales order No": "sales_order_no",
            "Picking No": "picking_no",
            "SKU": "sku",
            "Description": "description",
            "NW(MT)": "nw_mt",
            "GW(MT)": "gw_mt",
            "CT/PLT": "qty",
            "NO": "no",
            "PART NO": "part_no",
            "PART DESCRIPTION": "part_description",
            "IN DATE": "in_date",
            "INVOICE": "invoice_net_weight",
            "INSPCECTION": "inspection_net_weight",
            "BALANCE": "balance",
            "DAMAGE": "damage_weight",
            "DAMAGE 사유": "damage_reason",
            "CONT NO": "container_no",
            "Product": "product",
            "ETA BUSAN": "eta_busan",
            "Date in stock": "date_in_stock",
            "SC RCVD": "sc_rcvd",
            "Days": "days",
            "QTY (MT)": "qty_mt",
            "Lot No": "lot_no",
            "WH": "wh",
            "Salar Invoice no.": "salar_invoice_no",
            "SOLD TO": "sold_to",
            "SALE REF": "sale_ref",
            "Invoice date": "invoice_date",
            "Picked up Qty (MT)": "picked_up_qty_mt",
            "Balance": "balance",
            "GW": "gw",
            "Actual pick up": "actual_pick_up",
            "Old": "old",
            "Condition": "condition",
            "Remark": "remark",
        }
        for c in range(1, ws.max_column + 1):
            val = ws.cell(header_row, c).value
            if val:
                h = str(val).strip()
                headers.append({"column": c, "header": h, "default_field": defaults.get(h, "")})
        return ok_response(data={
            "report_type": report_type,
            "template": template,
            "display_name": _template_display_name(report_type, template),
            "sheet": ws.title,
            "header_row": header_row,
            "data_start_row": data_start,
            "headers": headers,
            "columns": _get_template_columns(report_type, template),
            "available_fields": _REPORT_FIELDS[report_type],
            "available_filters": ["all"] + [f["field"] for f in _REPORT_FIELDS[report_type]],
        })

    title_rows = sorted(rng.min_row for rng in ws.merged_cells.ranges if rng.min_col == 1 and rng.max_col == 6 and rng.min_row == rng.max_row)
    first = title_rows[0] if title_rows else 2
    header_row = first + 2
    headers = []
    defaults = {
        "Bag": "fixed_1",
        "Description of goods": "description",
        "Lot No.": "lot_no",
        "Q'ty": "qty",
        "Net Weight": "nw_mt",
        "Gross Weight": "gw_mt",
    }
    for c in range(1, 7):
        val = ws.cell(header_row, c).value
        if val:
            h = str(val).strip()
            headers.append({"column": c, "header": h, "default_field": defaults.get(h, "")})
    return ok_response(data={
        "report_type": report_type,
        "template": template,
        "display_name": _template_display_name(report_type, template),
        "sheet": ws.title,
        "block_title_rows": title_rows,
        "header_row": header_row,
        "data_start_row": header_row + 1,
        "headers": headers,
        "columns": _get_template_columns(report_type, template),
        "available_fields": _REPORT_FIELDS[report_type],
        "block_title_mapping": {
            "container_no": "container_info.container_no",
            "seal_no": "container_info.seal_no",
            "size_type": "container_info.size_type, 없으면 20FT",
        },
        "available_filters": ["all", "container_no", "lot_no", "seal_no", "size_type"],
    })


@router.get("/report-template-preview", summary="보고서 템플릿 미리보기")
def preview_report_template(
    report_type: str = QP(...),
    template: str = QP("template_1"),
    max_rows: int = QP(14, ge=1, le=40),
    max_cols: int = QP(16, ge=1, le=30),
):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 미설치")

    wb = load_workbook(_report_template_path(report_type, template), data_only=False)
    if report_type == "outbound_report" and "Outbound" in wb.sheetnames:
        ws = wb["Outbound"]
    elif report_type == "sales_order_dn" and "DN" in wb.sheetnames:
        ws = wb["DN"]
    else:
        ws = wb.active

    rows = []
    row_limit = min(ws.max_row or 1, max_rows)
    col_limit = min(ws.max_column or 1, max_cols)
    for r in range(1, row_limit + 1):
        vals = []
        for c in range(1, col_limit + 1):
            v = ws.cell(r, c).value
            vals.append("" if v is None else str(v))
        rows.append(vals)

    return ok_response(data={
        "report_type": report_type,
        "template": template,
        "display_name": _template_display_name(report_type, template),
        "sheet": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "rows": rows,
    })


@router.get("/report-template-file", summary="보고서 템플릿 원본 Excel 열기")
def download_report_template_file(report_type: str = QP(...), template: str = QP("template_1")):
    path = _report_template_path(report_type, template)
    out_name = f"{_safe_filename_part(_template_display_name(report_type, template))}_{template}.xlsx"
    return FileResponse(
        str(path),
        filename=out_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.post("/report-template-open", summary="보고서 템플릿 원본 Excel 로컬 열기")
def open_report_template_file(report_type: str = QP(...), template: str = QP("template_1")):
    path = _report_template_path(report_type, template)
    try:
        os.startfile(str(path))
    except AttributeError:
        raise HTTPException(status_code=500, detail="이 환경에서는 로컬 Excel 열기를 지원하지 않습니다.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"템플릿 파일 열기 실패: {e}")
    return ok_response(data={
        "report_type": report_type,
        "template": template,
        "path": str(path),
    }, message="원본 Excel 파일을 열었습니다.")


@router.get("/report-preview", summary="보고서 생성 대상 미리보기")
def preview_template_report(
    report_type: str = QP(...),
    filter_type: str = QP("all"),
    filter_value: str = QP(""),
    filter_values: str = QP(""),
    start_date: str = QP(""),
    end_date: str = QP(""),
):
    con = _db()
    try:
        if report_type == "outbound_report":
            rows = _filter_outbound_rows(_outbound_report_rows(con), filter_type, filter_value, start_date, end_date, filter_values)
            return ok_response(data={"summary": _summarize_outbound_rows(rows), "sample": rows[:12]})
        if report_type == "sales_order_dn":
            rows = _filter_sales_order_dn_rows(_sales_order_dn_report_rows(con), filter_type, filter_value, start_date, end_date, filter_values)
            validation_so = filter_value if filter_type == "sales_order_no" else ""
            try:
                from backend.api.sales_order_validation import validate_sales_order_no
                allocation_validation = validate_sales_order_no(validation_so)
            except Exception as ve:
                allocation_validation = {
                    "level": "warning",
                    "issues": [{"severity": "warning", "message": f"Allocation 검증 실패: {ve}"}],
                    "context": {},
                }
            return ok_response(data={
                "summary": _summarize_outbound_rows(rows),
                "sample": rows[:12],
                "allocation_validation": allocation_validation,
            })
        if report_type == "export_work_report":
            blocks = _filter_export_blocks(_export_work_blocks(con), filter_type, filter_value, filter_values)
            return ok_response(data={"summary": _summarize_export_blocks(blocks), "sample": blocks[:5]})
        if report_type == "storage_confirmation":
            rows = _filter_generic_rows(_storage_confirmation_rows(con), filter_type, filter_value, start_date, end_date, filter_values)
            return ok_response(data={"summary": _summarize_generic_rows(rows), "sample": rows[:12]})
        if report_type == "sold_inventory_report":
            rows = _filter_generic_rows(_sold_inventory_rows(con), filter_type, filter_value, start_date, end_date, filter_values)
            return ok_response(data={"summary": _summarize_generic_rows(rows), "sample": rows[:12]})
        raise HTTPException(status_code=400, detail="알 수 없는 보고서 유형입니다.")
    finally:
        con.close()


@router.get("/outbound-report-excel", summary="Outbound Report 템플릿 Excel 생성")
def export_outbound_report_excel(
    template: str = QP("template_1"),
    filter_type: str = QP("all"),
    filter_value: str = QP(""),
    filter_values: str = QP(""),
    start_date: str = QP(""),
    end_date: str = QP(""),
):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 미설치")

    con = _db()
    try:
        rows = _filter_outbound_rows(_outbound_report_rows(con), filter_type, filter_value, start_date, end_date, filter_values)
    finally:
        con.close()
    if not rows:
        raise HTTPException(status_code=404, detail="SOLD 데이터가 없습니다.")

    wb = load_workbook(_report_template_path("outbound_report", template))
    ws = wb["Outbound"] if "Outbound" in wb.sheetnames else wb.active
    ws["C3"] = date.today().isoformat()
    columns = [c for c in _get_template_columns("outbound_report", template) if c.get("enabled", True)]

    start_row = 6
    clear_to = max(ws.max_row, start_row + len(rows) + 5)
    clear_cols = max(13, 1 + len(columns))
    _copy_rows_style(ws, start_row, start_row, len(rows), 2, clear_cols)
    for r in range(start_row, clear_to + 1):
        for c in range(2, clear_cols + 1):
            ws.cell(r, c).value = None

    for offset, col in enumerate(columns):
        ws.cell(5, 2 + offset).value = col.get("header") or col.get("field")

    for idx, row in enumerate(rows):
        excel_row = start_row + idx
        for offset, col in enumerate(columns):
            field = col.get("field") or ""
            val = _outbound_value(row, field)
            if field in {"delivery_date", "sold_date"}:
                val = _fmt_date(val)
            elif field in {"nw_mt", "gw_mt"}:
                val = round(float(val or 0), 6)
            ws.cell(excel_row, 2 + offset).value = val if val is not None else ""

    for offset, col in enumerate(columns):
        field = col.get("field")
        if field in {"nw_mt", "gw_mt", "qty"}:
            cell = ws.cell(4, 2 + offset)
            letter = cell.column_letter
            cell.value = f"=SUBTOTAL(9,{letter}{start_row}:{letter}{start_row + len(rows) - 1})"

    out_dir = Path(tempfile.gettempdir()) / "sqm_exports"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_name = f"Outbound_Report_{date.today().isoformat()}_{_safe_filename_part(template)}.xlsx"
    out_path = out_dir / out_name
    wb.save(out_path)
    return FileResponse(str(out_path), filename=out_name, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.get("/sales-order-dn-report-excel", summary="Sales Order DN 템플릿 Excel 생성")
def export_sales_order_dn_report_excel(
    template: str = QP("template_1"),
    filter_type: str = QP("all"),
    filter_value: str = QP(""),
    filter_values: str = QP(""),
    start_date: str = QP(""),
    end_date: str = QP(""),
):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 미설치")

    con = _db()
    try:
        rows = _filter_sales_order_dn_rows(_sales_order_dn_report_rows(con), filter_type, filter_value, start_date, end_date, filter_values)
    finally:
        con.close()
    if not rows:
        raise HTTPException(status_code=404, detail="Sales Order DN 데이터가 없습니다.")

    wb = load_workbook(_report_template_path("sales_order_dn", template))
    ws = wb["DN"] if "DN" in wb.sheetnames else wb.active
    if filter_type == "sales_order_no" and filter_value:
        ws["B3"] = "Sales order No :"
        ws["C3"] = filter_value
    columns = [c for c in _get_template_columns("sales_order_dn", template) if c.get("enabled", True)]

    start_row = 6
    clear_to = max(ws.max_row, start_row + len(rows) + 5)
    clear_cols = max(13, len(columns) + 1)
    _copy_rows_style(ws, start_row, start_row, len(rows), 2, clear_cols)
    for r in range(start_row, clear_to + 1):
        for c in range(2, clear_cols + 1):
            ws.cell(r, c).value = None

    for offset, col in enumerate(columns):
        ws.cell(5, 2 + offset).value = col.get("header") or col.get("field")

    for idx, row in enumerate(rows):
        excel_row = start_row + idx
        for offset, col in enumerate(columns):
            field = col.get("field") or ""
            val = _outbound_value(row, field)
            if field == "delivery_date":
                val = _fmt_date(val)
            elif field in {"nw_mt", "gw_mt"}:
                val = round(float(val or 0), 6)
            ws.cell(excel_row, 2 + offset).value = val if val is not None else ""

    for offset, col in enumerate(columns):
        field = col.get("field")
        if field in {"nw_mt", "gw_mt", "qty"}:
            cell = ws.cell(4, 2 + offset)
            letter = cell.column_letter
            cell.value = f"=SUBTOTAL(9,{letter}{start_row}:{letter}{start_row + len(rows) - 1})"

    out_dir = Path(tempfile.gettempdir()) / "sqm_exports"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_name = f"Sales_Order_DN_{date.today().isoformat()}_{_safe_filename_part(template)}.xlsx"
    out_path = out_dir / out_name
    wb.save(out_path)
    return FileResponse(str(out_path), filename=out_name, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _export_flat_template_excel(
    report_type: str,
    rows: list[dict],
    template: str,
    sheet_name: str = "",
    header_row: int = 2,
    data_start: int = 3,
    start_col: int = 1,
):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 미설치")
    if not rows:
        raise HTTPException(status_code=404, detail=f"{_REPORT_TYPES[report_type]} 데이터가 없습니다.")
    wb = load_workbook(_report_template_path(report_type, template))
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    columns = [c for c in _get_template_columns(report_type, template) if c.get("enabled", True)]
    clear_to = max(ws.max_row, data_start + len(rows) + 5)
    clear_cols = max(ws.max_column, start_col + len(columns) - 1)
    _copy_rows_style(ws, data_start, data_start, len(rows), start_col, clear_cols)
    for r in range(data_start, clear_to + 1):
        for c in range(start_col, clear_cols + 1):
            ws.cell(r, c).value = None
    for offset, col in enumerate(columns):
        ws.cell(header_row, start_col + offset).value = col.get("header") or col.get("field")
    for idx, row in enumerate(rows):
        excel_row = data_start + idx
        for offset, col in enumerate(columns):
            val = row.get(col.get("field") or "")
            ws.cell(excel_row, start_col + offset).value = val if val is not None else ""
    out_dir = Path(tempfile.gettempdir()) / "sqm_exports"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_name = f"{_safe_filename_part(_REPORT_TYPES[report_type])}_{date.today().isoformat()}_{_safe_filename_part(template)}.xlsx"
    out_path = out_dir / out_name
    wb.save(out_path)
    return FileResponse(str(out_path), filename=out_name, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.get("/storage-confirmation-excel", summary="Storage Confirmation 템플릿 Excel 생성")
def export_storage_confirmation_excel(
    template: str = QP("template_1"),
    filter_type: str = QP("all"),
    filter_value: str = QP(""),
    filter_values: str = QP(""),
    start_date: str = QP(""),
    end_date: str = QP(""),
):
    con = _db()
    try:
        rows = _filter_generic_rows(_storage_confirmation_rows(con), filter_type, filter_value, start_date, end_date, filter_values)
    finally:
        con.close()
    return _export_flat_template_excel("storage_confirmation", rows, template, header_row=3, data_start=5, start_col=1)


@router.get("/sold-inventory-report-excel", summary="SOLD Inventory Report 템플릿 Excel 생성")
def export_sold_inventory_report_excel(
    template: str = QP("template_1"),
    filter_type: str = QP("all"),
    filter_value: str = QP(""),
    filter_values: str = QP(""),
    start_date: str = QP(""),
    end_date: str = QP(""),
):
    con = _db()
    try:
        rows = _filter_generic_rows(_sold_inventory_rows(con), filter_type, filter_value, start_date, end_date, filter_values)
    finally:
        con.close()
    return _export_flat_template_excel("sold_inventory_report", rows, template, sheet_name="통관요청", header_row=2, data_start=3, start_col=1)


@router.get("/export-work-report-excel", summary="수출 작업 리포트 템플릿 Excel 생성")
def export_work_report_excel(
    template: str = QP("template_1"),
    filter_type: str = QP("all"),
    filter_value: str = QP(""),
    filter_values: str = QP(""),
):
    try:
        from openpyxl import load_workbook
        from openpyxl.cell.cell import MergedCell
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 미설치")

    con = _db()
    try:
        blocks = _filter_export_blocks(_export_work_blocks(con), filter_type, filter_value, filter_values)
    finally:
        con.close()
    if not blocks:
        raise HTTPException(status_code=404, detail="컨테이너 데이터가 없습니다.")

    wb = load_workbook(_report_template_path("export_work_report", template))
    ws = wb.active
    columns = [c for c in _get_template_columns("export_work_report", template) if c.get("enabled", True)]
    title_rows = sorted(
        rng.min_row
        for rng in ws.merged_cells.ranges
        if rng.min_col == 1 and rng.max_col == 6 and rng.min_row == rng.max_row
    )
    if not title_rows:
        title_rows = [2 + i * 16 for i in range(len(blocks))]

    for idx, block in enumerate(blocks):
        if idx >= len(title_rows):
            break
        base = title_rows[idx]
        data_start = base + 3
        sum_row = base + 11
        seal = block.get("seal_no") or ""
        size = block.get("size_type") or "20FT"
        ws.cell(base, 1).value = f"SQM ( {block.get('container_no') or ''} / {seal} ) {size} * 1컨"
        clear_cols = max(6, len(columns))
        _copy_rows_style(ws, data_start, data_start, sum_row - data_start, 1, clear_cols)
        for offset, col in enumerate(columns):
            cell = ws.cell(base + 2, 1 + offset)
            if not isinstance(cell, MergedCell):
                cell.value = col.get("header") or col.get("field")
        for r in range(data_start, sum_row):
            for c in range(1, clear_cols + 1):
                cell = ws.cell(r, c)
                if not isinstance(cell, MergedCell):
                    cell.value = None
        out_row = data_start
        for lot in block.get("lots", []):
            if lot.get("normal_qty"):
                for offset, col in enumerate(columns):
                    cell = ws.cell(out_row, 1 + offset)
                    if not isinstance(cell, MergedCell):
                        cell.value = _export_lot_value(block, lot, col.get("field") or "", False)
                out_row += 1
        for lot in block.get("lots", []):
            if lot.get("sample_qty"):
                for offset, col in enumerate(columns):
                    cell = ws.cell(out_row, 1 + offset)
                    if not isinstance(cell, MergedCell):
                        cell.value = _export_lot_value(block, lot, col.get("field") or "", True)
                out_row += 1
        for offset, col in enumerate(columns):
            cell = ws.cell(sum_row, 1 + offset)
            if isinstance(cell, MergedCell):
                continue
            field = col.get("field")
            if field in {"qty", "nw_mt", "gw_mt"}:
                letter = cell.column_letter
                cell.value = f"=SUM({letter}{data_start}:{letter}{sum_row - 1})"
            elif field == "fixed_1":
                cell.value = None

    for base in title_rows[len(blocks):]:
        for r in range(base, min(base + 12, ws.max_row) + 1):
            for c in range(1, max(6, len(columns)) + 1):
                cell = ws.cell(r, c)
                if not isinstance(cell, MergedCell):
                    cell.value = None

    out_dir = Path(tempfile.gettempdir()) / "sqm_exports"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_name = f"Export_Work_Report_{date.today().isoformat()}_{_safe_filename_part(template)}.xlsx"
    out_path = out_dir / out_name
    wb.save(out_path)
    return FileResponse(str(out_path), filename=out_name, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── DN Cross Check ────────────────────────────────────────────────
@router.get("/dn-cross-check", summary="🔁 DN 교차검증 (F055-alt)")
def get_dn_cross_check():
    """
    inventory vs document_do 교차 비교
    — DO 있는데 재고 없는 LOT, 재고 있는데 DO 없는 LOT 검출
    """
    try:
        con = _db()

        # DO 있는데 inventory 없는 케이스
        do_no_inv = con.execute("""
            SELECT d.lot_no, d.do_no, d.bl_no, d.vessel,
                   d.arrival_date, d.gross_weight_kg,
                   'DO있음_재고없음' AS cross_status
            FROM document_do d
            LEFT JOIN inventory i ON i.lot_no = d.lot_no
            WHERE i.id IS NULL
            ORDER BY d.arrival_date DESC
        """).fetchall()

        # inventory 있는데 DO 없는 케이스
        inv_no_do = con.execute("""
            SELECT i.lot_no, i.sap_no, i.bl_no, i.product,
                   i.status, i.current_weight, i.inbound_date,
                   'DO없음_재고있음' AS cross_status
            FROM inventory i
            LEFT JOIN document_do d ON d.lot_no = i.lot_no
            WHERE d.id IS NULL
            ORDER BY i.inbound_date DESC
            LIMIT 100
        """).fetchall()

        # 매칭된 케이스 (정상)
        matched = con.execute("""
            SELECT COUNT(*) AS cnt
            FROM inventory i
            INNER JOIN document_do d ON d.lot_no = i.lot_no
        """).fetchone()

        con.close()
        return ok_response(data={
            "do_without_inventory": _rows(do_no_inv),
            "inventory_without_do": _rows(inv_no_do),
            "matched_count": matched["cnt"] if matched else 0,
            "issues_count": len(do_no_inv) + len(inv_no_do),
            "note": "교차검증 이슈가 없으면 issues_count=0",
        })
    except Exception as e:
        logger.error("dn-cross-check error: %s", e)
        return err_response(str(e))


# ── D/O 현황 ─────────────────────────────────────────────────────
@router.get("/do-status", summary="📄 D/O 현황 (F056-alt)")
def get_do_status(limit: int = QP(100, ge=1, le=500)):
    """document_do 전체 목록 + inventory JOIN 상태"""
    try:
        con = _db()
        rows = con.execute("""
            SELECT d.id, d.lot_no, d.do_no, d.bl_no, d.sap_no,
                   d.vessel, d.voyage, d.carrier_id,
                   d.port_of_loading, d.port_of_discharge,
                   d.arrival_date, d.stock_date, d.free_time,
                   d.con_return, d.total_containers,
                   d.gross_weight_kg, d.warehouse_name,
                   COALESCE(i.status, '미입고') AS inv_status,
                   i.current_weight, i.location,
                   d.created_at
            FROM document_do d
            LEFT JOIN inventory i ON i.lot_no = d.lot_no
            ORDER BY d.arrival_date DESC, d.lot_no
            LIMIT ?
        """, (limit,)).fetchall()

        # 상태별 집계
        summary = con.execute("""
            SELECT COALESCE(i.status, '미입고') AS status,
                   COUNT(*) AS cnt
            FROM document_do d
            LEFT JOIN inventory i ON i.lot_no = d.lot_no
            GROUP BY status
            ORDER BY cnt DESC
        """).fetchall()
        con.close()

        return ok_response(data={
            "items": _rows(rows),
            "total": len(rows),
            "summary_by_status": _rows(summary),
            "columns": ["lot_no", "do_no", "bl_no", "vessel",
                        "arrival_date", "stock_date", "inv_status",
                        "gross_weight_kg", "warehouse_name"],
        })
    except Exception as e:
        logger.error("do-status error: %s", e)
        return err_response(str(e))


# ── 거래명세서 목록 ─────────────────────────────────────────────
@router.get("/invoice-list", summary="🧾 거래명세서 목록 (F045-alt)")
def get_invoice_list(
    customer: str = QP(None, description="고객명 필터"),
    limit: int = QP(100, ge=1, le=500),
):
    """document_invoice 목록 — 거래명세서 조회"""
    try:
        con = _db()

        if customer:
            rows = con.execute("""
                SELECT id, lot_no, sap_no, invoice_no, salar_invoice_no,
                       invoice_date, customer_code, customer_name,
                       customer_ref, product_name, quantity_mt,
                       unit_price, total_amount, currency,
                       net_weight_kg, vessel, origin, destination,
                       incoterm, payment_term, bl_no, created_at
                FROM document_invoice
                WHERE customer_name LIKE ?
                ORDER BY invoice_date DESC, lot_no
                LIMIT ?
            """, (f"%{customer}%", limit)).fetchall()
        else:
            rows = con.execute("""
                SELECT id, lot_no, sap_no, invoice_no, salar_invoice_no,
                       invoice_date, customer_code, customer_name,
                       customer_ref, product_name, quantity_mt,
                       unit_price, total_amount, currency,
                       net_weight_kg, vessel, origin, destination,
                       incoterm, payment_term, bl_no, created_at
                FROM document_invoice
                ORDER BY invoice_date DESC, lot_no
                LIMIT ?
            """, (limit,)).fetchall()

        # 고객별 집계
        by_customer = con.execute("""
            SELECT customer_name,
                   COUNT(*) AS invoice_cnt,
                   ROUND(SUM(quantity_mt), 3) AS total_mt,
                   ROUND(SUM(total_amount), 2) AS total_amount,
                   MAX(currency) AS currency
            FROM document_invoice
            GROUP BY customer_name
            ORDER BY total_mt DESC
        """).fetchall()
        con.close()

        return ok_response(data={
            "items": _rows(rows),
            "total": len(rows),
            "by_customer": _rows(by_customer),
            "columns": ["invoice_no", "invoice_date", "customer_name",
                        "lot_no", "product_name", "quantity_mt",
                        "total_amount", "currency", "vessel"],
        })
    except Exception as e:
        logger.error("invoice-list error: %s", e)
        return err_response(str(e))


# ── 시스템 설정 정보 ──────────────────────────────────────────────
@router.get("/settings-info", summary="⚙️ 시스템 설정 정보 (F058-alt)")
def get_settings_info():
    """DB 메타데이터 + 테이블별 행수 + 시스템 정보"""
    try:
        con = _db()

        # 테이블별 행수
        tables = [
            "inventory", "inventory_tonbag", "stock_movement",
            "audit_log", "return_history", "allocation_plan",
            "document_do", "document_invoice", "document_bl",
        ]
        table_stats = []
        for t in tables:
            try:
                cnt = con.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
                table_stats.append({"table": t, "rows": cnt})
            except Exception:
                table_stats.append({"table": t, "rows": -1, "error": "테이블 없음"})

        # DB 파일 크기
        db_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
            "data", "db", "sqm_inventory.db"
        )
        try:
            db_size_mb = round(os.path.getsize(db_path) / 1024 / 1024, 3)
        except Exception:
            db_size_mb = -1

        # journal mode 확인
        jm = con.execute("PRAGMA journal_mode").fetchone()[0]
        wal_size = con.execute("PRAGMA page_count").fetchone()[0]

        con.close()
        return ok_response(data={
            "version": "SQM v864.3",
            "db_path": db_path,
            "db_size_mb": db_size_mb,
            "journal_mode": jm,
            "page_count": wal_size,
            "table_stats": table_stats,
            "note": "DB 메타데이터 및 테이블별 행수",
        })
    except Exception as e:
        logger.error("settings-info error: %s", e)
        return err_response(str(e))
