"""
SQM 재고관리 시스템 - Excel 내보내기 Mixin
==========================================

v3.6.0: Docstring 보강

모듈 개요:
    재고 데이터를 Excel 파일로 내보내는 Mixin 클래스입니다.
    다양한 출력 옵션을 지원합니다.

내보내기 옵션:
    - Option 1: 재고 현황 (LOT별)
    - Option 2: 상세 재고 (v5.5.3 P4 한글화)
    - Option 3: 재고리스트 Excel (v5.5.3 P1 18열)
    - Option 4: 톤백리스트 Excel (v5.5.3 P2 21열)
    - Option 5: LOT-톤백 리포트 (v5.5.3 P4 서식)
    - Option 6: 전체 재고 (v5.5.3 P4 5시트 서식)

"""
from engine_modules.constants import STATUS_AVAILABLE, STATUS_DEPLETED, STATUS_PICKED, STATUS_SOLD
import logging
from pathlib import Path

logger = logging.getLogger(__name__)


def _unique_excel_path(desired_path: str) -> str:
    """같은 이름 파일이 있으면 숫자 붙인 경로 반환 (기존 파일 덮어쓰지 않음)."""
    p = Path(desired_path)
    if not p.exists():
        return desired_path
    parent, stem, ext = p.parent, p.stem, p.suffix
    for n in range(1, 100):
        candidate = parent / f"{stem}_{n}{ext}"
        if not candidate.exists():
            logger.info(f"기존 파일 존재 → 저장: {candidate.name}")
            return str(candidate)
    return str(parent / f"{stem}_99{ext}")


class ExportMixin:
    """엑셀 내보내기 Mixin"""

    def export_to_excel(self, output_path: str, option: int = 1, include_sample: bool = True) -> str:
        """
        엑셀 내보내기.
        같은 이름의 파일이 있으면 파일명 끝에 _1, _2 ... 를 붙여 새 파일로 저장.

        Args:
            output_path: 출력 파일 경로
            option: 내보내기 옵션
                1: 기본 재고 목록
                2: 상세 재고
                3: Ruby 포맷
                4: 톤백 목록
                5: LOT-톤백 리포트
                6: 전체 재고
            include_sample: option=4(톤백리스트)에서 샘플 포함 여부

        Returns:
            실제 저장된 파일 경로
        """
        output_path = _unique_excel_path(output_path)
        exporters = {
            1: self._export_basic_inventory,
            2: self._export_detailed_inventory,
            3: self._export_ruby_format,
            4: (lambda p: self._export_tonbag_list(p, include_sample=include_sample)),
            5: self._export_lot_tonbag_report,
            6: self._export_full_inventory,
            8: self._export_return_history,
            9: self._export_integrity_report,
        }

        exporter = exporters.get(option, self._export_basic_inventory)
        return exporter(output_path)

    def _apply_excel_formatting(self, ws, df, title: str = '',
                                col_widths: dict = None, num_formats: dict = None) -> None:
        """v5.5.3 P1: 공통 Excel 서식 (타이틀, 헤더, 컬럼 너비, 숫자 포맷, 푸터)

        Args:
            ws: openpyxl Worksheet
            df: pandas DataFrame (컬럼명 참조용)
            title: Row 1 타이틀 텍스트
            col_widths: {헤더명: 너비} — 미지정 시 15
            num_formats: {헤더명: '#,##0' 등} — 숫자 컬럼 포맷
        """
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        col_count = len(df.columns)
        data_start_row = 3  # Row 1=타이틀, Row 2=빈줄, Row 3=헤더

        # ── Row 1: 타이틀 ──
        if title:
            ws.merge_cells(f'A1:{get_column_letter(col_count)}1')
            cell = ws['A1']
            cell.value = title
            cell.font = Font(name='맑은 고딕', bold=True, size=14, color='2C3E50')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 32

        # ── Row 2: 간격 ──
        ws.row_dimensions[2].height = 6

        # ── Row 3: 헤더 스타일 ──
        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        header_font = Font(name='맑은 고딕', bold=True, color='FFFFFF', size=10)
        thin_border = Border(
            bottom=Side(style='thin', color='95A5A6'),
        )
        for col_idx in range(1, col_count + 1):
            cell = ws.cell(row=data_start_row, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        ws.row_dimensions[data_start_row].height = 24

        # ── 컬럼 너비 ──
        col_widths = col_widths or {}
        for col_idx, col_name in enumerate(df.columns, 1):
            width = col_widths.get(col_name, 15)
            # No. 컬럼은 좁게
            if col_name == 'No.':
                width = 6
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # ── 숫자 포맷 + 데이터 행 정렬 ──
        num_formats = num_formats or {}
        col_format_map = {}
        for col_idx, col_name in enumerate(df.columns, 1):
            if col_name in num_formats:
                col_format_map[col_idx] = num_formats[col_name]

        data_font = Font(name='맑은 고딕', size=10)
        for row_idx in range(data_start_row + 1, data_start_row + 1 + len(df)):
            for col_idx in range(1, col_count + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                # 숫자 포맷 적용
                if col_idx in col_format_map:
                    cell.number_format = col_format_map[col_idx]
                    cell.alignment = Alignment(horizontal='right')
                elif col_idx == 1:  # No.
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='center')

        # ── 줄무늬 행 (짝수행 연회색) ──
        stripe_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        for row_idx in range(data_start_row + 1, data_start_row + 1 + len(df)):
            if (row_idx - data_start_row) % 2 == 0:
                for col_idx in range(1, col_count + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = stripe_fill

        # ── 푸터 ──
        try:
            from gui_app_modular.utils.report_footer import add_gy_logistics_footer
            add_gy_logistics_footer(ws)
        except (ImportError, ModuleNotFoundError) as _e:
            logger.debug(f"[export_mixin] 무시: {_e}")

    def _export_basic_inventory(self, output_path: str) -> str:
        """기본 재고 목록 내보내기"""
        import pandas as pd

        try:
            inventory = self.get_inventory()

            if not inventory:
                # 빈 파일 생성
                pd.DataFrame().to_excel(output_path, index=False)
                return output_path

            df = pd.DataFrame(inventory)

            # v4.19.1: 18개 전체 컬럼 포함
            columns = [
                'lot_no', 'sap_no', 'bl_no', 'product', 'arrival_date',
                'initial_weight', 'current_weight', 'mxbg_pallet', 'status',
                'container_no', 'vessel', 'warehouse', 'location',
                'free_time', 'customs', 'created_at', 'updated_at', 'remarks'
            ]
            columns = [c for c in columns if c in df.columns]
            df = df[columns]

            # 한글 컬럼명 (v4.19.1: 18열 전체)
            column_names = {
                'lot_no': 'LOT NO',
                'sap_no': 'SAP NO',
                'bl_no': 'BL NO',
                'product': 'PRODUCT',
                'arrival_date': 'ARRIVAL',
                'initial_weight': 'TOTAL(KG)',
                'current_weight': 'AVAILABLE(KG)',
                'mxbg_pallet': 'BAGS',
                'status': 'STATUS',
                'container_no': 'CONTAINER',
                'vessel': 'VESSEL',
                'warehouse': 'WAREHOUSE',
                'location': 'LOCATION',
                'free_time': 'FREE TIME',
                'customs': 'CUSTOMS',
                'created_at': 'CREATED AT',
                'updated_at': 'UPDATED AT',
                'remarks': 'REMARKS'
            }
            df = df.rename(columns=column_names)

            # 엑셀 저장
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='재고목록', index=False, startrow=2)

                worksheet = writer.sheets['재고목록']

                # v4.0.0 Q8: 타이틀행 삽입
                from datetime import datetime

                from openpyxl.styles import Alignment, Font, PatternFill
                from openpyxl.utils import get_column_letter

                # Row 1: 타이틀
                worksheet.merge_cells(f'A1:{get_column_letter(len(df.columns))}1')
                title_cell = worksheet['A1']
                title_cell.value = f'SQM 재고 현황 보고서 — {datetime.now().strftime("%Y-%m-%d")}'
                title_cell.font = Font(name='맑은 고딕', bold=True, size=14, color='2C3E50')
                title_cell.alignment = Alignment(horizontal='center')
                worksheet.row_dimensions[1].height = 32

                # Row 2: 빈 줄 (헤더와 타이틀 사이 간격)
                worksheet.row_dimensions[2].height = 8

                # 헤더 스타일 (Row 3)
                header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
                header_font = Font(name='맑은 고딕', bold=True, color='FFFFFF', size=10)
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=3, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')

                # 컬럼 너비 조정
                for i, col in enumerate(df.columns, 1):
                    worksheet.column_dimensions[get_column_letter(i)].width = 15

                # v4.0.0 Q9: GY Logistics 푸터
                try:
                    from gui_app_modular.utils.report_footer import (
                        add_gy_logistics_footer,
                    )
                    add_gy_logistics_footer(worksheet)
                except (ImportError, ModuleNotFoundError) as _e:
                    logger.debug(f'Suppressed: {_e}')

            logger.info(f"엑셀 내보내기 완료: {output_path}")
            return output_path

        except (ImportError, ModuleNotFoundError) as e:
            logger.error(f"엑셀 내보내기 오류: {e}")
            raise

    def _export_detailed_inventory(self, output_path: str) -> str:
        """v5.5.3 P4: 상세 재고 — 컬럼 한글화 + 서식"""
        from datetime import datetime as dt

        import pandas as pd

        try:
            inventory = self.get_inventory()

            # 주요 컬럼만 선택 + 한글 헤더
            rename_map = {
                'lot_no': 'LOT NO', 'sap_no': 'SAP NO', 'bl_no': 'BL NO',
                'container_no': 'CONTAINER', 'product': 'PRODUCT',
                'product_code': 'CODE', 'mxbg_pallet': 'MXBG',
                'net_weight': 'NET(Kg)', 'gross_weight': 'GROSS(Kg)',
                'salar_invoice_no': 'INVOICE NO',
                'ship_date': 'SHIP DATE', 'arrival_date': 'ARRIVAL',
                'free_time': 'FREE TIME', 'warehouse': 'WH',
                'status': 'STATUS', 'customs': 'CUSTOMS',
                'initial_weight': 'Inbound(Kg)', 'current_weight': 'Balance(Kg)',
                'picked_weight': 'Picked(Kg)',
                'sold_to': 'CUSTOMER', 'vessel': 'VESSEL', 'location': 'LOCATION',
                'inbound_date': 'INBOUND DATE',
                'created_at': 'CREATED', 'updated_at': 'UPDATED',
            }
            keep_cols = [c for c in rename_map if c in (inventory[0] if inventory else {})]

            df = pd.DataFrame(inventory)
            if not df.empty:
                df = df[[c for c in keep_cols if c in df.columns]]
                df = df.rename(columns=rename_map)

            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='상세재고', index=False, startrow=2)
                ws = writer.sheets['상세재고']
                self._apply_excel_formatting(
                    ws, df,
                    title=f'SQM 상세 재고 — {dt.now().strftime("%Y-%m-%d")}',
                    num_formats={'NET(Kg)': '#,##0', 'GROSS(Kg)': '#,##0',
                                 'Inbound(Kg)': '#,##0', 'Balance(Kg)': '#,##0',
                                 'Picked(Kg)': '#,##0'},
                )

                # 제품별 요약
                if not df.empty and 'PRODUCT' in df.columns:
                    agg_cols = {}
                    for c in ['Inbound(Kg)', 'Balance(Kg)', 'MXBG']:
                        if c in df.columns:
                            agg_cols[c] = 'sum'
                    if agg_cols:
                        summary = df.groupby('PRODUCT').agg(agg_cols).reset_index()
                    else:
                        summary = df[['PRODUCT']].drop_duplicates()
                    summary.to_excel(writer, sheet_name='제품별요약', index=False, startrow=2)
                    ws2 = writer.sheets['제품별요약']
                    self._apply_excel_formatting(
                        ws2, summary,
                        title='제품별 재고 요약',
                        num_formats={'Inbound(Kg)': '#,##0', 'Balance(Kg)': '#,##0'},
                    )

            return output_path

        except (ValueError, OSError, KeyError) as e:
            logger.error(f"상세 재고 내보내기 오류: {e}")
            raise

    def _export_ruby_format(self, output_path: str) -> str:
        """v5.5.3 P1: 재고리스트 Excel — 화면 18열 동일 출력 + 서식"""
        import pandas as pd

        try:
            inventory = self.get_inventory()

            # 화면 inventory_tab.INVENTORY_COLUMNS와 동일한 순서/이름 (재고리스트·입고 템플릿과 통일)
            column_map = [
                # (DB 키,           Excel 헤더,       너비, 숫자포맷)
                ('lot_no',          'LOT NO',          16, None),
                ('sap_no',          'SAP NO',          16, None),
                ('bl_no',           'BL NO',           18, None),
                ('product',         'PRODUCT',         22, None),
                ('status',          'STATUS',          12, None),
                ('current_weight',  'Balance(Kg)',     14, '#,##0'),
                ('net_weight',      'NET(Kg)',         13, '#,##0'),
                ('container_no',    'CONTAINER',       17, None),
                ('mxbg_pallet',     'MXBG',            8, None),
                ('avail_bags',      'Avail',           10, None),
                ('salar_invoice_no','INVOICE NO',      14, None),
                ('ship_date',       'SHIP DATE',       13, None),
                ('arrival_date',    'ARRIVAL',         13, None),
                ('con_return',      'CON RETURN',       13, None),
                ('free_time',       'FREE TIME',       12, None),
                ('warehouse',       'WH',              8, None),
                ('customs',         'CUSTOMS',         12, None),
                ('initial_weight',  'Inbound(Kg)',     14, '#,##0'),
                ('outbound_weight', 'Outbound(Kg)',    14, '#,##0'),
            ]

            rows = []
            for idx, item in enumerate(inventory, 1):
                row = {'No.': idx}
                for db_key, header, _, _ in column_map:
                    val = item.get(db_key, '')
                    # outbound_weight 계산: initial - current
                    if db_key == 'outbound_weight':
                        initial = item.get('initial_weight') or 0
                        current = item.get('current_weight') or 0
                        val = max(0, initial - current)
                    row[header] = val if val is not None else ''
                rows.append(row)

            headers = ['No.'] + [h for _, h, _, _ in column_map]
            df = pd.DataFrame(rows, columns=headers)

            # Excel 저장 + 서식
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='재고리스트', index=False, startrow=2)
                ws = writer.sheets['재고리스트']

                self._apply_excel_formatting(
                    ws, df,
                    title=f'SQM 재고리스트 — {__import__("datetime").datetime.now().strftime("%Y-%m-%d")}',
                    col_widths={h: w for _, h, w, _ in column_map},
                    num_formats={h: fmt for _, h, _, fmt in column_map if fmt},
                )

            logger.info(f"재고리스트 Excel 내보내기 완료: {output_path}")
            return output_path

        except (RuntimeError, ValueError, OSError) as e:
            logger.error(f"재고리스트 Excel 내보내기 오류: {e}")
            raise

    def _export_tonbag_list(self, output_path: str, include_sample: bool = True) -> str:
        """v5.5.3 P2: 톤백리스트 Excel — 화면 20열 동일 (v5.6.3: MXBG 제거, 개별 무게)"""
        import pandas as pd

        try:
            # JOIN 쿼리로 LOT 정보 포함 조회
            all_tonbags = self.get_tonbags_with_inventory()
            sample_count = 0
            tonbags = []
            for item in all_tonbags:
                sub_lt_raw = item.get('sub_lt')
                status_raw = (item.get('tonbag_status') or item.get('status') or '').strip().upper()
                try:
                    is_sample = int(float(sub_lt_raw)) == 0
                except (TypeError, ValueError):
                    is_sample = False
                if status_raw == 'SAMPLE':
                    is_sample = True
                if is_sample:
                    sample_count += 1
                if include_sample or (not is_sample):
                    tonbags.append(item)

            if not tonbags:
                pd.DataFrame().to_excel(output_path, index=False)
                return output_path

            # 화면 tonbag_tab._tonbag_columns와 동일한 순서 (톤백리스트·출고 템플릿과 통일)
            # NET/Balance/Inbound = 톤백 개별 무게(tonbag_weight) 계산값
            column_map = [
                # (DB/계산 키,       Excel 헤더,       너비, 숫자포맷)
                ('lot_no',           'LOT NO',         16, None),
                ('tonbag_no_print',  'TONBAG NO',      12, None),
                ('sap_no',           'SAP NO',         16, None),
                ('bl_no',            'BL NO',          18, None),
                ('product',          'PRODUCT',        22, None),
                ('tonbag_status',    'STATUS',         12, None),
                ('balance_tonbag',   'Balance(Kg)',    14, '#,##0'),
                ('tonbag_uid',       'UID',            20, None),
                ('container_no',     'CONTAINER',      17, None),
                ('location',         'LOCATION',       12, None),
                ('net_weight_tonbag','NET(Kg)',        13, '#,##0'),
                ('salar_invoice_no', 'INVOICE NO',     14, None),
                ('ship_date',        'SHIP DATE',      13, None),
                ('arrival_date',     'ARRIVAL',        13, None),
                ('con_return',       'CON RETURN',     13, None),
                ('free_time',        'FREE TIME',      12, None),
                ('warehouse',        'WH',             8, None),
                ('customs',          'CUSTOMS',        12, None),
                ('inbound_tonbag',   'Inbound(Kg)',    14, '#,##0'),
                ('outbound_weight',  'Outbound(Kg)',   14, '#,##0'),
            ]

            # tonbag_no_print, tonbag_uid 계산
            from engine_modules.tonbag_compat import (
                get_tonbag_display_no,
                get_tonbag_uid,
            )

            rows = []
            for idx, item in enumerate(tonbags, 1):
                # v5.6.3: 톤백 개별 무게 (샘플 1kg, 톤백 500kg) — 재고리스트 LOT 합계와 정합성
                tw = float(item.get('tonbag_weight') or item.get('weight') or 0)
                status = item.get('tonbag_status') or item.get('status') or STATUS_AVAILABLE
                if status in (STATUS_PICKED, STATUS_SOLD, 'SHIPPED', STATUS_DEPLETED):
                    balance_tb = 0.0
                    outbound_tb = tw
                else:
                    balance_tb = tw
                    outbound_tb = 0.0

                row = {'No.': idx}
                for db_key, header, _, _ in column_map:
                    if db_key == 'tonbag_no_print':
                        val = get_tonbag_display_no(item)
                    elif db_key == 'tonbag_uid':
                        val = get_tonbag_uid(item)
                    elif db_key == 'net_weight_tonbag':
                        val = tw
                    elif db_key == 'balance_tonbag':
                        val = balance_tb
                    elif db_key == 'inbound_tonbag':
                        val = tw
                    elif db_key == 'outbound_weight':
                        val = outbound_tb
                    else:
                        val = item.get(db_key, '')
                    row[header] = val if val is not None else ''
                rows.append(row)

            headers = ['No.'] + [h for _, h, _, _ in column_map]
            df = pd.DataFrame(rows, columns=headers)

            # Excel 저장 + 서식 (P1 공통 함수 재사용)
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='톤백리스트', index=False, startrow=2)
                ws = writer.sheets['톤백리스트']

                self._apply_excel_formatting(
                    ws, df,
                    title=(
                        f"SQM 톤백리스트 — {__import__('datetime').datetime.now().strftime('%Y-%m-%d')} "
                        f"(기준: {'샘플 포함' if include_sample else '샘플 제외'} | "
                        f"총 {len(tonbags)}건 | 샘플 {sample_count}건)"
                    ),
                    col_widths={h: w for _, h, w, _ in column_map},
                    num_formats={h: fmt for _, h, _, fmt in column_map if fmt},
                )

            logger.info(f"톤백리스트 Excel 내보내기 완료: {output_path}")
            return output_path

        except (RuntimeError, ValueError, OSError) as e:
            logger.error(f"톤백리스트 Excel 내보내기 오류: {e}")
            raise

    def _export_lot_tonbag_report(self, output_path: str) -> str:
        """v5.5.3 P4: LOT-톤백 리포트 — 컬럼 한글화 + 서식"""
        from datetime import datetime as dt

        import pandas as pd

        # LOT 컬럼 한글 매핑
        lot_rename = {
            'lot_no': 'LOT NO', 'sap_no': 'SAP NO', 'bl_no': 'BL NO',
            'container_no': 'CONTAINER', 'product': 'PRODUCT',
            'mxbg_pallet': 'MXBG', 'net_weight': 'NET(Kg)',
            'initial_weight': 'Inbound(Kg)', 'current_weight': 'Balance(Kg)',
            'status': 'STATUS', 'warehouse': 'WH',
            'arrival_date': 'ARRIVAL', 'ship_date': 'SHIP DATE',
        }
        # 톤백 컬럼 한글 매핑
        tb_rename = {
            'lot_no': 'LOT NO', 'tonbag_no': 'TONBAG NO',
            'sap_no': 'SAP NO', 'bl_no': 'BL NO',
            'weight': 'Weight(Kg)', 'status': 'STATUS',
            'location': 'LOCATION', 'picked_to': 'CUSTOMER',
            'inbound_date': 'INBOUND', 'picked_date': 'PICKED DATE',
        }

        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # LOT 목록 (한글 + 서식)
                lots = self.get_inventory()
                df_lot = pd.DataFrame(lots)
                if not df_lot.empty:
                    keep = [c for c in lot_rename if c in df_lot.columns]
                    df_lot = df_lot[keep].rename(columns=lot_rename)
                df_lot.to_excel(writer, sheet_name='LOT목록', index=False, startrow=2)
                self._apply_excel_formatting(
                    writer.sheets['LOT목록'], df_lot,
                    title=f'LOT 목록 — {dt.now().strftime("%Y-%m-%d")}',
                    num_formats={'NET(Kg)': '#,##0', 'Inbound(Kg)': '#,##0', 'Balance(Kg)': '#,##0'},
                )

                # LOT별 톤백 시트 (상위 20개)
                for lot in lots[:20]:
                    lot_no = str(lot.get('lot_no') or '').strip()
                    if not lot_no:
                        continue
                    tonbags = self.get_tonbags(lot_no=lot_no)
                    if not tonbags:
                        continue

                    df_tb = pd.DataFrame(tonbags)
                    keep_tb = [c for c in tb_rename if c in df_tb.columns]
                    if keep_tb:
                        df_tb = df_tb[keep_tb].rename(columns=tb_rename)

                    # 시트명: LOT번호 (최대 31자, 특수문자 제거)
                    safe_name = lot_no[:28].replace('/', '-').replace('\\', '-')
                    df_tb.to_excel(writer, sheet_name=safe_name, index=False, startrow=2)
                    self._apply_excel_formatting(
                        writer.sheets[safe_name], df_tb,
                        title=f'LOT: {lot_no}',
                        num_formats={'Weight(Kg)': '#,##0'},
                    )

            return output_path

        except (RuntimeError, ValueError, OSError) as e:
            logger.error(f"LOT-톤백 리포트 내보내기 오류: {e}")
            raise

    def _export_full_inventory(self, output_path: str) -> str:
        """v5.5.3 P4: 전체 재고 — 5개 시트 서식 + 한글화"""
        from datetime import datetime as dt

        import pandas as pd

        try:
            today = dt.now().strftime("%Y-%m-%d")

            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # 1. 재고 요약 (한글 키)
                summary = self.get_inventory_summary()
                summary_rename = {
                    'total_lots': '총 LOT수', 'total_bags': '총 톤백수',
                    'total_weight_kg': '총입고(Kg)', 'total_weight_mt': '총입고(MT)',
                    'available_weight_kg': '잔량(Kg)', 'available_weight_mt': '잔량(MT)',
                    'picked_weight_kg': '출고중(Kg)', 'picked_weight_mt': '출고중(MT)',
                    'sold_weight_kg': '출고완료(Kg)', 'sold_weight_mt': '출고완료(MT)',
                }
                df_sum = pd.DataFrame([summary])
                keep_s = [c for c in summary_rename if c in df_sum.columns]
                if keep_s:
                    df_sum = df_sum[keep_s].rename(columns=summary_rename)
                df_sum.to_excel(writer, sheet_name='요약', index=False, startrow=2)
                self._apply_excel_formatting(
                    writer.sheets['요약'], df_sum,
                    title=f'SQM 재고 요약 — {today}',
                    num_formats={'총입고(Kg)': '#,##0', '잔량(Kg)': '#,##0',
                                 '출고중(Kg)': '#,##0', '출고완료(Kg)': '#,##0'},
                )

                # 2. LOT 목록 (한글화)
                lot_rename = {
                    'lot_no': 'LOT NO', 'sap_no': 'SAP NO', 'bl_no': 'BL NO',
                    'container_no': 'CONTAINER', 'product': 'PRODUCT',
                    'mxbg_pallet': 'MXBG', 'net_weight': 'NET(Kg)',
                    'initial_weight': 'Inbound(Kg)', 'current_weight': 'Balance(Kg)',
                    'status': 'STATUS', 'warehouse': 'WH',
                    'arrival_date': 'ARRIVAL', 'customs': 'CUSTOMS',
                }
                lots = self.get_inventory()
                df_lot = pd.DataFrame(lots)
                if not df_lot.empty:
                    keep_l = [c for c in lot_rename if c in df_lot.columns]
                    df_lot = df_lot[keep_l].rename(columns=lot_rename)
                df_lot.to_excel(writer, sheet_name='LOT목록', index=False, startrow=2)
                self._apply_excel_formatting(
                    writer.sheets['LOT목록'], df_lot,
                    title=f'LOT 목록 — {today}',
                    num_formats={'NET(Kg)': '#,##0', 'Inbound(Kg)': '#,##0', 'Balance(Kg)': '#,##0'},
                )

                # 3. 톤백 전체 (JOIN + 한글화)
                tb_rename = {
                    'lot_no': 'LOT NO', 'tonbag_no': 'TONBAG NO',
                    'sap_no': 'SAP NO', 'product': 'PRODUCT',
                    'location': 'LOCATION', 'tonbag_status': 'STATUS',
                    'tonbag_weight': 'Weight(Kg)',
                    'initial_weight': 'LOT Inbound(Kg)', 'current_weight': 'LOT Balance(Kg)',
                    'picked_to': 'CUSTOMER',
                }
                tonbags = self.get_tonbags_with_inventory()
                df_tb = pd.DataFrame(tonbags)
                if not df_tb.empty:
                    keep_t = [c for c in tb_rename if c in df_tb.columns]
                    df_tb = df_tb[keep_t].rename(columns=tb_rename)
                df_tb.to_excel(writer, sheet_name='톤백전체', index=False, startrow=2)
                self._apply_excel_formatting(
                    writer.sheets['톤백전체'], df_tb,
                    title=f'톤백 전체 — {today}',
                    num_formats={'Weight(Kg)': '#,##0', 'LOT Inbound(Kg)': '#,##0',
                                 'LOT Balance(Kg)': '#,##0'},
                )

                # 4. 제품별 (한글화)
                prod_rename = {
                    'product': 'PRODUCT', 'lot_count': 'LOT수',
                    'total_kg': '입고(Kg)', 'available_kg': '잔량(Kg)',
                    'mxbg_pallet': '톤백수',
                }
                by_product = self.get_inventory_by_product()
                df_prod = pd.DataFrame(by_product)
                if not df_prod.empty:
                    keep_p = [c for c in prod_rename if c in df_prod.columns]
                    df_prod = df_prod[keep_p].rename(columns=prod_rename)
                df_prod.to_excel(writer, sheet_name='제품별', index=False, startrow=2)
                self._apply_excel_formatting(
                    writer.sheets['제품별'], df_prod,
                    title=f'제품별 재고 — {today}',
                    num_formats={'입고(Kg)': '#,##0', '잔량(Kg)': '#,##0'},
                )

                # 5. 고객별 (한글화)
                cust_rename = {
                    'customer': '고객', 'lot_count': 'LOT수',
                    'total_kg': '출고(Kg)', 'mxbg_pallet': '톤백수',
                }
                by_customer = self.get_inventory_by_customer()
                df_cust = pd.DataFrame(by_customer)
                if not df_cust.empty:
                    keep_c = [c for c in cust_rename if c in df_cust.columns]
                    df_cust = df_cust[keep_c].rename(columns=cust_rename)
                df_cust.to_excel(writer, sheet_name='고객별', index=False, startrow=2)
                self._apply_excel_formatting(
                    writer.sheets['고객별'], df_cust,
                    title=f'고객별 출고 — {today}',
                    num_formats={'출고(Kg)': '#,##0'},
                )

            return output_path

        except (RuntimeError, ValueError, OSError) as e:
            logger.error(f"전체 재고 내보내기 오류: {e}")
            raise

    def _export_return_history(self, output_path: str) -> str:
        """v6.12.2: 반품 이력 전체 내보내기 (3시트: 이력/사유별/고객별)."""
        from datetime import date as _date

        import pandas as pd

        today = _date.today().strftime('%Y-%m-%d')

        try:
            # Sheet 1: 반품 이력 전체
            rows = self.db.fetchall("""
                SELECT r.lot_no, r.sub_lt, r.return_date, r.original_customer,
                       r.original_sale_ref, r.reason, r.remark, r.weight_kg,
                       r.created_at, i.sap_no, i.product
                FROM return_history r
                LEFT JOIN inventory i ON r.lot_no = i.lot_no
                ORDER BY r.created_at DESC
            """)
            df_hist = pd.DataFrame([dict(r) if isinstance(r, dict) else {
                'lot_no': r[0], 'sub_lt': r[1], 'return_date': r[2],
                'original_customer': r[3], 'original_sale_ref': r[4],
                'reason': r[5], 'remark': r[6], 'weight_kg': r[7],
                'created_at': r[8], 'sap_no': r[9], 'product': r[10]
            } for r in (rows or [])])
            if df_hist.empty:
                df_hist = pd.DataFrame(columns=[
                    'lot_no', 'sub_lt', 'return_date', 'original_customer',
                    'original_sale_ref', 'reason', 'remark', 'weight_kg',
                    'created_at', 'sap_no', 'product'])
            df_hist.columns = [
                'LOT NO', 'Sub LT', '반품일', '고객', 'Sale Ref',
                '사유', '비고', '중량(kg)', '생성일', 'SAP NO', '제품']

            # Sheet 2: 사유별 집계
            rows2 = self.db.fetchall("""
                SELECT COALESCE(reason, '미기재') AS reason, COUNT(*) AS cnt,
                       COALESCE(SUM(weight_kg), 0) AS total
                FROM return_history
                GROUP BY COALESCE(reason, '미기재')
                ORDER BY cnt DESC
            """)
            df_reason = pd.DataFrame([{
                '사유': (r['reason'] if isinstance(r, dict) else r[0]),
                '건수': (r['cnt'] if isinstance(r, dict) else r[1]),
                '중량(kg)': float(r['total'] if isinstance(r, dict) else r[2]),
            } for r in (rows2 or [])])

            # Sheet 3: 고객별 집계
            rows3 = self.db.fetchall("""
                SELECT COALESCE(original_customer, '미기재') AS customer,
                       COUNT(*) AS cnt, COALESCE(SUM(weight_kg), 0) AS total
                FROM return_history
                GROUP BY COALESCE(original_customer, '미기재')
                ORDER BY cnt DESC
            """)
            df_cust = pd.DataFrame([{
                '고객': (r['customer'] if isinstance(r, dict) else r[0]),
                '건수': (r['cnt'] if isinstance(r, dict) else r[1]),
                '중량(kg)': float(r['total'] if isinstance(r, dict) else r[2]),
            } for r in (rows3 or [])])

            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df_hist.to_excel(writer, sheet_name='반품 이력', index=False, startrow=2)
                self._apply_excel_formatting(
                    writer.sheets['반품 이력'], df_hist,
                    title=f'SQM 반품 이력 — {today}',
                    num_formats={'중량(kg)': '#,##0'},
                )

                if not df_reason.empty:
                    df_reason.to_excel(writer, sheet_name='사유별', index=False, startrow=2)
                    self._apply_excel_formatting(
                        writer.sheets['사유별'], df_reason,
                        title=f'사유별 집계 — {today}',
                        num_formats={'중량(kg)': '#,##0'},
                    )

                if not df_cust.empty:
                    df_cust.to_excel(writer, sheet_name='고객별', index=False, startrow=2)
                    self._apply_excel_formatting(
                        writer.sheets['고객별'], df_cust,
                        title=f'고객별 집계 — {today}',
                        num_formats={'중량(kg)': '#,##0'},
                    )

            logger.info(f"[ExportReturn] 반품 이력 내보내기: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"반품 이력 내보내기 오류: {e}", exc_info=True)
            raise


    def _export_integrity_report(self, output_path: str) -> str:
        """v7.0.1: 정합성 검증 리포트 Excel 내보내기"""
        from features.reports.integrity_report import generate_integrity_report_excel
        result = generate_integrity_report_excel(self, output_path)
        if not result:
            raise RuntimeError("정합성 리포트 생성 실패")
        return result
