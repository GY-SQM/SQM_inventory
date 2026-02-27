# -*- coding: utf-8 -*-
"""
SQM 재고관리 - 고급 다이얼로그 Mixin
======================================
v3.8.4 - advanced_features_mixin에서 분리

기능:
- 반품 처리 다이얼로그
- 수동 입고 다이얼로그
- 문서 변환 다이얼로그
- 출고 이력 조회
- 스냅샷 차트
- 출고 인보이스 생성
"""

import os
import sqlite3
import logging
from ..utils.ui_constants import CustomMessageBox, apply_modal_window_options

logger = logging.getLogger(__name__)


class AdvancedDialogsMixin:
    """고급 다이얼로그 Mixin"""

    def _on_return_inbound_upload(self) -> None:
        """v6.0: 반품 입고 — [템플릿 열기(붙여넣기)] vs [파일 업로드] 선택 후 파싱·DB 반영 (입고 형식과 동일)."""
        choice = self._show_template_or_upload_choice("반품 입고", "return_inbound")
        if choice is None:
            return
        if choice == "template":
            self._show_return_inbound_spreadsheet_dialog()
            return
        from ..utils.constants import filedialog
        path = filedialog.askopenfilename(
            parent=self.root,
            title="반품 입고 Excel 선택",
            filetypes=[("Excel files", "*.xlsx;*.xls"), ("All files", "*.*")],
        )
        if not path or not path.strip():
            return
        self._apply_return_inbound_after_parse(path, source_file=path)

    def _on_return_inbound_paste_confirm(self, rows: list) -> None:
        """반품 입고 붙여넣기 확인 시 파싱 후 DB 반영."""
        from features.parsers.return_inbound_parser import parse_return_inbound_from_rows
        parse_result = parse_return_inbound_from_rows(rows)
        if not parse_result.get("parse_ok") or parse_result.get("errors"):
            err_msg = "\n".join(parse_result.get("errors", ["파싱 실패"]))
            CustomMessageBox.showerror(
                self.root,
                "반품 입고 — 검증 실패",
                "필수 항목(LOT NO, NET(Kg), PICKING NO, 반품사유) 누락 또는 오류.\n\n" + err_msg,
            )
            return
        n = len(parse_result.get("items", []))
        if n == 0:
            CustomMessageBox.showwarning(self.root, "반품 입고", "처리할 행이 없습니다.")
            return
        if not CustomMessageBox.askyesno(
            self.root,
            "반품 입고 확인",
            f"총 {n}건을 반품 입고 처리합니다.\nPICKING NO 매칭 실패 시 전체가 중단됩니다.\n계속하시겠습니까?",
        ):
            return
        self._apply_return_inbound_after_parse(parse_result, source_file="(붙여넣기)", skip_confirm=True)

    def _apply_return_inbound_after_parse(
        self, parse_result_or_path, source_file: str = "", skip_confirm: bool = False
    ) -> None:
        """반품 파싱 결과 또는 Excel 경로를 받아 검증 후 DB 반영. skip_confirm=True면 확인 대화상자 생략(붙여넣기에서 이미 표시)."""
        if isinstance(parse_result_or_path, str):
            path = parse_result_or_path
            try:
                from features.parsers.return_inbound_parser import parse_return_inbound_excel
                parse_result = parse_return_inbound_excel(path)
            except ImportError:
                CustomMessageBox.showerror(
                    self.root,
                    "반품 입고",
                    "features.parsers.return_inbound_parser를 불러올 수 없습니다.",
                )
                return
            source_file = path
        else:
            parse_result = parse_result_or_path
        if not parse_result.get("parse_ok") or parse_result.get("errors"):
            err_msg = "\n".join(parse_result.get("errors", ["파싱 실패"]))
            CustomMessageBox.showerror(
                self.root,
                "반품 입고 — 검증 실패",
                "필수 항목(LOT NO, WEIGHT(MT) 또는 NET(Kg), PICKING NO, REASON) 누락 또는 오류.\n\n" + err_msg,
            )
            return
        n = len(parse_result.get("items", []))
        if n == 0:
            CustomMessageBox.showwarning(self.root, "반품 입고", "처리할 행이 없습니다.")
            return
        if not skip_confirm:
            # v6.12.1: 미리보기 편집 다이얼로그
            _confirmed_items = [None]

            def _on_preview_confirm(edited_items):
                _confirmed_items[0] = edited_items

            try:
                from ..dialogs.return_inbound_preview import ReturnInboundPreviewDialog
                current_theme = getattr(self, '_current_theme', 'flatly')
                ReturnInboundPreviewDialog(
                    self.root,
                    parse_result.get('items', []),
                    on_confirm=_on_preview_confirm,
                    current_theme=current_theme,
                )
            except ImportError:
                # fallback: 기존 텍스트 확인
                if not CustomMessageBox.askyesno(
                    self.root,
                    "반품 입고 확인",
                    f"총 {n}건을 반품 입고 처리합니다.\nPICKING NO 매칭 실패 시 전체가 중단됩니다.\n계속하시겠습니까?",
                ):
                    return
                _confirmed_items[0] = parse_result.get('items', [])

            if _confirmed_items[0] is None:
                return  # 취소
            parse_result['items'] = _confirmed_items[0]  # 편집된 데이터로 교체

            # v6.12.2: 대량 반품 관리자 확인
            try:
                from engine_modules.constants import RETURN_AUTO_APPROVE_MAX_TONBAGS
                _max_auto = RETURN_AUTO_APPROVE_MAX_TONBAGS
            except ImportError:
                _max_auto = 5
            total_items = len(parse_result.get('items', []))
            total_tb = sum(it.get('tonbag_count', 1) for it in parse_result.get('items', []))
            if total_tb > _max_auto:
                if not CustomMessageBox.askyesno(
                    self.root, "⚠️ 대량 반품 관리자 확인",
                    f"반품 톤백 {total_tb}개 (자동승인 기준 {_max_auto}개 초과)\n\n"
                    f"대량 반품은 되돌릴 수 없습니다.\n정말 진행하시겠습니까?"
                ):
                    return
        try:
            from features.parsers.return_inbound_engine import apply_return_inbound_to_db
            r = apply_return_inbound_to_db(self.engine, parse_result, source_file)
            self._log(f"✅ 반품 입고 완료: {r.get('returned', 0)}건")
            CustomMessageBox.showinfo(
                self.root,
                "반품 입고 완료",
                f"반품 입고 처리 완료\n\n반영: {r.get('returned', 0)}건",
            )
            if hasattr(self, '_deferred_refresh_main_tabs'):
                self._deferred_refresh_main_tabs(delay_ms=50)
            elif hasattr(self, '_refresh_main_tabs'):
                self._refresh_main_tabs()
            else:
                if hasattr(self, "_refresh_inventory"):
                    self._refresh_inventory()
                if hasattr(self, "_refresh_tonbag"):
                    self._refresh_tonbag()
            if hasattr(self, "_refresh_dashboard"):
                self._refresh_dashboard()
        except RuntimeError as e:
            CustomMessageBox.showerror(self.root, "반품 입고 오류", str(e))
        except Exception as e:
            logger.exception("반품 입고 처리 중 오류")
            CustomMessageBox.showerror(self.root, "반품 입고 오류", str(e))

    def _show_return_dialog(self, initial_tab: int = 0) -> None:
        """v4.1.4: 반품 처리 다이얼로그 — initial_tab: 0=소량(단건), 1=다량(Excel)."""
        from ..utils.constants import tk, ttk, BOTH

        dialog = tk.Toplevel(self.root)
        dialog.title("🔄 반품 처리")
        dialog.geometry("780x650")
        apply_modal_window_options(dialog)
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text="반품 처리", font=('맑은 고딕', 18, 'bold')).pack(pady=8)

        nb = ttk.Notebook(dialog)
        nb.pack(fill=BOTH, expand=True, padx=10, pady=5)

        # TAB 0: 단건 반품 (소량)
        self._build_return_single_tab(nb, dialog)
        # TAB 1: Excel 일괄 반품 (다량)
        self._build_return_excel_tab(nb, dialog)

        try:
            nb.select(initial_tab)
        except (tk.TclError, TypeError, IndexError) as _e:
            logger.debug(f"Notebook select: {_e}")

    def _build_return_single_tab(self, nb, dialog) -> None:
        """반품 다이얼로그 — TAB 1: 단건 입력"""
        from ..utils.constants import tk, ttk, X, LEFT, W, END
        from ..utils.custom_messagebox import CustomMessageBox
        tab_single = ttk.Frame(nb)
        nb.add(tab_single, text="  📝 단건 입력  ")
        
        frame = ttk.LabelFrame(tab_single, text="반품 정보")
        frame.pack(fill=X, padx=15, pady=10)
        
        ttk.Label(frame, text="LOT 번호:").grid(row=0, column=0, sticky='e', padx=5, pady=5)
        lot_combo = ttk.Combobox(frame, width=27, state='readonly')
        lot_combo.grid(row=0, column=1, padx=5, pady=5)
        # LOT 번호 목록 DB에서 로드
        def _load_lot_list():
            if not hasattr(self, 'engine') or not self.engine:
                return
            try:
                rows = self.engine.db.fetchall(
                    "SELECT DISTINCT lot_no FROM inventory_tonbag ORDER BY lot_no")
                lots = [r['lot_no'] if isinstance(r, dict) else r[0] for r in (rows or [])]
                lot_combo['values'] = lots if lots else ["(등록된 LOT 없음)"]
                if lots:
                    lot_combo.set(lots[0])
            except (ValueError, TypeError, AttributeError) as e:
                logger.debug(f"LOT 목록 조회 오류: {e}")
        _load_lot_list()
        
        ttk.Label(frame, text="Tonbag No:").grid(row=1, column=0, sticky='e', padx=5, pady=5)
        tonbag_combo = ttk.Combobox(frame, width=27, state='readonly')
        tonbag_combo.grid(row=1, column=1, padx=5, pady=5)
        tonbag_combo.set("← LOT 번호 선택 후 자동 조회")
        
        def _on_lot_change(event=None):
            lot_no = (lot_combo.get() or '').strip()
            if not lot_no or not hasattr(self, 'engine') or lot_no == "(등록된 LOT 없음)":
                return
            try:
                rows = self.engine.db.fetchall(
                    "SELECT sub_lt, weight, status FROM inventory_tonbag "
                    "WHERE lot_no = ? ORDER BY sub_lt", (lot_no,))
                if rows:
                    values = []
                    for r in rows:
                        sub = r['sub_lt'] if isinstance(r, dict) else r[0]
                        w = r['weight'] if isinstance(r, dict) else r[1]
                        st = r['status'] if isinstance(r, dict) else r[2]
                        values.append(f"{sub} ({w:.1f}kg, {st})")
                    tonbag_combo['values'] = values
                    tonbag_combo.set(values[0])
                    first_w = rows[0]['weight'] if isinstance(rows[0], dict) else rows[0][1]
                    qty_entry.delete(0, 'end')
                    qty_entry.insert(0, f"{first_w:.1f}")
                else:
                    tonbag_combo['values'] = ["톤백 없음"]
                    tonbag_combo.set("톤백 없음")
            except (ValueError, TypeError, AttributeError) as e:
                logger.debug(f"톤백 조회 오류: {e}")
        
        lot_combo.bind('<<ComboboxSelected>>', _on_lot_change)
        lot_combo.bind('<FocusOut>', _on_lot_change)
        lot_combo.bind('<Return>', _on_lot_change)
        
        def _on_tonbag_select(event=None):
            sel = tonbag_combo.get()
            if '(' in sel and 'kg' in sel:
                try:
                    w = sel.split('(')[1].split('kg')[0]
                    qty_entry.delete(0, 'end')
                    qty_entry.insert(0, w)
                except (ValueError, TypeError, KeyError) as _e:
                    logger.debug(f"Suppressed: {_e}")
        
        tonbag_combo.bind('<<ComboboxSelected>>', _on_tonbag_select)
        
        ttk.Label(frame, text="반품 수량 (kg):").grid(row=2, column=0, sticky='e', padx=5, pady=5)
        qty_entry = ttk.Entry(frame, width=30)
        qty_entry.grid(row=2, column=1, padx=5, pady=5)
        
        ttk.Label(frame, text="반품 사유:").grid(row=3, column=0, sticky='e', padx=5, pady=5)
        # v6.12.1: 표준 사유 드롭다운 + 직접 입력 가능 (Combobox)
        _reason_codes = [
            "품질 불량", "수량 오류", "고객 취소", "배송 문제",
            "파손/변질", "규격 불일치", "기타"]
        reason_combo = ttk.Combobox(frame, width=27, values=_reason_codes)
        reason_combo.grid(row=3, column=1, padx=5, pady=5)
        reason_combo.set("품질 불량")
        
        ttk.Label(frame, text="비고:").grid(row=4, column=0, sticky='ne', padx=5, pady=5)
        note_text = tk.Text(frame, width=30, height=3)
        note_text.grid(row=4, column=1, padx=5, pady=5)
        
        def _process_single_return():
            lot_no = (lot_combo.get() or '').strip()
            tonbag_sel = tonbag_combo.get().strip()
            qty_str = qty_entry.get().strip()
            reason = reason_combo.get()
            note = note_text.get("1.0", "end").strip()

            # 최소 데이터: LOT 번호, Tonbag 선택, 반품 수량(kg)
            if not lot_no or lot_no == "(등록된 LOT 없음)":
                CustomMessageBox.showwarning(dialog, "입력 필요",
                    "LOT 번호를 목록에서 선택하세요.\n\n화물 정합성을 위해 LOT 번호는 필수입니다.")
                lot_combo.focus_set()
                return
            if not tonbag_sel or tonbag_sel == "← LOT 번호 선택 후 자동 조회" or tonbag_sel == "톤백 없음":
                CustomMessageBox.showwarning(dialog, "입력 필요",
                    "LOT 번호 선택 후 반품할 톤백을 선택하세요.\n\nTonbag No는 필수입니다.")
                lot_combo.focus_set()
                return
            if not qty_str:
                CustomMessageBox.showwarning(dialog, "입력 필요",
                    "반품 수량(kg)을 입력하세요.\n\n화물 무게 정합성을 위해 반품 수량은 필수입니다.")
                qty_entry.focus_set()
                return
            try:
                qty = float(qty_str)
                if qty <= 0:
                    raise ValueError
            except ValueError:
                CustomMessageBox.showwarning(dialog, "입력 오류",
                    "반품 수량에는 0보다 큰 숫자(kg)를 입력하세요.")
                qty_entry.focus_set()
                return

            # 무게 정합성: 반품 수량이 해당 톤백 무게를 초과하면 경고
            tonbag_kg = None
            if '(' in tonbag_sel and 'kg' in tonbag_sel:
                try:
                    tonbag_kg = float(tonbag_sel.split('(')[1].split('kg')[0].strip())
                except (ValueError, TypeError, IndexError) as _e:
                    logger.debug(f"톤백 무게 파싱: {_e}")
            if tonbag_kg is not None and qty > tonbag_kg:
                CustomMessageBox.showwarning(dialog, "무게 정합성 경고",
                    f"반품 수량({qty:,.2f} kg)이 해당 톤백 무게({tonbag_kg:,.2f} kg)를 초과합니다.\n\n"
                    "수정해 주세요. (반품 수량 ≤ 톤백 무게)")
                qty_entry.focus_set()
                return

            if not CustomMessageBox.askyesno(dialog, "반품 확인",
                f"LOT: {lot_no}\n수량: {qty:,.2f} kg\n사유: {reason}\n\n반품 처리하시겠습니까?"):
                return
            sub_lt_val = 1
            if tonbag_sel and tonbag_sel[0].isdigit():
                try:
                    sub_lt_val = int(tonbag_sel.split(' ')[0])
                except (ValueError, IndexError):
                    sub_lt_val = 1
            if hasattr(self.engine, 'return_single_tonbag'):
                try:
                    result = self.engine.return_single_tonbag(
                        lot_no=lot_no, sub_lt=sub_lt_val, reason=reason, remark=note)
                except (ValueError, RuntimeError, sqlite3.OperationalError, sqlite3.IntegrityError) as e:
                    CustomMessageBox.showerror(dialog, "반품 오류", f"반품 처리 중 오류:\n{str(e)[:500]}")
                    return
            elif hasattr(self.engine, 'process_return'):
                try:
                    result = self.engine.process_return([{
                        'lot_no': lot_no, 'sub_lt': sub_lt_val,
                        'reason': reason, 'remark': note}],
                        source_type='RETURN_SINGLE')
                except (ValueError, RuntimeError, sqlite3.OperationalError, sqlite3.IntegrityError) as e:
                    CustomMessageBox.showerror(dialog, "반품 오류", f"반품 처리 중 오류:\n{str(e)[:500]}")
                    return
            else:
                CustomMessageBox.showwarning(dialog, "안내", "반품 엔진을 찾을 수 없습니다.")
                return
            if result.get('success'):
                self._log(f"✅ 반품 완료: {lot_no}-{sub_lt_val}")
                CustomMessageBox.showinfo(dialog, "완료",
                    f"반품 처리 완료\n\nLOT: {lot_no}\n톤백: {sub_lt_val}\n사유: {reason}")
                dialog.destroy()
                if hasattr(self, '_deferred_refresh_main_tabs'):
                    self._deferred_refresh_main_tabs(delay_ms=50)
                elif hasattr(self, '_refresh_main_tabs'):
                    self._refresh_main_tabs()
                else:
                    self._refresh_inventory()
                    if hasattr(self, '_refresh_tonbag'):
                        self._refresh_tonbag()
                if hasattr(self, '_refresh_dashboard'):
                    self._refresh_dashboard()
            else:
                errs = '\n'.join(result.get('errors', ['알 수 없는 오류']))
                CustomMessageBox.showerror(dialog, "오류", f"반품 실패:\n{errs}")
        
        s_btn = ttk.Frame(tab_single)
        s_btn.pack(pady=15)
        try:
            ttk.Button(s_btn, text="반품 처리", command=_process_single_return,
                       bootstyle="primary").pack(side='left', padx=10)
        except TypeError:
            ttk.Button(s_btn, text="반품 처리", command=_process_single_return).pack(side='left', padx=10)
        ttk.Button(s_btn, text="취소", command=dialog.destroy).pack(side='left', padx=10)

    def _build_return_excel_tab(self, nb, dialog) -> None:
        """반품 다이얼로그 — TAB 2: Excel 일괄 반품"""
        from ..utils.constants import tk, ttk, BOTH, X, Y, LEFT, RIGHT, END, VERTICAL
        from ..utils.custom_messagebox import CustomMessageBox
        tab_excel = ttk.Frame(nb)
        nb.add(tab_excel, text="  📂 Excel 일괄 반품  ")
        
        top_bar = ttk.Frame(tab_excel)
        top_bar.pack(fill=X, padx=10, pady=8)
        
        def _download_return_template():
            """반품 템플릿 = 재고 리스트와 동일 형식 + return_qty_kg, return_reason. 필수(lot_no, return_qty_kg, return_reason)만 색상."""
            from ..utils.constants import filedialog
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            file_path = filedialog.asksaveasfilename(
                title="반품 양식 저장", defaultextension=".xlsx",
                initialfile="반품_양식_템플릿.xlsx",
                filetypes=[("Excel files", "*.xlsx")])
            if not file_path:
                return
            try:
                inv_cols = getattr(self, 'INVENTORY_TEMPLATE_COLUMNS', None)
                if inv_cols is None:
                    inv_cols = [
                        ('lot_no', 'LOT NO', 120, True), ('sap_no', 'SAP NO', 120, False),
                        ('bl_no', 'BL NO', 140, False), ('product', 'PRODUCT', 160, False),
                        ('net_weight', 'NET(Kg)', 100, False), ('container_no', 'CONTAINER', 130, False),
                        ('mxbg_pallet', 'MXBG', 70, False),
                    ]
                return_extra = [
                    ('return_qty_kg', '반품수량(갯수)', 16, True),
                    ('return_reason', 'RETURN REASON', 20, True),
                ]
                headers = list(inv_cols) + return_extra
                required_return = {'lot_no', 'bl_no', 'return_qty_kg', 'return_reason'}
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "반품 데이터"
                hfont = Font(bold=True, color="FFFFFF", size=11)
                req_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
                opt_fill = PatternFill(start_color="7F8C8D", end_color="7F8C8D", fill_type="solid")
                smp_fill = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
                thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                ncols = len(headers)
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
                ws['A1'] = "🔄 반품 템플릿 — 필수: Lot No, BL NO, 톤백중량(DB조회), 반품수량(갯수), 사유. 2행=DB필드명."
                ws['A1'].font = Font(bold=True, size=11, color="C0392B")
                ws.row_dimensions[1].height = 28
                for col, h in enumerate(headers, 1):
                    db_f = h[0]
                    ws.cell(row=2, column=col, value=db_f).font = Font(size=8, color="999999")
                ws.row_dimensions[2].height = 14
                for col, h in enumerate(headers, 1):
                    db_f, disp, w, _ = h[0], h[1], h[2], (h[3] if len(h) > 3 else False)
                    c = ws.cell(row=3, column=col, value=disp + (' *' if db_f in required_return else ''))
                    c.font = hfont
                    c.fill = req_fill if db_f in required_return else opt_fill
                    c.alignment = Alignment(horizontal='center')
                    c.border = thin
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(w, 24)
                for r in range(4, 8):
                    for c in range(1, ncols + 1):
                        cell = ws.cell(row=r, column=c, value='')
                        cell.fill = smp_fill
                        cell.border = thin
                wb.save(file_path)
                CustomMessageBox.showinfo(dialog, "완료", f"반품 양식 저장 완료\n\n{file_path}\n\n재고 리스트와 동일 형식 + 반품수량·사유. 필수만 채우면 됩니다.")
                self._log(f"📥 반품 양식 다운로드: {file_path}")
            except (FileNotFoundError, OSError, PermissionError) as e:
                CustomMessageBox.showerror(dialog, "오류", f"파일 저장 실패: {e}")
        
        try:
            ttk.Button(top_bar, text="📥 반품 양식 다운로드", command=_download_return_template,
                       bootstyle="info").pack(side='left', padx=5)
        except TypeError:
            ttk.Button(top_bar, text="📥 반품 양식 다운로드", command=_download_return_template).pack(side='left', padx=5)
        
        file_var = tk.StringVar(value="파일을 선택하세요...")
        ttk.Label(top_bar, textvariable=file_var, foreground='gray').pack(side='left', padx=10, fill=X, expand=True)
        
        # 미리보기 Treeview — 필수 4열: Lot No, BL NO, 톤백중량, 반품수량(갯수)
        pv_frame = ttk.LabelFrame(tab_excel, text="반품 미리보기 (DB 자동 조회)")
        pv_frame.pack(fill=BOTH, expand=True, padx=10, pady=5)
        ttk.Label(pv_frame, text="※ 필수: Lot No, BL NO, 톤백중량, 반품수량(갯수) — 네 열 중 하나라도 비면 반품 업로드 시 오류 표시", font=('맑은 고딕', 9)).pack(anchor='w')
        cols = ('lot_no', 'bl_no', 'tonbag_no', 'product', 'weight_kg',
                'return_qty', 'reason', 'status', 'remark')
        pv_tree = ttk.Treeview(pv_frame, columns=cols, show='headings', height=10)
        for cid, txt, w in [
            ('lot_no', 'LOT NO *', 100), ('bl_no', 'BL NO *', 100),
            ('tonbag_no', 'Tonbag#', 65), ('product', 'Product', 90),
            ('weight_kg', '톤백중량(kg) *', 90), ('return_qty', '반품수량(갯수) *', 90),
            ('reason', '사유', 100), ('status', '상태', 70), ('remark', '비고', 100)]:
            pv_tree.heading(cid, text=txt)
            pv_tree.column(cid, width=w, anchor='center' if cid in ('tonbag_no','status') else 'w')
        pv_sb = ttk.Scrollbar(pv_frame, orient=VERTICAL, command=pv_tree.yview)
        pv_tree.configure(yscrollcommand=pv_sb.set)
        pv_tree.pack(side=LEFT, fill=BOTH, expand=True)
        pv_sb.pack(side=RIGHT, fill=Y)
        
        summary_var = tk.StringVar(value="데이터 붙여넣기 또는 파일 업로드하세요")
        ttk.Label(tab_excel, textvariable=summary_var, font=('맑은 고딕', 11, 'bold')).pack(pady=3)
        
        # 파싱된 반품 데이터 저장
        parsed_returns = []
        
        def _show_bulk_return_input_choice():
            """Excel/데이터 입력 원칙: 데이터 붙여넣기 vs 파일 업로드 선택 후 진행"""
            from ..utils.ui_constants import (
                UPLOAD_CHOICE_HEADER, UPLOAD_CHOICE_PASTE, UPLOAD_CHOICE_UPLOAD,
                UPLOAD_CHOICE_BTN_PASTE, UPLOAD_CHOICE_BTN_UPLOAD,
            )
            from ..utils.ui_constants import center_dialog, DialogSize, apply_modal_window_options
            result = [None]
            win = tk.Toplevel(dialog)
            win.title("다량 반품 데이터 입력")
            apply_modal_window_options(win)
            win.transient(dialog)
            win.grab_set()
            win.geometry(DialogSize.get_geometry(dialog, 'small'))
            win.minsize(400, 260)
            center_dialog(win, dialog)
            f = ttk.Frame(win, padding=(20, 20, 20, 32))
            f.pack(fill=tk.BOTH, expand=True)
            ttk.Label(f, text=UPLOAD_CHOICE_HEADER, font=('맑은 고딕', 12, 'bold')).pack(anchor='w', pady=(0, 12))
            ttk.Label(f, text=UPLOAD_CHOICE_PASTE, font=('맑은 고딕', 10), wraplength=400, justify=tk.LEFT).pack(anchor='w', pady=(0, 10))
            ttk.Label(f, text=UPLOAD_CHOICE_UPLOAD, font=('맑은 고딕', 10), wraplength=400, justify=tk.LEFT).pack(anchor='w', pady=(0, 24))
            btn_f = ttk.Frame(f)
            btn_f.pack(anchor='center')
            def on_paste():
                result[0] = 'paste'
                win.destroy()
            def on_upload():
                result[0] = 'upload'
                win.destroy()
            ttk.Button(btn_f, text=UPLOAD_CHOICE_BTN_PASTE, command=on_paste, width=22).pack(side=tk.LEFT, padx=(0, 8))
            ttk.Button(btn_f, text=UPLOAD_CHOICE_BTN_UPLOAD, command=on_upload, width=22).pack(side=tk.LEFT)
            win.protocol("WM_DELETE_WINDOW", win.destroy)
            win.wait_window(win)
            return result[0]
        
        def _upload_return_excel():
            """Excel/데이터 입력 원칙: 선택 후 파일 업로드 또는 붙여넣기 → DB 조회 → 미리보기"""
            from ..utils.constants import filedialog
            import pandas as pd
            
            choice = _show_bulk_return_input_choice()
            if not choice:
                return
            if choice == "paste":
                CustomMessageBox.showinfo(
                    dialog, "다량 반품 데이터 입력",
                    "다량 반품은 프로그램 내장 형식과 동일합니다.\n"
                    "반품 양식 다운로드 후 엑셀에 채워 [파일 업로드]를 선택하거나,\n"
                    "입고 메뉴의 [반품 입고 (Excel)]에서 데이터 붙여넣기를 사용하세요.")
                return
            fp = filedialog.askopenfilename(
                title="반품 Excel 선택", filetypes=[("Excel files", "*.xlsx *.xls")])
            if not fp:
                return
            file_var.set(os.path.basename(fp))
            
            try:
                df = pd.read_excel(fp, header=None)
                # 헤더 행 찾기 (lot_no 또는 LOT NO 포함 행)
                header_row = None
                for idx in range(min(5, len(df))):
                    row_vals = [str(v).lower().strip() for v in df.iloc[idx]]
                    if any('lot_no' in v or 'lot no' in v for v in row_vals):
                        header_row = idx
                        break
                if header_row is None:
                    header_row = 2  # 기본: 3행 (0-indexed 2)
                
                df.columns = df.iloc[header_row].astype(str).str.strip().str.lower()
                df = df.iloc[header_row+1:].reset_index(drop=True)
                df = df.dropna(how='all')
                
                # 컬럼 매핑
                col_map = {}
                for c in df.columns:
                    cl = str(c).lower().replace(' ', '_').replace('*', '').strip()
                    if 'lot' in cl and 'no' in cl:
                        col_map['lot_no'] = c
                    elif 'bl' in cl and 'no' in cl:
                        col_map['bl_no'] = c
                    elif 'tonbag' in cl or ('ton' in cl and 'bag' in cl):
                        col_map['tonbag_no'] = c
                    elif 'return' in cl and ('qty' in cl or 'kg' in cl):
                        col_map['return_qty'] = c
                    elif '반품수량' in str(c) or 'return_qty' in cl:
                        col_map['return_qty'] = c
                    elif 'reason' in cl or cl == 'return_reason':
                        col_map['reason'] = c
                    elif 'remark' in cl:
                        col_map['remark'] = c
                
                if 'lot_no' not in col_map:
                    CustomMessageBox.showerror(dialog, "오류", "LOT NO 컬럼을 찾을 수 없습니다.")
                    return
                if 'return_qty' not in col_map and 'reason' not in col_map:
                    CustomMessageBox.showwarning(dialog, "안내", "RETURN QTY (KG) 또는 RETURN REASON 컬럼을 권장합니다.")
                
                # 미리보기 구성 + 필수 4열 검증 (Lot No, BL NO, 톤백중량, 반품수량(갯수))
                pv_tree.delete(*pv_tree.get_children())
                parsed_returns.clear()
                ok_count = 0
                err_count = 0
                required_missing_rows = []  # 필수 누락 행 번호(Excel 1-based)
                returnable_statuses = ('PICKED', 'CONFIRMED', 'SHIPPED', 'SOLD', 'RESERVED')
                data_start_row = header_row + 2  # Excel 1-based 첫 데이터 행

                for idx, row in df.iterrows():
                    lot_no = str(row.get(col_map.get('lot_no', ''), '')).strip()
                    if not lot_no or lot_no == 'nan':
                        continue
                    excel_row = data_start_row + int(idx)

                    bl_no = str(row.get(col_map.get('bl_no', ''), '')).strip()
                    tonbag_str = str(row.get(col_map.get('tonbag_no', ''), '')).strip()
                    qty_str = str(row.get(col_map.get('return_qty', ''), '')).strip()
                    reason = str(row.get(col_map.get('reason', ''), '품질 불량')).strip()
                    remark = str(row.get(col_map.get('remark', ''), '')).strip()
                    if reason == 'nan': reason = '품질 불량'
                    if remark == 'nan': remark = ''
                    if bl_no == 'nan': bl_no = ''
                    if tonbag_str == 'nan': tonbag_str = ''

                    try:
                        return_qty_kg = float(qty_str) if qty_str and qty_str != 'nan' else 0.0
                    except (ValueError, TypeError):
                        return_qty_kg = 0.0

                    product = ''
                    if hasattr(self, 'engine') and self.engine:
                        try:
                            inv = self.engine.db.fetchone(
                                "SELECT product, bl_no FROM inventory WHERE lot_no = ?", (lot_no,))
                            if inv:
                                product = inv['product'] if isinstance(inv, dict) else inv[0]
                                if not bl_no:
                                    bl_no = (inv['bl_no'] if isinstance(inv, dict) else inv[1]) or ''
                        except (sqlite3.OperationalError, sqlite3.IntegrityError, OSError):
                            pass

                    # 필수 4열 검증: Lot No, BL NO, 톤백중량, 반품수량(갯수)
                    missing_this_row = []
                    if not lot_no or lot_no == 'nan':
                        missing_this_row.append('Lot No')
                    if not bl_no or bl_no == 'nan':
                        missing_this_row.append('BL NO')
                    if not qty_str or qty_str == 'nan' or return_qty_kg <= 0:
                        missing_this_row.append('반품수량(갯수)')

                    items_to_add = []
                    if tonbag_str:
                        sub_lt = 1
                        try:
                            sub_lt = int(float(tonbag_str))
                        except (ValueError, TypeError):
                            sub_lt = 1
                        if hasattr(self, 'engine') and self.engine:
                            tb = self.engine.db.fetchone(
                                "SELECT sub_lt, weight, status FROM inventory_tonbag "
                                "WHERE lot_no = ? AND sub_lt = ?", (lot_no, sub_lt))
                            if tb:
                                sub_lt = tb['sub_lt'] if isinstance(tb, dict) else tb[0]
                                weight_kg = tb['weight'] if isinstance(tb, dict) else tb[1]
                                status = tb['status'] if isinstance(tb, dict) else tb[2]
                                if status in returnable_statuses:
                                    items_to_add = [{'sub_lt': sub_lt, 'weight_kg': weight_kg, 'status': status}]
                            if not items_to_add:
                                items_to_add = [{'sub_lt': sub_lt, 'weight_kg': 0, 'status': 'NOT FOUND'}]
                        else:
                            items_to_add = [{'sub_lt': sub_lt, 'weight_kg': 0, 'status': '?'}]
                    elif return_qty_kg > 0 and hasattr(self, 'engine') and self.engine:
                        tonbags = self.engine.db.fetchall(
                            """SELECT sub_lt, weight, status FROM inventory_tonbag
                               WHERE lot_no = ? AND status IN (?,?,?,?,?)
                               ORDER BY sub_lt DESC""",
                            (lot_no,) + returnable_statuses)
                        remaining_kg = return_qty_kg
                        for tb in (tonbags or []):
                            if remaining_kg <= 0.01:
                                break
                            w = float(tb['weight'] if isinstance(tb, dict) else tb[1] or 0)
                            if w <= 0:
                                continue
                            items_to_add.append({
                                'sub_lt': tb['sub_lt'] if isinstance(tb, dict) else tb[0],
                                'weight_kg': w,
                                'status': tb['status'] if isinstance(tb, dict) else tb[2],
                            })
                            remaining_kg -= w
                        if not items_to_add:
                            items_to_add = [{'sub_lt': 0, 'weight_kg': 0, 'status': 'NOT FOUND'}]
                    else:
                        if hasattr(self, 'engine') and self.engine:
                            tb = self.engine.db.fetchone(
                                "SELECT sub_lt, weight, status FROM inventory_tonbag "
                                "WHERE lot_no = ? AND status IN (?,?,?,?,?) "
                                "ORDER BY sub_lt LIMIT 1", (lot_no,) + returnable_statuses)
                            if tb:
                                items_to_add = [{
                                    'sub_lt': tb['sub_lt'] if isinstance(tb, dict) else tb[0],
                                    'weight_kg': tb['weight'] if isinstance(tb, dict) else tb[1],
                                    'status': tb['status'] if isinstance(tb, dict) else tb[2],
                                }]
                        if not items_to_add:
                            items_to_add = [{'sub_lt': 0, 'weight_kg': 0, 'status': 'NOT FOUND'}]

                    has_tonbag_weight = any(float(it.get('weight_kg', 0) or 0) > 0 for it in items_to_add)
                    if not has_tonbag_weight:
                        missing_this_row.append('톤백중량')
                    if missing_this_row:
                        required_missing_rows.append((excel_row, missing_this_row))

                    for it in items_to_add:
                        sub_lt = it['sub_lt']
                        weight_kg = it['weight_kg']
                        status = it['status']
                        is_ok = status in returnable_statuses and not missing_this_row
                        tag = 'ok' if is_ok else 'err'
                        if is_ok:
                            ok_count += 1
                        else:
                            err_count += 1
                        disp_qty = qty_str if qty_str and qty_str != 'nan' else (f"{weight_kg:.1f}" if weight_kg else '')
                        pv_tree.insert('', 'end', values=(
                            lot_no, bl_no, sub_lt, product,
                            f"{weight_kg:.1f}" if isinstance(weight_kg, (int, float)) else weight_kg,
                            disp_qty, reason, status, remark
                        ), tags=(tag,))
                        parsed_returns.append({
                            'lot_no': lot_no, 'sub_lt': sub_lt,
                            'reason': reason, 'remark': remark,
                            'status': status, 'valid': is_ok,
                        })
                
                pv_tree.tag_configure('ok', foreground='#27ae60')
                pv_tree.tag_configure('err', foreground='#e74c3c')
                
                if required_missing_rows:
                    lines = [f"  행 {r}: {', '.join(m)}" for r, m in required_missing_rows[:20]]
                    if len(required_missing_rows) > 20:
                        lines.append(f"  ... 외 {len(required_missing_rows) - 20}행")
                    CustomMessageBox.showerror(dialog, "필수 항목 누락",
                        "반품 업로드 시 Lot No, BL NO, 톤백중량, 반품수량(갯수) 네 열은 필수입니다.\n\n"
                        "다음 행에 필수 항목이 비어 있습니다:\n\n" + "\n".join(lines) + "\n\n수정 후 다시 업로드하세요.")
                
                summary_var.set(f"✅ 반품 가능: {ok_count}건  |  ❌ 불가: {err_count}건  |  총: {ok_count + err_count}건")
                
            except (FileNotFoundError, OSError, PermissionError) as e:
                CustomMessageBox.showerror(dialog, "오류", f"파일 읽기 실패: {e}")
        
        def _execute_bulk_return():
            """일괄 반품 실행 — v6.12.2: 자동승인/관리자확인 워크플로우."""
            valid_items = [r for r in parsed_returns if r.get('valid')]
            if not valid_items:
                CustomMessageBox.showwarning(dialog, "안내", "반품 가능한 항목이 없습니다.")
                return

            # v6.12.2: 톤백 수 기준 자동승인 / 관리자확인 분기
            try:
                from engine_modules.constants import RETURN_AUTO_APPROVE_MAX_TONBAGS
                _max_auto = RETURN_AUTO_APPROVE_MAX_TONBAGS
            except ImportError:
                _max_auto = 5

            total_tonbags = len(valid_items)
            if total_tonbags > _max_auto:
                # 대량 반품 → 관리자 확인 강화 (사유 + 건수 명시)
                reasons = set(r.get('reason', '미기재') for r in valid_items)
                reason_summary = ', '.join(reasons) if len(reasons) <= 3 else f"{len(reasons)}종"
                confirm_msg = (
                    f"⚠️ 대량 반품 — 관리자 확인 필요\n\n"
                    f"반품 건수: {total_tonbags}건 (자동승인 기준 {_max_auto}건 초과)\n"
                    f"반품 사유: {reason_summary}\n\n"
                    f"대량 반품은 되돌릴 수 없습니다.\n"
                    f"정말 진행하시겠습니까?"
                )
                if not CustomMessageBox.askyesno(dialog, "⚠️ 대량 반품 관리자 확인", confirm_msg):
                    return
                # 이중 확인
                if not CustomMessageBox.askyesno(dialog, "최종 확인",
                    f"총 {total_tonbags}건 반품을 최종 실행합니다.\n\n확인하셨습니까?"):
                    return
            else:
                # 소량 반품 → 간단 확인
                if not CustomMessageBox.askyesno(dialog, "일괄 반품 확인",
                    f"총 {total_tonbags}건을 반품 처리합니다.\n\n계속하시겠습니까?"):
                    return
            
            if hasattr(self.engine, 'process_return'):
                try:
                    result = self.engine.process_return(valid_items,
                        source_type='RETURN_EXCEL')
                except (ValueError, RuntimeError, sqlite3.OperationalError, sqlite3.IntegrityError) as e:
                    CustomMessageBox.showerror(dialog, "반품 오류", f"일괄 반품 처리 중 오류:\n{str(e)[:500]}")
                    return
                if result.get('success'):
                    self._log(f"✅ 일괄 반품 완료: {result.get('returned', 0)}건")
                    CustomMessageBox.showinfo(dialog, "완료",
                        f"일괄 반품 처리 완료\n\n"
                        f"성공: {result.get('returned', 0)}건\n"
                        f"스킵: {result.get('skipped', 0)}건")
                    dialog.destroy()
                    if hasattr(self, '_deferred_refresh_main_tabs'):
                        self._deferred_refresh_main_tabs(delay_ms=50)
                    elif hasattr(self, '_refresh_main_tabs'):
                        self._refresh_main_tabs()
                    else:
                        self._refresh_inventory()
                        if hasattr(self, '_refresh_tonbag'):
                            self._refresh_tonbag()
                    if hasattr(self, '_refresh_dashboard'):
                        self._refresh_dashboard()
                else:
                    errs = '\n'.join(result.get('errors', ['알 수 없는 오류'])[:5])
                    CustomMessageBox.showerror(dialog, "오류", f"반품 처리 오류:\n{errs}")
            else:
                CustomMessageBox.showwarning(dialog, "안내", "반품 엔진을 찾을 수 없습니다.")
        
        try:
            ttk.Button(top_bar, text="📂 데이터 입력", command=_upload_return_excel,
                       bootstyle="warning").pack(side='left', padx=5)
        except TypeError:
            ttk.Button(top_bar, text="📂 데이터 입력", command=_upload_return_excel).pack(side='left', padx=5)
        
        ex_btn = ttk.Frame(tab_excel)
        ex_btn.pack(pady=8)
        try:
            ttk.Button(ex_btn, text="🔄 일괄 반품 실행", command=_execute_bulk_return,
                       bootstyle="danger").pack(side='left', padx=10)
        except TypeError:
            ttk.Button(ex_btn, text="🔄 일괄 반품 실행", command=_execute_bulk_return).pack(side='left', padx=10)
        ttk.Button(ex_btn, text="취소", command=dialog.destroy).pack(side='left', padx=10)
    
    # =========================================================================
    # v3.8.4: 수동 입고 입력 다이얼로그
    # =========================================================================
    
    # =========================================================================
    # v3.8.4: 문서 변환 (OCR/PDF)
    # =========================================================================
    
    def _show_document_convert_dialog(self) -> None:
        """v3.8.4: 문서 변환 (OCR 스캔 / PDF → Excel/Word)"""
        from ..utils.constants import tk, ttk, BOTH, X
        from ..utils.custom_messagebox import CustomMessageBox
        
        dialog = tk.Toplevel(self.root)
        dialog.title("📄 문서 변환 (OCR/PDF)")
        dialog.geometry("500x400")
        apply_modal_window_options(dialog)
        dialog.transient(self.root)
        dialog.grab_set()
        
        ttk.Label(dialog, text="문서 변환", font=('맑은 고딕', 16, 'bold')).pack(pady=10)
        
        # 변환 모드 선택
        mode_frame = ttk.LabelFrame(dialog, text="변환 모드 선택")
        mode_frame.pack(fill=X, padx=20, pady=5)
        
        mode_var = tk.StringVar(value='ocr_scan')
        modes = [
            ('ocr_scan', '📷 OCR 스캔 (이미지/스캔 PDF → 텍스트 추출)'),
            ('pdf_convert', '📄 PDF → Excel/Word 변환'),
        ]
        for val, text in modes:
            ttk.Radiobutton(mode_frame, text=text, variable=mode_var, 
                           value=val).pack(anchor='w', padx=10, pady=3)
        
        # 출력 형식
        out_frame = ttk.LabelFrame(dialog, text="출력 형식")
        out_frame.pack(fill=X, padx=20, pady=5)
        
        out_var = tk.StringVar(value='excel')
        ttk.Radiobutton(out_frame, text='📊 Excel (.xlsx)', variable=out_var, 
                        value='excel').pack(anchor='w', padx=10, pady=3)
        ttk.Radiobutton(out_frame, text='📝 Word (.docx)', variable=out_var, 
                        value='word').pack(anchor='w', padx=10, pady=3)
        
        # 파일 선택
        file_frame = ttk.LabelFrame(dialog, text="파일 선택")
        file_frame.pack(fill=X, padx=20, pady=5)
        
        file_path_var = tk.StringVar()
        ttk.Entry(file_frame, textvariable=file_path_var, width=45).pack(side='left', padx=5, pady=5)
        
        def browse_file():
            from tkinter import filedialog
            filetypes = [
                ("지원 파일", "*.pdf;*.png;*.jpg;*.jpeg;*.tif;*.tiff;*.bmp"),
                ("PDF 파일", "*.pdf"),
                ("이미지 파일", "*.png;*.jpg;*.jpeg;*.tif;*.tiff;*.bmp"),
            ]
            path = filedialog.askopenfilename(parent=dialog, filetypes=filetypes)
            if path:
                file_path_var.set(path)
        
        ttk.Button(file_frame, text="찾아보기", command=browse_file).pack(side='left', padx=5, pady=5)
        
        def process_convert():
            filepath = file_path_var.get().strip()
            if not filepath or not os.path.exists(filepath):
                CustomMessageBox.showwarning(dialog, "파일 선택", "변환할 파일을 선택하세요.")
                return
            
            mode = mode_var.get()
            out_fmt = out_var.get()
            
            api_ok = getattr(self, '_api_connected', False)
            if mode == 'ocr_scan' and not api_ok:
                CustomMessageBox.showwarning(dialog, "API 필요",
                    "OCR 스캔에는 Gemini API가 필요합니다.\n\n"
                    "도구 > Gemini API 설정에서 API Key를 설정하세요.\n"
                    "https://aistudio.google.com에서 무료 발급 가능합니다.")
                return
            
            CustomMessageBox.showinfo(dialog, "변환 시작",
                f"변환 모드: {'OCR 스캔' if mode == 'ocr_scan' else 'PDF 변환'}\n"
                f"출력 형식: {'Excel' if out_fmt == 'excel' else 'Word'}\n"
                f"파일: {os.path.basename(filepath)}\n\n"
                f"변환 기능은 다음 업데이트에서 구현됩니다.\n"
                f"(Gemini Vision API 연동 예정)")
        
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=15)
        try:
            ttk.Button(btn_frame, text="🔄 변환 시작", command=process_convert,
                       bootstyle="info").pack(side='left', padx=10)
        except TypeError:
            ttk.Button(btn_frame, text="🔄 변환 시작", command=process_convert).pack(side='left', padx=10)
        ttk.Button(btn_frame, text="닫기", command=dialog.destroy).pack(side='left', padx=10)

    # ═══════════════════════════════════════════════════════
    # v3.8.4: 출고 이력 조회
    # ═══════════════════════════════════════════════════════
    
    def _show_outbound_history(self) -> None:
        """출고 이력(stock_movement) 조회 팝업"""
        from ..utils.constants import tk, ttk, BOTH, X, Y, LEFT, RIGHT, END
        from ..utils.custom_messagebox import CustomMessageBox
        
        dialog = tk.Toplevel(self.root)
        dialog.title("📋 출고 이력 조회")
        dialog.geometry("900x500")
        apply_modal_window_options(dialog)
        dialog.transient(self.root)
        
        # 필터
        filter_frame = ttk.Frame(dialog)
        filter_frame.pack(fill=X, padx=10, pady=5)
        
        ttk.Label(filter_frame, text="유형:").pack(side=LEFT, padx=5)
        type_var = tk.StringVar(value='전체')
        type_cb = ttk.Combobox(filter_frame, textvariable=type_var, state='readonly', width=15,
                               values=['전체', 'OUTBOUND', 'CANCEL_OUTBOUND', 'INBOUND', 'RETURN'])
        type_cb.pack(side=LEFT, padx=5)
        
        ttk.Label(filter_frame, text="LOT:").pack(side=LEFT, padx=(15, 5))
        lot_var = tk.StringVar()
        ttk.Entry(filter_frame, textvariable=lot_var, width=15).pack(side=LEFT, padx=5)
        
        # 트리뷰
        tree_frame = ttk.Frame(dialog)
        tree_frame.pack(fill=BOTH, expand=True, padx=10, pady=5)
        
        cols = ('id', 'lot_no', 'type', 'qty_kg', 'customer', 'date', 'created')
        tree = ttk.Treeview(tree_frame, columns=cols, show='headings', height=15)
        
        for col, text, w in [
            ('id', 'ID', 50), ('lot_no', 'LOT NO', 120), ('type', '유형', 120),
            ('qty_kg', '수량(kg)', 100), ('customer', '고객', 120),
            ('date', '날짜', 100), ('created', '생성일', 140)
        ]:
            tree.heading(col, text=text)
            tree.column(col, width=w, anchor='e' if col == 'qty_kg' else 'w')
        
        scrollbar = ttk.Scrollbar(tree_frame, orient='vertical', command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side=LEFT, fill=BOTH, expand=True)
        scrollbar.pack(side=RIGHT, fill=Y)
        
        # 합계
        summary_var = tk.StringVar(value="조회 버튼을 클릭하세요")
        ttk.Label(dialog, textvariable=summary_var, font=('', 13, 'bold')).pack(pady=5)
        
        def do_search():
            tree.delete(*tree.get_children())
            try:
                query = "SELECT id, lot_no, movement_type, qty_kg, '' AS customer, created_at AS movement_date, created_at FROM stock_movement WHERE 1=1"
                params = []
                
                mv_type = type_var.get()
                if mv_type != '전체':
                    query += " AND movement_type = ?"
                    params.append(mv_type)
                
                lot_filter = lot_var.get().strip()
                if lot_filter:
                    query += " AND lot_no LIKE ?"
                    params.append(f"%{lot_filter}%")
                
                query += " ORDER BY created_at DESC LIMIT 500"
                
                rows = self.engine.db.fetchall(query, tuple(params))
                total_kg = 0
                for r in rows:
                    row_id = r['id'] if isinstance(r, dict) else r[0]
                    lot = r['lot_no'] if isinstance(r, dict) else r[1]
                    mtype = r['movement_type'] if isinstance(r, dict) else r[2]
                    qty = r['qty_kg'] if isinstance(r, dict) else r[3]
                    cust = r['customer'] if isinstance(r, dict) else r[4]
                    mdate = r['movement_date'] if isinstance(r, dict) else r[5]
                    created = r['created_at'] if isinstance(r, dict) else r[6]
                    
                    qty_val = float(qty) if qty else 0
                    total_kg += qty_val
                    
                    tree.insert('', END, values=(
                        row_id, lot, mtype, f"{qty_val:,.0f}",
                        cust or '', str(mdate or '')[:10], str(created or '')[:19]
                    ))
                
                summary_var.set(f"조회: {len(rows)}건 | 총 수량: {total_kg:,.0f} kg")
                
            except (ValueError, TypeError, KeyError) as e:
                summary_var.set(f"오류: {e}")
        
        ttk.Button(filter_frame, text="🔍 조회", command=do_search).pack(side=LEFT, padx=15)
        ttk.Button(filter_frame, text="❌ 닫기", command=dialog.destroy).pack(side=RIGHT, padx=5)
        
        # 초기 로드
        do_search()

    # ═══════════════════════════════════════════════════════
    # v3.8.4 A6: 재고 추이 차트
    # ═══════════════════════════════════════════════════════
    
    def _show_snapshot_chart(self) -> None:
        """재고 스냅샷 추이 차트"""
        from ..utils.constants import tk, ttk, BOTH, X, LEFT, END
        from ..utils.custom_messagebox import CustomMessageBox
        
        try:
            rows = self.engine.db.fetchall("""
                SELECT snapshot_date, total_lots, total_weight_kg, 
                       available_weight_kg, picked_weight_kg
                FROM inventory_snapshot 
                ORDER BY snapshot_date DESC LIMIT 30
            """)
            
            if not rows:
                CustomMessageBox.showinfo(self.root, "재고 추이",
                    "스냅샷 데이터가 아직 없습니다.\n\n프로그램을 매일 실행하면 자동으로 축적됩니다.")
                return
            
            rows = list(reversed(rows))
            
            dialog = tk.Toplevel(self.root)
            dialog.title("📊 재고 추이 (최근 30일)")
            dialog.geometry("800x400")
            apply_modal_window_options(dialog)
            dialog.transient(self.root)
            
            # 표 형태
            tree = ttk.Treeview(dialog, columns=('date', 'lots', 'total', 'avail', 'picked'),
                               show='headings', height=15)
            
            for col, text, w in [('date','날짜',100), ('lots','LOT수',60),
                                 ('total','총재고(MT)',100), ('avail','판매가능(MT)',100),
                                 ('picked','출고(MT)',100)]:
                tree.heading(col, text=text)
                tree.column(col, width=w, anchor='e' if col != 'date' else 'w')
            
            for r in rows:
                tree.insert('', END, values=(
                    r['snapshot_date'],
                    r['total_lots'],
                    f"{(r['total_weight_kg'] or 0)/1000:,.1f}",
                    f"{(r['available_weight_kg'] or 0)/1000:,.1f}",
                    f"{(r['picked_weight_kg'] or 0)/1000:,.1f}",
                ))
            
            tree.pack(fill=BOTH, expand=True, padx=10, pady=10)
            ttk.Button(dialog, text="닫기", command=dialog.destroy).pack(pady=5)
            
        except (RuntimeError, ValueError) as e:
            CustomMessageBox.showerror(self.root, "오류", f"스냅샷 조회 오류:\n{e}")

    # ═══════════════════════════════════════════════════════
    # v3.8.4 A7: 출고 거래명세서 PDF/Excel
    # ═══════════════════════════════════════════════════════
    
    def _generate_outbound_invoice(self) -> None:
        """출고 거래명세서 Excel 생성"""
        from ..utils.constants import tk, ttk, filedialog, BOTH, X, LEFT, W, END
        from ..utils.custom_messagebox import CustomMessageBox
        
        # 고객 + 기간 선택 다이얼로그
        dialog = tk.Toplevel(self.root)
        dialog.title("📄 거래명세서 생성")
        dialog.geometry("400x250")
        apply_modal_window_options(dialog)
        dialog.transient(self.root)
        dialog.grab_set()
        
        frame = ttk.Frame(dialog, padding=15)
        frame.pack(fill=BOTH, expand=True)
        
        ttk.Label(frame, text="고객명:").grid(row=0, column=0, sticky=W, pady=5)
        cust_var = tk.StringVar()
        
        # 고객 목록 조회
        try:
            customers = self.engine.db.fetchall(
                "SELECT DISTINCT customer FROM stock_movement WHERE customer != '' ORDER BY customer")
            cust_list = [c['customer'] for c in customers if c['customer']]
        except (sqlite3.OperationalError, sqlite3.IntegrityError, OSError):
            cust_list = []
        
        ttk.Combobox(frame, textvariable=cust_var, values=cust_list, width=30).grid(
            row=0, column=1, sticky=W, pady=5)
        
        ttk.Label(frame, text="시작일:").grid(row=1, column=0, sticky=W, pady=5)
        from_var = tk.StringVar(value="2025-01-01")
        ttk.Entry(frame, textvariable=from_var, width=15).grid(row=1, column=1, sticky=W, pady=5)
        
        ttk.Label(frame, text="종료일:").grid(row=2, column=0, sticky=W, pady=5)
        from datetime import date
        to_var = tk.StringVar(value=date.today().isoformat())
        ttk.Entry(frame, textvariable=to_var, width=15).grid(row=2, column=1, sticky=W, pady=5)
        
        def do_generate():
            customer = cust_var.get().strip()
            date_from = from_var.get().strip()
            date_to = to_var.get().strip()
            
            if not customer:
                CustomMessageBox.showwarning(dialog, "입력 필요", "고객명을 선택하세요.")
                return
            
            # 출고 데이터 조회 (customer 컬럼 없어도 동작)
            try:
                for q, p in [
                    ("""SELECT lot_no, movement_type, qty_kg, customer, 
                            COALESCE(movement_date, created_at) AS movement_date, created_at
                        FROM stock_movement 
                        WHERE customer = ? AND movement_type = 'OUTBOUND'
                          AND COALESCE(movement_date, created_at) >= ? AND COALESCE(movement_date, created_at) <= ?
                        ORDER BY created_at""",
                     (customer, date_from, date_to + ' 23:59:59')),
                    ("""SELECT lot_no, movement_type, qty_kg, '' AS customer, 
                            created_at AS movement_date, created_at
                        FROM stock_movement 
                        WHERE movement_type = 'OUTBOUND'
                          AND created_at >= ? AND created_at <= ?
                        ORDER BY created_at""",
                     (date_from, date_to + ' 23:59:59')),
                ]:
                    try:
                        movements = self.engine.db.fetchall(q, p)
                        break
                    except (sqlite3.OperationalError,) as e:
                        if "no such column" in str(e).lower():
                            continue
                        raise
                else:
                    movements = []
                
                if not movements:
                    CustomMessageBox.showinfo(dialog, "결과 없음", "해당 기간 출고 이력이 없습니다.")
                    return
                
                # Excel 저장
                save_path = filedialog.asksaveasfilename(
                    title="거래명세서 저장",
                    defaultextension=".xlsx",
                    initialfile=f"거래명세서_{customer}_{date_from}_{date_to}.xlsx",
                    filetypes=[("Excel files", "*.xlsx")]
                )
                
                if not save_path:
                    return
                try:
                    from ..utils.excel_file_helper import get_unique_excel_path
                    save_path = get_unique_excel_path(save_path)
                except ImportError:
                    pass
                
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "거래명세서"
                
                # 스타일
                title_font = Font(bold=True, size=16)
                header_font = Font(bold=True, color="FFFFFF", size=10)
                header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
                border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
                
                # 타이틀
                ws.merge_cells('A1:F1')
                ws['A1'] = f"거래명세서 — {customer}"
                ws['A1'].font = title_font
                
                ws['A2'] = f"기간: {date_from} ~ {date_to}"
                ws['A2'].font = Font(size=10, color='666666')
                
                # 헤더
                headers = ['No', 'LOT NO', '수량(kg)', '수량(MT)', '출고일', '비고']
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(row=4, column=col, value=h)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                
                # 데이터
                total_kg = 0
                for i, mv in enumerate(movements, 1):
                    qty = mv['qty_kg'] or 0
                    total_kg += qty
                    
                    row_data = [
                        i,
                        mv['lot_no'],
                        f"{qty:,.0f}",
                        f"{qty/1000:.3f}",
                        str(mv['movement_date'] or '')[:10],
                        ''
                    ]
                    for col, val in enumerate(row_data, 1):
                        cell = ws.cell(row=4+i, column=col, value=val)
                        cell.border = border
                        if col in (3, 4):
                            cell.alignment = Alignment(horizontal='right')
                
                # 합계
                sum_row = 5 + len(movements)
                ws.cell(row=sum_row, column=1, value="합계").font = Font(bold=True)
                ws.cell(row=sum_row, column=3, value=f"{total_kg:,.0f}").font = Font(bold=True)
                ws.cell(row=sum_row, column=4, value=f"{total_kg/1000:.3f}").font = Font(bold=True)
                
                ws.column_dimensions['A'].width = 6
                ws.column_dimensions['B'].width = 16
                ws.column_dimensions['C'].width = 14
                ws.column_dimensions['D'].width = 12
                ws.column_dimensions['E'].width = 14
                ws.column_dimensions['F'].width = 15
                
                wb.save(save_path)
                
                dialog.destroy()
                self._log(f"✅ 거래명세서 저장: {save_path}")
                CustomMessageBox.showinfo(self.root, "완료",
                    f"거래명세서가 저장되었습니다.\n\n"
                    f"고객: {customer}\n"
                    f"건수: {len(movements)}건\n"
                    f"총량: {total_kg/1000:.3f} MT\n\n"
                    f"파일: {save_path}")
                    
            except ImportError:
                CustomMessageBox.showerror(dialog, "오류", "openpyxl이 필요합니다.")
            except (RuntimeError, ValueError) as e:
                CustomMessageBox.showerror(dialog, "오류", f"거래명세서 생성 오류:\n{e}")
        
        ttk.Button(frame, text="📄 생성", command=do_generate).grid(row=3, column=1, sticky=W, pady=15)
        ttk.Button(frame, text="취소", command=dialog.destroy).grid(row=3, column=0, sticky=W, pady=15)

    def _show_return_statistics(self) -> None:
        """v6.12.1: 반품 사유 통계 리포트 다이얼로그."""
        try:
            from ..dialogs.return_statistics_dialog import ReturnStatisticsDialog
            current_theme = getattr(self, '_current_theme', 'flatly')
            ReturnStatisticsDialog(self.root, self.engine, current_theme=current_theme)
        except ImportError as e:
            logger.debug(f"반품 통계 다이얼로그 로드 실패: {e}")
            CustomMessageBox.showerror(self.root, "오류", "반품 통계 다이얼로그를 불러올 수 없습니다.")

    def _send_return_alert_email(self) -> None:
        """v6.12.2: 반품 경고 이메일 수동 발송."""
        try:
            from features.notifications.return_alert_email import (
                send_return_alert_email, check_return_alerts, load_email_config,
                create_default_email_config
            )
        except ImportError as e:
            CustomMessageBox.showerror(self.root, "오류", f"알림 모듈 로드 실패:\n{e}")
            return

        config = load_email_config()
        if not config.get('enabled'):
            if CustomMessageBox.askyesno(self.root, "이메일 알림 미설정",
                "이메일 알림이 비활성화 상태입니다.\n\n"
                "config_email.json 템플릿을 바탕화면에 생성하시겠습니까?\n"
                "(생성 후 SMTP 정보를 입력하고 enabled=true로 변경하세요)"):
                import os
                desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
                path = create_default_email_config(desktop)
                CustomMessageBox.showinfo(self.root, "완료",
                    f"config_email.json 템플릿이 생성되었습니다.\n\n{path}\n\n"
                    f"파일을 편집한 후 앱 루트 폴더에 복사하세요.")
            return

        alerts = check_return_alerts(self.engine)
        if not alerts:
            CustomMessageBox.showinfo(self.root, "안내", "현재 반품 경고 대상 LOT이 없습니다.")
            return

        lot_summary = '\n'.join(f"  • {a['lot_no']}: {a['count']}회" for a in alerts[:5])
        if not CustomMessageBox.askyesno(self.root, "반품 경고 이메일 발송",
            f"반복 반품 LOT {len(alerts)}건:\n{lot_summary}\n\n"
            f"수신자: {', '.join(config.get('recipients', []))}\n\n발송하시겠습니까?"):
            return

        result = send_return_alert_email(self.engine, config)
        if result['sent']:
            self._log(f"📧 반품 경고 이메일 발송 완료: {result['alert_count']}건")
            CustomMessageBox.showinfo(self.root, "발송 완료",
                f"반품 경고 이메일이 발송되었습니다.\n\n대상: {result['alert_count']}건")
        else:
            CustomMessageBox.showerror(self.root, "발송 실패",
                f"이메일 발송에 실패했습니다.\n\n{result.get('error', '알 수 없는 오류')}")
    def _show_email_config(self) -> None:
        """v6.12.2: 이메일 설정 GUI 편집기."""
        try:
            from ..dialogs.email_config_dialog import EmailConfigDialog
            current_theme = getattr(self, '_current_theme', 'flatly')
            EmailConfigDialog(self.root, current_theme=current_theme)
        except ImportError as e:
            logger.debug(f"이메일 설정 다이얼로그 로드 실패: {e}")
            CustomMessageBox.showerror(self.root, "오류", "이메일 설정 다이얼로그를 불러올 수 없습니다.")

    def _on_integrity_report(self) -> None:
        """v7.0.1: 정합성 검증 리포트 생성 (PDF + Excel)"""
        from ..utils.constants import filedialog
        from datetime import datetime
        
        today = datetime.now().strftime('%Y%m%d')
        
        file_path = filedialog.asksaveasfilename(
            title="정합성 검증 리포트 저장",
            defaultextension=".xlsx",
            initialfile=f"SQM-IntegrityReport-{today}.xlsx",
            filetypes=[
                ("Excel files", "*.xlsx"),
                ("PDF files", "*.pdf"),
                ("All files", "*.*")
            ]
        )
        
        if not file_path:
            return
        
        try:
            if file_path.lower().endswith('.pdf'):
                from features.reports.integrity_report import generate_integrity_report_pdf
                result = generate_integrity_report_pdf(self.engine, file_path)
            else:
                from features.reports.integrity_report import generate_integrity_report_excel
                result = generate_integrity_report_excel(self.engine, file_path)
            
            if result:
                CustomMessageBox.showinfo(
                    self.root, "완료",
                    f"정합성 검증 리포트 생성 완료\n\n{os.path.basename(file_path)}"
                )
                # 탐색기에서 열기
                import subprocess, platform
                if platform.system() == 'Windows':
                    subprocess.Popen(['explorer', '/select,', file_path])
            else:
                CustomMessageBox.showerror(self.root, "오류", "리포트 생성 실패")
        except Exception as e:
            logger.error(f"정합성 리포트 생성 오류: {e}", exc_info=True)
            CustomMessageBox.showerror(self.root, "오류", f"리포트 생성 오류:\n{e}")
