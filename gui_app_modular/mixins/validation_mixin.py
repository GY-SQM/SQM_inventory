# -*- coding: utf-8 -*-
"""
SQM Inventory - Validation Mixin
================================

v2.9.91 - Extracted from gui_app.py

Preflight validation, data validation, and validation result dialogs
"""

import sqlite3
import logging
from ..utils.ui_constants import CustomMessageBox
from datetime import datetime
from typing import Optional, List, Dict, Any

logger = logging.getLogger(__name__)


class ValidationMixin:
    """
    Validation mixin
    
    Mixed into SQMInventoryApp class
    """
    
    def _validate_inbound_file(self) -> None:
        """Validate inbound file before processing"""
        from ..utils.constants import filedialog, HAS_VALIDATOR
        
        if not HAS_VALIDATOR:
            CustomMessageBox.showwarning(self.root, "Warning", "Validator module not available")
            return
        
        file_path = filedialog.askopenfilename(
            title="Select Inbound File to Validate",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if not file_path:
            return
        
        self._log(f"Validating inbound file: {file_path}")
        
        try:
            from validators import DataValidator, read_excel_for_validation
            
            data, headers = read_excel_for_validation(file_path)
            
            if not data:
                CustomMessageBox.showwarning(self.root, "Warning", "No data in file")
                return
            
            validator = DataValidator(self.engine.db)
            result = validator.validate_inbound(data, check_db=True)
            
            self._show_validation_result(
                f"Inbound Validation: {file_path.split('/')[-1]}", 
                result, 
                validator
            )
            
        except (OSError, RuntimeError) as e:
            self._log(f"X Validation error: {e}")
            CustomMessageBox.showerror(self.root, "Error", f"Validation failed: {e}")
    
    def _validate_outbound_file(self) -> None:
        """Validate outbound file before processing"""
        from ..utils.constants import filedialog, HAS_VALIDATOR
        
        if not HAS_VALIDATOR:
            CustomMessageBox.showwarning(self.root, "Warning", "Validator module not available")
            return
        
        file_path = filedialog.askopenfilename(
            title="Select Outbound File to Validate",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if not file_path:
            return
        
        self._log(f"Validating outbound file: {file_path}")
        
        try:
            from validators import DataValidator, read_excel_for_validation
            
            data, headers = read_excel_for_validation(file_path)
            
            if not data:
                CustomMessageBox.showwarning(self.root, "Warning", "No data in file")
                return
            
            validator = DataValidator(self.engine.db)
            result = validator.validate_outbound(data, check_db=True)
            
            self._show_validation_result(
                f"Outbound Validation: {file_path.split('/')[-1]}",
                result,
                validator
            )
            
        except (OSError, RuntimeError) as e:
            self._log(f"X Validation error: {e}")
            CustomMessageBox.showerror(self.root, "Error", f"Validation failed: {e}")
    
    def _show_validation_result(self, title: str, result, validator) -> None:
        """Show validation result dialog"""
        from ..utils.constants import tk, ttk, filedialog
        from ..utils.constants import BOTH, LEFT, RIGHT, X, Y, VERTICAL, END, W
        
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Validation: {title}")
        dialog.geometry("800x600")
        dialog.transient(self.root)
        
        # Summary frame
        summary_frame = ttk.LabelFrame(dialog, text="Summary")
        summary_frame.pack(fill=X, padx=10, pady=10)
        
        status_text = "PASS" if result.is_valid else "FAIL"
        status_color = "green" if result.is_valid else "red"
        
        ttk.Label(summary_frame, text=f"Result: {status_text}",
                  font=('', 16, 'bold')).grid(row=0, column=0, sticky=W, padx=10)
        ttk.Label(summary_frame, text=f"Rows: {result.total_rows}").grid(row=0, column=1, padx=20)
        ttk.Label(summary_frame, text=f"Errors: {result.error_count}",
                  foreground='red' if result.error_count > 0 else 'black').grid(row=0, column=2, padx=20)
        ttk.Label(summary_frame, text=f"Warnings: {result.warning_count}",
                  foreground='orange' if result.warning_count > 0 else 'black').grid(row=0, column=3, padx=20)
        
        # Detail frame
        detail_frame = ttk.LabelFrame(dialog, text="Details")
        detail_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)
        
        # Treeview
        columns = ("level", "row", "column", "value", "message", "suggestion")
        tree = ttk.Treeview(detail_frame, columns=columns, show="headings", height=15)
        
        tree.heading("level", text="Level")
        tree.heading("row", text="Row")
        tree.heading("column", text="Column")
        tree.heading("value", text="Value")
        tree.heading("message", text="Message")
        tree.heading("suggestion", text="Suggestion")
        
        tree.column("level", width=60, anchor="center")
        tree.column("row", width=50, anchor="center")
        tree.column("column", width=100)
        tree.column("value", width=100)
        tree.column("message", width=250)
        tree.column("suggestion", width=200)
        
        scrollbar = ttk.Scrollbar(detail_frame, orient=VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        tree.pack(side=LEFT, fill=BOTH, expand=True)
        scrollbar.pack(side=RIGHT, fill=Y)
        
        # Populate errors
        for err in result.errors:
            level_text = "ERROR" if err.level.value == "ERROR" else "WARNING"
            tree.insert('', END, values=(
                level_text,
                err.row if err.row > 0 else "-",
                err.column or "-",
                str(err.value)[:30] if err.value else "-",
                err.message,
                err.suggestion or "-"
            ))
        
        # Button frame
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(fill=X, padx=10, pady=10)
        
        def export_errors():
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"validation_errors_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            if not save_path:
                return
            
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Validation Errors"
                
                headers = ["Level", "Row", "Column", "Value", "Message", "Suggestion"]
                header_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=h)
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
                
                for row_idx, err in enumerate(result.errors, 2):
                    ws.cell(row=row_idx, column=1, value=err.level.value)
                    ws.cell(row=row_idx, column=2, value=err.row if err.row > 0 else "")
                    ws.cell(row=row_idx, column=3, value=err.column or "")
                    ws.cell(row=row_idx, column=4, value=str(err.value) if err.value else "")
                    ws.cell(row=row_idx, column=5, value=err.message)
                    ws.cell(row=row_idx, column=6, value=err.suggestion or "")
                
                wb.save(save_path)
                CustomMessageBox.showinfo(self.root, "Complete", f"Errors exported to:\n{save_path}")
                
            except (RuntimeError, ValueError) as e:
                CustomMessageBox.showerror(self.root, "Error", f"Export failed: {e}")
        
        ttk.Button(btn_frame, text="Export Errors", command=export_errors).pack(side=LEFT, padx=5)
        ttk.Button(btn_frame, text="Close", command=dialog.destroy).pack(side=RIGHT, padx=5)
        
        self._log(f"Validation {title}: {'PASS' if result.is_valid else f'{result.error_count} errors, {result.warning_count} warnings'}")
    
    def _run_preflight_check(self, data: List[Dict], operation: str = 'INBOUND') -> Optional[Any]:
        """
        Run preflight validation check
        
        Args:
            data: Data to validate
            operation: 'INBOUND' or 'OUTBOUND'
            
        Returns:
            PreflightResult or None if check passes
        """
        from ..utils.constants import HAS_PREFLIGHT
        
        if not HAS_PREFLIGHT:
            return None
        
        try:
            from engine_modules.preflight import PreflightValidator
            
            validator = PreflightValidator(self.engine.db)
            
            if operation == 'OUTBOUND':
                result = validator.validate_outbound(data, check_db=True)
            else:
                result = validator.validate_inbound(data, check_db=True)
            
            if result.has_blocking_errors():
                return result
            
            return None
            
        except (OSError, RuntimeError) as e:
            logger.error(f"Preflight check error: {e}")
            return None
    
    def _show_preflight_errors(self, result) -> None:
        """Show preflight validation errors"""

        
        error_msg = f"Preflight Validation Failed!\n\nErrors: {result.error_count}\n\nFix these issues:\n\n"
        
        for err in result.issues[:10]:
            if err.level.value in ('FATAL', 'ERROR'):
                cell_loc = err.get_cell_location() if hasattr(err, 'get_cell_location') else f"Row {err.row}"
                error_msg += f"• [{cell_loc}] {err.message}\n"
        
        if result.error_count > 10:
            error_msg += f"\n... and {result.error_count - 10} more errors"
        
        CustomMessageBox.showerror(self.root, "Validation Failed", error_msg)
    
    def _check_lot_exists(self, lot_no: str) -> bool:
        """Check if LOT exists in database"""
        try:
            result = self.engine.db.fetchone(
                "SELECT 1 FROM inventory WHERE lot_no = ?",
                (lot_no,)
            )
            return result is not None
        except (sqlite3.OperationalError, sqlite3.IntegrityError, OSError):
            return False
    
    def _check_tonbag_available(self, lot_no: str, sub_lt: int) -> bool:
        """Check if tonbag is available for outbound"""
        try:
            result = self.engine.db.fetchone(
                "SELECT status FROM inventory_tonbag WHERE lot_no = ? AND sub_lt = ?",
                (lot_no, sub_lt)
            )
            return result and result.get('status') == 'AVAILABLE'
        except (sqlite3.OperationalError, sqlite3.IntegrityError, OSError):
            return False
