# -*- coding: utf-8 -*-
"""
SQM v4.0.1 — 출고 템플릿/Allocation Mixin
===========================================

outbound_handlers.py에서 분리:
- 출고 양식 다운로드
- Allocation Table 생성 (샘플)
- Virtual Allocation 생성
"""
import logging
from ..utils.custom_messagebox import CustomMessageBox
from datetime import datetime

logger = logging.getLogger(__name__)


class OutboundTemplateMixin:
    """출고 템플릿 및 Allocation Table Mixin"""

    def _download_outbound_template(self) -> None:
        """화주 Allocation 양식 템플릿 생성 — 붙여넣기 또는 업로드용
        
        화주 표준 양식:
        Row 1: 타이틀 (Allocation - PT LBM - September / CIF Semarang - 300MT of MIc9000)
        Row 2: 합계 QTY
        Row 3: Product | SAP NO | ETA BUSAN | Date in stock | QTY (MT) | Lot No | WH | Customs | GW | SALE REF
        Row 4~: 데이터 (템플릿에 붙여넣기하거나, 이 양식으로 파일 작성 후 업로드)
        """
        from ..utils.constants import filedialog

        file_path = filedialog.asksaveasfilename(
            title="출고 Allocation Table 템플릿 저장",
            defaultextension=".xlsx",
            initialfile="Allocation_Table_템플릿.xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )

        if not file_path:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Allocation Table"

            # === 스타일 정의 ===
            title_font = Font(bold=True, size=14, color="2C3E50")
            header_font = Font(bold=True, color="FFFFFF", size=10)
            data_font = Font(size=10)
            header_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")  # 초록 배경
            sale_ref_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # SALE REF 연한 초록
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            center = Alignment(horizontal='center', vertical='center')
            right_align = Alignment(horizontal='right', vertical='center')

            # === Row 1: 타이틀 (화주 양식) ===
            ws.merge_cells('A1:J1')
            ws['A1'] = "Allocation - PT LBM - September / CIF Semarang - 300MT of MIc9000"
            ws['A1'].font = title_font
            ws.row_dimensions[1].height = 30

            # === Row 2: 합계 QTY (E열) ===
            ws.row_dimensions[2].height = 20
            ws.cell(row=2, column=5, value=300.06)
            ws['E2'].number_format = '#,##0.0000'

            # === Row 3: 헤더 (화주 양식 — 10컬럼) ===
            headers = [
                ('Product', 16),
                ('SAP NO', 14),
                ('ETA BUSAN', 14),
                ('Date in stock', 14),
                ('QTY (MT)', 12),
                ('Lot No', 14),
                ('WH', 8),
                ('Customs', 12),
                ('GW', 12),
                ('SALE REF', 12),
            ]
            for col, (text, width) in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=text)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
                cell.border = thin_border
                ws.column_dimensions[get_column_letter(col)].width = width

            # === Row 4~: 샘플 데이터 (화주 양식 기준, 붙여넣기 시 참고용) ===
            sample_lots = [
                '1125052654', '1125052707', '1125052708', '1125052709', '1125052710',
                '1125052711', '1125052712', '1125052713', '1125052714', '1125052715',
                '1125052716', '1125052717', '1125052718',
            ]
            for i, lot in enumerate(sample_lots, 4):
                vals = ['MIC9000', '2200032552', '', '2025-07-29', 5, lot, 'GY', 'Cleared', 5.13, '1955']
                for col, val in enumerate(vals, 1):
                    cell = ws.cell(row=i, column=col, value=val)
                    cell.font = data_font
                    cell.border = thin_border
                    cell.fill = sale_ref_fill if col == 10 else None
                    if col in (5, 9):  # QTY (MT), GW
                        cell.number_format = '#,##0.000'
                        cell.alignment = right_align
                    else:
                        cell.alignment = center

            # === Row 2: 합계 수식 (데이터 추가 시 자동 반영) ===
            last_row = 3 + len(sample_lots)
            ws['E2'].value = f"=SUM(E4:E{last_row})"
            ws['E2'].number_format = '#,##0.0000'

            # === 작성 안내 시트 ===
            ws2 = wb.create_sheet("📋 작성 안내")
            guides = [
                ("📋 Allocation Table (화주 양식) 작성 안내", ""),
                ("", ""),
                ("구분", "설명"),
                ("Row 1", "타이틀: 'Allocation - 고객 - 기간 / 목적지 - 수량MT of 제품'"),
                ("Row 2", "합계 QTY (E열, 수식 또는 직접 입력)"),
                ("Row 3", "헤더: Product | SAP NO | ETA BUSAN | Date in stock | QTY (MT) | Lot No | WH | Customs | GW | SALE REF"),
                ("Row 4~", "데이터 행 — 이 템플릿에 붙여넣기하거나, 이 양식으로 파일 작성 후 업로드"),
                ("", ""),
                ("★ 사용 방법", ""),
                ("방법 1", "템플릿 다운로드 → 4행부터 데이터 붙여넣기 → 저장 → 업로드"),
                ("방법 2", "화주에서 받은 Allocation Excel을 그대로 업로드"),
                ("", ""),
                ("★ 필수 항목", "Product, QTY (MT), Lot No"),
            ]
            for r, (a, b) in enumerate(guides, 1):
                ws2.cell(row=r, column=1, value=a)
                ws2.cell(row=r, column=2, value=b)
                if r == 1:
                    ws2.cell(row=1, column=1).font = Font(bold=True, size=13)
                elif r == 3:
                    ws2.cell(row=r, column=1).font = Font(bold=True)
                    ws2.cell(row=r, column=2).font = Font(bold=True)
            ws2.column_dimensions['A'].width = 28
            ws2.column_dimensions['B'].width = 70

            try:
                from gui_app_modular.utils.report_footer import add_gy_logistics_footer
                add_gy_logistics_footer(ws)
            except (ImportError, ModuleNotFoundError) as _e:
                logger.debug(f'Suppressed: {_e}')
            wb.save(file_path)

            self._log(f"✅ Allocation Table 템플릿 저장: {file_path}")
            CustomMessageBox.showinfo(self.root, "완료",
                f"Allocation Table 템플릿(화주 양식)이 저장되었습니다.\n\n"
                f"파일: {file_path}\n\n"
                "★ 4행부터 데이터 붙여넣기 후 저장하거나,\n"
                "★ 화주에서 받은 동일 양식 Excel을 그대로 업로드하세요.")
                
        except ImportError:
            CustomMessageBox.showerror(self.root, "오류", "openpyxl 패키지가 필요합니다.\npip install openpyxl")
        except (RuntimeError, ValueError) as e:
            logger.error(f"Allocation Table 템플릿 생성 오류: {e}")
            CustomMessageBox.show_detailed_error(self.root, "오류", "Allocation Table 생성 실패", exception=e)

    def _generate_virtual_allocation(self) -> None:
        """v3.9.3: 가상 출고 Allocation Table 생성
        
        현재 DB 재고 기준:
        - 60% → 출고 (분할/반송, 입고일~2026-02-08 사이 랜덤 출고)
        - 20% → 반품 (반송)
        - 20% → 재고 유지
        - 일반 + 샘플 톤백 모두 포함
        """
        from ..utils.constants import filedialog
        
        file_path = filedialog.asksaveasfilename(
            title="가상 출고 Allocation Table 저장",
            defaultextension=".xlsx",
            initialfile="출고_Allocation_Table_가상.xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not file_path:
            return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            import random
            from datetime import datetime, timedelta
            
            # DB에서 재고 로드
            all_lots = self.engine.get_inventory()
            if not all_lots:
                CustomMessageBox.showwarning(self.root, "데이터 없음",
                    "DB에 재고 데이터가 없습니다.\n먼저 입고를 진행해주세요.")
                return
            
            random.seed(42)
            random.shuffle(all_lots)
            
            total = len(all_lots)
            n_out = int(total * 0.6)
            n_ret = int(total * 0.2)
            
            out_lots = sorted(all_lots[:n_out], key=lambda x: x.get('lot_no', ''))
            ret_lots = sorted(all_lots[n_out:n_out+n_ret], key=lambda x: x.get('lot_no', ''))
            stk_lots = sorted(all_lots[n_out+n_ret:], key=lambda x: x.get('lot_no', ''))
            
            # 스타일
            hdr_font = Font(bold=True, color="FFFFFF", size=10)
            hdr_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            out_fill = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
            ret_fill = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")
            smp_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
            stk_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            sum_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
            ctr = Alignment(horizontal='center', vertical='center')
            rgt = Alignment(horizontal='right', vertical='center')
            
            sold_tos = ['LBM AP - Q4 2025', 'PT ABC - Q1 2026', 'Samsung SDI', 'LG Energy', 'CATL']
            
            def rand_date(start_str):
                try:
                    s = datetime.strptime(str(start_str)[:10], '%Y-%m-%d')
                except (ValueError, TypeError, KeyError):
                    s = datetime(2025, 9, 1)
                e = datetime(2026, 2, 8)
                d = max((e - s).days, 7)
                return (s + timedelta(days=random.randint(7, d))).strftime('%Y-%m-%d')
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Allocation Table"
            
            # 타이틀
            ws.merge_cells('A1:K1')
            ws['A1'] = f"Allocation - {total} LOTs (가상 60/20/20)"
            ws['A1'].font = Font(bold=True, size=14)
            
            # 헤더
            headers = ['Product','SAP NO','Date in stock','QTY (MT)','Lot No',
                       'WH','Customs','Export','SOLD TO','SALE REF','GW']
            widths = [16,14,14,12,14,8,12,12,28,14,12]
            for i, (h, w) in enumerate(zip(headers, widths), 1):
                c = ws.cell(row=3, column=i, value=h)
                c.font = hdr_font; c.fill = hdr_fill; c.alignment = ctr; c.border = border
                ws.column_dimensions[get_column_letter(i)].width = w
            
            row = 4
            out_qty = out_gw = ret_qty = ret_gw = 0
            
            def write_row(r, vals, fill, is_sample=False):
                for i, v in enumerate(vals, 1):
                    c = ws.cell(row=r, column=i, value=v)
                    c.font = Font(size=10, color="0066CC") if is_sample else Font(size=10)
                    c.fill = smp_fill if is_sample else fill
                    c.border = border
                    c.alignment = rgt if i in (4, 11) else ctr
                    if i in (4, 11):
                        c.number_format = '#,##0.00000' if is_sample else '#,##0.000'
            
            # 출고 (60%)
            for lot in out_lots:
                qty_mt = (lot.get('net_weight', 5000) or 5000) / 1000
                gw_mt = qty_mt * 1.026
                write_row(row, [lot.get('product',''), lot.get('sap_no',''),
                    lot.get('arrival_date',''), qty_mt, lot.get('lot_no',''),
                    lot.get('warehouse','GY'), lot.get('customs','Cleared'), '분할/반송',
                    random.choice(sold_tos), str(2900+random.randint(1,99)), gw_mt], out_fill)
                row += 1; out_qty += qty_mt; out_gw += gw_mt
            
            for lot in out_lots:
                write_row(row, [f"{lot.get('product','')}_sample", lot.get('sap_no',''),
                    lot.get('arrival_date',''), 0.001, lot.get('lot_no',''),
                    lot.get('warehouse','GY'), lot.get('customs','Cleared'), '분할/반송',
                    random.choice(sold_tos), '', 0.00125], out_fill, is_sample=True)
                row += 1; out_qty += 0.001; out_gw += 0.00125
            
            # 반품 (20%)
            for lot in ret_lots:
                qty_mt = (lot.get('net_weight', 5000) or 5000) / 1000
                gw_mt = qty_mt * 1.026
                write_row(row, [lot.get('product',''), lot.get('sap_no',''),
                    lot.get('arrival_date',''), qty_mt, lot.get('lot_no',''),
                    lot.get('warehouse','GY'), 'Uncleared', '반송',
                    'RETURN - 반품', '', gw_mt], ret_fill)
                row += 1; ret_qty += qty_mt; ret_gw += gw_mt
            
            for lot in ret_lots:
                write_row(row, [f"{lot.get('product','')}_sample", lot.get('sap_no',''),
                    lot.get('arrival_date',''), 0.001, lot.get('lot_no',''),
                    lot.get('warehouse','GY'), 'Uncleared', '반송',
                    'RETURN - 반품', '', 0.00125], ret_fill, is_sample=True)
                row += 1; ret_qty += 0.001; ret_gw += 0.00125
            
            last = row - 1
            ws.cell(row=2, column=3, value="합계 QTY").font = Font(bold=True)
            ws.cell(row=2, column=4, value=f"=SUM(D4:D{last})").font = Font(bold=True)
            ws['D2'].number_format = '#,##0.000'
            ws.cell(row=2, column=10, value="합계 GW").font = Font(bold=True)
            ws.cell(row=2, column=11, value=f"=SUM(K4:K{last})").font = Font(bold=True)
            ws['K2'].number_format = '#,##0.000'
            
            # 요약
            sc = 14
            for i, h in enumerate(['Export','LOTs','합계 QTY(MT)','합계 GW']):
                c = ws.cell(row=3, column=sc+i, value=h)
                c.font = hdr_font; c.fill = hdr_fill; c.alignment = ctr; c.border = border
                ws.column_dimensions[get_column_letter(sc+i)].width = 16
            
            stk_qty = sum((l.get('net_weight', 5000) or 5000)/1000 for l in stk_lots)
            for r, d in enumerate([
                ['분할/반송 (출고)', f'{len(out_lots)}', out_qty, out_gw],
                ['반송 (반품)', f'{len(ret_lots)}', ret_qty, ret_gw],
                ['재고 유지', f'{len(stk_lots)}', stk_qty, stk_qty*1.026],
                ['총합계', f'{total}', out_qty+ret_qty+stk_qty, out_gw+ret_gw+stk_qty*1.026],
            ], 4):
                for c, v in enumerate(d):
                    cell = ws.cell(row=r, column=sc+c, value=v)
                    cell.border = border; cell.alignment = rgt if c >= 2 else ctr
                    if c >= 2: cell.number_format = '#,##0.000'
                    if '총합계' in str(d[0]):
                        cell.font = Font(bold=True); cell.fill = sum_fill
            
            # 재고 유지 시트
            ws2 = wb.create_sheet("재고 유지 (20%)")
            ws2.cell(row=1, column=1, value="재고 유지 LOT (미출고)").font = Font(bold=True, size=13)
            for i, h in enumerate(['No.','Product','SAP NO','Lot No','QTY(MT)','WH','STATUS'], 1):
                c = ws2.cell(row=3, column=i, value=h)
                c.font = hdr_font; c.fill = hdr_fill; c.alignment = ctr; c.border = border
                ws2.column_dimensions[get_column_letter(i)].width = 16
            for idx, lot in enumerate(stk_lots, 1):
                qty_mt = (lot.get('net_weight', 5000) or 5000)/1000
                for j, v in enumerate([idx, lot.get('product',''), lot.get('sap_no',''),
                    lot.get('lot_no',''), qty_mt, lot.get('warehouse','GY'), 'AVAILABLE'], 1):
                    c = ws2.cell(row=3+idx, column=j, value=v)
                    c.fill = stk_fill; c.border = border; c.alignment = ctr
                    if j == 5: c.number_format = '#,##0.000'
            
            wb.save(file_path)
            self._log(f"✅ 가상 Allocation Table 저장: {file_path}")
            self._log(f"  출고: {len(out_lots)} LOTs | 반품: {len(ret_lots)} LOTs | 재고: {len(stk_lots)} LOTs")
            CustomMessageBox.showinfo(self.root, "완료",
                f"가상 Allocation Table 생성 완료\n\n"
                f"출고 (60%): {len(out_lots)} LOTs\n"
                f"반품 (20%): {len(ret_lots)} LOTs\n"
                f"재고 유지 (20%): {len(stk_lots)} LOTs\n\n"
                f"파일: {file_path}")
            
        except ImportError:
            CustomMessageBox.showerror(self.root, "오류", "openpyxl 필요: pip install openpyxl")
        except (RuntimeError, ValueError) as e:
            logger.error(f"가상 Allocation 생성 오류: {e}", exc_info=True)
            CustomMessageBox.show_detailed_error(self.root, "오류", "생성 실패", exception=e)

    def _generate_allocation_samples(self) -> None:
        """화주 양식(PT LBM / CN Semarang) Allocation 샘플 Excel 3개 생성"""
        import subprocess
        import sys
        from pathlib import Path

        project_root = Path(__file__).resolve().parent.parent.parent
        script_path = project_root / "scripts" / "generate_allocation_from_tonbag.py"
        out_dir = project_root / "generated_allocation"

        if not script_path.exists():
            CustomMessageBox.showerror(self.root, "오류", f"스크립트 없음: {script_path}")
            return

        try:
            result = subprocess.run(
                [sys.executable, str(script_path)],
                cwd=str(project_root),
                capture_output=True,
                text=True,
                encoding='utf-8',
                errors='replace',
                timeout=60,
            )
            out = (result.stdout or '') + (result.stderr or '')
            if result.returncode == 0:
                self._log("✅ Allocation 샘플 3개 생성 완료")
                CustomMessageBox.showinfo(
                    self.root, "완료",
                    f"Allocation Table 샘플 3개 생성 완료\n\n"
                    f"출력 폴더:\n{out_dir}\n\n"
                    f"파일: Allocation_샘플_1.xlsx, 2.xlsx, 3.xlsx"
                )
            else:
                logger.warning(f"Allocation 샘플 생성 비정상 종료: {result.returncode}\n{out}")
                CustomMessageBox.showwarning(
                    self.root, "경고",
                    f"생성 중 오류 발생 (코드 {result.returncode})\n\n{out[:500]}"
                )
        except subprocess.TimeoutExpired:
            CustomMessageBox.showerror(self.root, "오류", "생성 시간 초과(60초)")
        except Exception as e:
            logger.error(f"Allocation 샘플 생성 오류: {e}", exc_info=True)
            CustomMessageBox.show_detailed_error(self.root, "오류", "Allocation 샘플 생성 실패", exception=e)

    def _load_allocation_sample(self, sample_num: int) -> None:
        """Allocation 샘플 파일 불러오기 → Allocation 출고 예약 다이얼로그에 로드"""
        from pathlib import Path

        project_root = Path(__file__).resolve().parent.parent.parent
        out_dir = project_root / "generated_allocation"
        fname = f"Allocation_샘플_{sample_num}.xlsx"
        file_path = out_dir / fname

        if not file_path.exists():
            ok = CustomMessageBox.askyesno(
                self.root, "샘플 없음",
                f"샘플 파일이 없습니다.\n{fname}\n\n"
                "먼저 샘플 3개를 생성할까요?"
            )
            if ok:
                self._generate_allocation_samples()
                if not file_path.exists():
                    return
            else:
                return

        try:
            from ..dialogs.allocation_dialog import AllocationDialog
            dlg = AllocationDialog(self, self.engine)
            dlg.show(initial_file=str(file_path))
        except (ImportError, AttributeError) as e:
            logger.error(f"Allocation 다이얼로그 오류: {e}", exc_info=True)
            CustomMessageBox.showerror(
                self.root, "오류",
                f"Allocation 다이얼로그를 열 수 없습니다:\n{e}"
            )
