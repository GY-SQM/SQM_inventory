# -*- coding: utf-8 -*-
"""
SQM 재고관리 - Excel 입고 처리 핸들러
=====================================

v2.9.91 - gui_app.py에서 분리

Excel 파일 입고 처리, 컬럼 자동 인식, 데이터 변환
"""

import logging
from ..utils.ui_constants import CustomMessageBox
from typing import Dict

logger = logging.getLogger(__name__)


class ImportHandlersMixin:
    """
    Excel 입고 처리 Mixin
    
    SQMInventoryApp 클래스에 mix-in 됩니다.
    """
    
    def _bulk_import_inventory_simple(self, file_path: str = None) -> None:
        """
        간단한 Excel 입고 처리 (파일 선택 다이얼로그 포함)
        
        Args:
            file_path: Excel 파일 경로 (None이면 파일 선택 다이얼로그 표시)
        """
        from ..utils.constants import filedialog
        
        # 파일 경로가 없으면 다이얼로그로 선택
        if file_path is None:
            file_path = filedialog.askopenfilename(
                title="입고 Excel 파일 선택",
                filetypes=[
                    ("Excel files", "*.xlsx *.xls"),
                    ("All files", "*.*")
                ]
            )
        
        if not file_path:
            return  # 취소됨
        
        # 파일 타입 자동 감지
        file_type = self._detect_excel_type(file_path)
        
        if file_type == 'inbound':
            self._import_inbound_excel_auto(file_path)
        elif file_type == 'outbound':
            self._import_outbound_excel_auto(file_path)
        elif file_type == 'location':
            # 위치 업데이트는 status_import_handlers에서 처리
            if hasattr(self, '_import_location_excel'):
                self._import_location_excel(file_path)
            else:
                CustomMessageBox.showinfo(self.root, "안내", 
                    "위치 업데이트 Excel 파일입니다.\n메뉴 > 업로드 > 위치 업데이트를 사용하세요.")
        else:
            # 타입 불명 - 사용자에게 선택 요청
            if CustomMessageBox.askyesno(self.root, "파일 타입 확인",
                f"파일 타입을 자동 감지할 수 없습니다.\n\n"
                f"파일: {file_path}\n\n"
                "입고 Excel로 처리할까요?\n"
                "(아니오를 선택하면 출고로 처리합니다)"):
                self._import_inbound_excel_auto(file_path)
            else:
                self._import_outbound_excel_auto(file_path)
    
    def _detect_excel_type(self, file_path: str) -> str:
        """
        Excel 파일 타입 자동 감지
        
        Returns:
            'inbound' | 'outbound' | 'location' | 'unknown'
        """
        from ..utils.constants import pd, HAS_PANDAS
        
        if not HAS_PANDAS:
            return 'unknown'
        
        try:
            df = pd.read_excel(file_path, nrows=10)
            columns_lower = [str(c).lower() for c in df.columns]
            
            # 입고 Excel 특징: lot_no, sap_no, product, weight
            inbound_keywords = ['lot', 'sap', 'product', 'weight', 'qty']
            inbound_score = sum(1 for kw in inbound_keywords 
                               if any(kw in c for c in columns_lower))
            
            # 출고 Excel 특징: outbound, customer, destination
            outbound_keywords = ['outbound', 'customer', 'destination', 'ship', 'deliver']
            outbound_score = sum(1 for kw in outbound_keywords 
                                if any(kw in c for c in columns_lower))
            
            # 위치 Excel 특징: location, zone, rack
            location_keywords = ['location', 'zone', 'rack', 'position', 'area']
            location_score = sum(1 for kw in location_keywords 
                                if any(kw in c for c in columns_lower))
            
            scores = {
                'inbound': inbound_score,
                'outbound': outbound_score,
                'location': location_score
            }
            
            max_type = max(scores, key=scores.get)
            if scores[max_type] >= 2:
                return max_type
            
            return 'unknown'
            
        except (FileNotFoundError, OSError, PermissionError) as e:
            logger.error(f"Excel 타입 감지 오류: {e}")
            return 'unknown'
    
    def _import_inbound_excel_auto(self, file_path: str) -> None:
        """
        입고 Excel 자동 처리
        
        컬럼 자동 인식 + 데이터 변환 + DB 저장
        """
        from ..utils.constants import pd, HAS_PANDAS, HAS_COLUMN_ALIASES, ColumnMapper
        from ..utils.safe_utils import safe_str, safe_float, safe_date
        
        if not HAS_PANDAS:
            CustomMessageBox.showerror(self.root, "오류", "pandas가 설치되지 않았습니다.")
            return
        
        try:
            df = pd.read_excel(file_path)
            
            if df.empty:
                CustomMessageBox.showwarning(self.root, "경고", "빈 Excel 파일입니다.")
                return
            
            self._log(f"📥 입고 Excel 로드: {len(df)}행")
            
            # 컬럼 매핑 (Column Alias 시스템 사용)
            if HAS_COLUMN_ALIASES and ColumnMapper:
                mapper = ColumnMapper()
                col_map = {}
                for col in df.columns:
                    std_key = mapper.get_standard_key(str(col))
                    if std_key:
                        col_map[std_key] = col
            else:
                # 기본 매핑
                col_map = self._get_basic_column_mapping(df.columns)
            
            self._log(f"   컬럼 매핑: {col_map}")
            
            # 필수 컬럼 확인 → 누락 시 매핑 다이얼로그
            required = ['lot_no']
            missing = [r for r in required if r not in col_map]
            if missing:
                try:
                    from ..dialogs.column_mapper_dialog import ColumnMapperDialog
                    sample_data = [list(df.iloc[i].values) for i in range(min(3, len(df)))]
                    mapper_dlg = ColumnMapperDialog(
                        self.root, list(df.columns), sample_data)
                    manual_map = mapper_dlg.get_result()
                    if manual_map:
                        col_map = manual_map
                        self._log(f"   수동 매핑: {col_map}")
                    else:
                        return  # 사용자 취소
                except (ImportError, ModuleNotFoundError) as e:
                    logger.debug(f"매핑 다이얼로그 오류: {e}")
                    CustomMessageBox.showerror(self.root, "오류", f"필수 컬럼 누락: {missing}")
                    return
            
            # 데이터 변환 및 저장
            added_lots = 0
            added_tonbags = 0
            skipped = 0
            errors = []
            
            for idx, row in df.iterrows():
                try:
                    lot_no = safe_str(row.get(col_map.get('lot_no', 'lot_no')))
                    if not lot_no:
                        skipped += 1
                        continue
                    
                    # 데이터 준비 (v3.8.8: 18열 전체 매핑)
                    # 중량: net_weight 우선, 없으면 qty_mt * 1000
                    _net_raw = row.get(col_map.get('net_weight', ''), None)
                    if _net_raw is not None and safe_float(_net_raw) > 0:
                        _net_kg = safe_float(_net_raw)
                    else:
                        _net_kg = safe_float(row.get(col_map.get('qty_mt', 'weight'), 0)) * 1000
                    
                    _gross_raw = row.get(col_map.get('gross_weight', ''), None)
                    _gross_kg = safe_float(_gross_raw) if _gross_raw is not None and safe_float(_gross_raw) > 0 else _net_kg
                    
                    data = {
                        'lot_no': lot_no,
                        'sap_no': safe_str(row.get(col_map.get('sap_no', 'sap_no'), '')),
                        'bl_no': safe_str(row.get(col_map.get('bl_no', 'bl_no'), '')),
                        'container_no': safe_str(row.get(col_map.get('container_no', 'container'), '')),
                        'product': safe_str(row.get(col_map.get('product', 'product'), '')),
                        'product_code': safe_str(row.get(col_map.get('product_code', ''), '')),
                        'lot_sqm': safe_str(row.get(col_map.get('lot_sqm', ''), '')),
                        'mxbg_pallet': int(safe_float(row.get(col_map.get('mxbg_pallet', 'mxbg_pallet'), 10))),
                        'net_weight': _net_kg,
                        'gross_weight': _gross_kg,
                        'initial_weight': _net_kg,
                        'current_weight': _net_kg,
                        'salar_invoice_no': safe_str(row.get(col_map.get('salar_invoice_no', ''), '')),
                        'ship_date': safe_date(row.get(col_map.get('ship_date', ''), '')),
                        'arrival_date': safe_date(row.get(col_map.get('arrival_date', 'arrival_date'), '')),
                        'free_time': int(safe_float(row.get(col_map.get('free_time', ''), 0))),
                        'warehouse': safe_str(row.get(col_map.get('warehouse', 'warehouse'), '광양')),
                        'stock_date': safe_date(row.get(col_map.get('stock_date', 'stock_date'), '')),
                        'location': safe_str(row.get(col_map.get('location', ''), '')),
                        'remark': safe_str(row.get(col_map.get('remark', ''), '')),
                        'status': 'AVAILABLE',
                    }
                    
                    # DB 저장
                    result = self.engine.add_inventory_from_dict(data)
                    if result.get('success'):
                        added_lots += 1
                        added_tonbags += result.get('tonbags', 0)
                    else:
                        errors.append(f"행 {idx+2}: {result.get('message', '알 수 없는 오류')}")
                    
                except (ValueError, TypeError, AttributeError) as e:
                    errors.append(f"행 {idx+2}: {str(e)}")
            
            # 결과 보고
            self._log(f"✅ 입고 완료: {added_lots}개 LOT, {added_tonbags}개 톤백")
            
            # v3.8.4: 처리 완료 파일 아카이브
            if added_lots > 0 and hasattr(self, '_archive_processed_file'):
                self._archive_processed_file(file_path, 'inbound')
            
            if errors:
                error_msg = '\n'.join(errors[:10])
                if len(errors) > 10:
                    error_msg += f"\n... 외 {len(errors)-10}개 오류"
                CustomMessageBox.showwarning(self.root, "일부 오류", 
                    f"입고 완료: {added_lots}개 LOT\n\n오류:\n{error_msg}")
            else:
                # 완료 알림
                try:
                    import winsound
                    winsound.MessageBeep(winsound.MB_ICONASTERISK)
                except (ImportError, RuntimeError) as e:
                    logger.debug(f"{type(e).__name__}: {e}")
                
                CustomMessageBox.showinfo(self.root, "✅ 입고 완료",
                    f"입고 처리가 완료되었습니다!\n\n"
                    f"📦 추가된 LOT: {added_lots:,}개\n"
                    f"📦 생성된 톤백: {added_tonbags:,}개\n"
                    f"⏭️ 스킵: {skipped:,}개")
            
            # UI 새로고침
            self._refresh_inventory()
            self._refresh_tonbag()
            
        except (RuntimeError, ValueError) as e:
            logger.error(f"입고 Excel 처리 오류: {e}")
            CustomMessageBox.showerror(self.root, "오류", f"입고 처리 오류: {e}")
    
    def _get_basic_column_mapping(self, columns) -> Dict[str, str]:
        """기본 컬럼 매핑 (Column Alias 없을 때)"""
        col_map = {}
        columns_lower = {str(c).lower().strip(): c for c in columns}
        
        mappings = {
            'lot_no': ['lot_no', 'lot no', 'lot', 'lotno'],
            'sap_no': ['sap_no', 'sap no', 'sap', 'sapno'],
            'bl_no': ['bl_no', 'bl no', 'bl', 'b/l', 'blno'],
            'product': ['product', 'material', '제품', '품목'],
            'container_no': ['container', 'container_no', 'cntr', '컨테이너'],
            'qty_mt': ['qty', 'weight', 'qty_mt', 'qty(mt)', '중량', '수량'],
            'arrival_date': ['arrival_date', 'arrival', '입항일'],
            'stock_date': ['stock_date', 'inbound_date', '입고일'],
            'warehouse': ['warehouse', 'wh', '창고'],
        }
        
        for std_key, candidates in mappings.items():
            for candidate in candidates:
                if candidate in columns_lower:
                    col_map[std_key] = columns_lower[candidate]
                    break
        
        return col_map
    
    def _import_outbound_excel_auto(self, file_path: str) -> None:
        """출고 Excel 자동 처리"""
        from ..utils.constants import pd, HAS_PANDAS
        from ..utils.safe_utils import safe_str
        
        if not HAS_PANDAS:
            CustomMessageBox.showerror(self.root, "오류", "pandas가 설치되지 않았습니다.")
            return
        
        try:
            df = pd.read_excel(file_path)
            
            if df.empty:
                CustomMessageBox.showwarning(self.root, "경고", "빈 Excel 파일입니다.")
                return
            
            self._log(f"📤 출고 Excel 로드: {len(df)}행")
            
            # 컬럼 매핑
            col_map = self._get_basic_column_mapping(df.columns)
            
            # 출고 처리
            processed = 0
            errors = []
            
            for idx, row in df.iterrows():
                try:
                    lot_no = safe_str(row.get(col_map.get('lot_no', 'lot_no')))
                    if not lot_no:
                        continue
                    
                    destination = safe_str(row.get('destination', row.get('customer', '')))

                    # v5.7.6: process_outbound(allocation_data) 시그니처 통일 — Excel에 무게 없으면 LOT 전량 출고
                    lot_row = self.engine.db.fetchone(
                        "SELECT current_weight FROM inventory WHERE lot_no = ?", (lot_no,)
                    )
                    if not lot_row:
                        errors.append(f"행 {idx+2}: LOT 없음 — {lot_no}")
                        continue
                    weight_kg = float(lot_row.get('current_weight') or 0)
                    if weight_kg <= 0:
                        errors.append(f"행 {idx+2}: 가용 재고 0 — {lot_no}")
                        continue
                    allocation_data = [{
                        'lot_no': lot_no,
                        'weight_kg': weight_kg,
                        'customer': destination,
                    }]
                    result = self.engine.process_outbound(allocation_data)
                    if result.get('success'):
                        processed += 1
                    else:
                        errors.append(f"행 {idx+2}: {result.get('message')}")
                    
                except (ValueError, TypeError, AttributeError) as e:
                    errors.append(f"행 {idx+2}: {str(e)}")
            
            # 결과 보고
            self._log(f"✅ 출고 완료: {processed}건")
            
            if errors:
                error_msg = '\n'.join(errors[:10])
                CustomMessageBox.showwarning(self.root, "일부 오류", 
                    f"출고 완료: {processed}건\n\n오류:\n{error_msg}")
            else:
                CustomMessageBox.showinfo(self.root, "✅ 출고 완료", f"출고 처리 완료: {processed}건")
            
            # UI 새로고침
            self._refresh_inventory()
            self._refresh_tonbag()
            
        except (RuntimeError, ValueError) as e:
            logger.error(f"출고 Excel 처리 오류: {e}")
            CustomMessageBox.showerror(self.root, "오류", f"출고 처리 오류: {e}")

    # ═══════════════════════════════════════════════════════
    # v3.8.4: 입고 샘플 Excel 템플릿 다운로드
    # ═══════════════════════════════════════════════════════
    
    def _download_inbound_template(self) -> None:
        """입고용 샘플 Excel 템플릿 생성 및 다운로드"""
        from ..utils.constants import filedialog
        
        file_path = filedialog.asksaveasfilename(
            title="샘플 Excel 템플릿 저장",
            defaultextension=".xlsx",
            initialfile="입고_샘플_템플릿.xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        if not file_path:
            return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "입고 데이터"
            
            # 헤더 정의 (v3.8.8: 18열 통일, 필수/선택 구분)
            headers = [
                ('lot_no',            'LOT NO *',         '필수', 18),
                ('sap_no',            'SAP NO',           '선택', 15),
                ('bl_no',             'BL NO',            '선택', 18),
                ('container_no',      'CONTAINER',        '선택', 16),
                ('product',           'PRODUCT *',        '필수', 15),
                ('product_code',      'CODE',             '선택', 10),
                ('lot_sqm',           'LOT SQM',          '선택', 12),
                ('mxbg_pallet',       'MXBG *',           '필수', 8),
                ('net_weight',        'NET(Kg) *',        '필수', 12),
                ('gross_weight',      'GROSS(Kg)',        '선택', 12),
                ('salar_invoice_no',  'INVOICE NO',       '선택', 18),
                ('ship_date',         'SHIP DATE',        '선택', 13),
                ('arrival_date',      'ARRIVAL',          '선택', 13),
                ('free_time',         'FREE TIME',        '선택', 10),
                ('warehouse',         'WH',               '선택', 8),
                ('stock_date',        'STOCK DATE',       '선택', 13),
                ('location',          'LOCATION',         '선택', 10),
                ('remark',            'REMARK',           '선택', 20),
            ]
            
            # 스타일
            header_font = Font(bold=True, color="FFFFFF", size=11)
            required_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            optional_fill = PatternFill(start_color="7F8C8D", end_color="7F8C8D", fill_type="solid")
            sample_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # 1행: 설명
            ws.merge_cells('A1:R1')
            ws['A1'] = "📥 SQM v3.8.8 입고 데이터 템플릿 — * 표시는 필수 항목입니다 (18열)"
            ws['A1'].font = Font(bold=True, size=12, color="2C3E50")
            ws.row_dimensions[1].height = 30
            
            # 2행: DB 필드명 (숨겨진 매핑 키)
            for col, (db_field, _, _, _) in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col, value=db_field)
                cell.font = Font(size=8, color="999999")
            ws.row_dimensions[2].height = 15
            
            # 3행: 표시 헤더
            for col, (_, display, req, width) in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=display)
                cell.font = header_font
                cell.fill = required_fill if req == '필수' else optional_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
                col_letter = openpyxl.utils.get_column_letter(col)
                ws.column_dimensions[col_letter].width = width
            
            # 4~5행: 샘플 데이터 (18열)
            samples = [
                ['SQM20250001', '4500012345', 'MAEU2500001', 'FFAU5000001-1',
                 'MIC9000', 'LC-B', 'SQM-001', 10, 20000, 20500,
                 'FA-2025-001', '2025-09-06', '2025-10-17', 25, '광양',
                 '2025-10-20', 'A-01', '1차 입고'],
                ['SQM20250002', '4500012346', 'MAEU2500002', 'TCLU3000001-1',
                 'NICKEL', 'NI-A', 'SQM-002', 8, 15000, 15300,
                 'FA-2025-002', '2025-09-06', '2025-10-17', 25, '광양',
                 '2025-10-22', 'B-02', ''],
            ]
            
            for row_idx, sample in enumerate(samples, 4):
                for col, val in enumerate(sample, 1):
                    cell = ws.cell(row=row_idx, column=col, value=val)
                    cell.fill = sample_fill
                    cell.border = thin_border
            
            # 안내 시트 (v3.8.8: 18열 가이드)
            ws2 = wb.create_sheet("안내")
            guides = [
                ("필드", "설명", "형식", "예시"),
                ("lot_no", "LOT 번호 (필수, 고유값)", "텍스트", "SQM20250001"),
                ("sap_no", "SAP 주문번호", "텍스트", "4500012345"),
                ("bl_no", "선하증권 번호", "텍스트", "MAEU2500001"),
                ("container_no", "컨테이너 번호", "텍스트", "FFAU5000001-1"),
                ("product", "제품명 (필수)", "텍스트", "MIC9000"),
                ("product_code", "제품 코드", "텍스트", "LC-B"),
                ("lot_sqm", "LOT SQM 번호", "텍스트", "SQM-001"),
                ("mxbg_pallet", "맥시백(톤백) 수 (필수)", "정수", "10"),
                ("net_weight", "순중량 kg (필수)", "숫자", "20000"),
                ("gross_weight", "총중량 kg", "숫자", "20500"),
                ("salar_invoice_no", "인보이스 번호", "텍스트", "FA-2025-001"),
                ("ship_date", "선적일", "날짜", "2025-09-06"),
                ("arrival_date", "입항일", "날짜", "2025-10-17"),
                ("free_time", "프리타임 (일)", "정수", "25"),
                ("warehouse", "창고", "텍스트", "광양"),
                ("stock_date", "입고일", "날짜", "2025-10-20"),
                ("location", "적치 위치", "텍스트", "A-01"),
                ("remark", "비고", "텍스트", ""),
            ]
            for row_idx, (a, b, c, d) in enumerate(guides, 1):
                ws2.cell(row=row_idx, column=1, value=a)
                ws2.cell(row=row_idx, column=2, value=b)
                ws2.cell(row=row_idx, column=3, value=c)
                ws2.cell(row=row_idx, column=4, value=d)
                if row_idx == 1:
                    for col in range(1, 5):
                        ws2.cell(row=1, column=col).font = Font(bold=True)
            
            ws2.column_dimensions['A'].width = 18
            ws2.column_dimensions['B'].width = 30
            ws2.column_dimensions['C'].width = 10
            ws2.column_dimensions['D'].width = 20
            
            try:
                from gui_app_modular.utils.report_footer import add_gy_logistics_footer
                add_gy_logistics_footer(ws)
            except (ImportError, ModuleNotFoundError) as _e:
                logger.debug(f'Suppressed: {_e}')
            wb.save(file_path)
            
            self._log(f"✅ 샘플 템플릿 저장: {file_path}")
            CustomMessageBox.showinfo(self.root, "완료",
                f"샘플 Excel 템플릿이 저장되었습니다.\n\n"
                f"파일: {file_path}\n\n"
                "* 표시 항목은 필수입니다.\n"
                "2행의 DB 필드명은 자동 매핑에 사용됩니다.")
                
        except ImportError:
            CustomMessageBox.showerror(self.root, "오류", "openpyxl 패키지가 필요합니다.\npip install openpyxl")
        except (ImportError, ModuleNotFoundError) as e:
            logger.error(f"템플릿 생성 오류: {e}")
            CustomMessageBox.showerror(self.root, "오류", f"템플릿 생성 실패:\n{e}")

    # ═══════════════════════════════════════════════════════
    # v3.8.4 A4: Excel 자동 아카이브
    # ═══════════════════════════════════════════════════════
    
    def _archive_processed_file(self, file_path: str, category: str = 'inbound') -> str:
        """
        처리 완료된 Excel 파일을 archive 폴더로 복사
        
        Args:
            file_path: 원본 파일 경로
            category: 'inbound' | 'outbound'
        
        Returns:
            아카이브 경로
        """
        import os
        import shutil
        from datetime import datetime
        
        try:
            db_dir = os.path.dirname(getattr(self, 'db_path', '') or '')
            if not db_dir:
                db_dir = os.path.dirname(file_path)
            
            archive_dir = os.path.join(db_dir, 'archive', category)
            os.makedirs(archive_dir, exist_ok=True)
            
            base_name = os.path.basename(file_path)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            name, ext = os.path.splitext(base_name)
            archive_name = f"{name}_{timestamp}{ext}"
            archive_path = os.path.join(archive_dir, archive_name)
            
            shutil.copy2(file_path, archive_path)
            self._log(f"📁 아카이브: {archive_name}")
            return archive_path
            
        except (OSError, IOError, PermissionError) as e:
            logger.debug(f"아카이브 실패 (무시): {e}")
            return ''
