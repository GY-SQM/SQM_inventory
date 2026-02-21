# -*- coding: utf-8 -*-
"""
SQM v5.9.4 — 원스톱 입고: DB 업로드 + Excel 내보내기 Mixin
==========================================================

onestop_inbound.py에서 분리 (1869줄 → ~1300 + ~500).
DB 저장 로직(_save_to_db), 업로드 스레드(_upload_thread),
중복 체크(_on_upload), Excel 내보내기(_export_to_excel)를 담당.
"""
import sqlite3
import logging
import threading
from datetime import datetime
from tkinter import filedialog

from core.constants import DEFAULT_WAREHOUSE

logger = logging.getLogger(__name__)


class InboundUploadMixin:
    """DB 업로드 + Excel 내보내기 로직 (OneStopInboundDialog에 MRO 합성)"""

    def _on_upload(self) -> None:
        """DB 업로드 (v3.8.8: 중복 LOT 사전 경고 + 위젯 안전 처리)"""
        if not self.preview_data:
            return
        from .onestop_inbound import DOC_TYPES
        if not self._has_required_docs():
            missing = [name for (dt, name, req) in DOC_TYPES if req and dt not in self.file_paths]
            try:
                from ..utils.custom_messagebox import CustomMessageBox
                CustomMessageBox.showwarning(
                    self.dialog, "필수 서류 누락",
                    "DB 업로드를 하려면 다음 3종 서류가 모두 필요합니다:\n\n"
                    "  • ① Packing List (포장명세서)\n"
                    "  • ② Invoice, FA (송장)\n"
                    "  • ③ Bill of Loading (선하증권)\n\n"
                    f"누락: {', '.join(missing)}\n\n"
                    "Delivery Order(인도지시서)는 선택사항이며, 나중에 [📋 D/O 후속 연결] 메뉴로 보충할 수 있습니다."
                )
            except (ImportError, ModuleNotFoundError):
                from ..utils.ui_constants import CustomMessageBox
                CustomMessageBox.showwarning(
                    self.dialog, "필수 서류 누락",
                    "Packing List, Invoice/FA, Bill of Loading 3종 모두 필요합니다."
                )
            return

        dup_lots = []
        if hasattr(self.engine, '_check_lot_exists') or hasattr(self.engine, 'db'):
            try:
                db = getattr(self.engine, 'db', None)
                if db:
                    for row in self.preview_data:
                        lot_no = row.get('lot_no', '')
                        if lot_no:
                            existing = db.fetchone(
                                "SELECT 1 FROM inventory WHERE lot_no = ?", (lot_no,))
                            if existing:
                                dup_lots.append(lot_no)
            except (sqlite3.OperationalError, sqlite3.IntegrityError, OSError) as e:
                logger.debug(f"중복 체크 오류: {e}")

        if dup_lots:
            try:
                from ..utils.custom_messagebox import CustomMessageBox
                dup_msg = ', '.join(dup_lots[:5])
                if len(dup_lots) > 5:
                    dup_msg += f" 외 {len(dup_lots) - 5}건"
                ok = CustomMessageBox.askyesno(
                    self.dialog, "⚠️ 중복 LOT 경고",
                    f"다음 {len(dup_lots)}개 LOT가 이미 DB에 존재합니다:\n\n"
                    f"{dup_msg}\n\n"
                    f"중복 LOT는 건너뛰고 나머지만 입고합니다.\n계속하시겠습니까?"
                )
            except (ImportError, ModuleNotFoundError):
                from tkinter import messagebox as msgbox
                ok = msgbox.askyesno("⚠️ 중복 LOT 경고",
                    f"{len(dup_lots)}개 LOT 중복! 건너뛰고 계속?")
            if not ok:
                return

        try:
            from ..utils.custom_messagebox import CustomMessageBox
            ok = CustomMessageBox.askyesno(
                self.dialog, "DB 업로드 확인",
                f"{len(self.preview_data)}개 LOT를 데이터베이스에 저장합니다.\n\n"
                f"이 작업은 되돌릴 수 없습니다.\n계속하시겠습니까?"
            )
        except (ImportError, ModuleNotFoundError):
            from tkinter import messagebox as msgbox
            ok = msgbox.askyesno("DB 업로드 확인",
                f"{len(self.preview_data)}개 LOT 저장?")

        if not ok:
            return

        try:
            if self.btn_upload and self.btn_upload.winfo_exists():
                self.btn_upload.config(state='disabled')
            if self.btn_excel and self.btn_excel.winfo_exists():
                self.btn_excel.config(state='disabled')
        except (RuntimeError, ValueError) as _e:
            logger.debug(f'Suppressed: {_e}')

        self._show_progress_inline()
        thread = threading.Thread(target=self._upload_thread, daemon=True)
        thread.start()

    def _upload_thread(self) -> None:
        """백그라운드 DB 업로드"""
        try:
            self._update_progress(0, "📤 DB 업로드 시작...")

            pl = self.parsed_results.get('packing_list')
            invoice = self.parsed_results.get('invoice')
            bl = self.parsed_results.get('bl')
            do = self.parsed_results.get('do')

            if not pl or not getattr(pl, 'lots', None):
                self._update_progress(0, "❌ Packing List 없음")
                self._enable_buttons()
                return
            if not invoice:
                self._update_progress(0, "❌ FA(송장) 필수 — 3종(PL+FA+BL) 모두 필요")
                self._enable_buttons()
                return
            if not bl:
                self._update_progress(0, "❌ B/L(선하증권) 필수 — 3종(PL+FA+BL) 모두 필요")
                self._enable_buttons()
                return

            success, failed_rows = self._save_to_db(pl, invoice, bl, do)

            if success:
                total = len(self.preview_data)
                self._update_progress(100, f"✅ 업로드 완료: {total} LOT")
                self._log_safe(f"✅ DB 업로드 완료: {total} LOT")
                self.upload_success = True
                self._show_success_and_close(total)
            else:
                self._update_progress(0, "❌ 업로드 실패")
                try:
                    from ..utils.upload_error_dialog import show_upload_error_dialog
                    from ..utils.upload_error_template import UploadErrorTemplate
                    rows_for_msg = failed_rows if failed_rows else [{'row': '?', 'value': '업로드 실패', 'column': ''}]
                    err_type = (rows_for_msg[0].get('type', 'missing_required') if rows_for_msg else 'missing_required')
                    error_msg = UploadErrorTemplate.format_multiple_errors(
                        errors=[{'type': err_type, 'rows': rows_for_msg}],
                        total_rows=len(self.preview_data)
                    )
                    show_upload_error_dialog(self.dialog, "입고 업로드 실패", error_msg)
                except (ImportError, Exception):
                    from ..utils.ui_constants import CustomMessageBox
                    CustomMessageBox.showerror(
                        self.dialog, "업로드 실패",
                        "입고 처리 중 오류가 발생했습니다.\n로그를 확인하세요."
                    )
                self._enable_buttons()

        except (ValueError, TypeError, AttributeError) as e:
            self._update_progress(0, f"❌ 오류: {e}")
            self._log_safe(f"❌ 업로드 오류: {e}")
            logger.error(f"업로드 오류: {e}", exc_info=True)
            try:
                from ..utils.upload_error_dialog import show_upload_error_dialog
                from ..utils.upload_error_template import UploadErrorTemplate
                error_msg = UploadErrorTemplate.format_multiple_errors(
                    errors=[{'type': 'file_format', 'rows': [{'row': '?', 'value': str(e), 'column': ''}]}],
                    total_rows=len(self.preview_data) if hasattr(self, 'preview_data') else 0
                )
                show_upload_error_dialog(self.dialog, "입고 처리 오류", error_msg)
            except (ImportError, Exception):
                from ..utils.ui_constants import CustomMessageBox
                CustomMessageBox.showerror(self.dialog, "오류", f"입고 처리 오류:\n{e}")
            self._enable_buttons()

    def _save_to_db(self, pl, invoice, bl, do):
        """engine.process_inbound를 LOT별로 호출하여 DB 저장

        Returns:
            (success: bool, failed_rows: list)
        """
        try:
            if not hasattr(self.engine, 'process_inbound'):
                self._log_safe("❌ engine.process_inbound 메서드 없음")
                return False, []

            _lots = getattr(pl, 'lots', []) or []
            total = len(_lots)
            if total == 0:
                return False, []

            created_lots = []
            skipped_lots = []
            errors = []
            failed_rows = []
            _last_idx = -1  # DB 예외 시 행 번호 표시용

            for idx, lot in enumerate(_lots):
                _last_idx = idx
                pct = 10 + int(80 * (idx + 1) / total)
                lot_no = getattr(lot, 'lot_no', '') or ''
                self._update_progress(pct, f"📦 LOT {idx+1}/{total}: {lot_no}")

                if lot_no:
                    try:
                        existing = self.engine.db.fetchone(
                            "SELECT 1 FROM inventory WHERE lot_no = ?", (lot_no,))
                        if existing:
                            self._log_safe(f"  ⏭ LOT {lot_no}: 이미 존재 (건너뜀)")
                            skipped_lots.append(lot_no)
                            continue
                    except (sqlite3.OperationalError, sqlite3.IntegrityError, OSError) as _e:
                        logger.debug(f'Suppressed: {_e}')

                _tonbag = getattr(lot, 'tonbag_count', None)
                if _tonbag is None or (isinstance(_tonbag, str) and str(_tonbag).strip() == ''):
                    _tonbag = getattr(lot, 'mxbg_pallet', 10) or 10
                try:
                    _tonbag = int(float(_tonbag))
                except (TypeError, ValueError):
                    _tonbag = getattr(lot, 'mxbg_pallet', 10) or 10

                _arrival_raw = getattr(do, 'arrival_date', None) if do else None
                _arrival = str(_arrival_raw) if _arrival_raw and str(_arrival_raw) != 'None' else ''
                if _arrival:
                    _a10 = (_arrival[:10] if len(_arrival) >= 10 else _arrival)
                    if not (len(_a10) == 10 and _a10.count('-') == 2 and _a10.replace('-', '').isdigit()):
                        _arrival = ''
                _free_time = 0
                _free_time_date = ''

                if not _arrival and self.preview_data and idx < len(self.preview_data):
                    _user_arr = self.preview_data[idx].get('arrival_date', '')
                    if _user_arr:
                        _ua = str(_user_arr)[:10]
                        if len(_ua) == 10 and _ua.count('-') == 2 and _ua.replace('-', '').isdigit():
                            _arrival = _ua
                    _user_ft = self.preview_data[idx].get('free_time', '')
                    if _user_ft:
                        try:
                            _free_time = int(_user_ft)
                        except (ValueError, TypeError):
                            pass

                if do:
                    _free_time_date = str(getattr(do, 'free_time_date', '') or '')
                    if not _free_time_date:
                        ft_infos = getattr(do, 'free_time_info', []) or []
                        for ft in ft_infos:
                            ftd = getattr(ft, 'free_time_date', '') or (ft.get('free_time_date', '') if isinstance(ft, dict) else '')
                            if ftd:
                                _free_time_date = str(ftd)
                                break

                    if _free_time_date and _arrival:
                        try:
                            _ft_dt = datetime.strptime(str(_free_time_date)[:10], '%Y-%m-%d').date()
                            _arr_dt = datetime.strptime(str(_arrival)[:10], '%Y-%m-%d').date()
                            _free_time = (_ft_dt - _arr_dt).days
                            if _free_time < 0:
                                _free_time = 0
                        except (ValueError, TypeError):
                            _free_time = 0

                packing_dict = {
                    'lot_no': getattr(lot, 'lot_no', '') or '',
                    'lot_sqm': getattr(lot, 'lot_sqm', '') or '',
                    'sap_no': getattr(pl, 'sap_no', '') or (getattr(invoice, 'sap_no', '') if invoice else '') or '',
                    'bl_no': self._format_bl(
                        (getattr(bl, 'bl_no', '') if bl else '') or
                        (getattr(do, 'bl_no', '') if do else '') or ''
                    ),
                    'container_no': getattr(lot, 'container_no', '') or '',
                    'product': getattr(pl, 'product', '') or 'LITHIUM CARBONATE',
                    'product_code': getattr(pl, 'code', '') or '',
                    'net_weight': getattr(lot, 'net_weight_kg', 0) or 0,
                    'gross_weight': getattr(lot, 'gross_weight_kg', 0) or 0,
                    'mxbg_pallet': getattr(lot, 'mxbg_pallet', 10) or 10,
                    'tonbag_count': _tonbag,
                    'salar_invoice_no': getattr(invoice, 'salar_invoice_no', '') if invoice else '',
                    'ship_date': self._date_str(getattr(bl, 'ship_date', None) if bl else None) or self._date_str(getattr(invoice, 'invoice_date', None) if invoice else None) or '',
                    'arrival_date': _arrival,
                    'free_time': _free_time,
                    'free_time_date': _free_time_date,
                    'con_return': _free_time_date[:10] if _free_time_date else '',
                    'warehouse': str(getattr(do, 'warehouse', DEFAULT_WAREHOUSE)) if do else DEFAULT_WAREHOUSE,
                    'vessel': getattr(pl, 'vessel', '') or '',
                }

                missing_display = []
                if not (str(packing_dict.get('lot_no', '') or '').strip()):
                    missing_display.append('LOT NO')
                if not (str(packing_dict.get('product', '') or '').strip()):
                    missing_display.append('PRODUCT')
                try:
                    nw = packing_dict.get('net_weight', 0)
                    if nw is None or (isinstance(nw, (int, float)) and float(nw) <= 0):
                        missing_display.append('NET(Kg)')
                except (TypeError, ValueError):
                    missing_display.append('NET(Kg)')
                try:
                    mx = packing_dict.get('mxbg_pallet', 0)
                    if mx is None or (isinstance(mx, (int, float)) and int(float(mx)) <= 0):
                        missing_display.append('MXBG')
                except (TypeError, ValueError):
                    missing_display.append('MXBG')
                if missing_display:
                    display_row = idx + 2  # Excel/미리보기 1-based 행 번호
                    failed_rows.append({
                        'row': display_row, 'row_num': display_row,
                        'value': '비어 있음',
                        'column': ', '.join(missing_display),
                        'missing_columns': missing_display,
                    })
                    errors.append(f"행 {idx + 2}: {', '.join(missing_display)} 누락")
                    continue

                inv_dict = None
                if invoice:
                    inv_dict = {
                        'sap_no': getattr(invoice, 'sap_no', '') or '',
                        'salar_invoice_no': getattr(invoice, 'salar_invoice_no', '') or '',
                        'invoice_date': str(getattr(invoice, 'invoice_date', '')) if getattr(invoice, 'invoice_date', None) else '',
                    }

                bl_dict = None
                if bl:
                    bl_dict = {
                        'bl_no': self._format_bl(getattr(bl, 'bl_no', '') or ''),
                        'ship_date': self._date_str(getattr(bl, 'ship_date', None)) or self._date_str(getattr(bl, 'shipped_date', None)) or '',
                        'vessel': getattr(bl, 'vessel', '') or '',
                    }

                do_dict = None
                if do:
                    _con_return = ''
                    ft_infos = getattr(do, 'free_time_info', []) or []
                    for ft in ft_infos:
                        ftd = getattr(ft, 'free_time_date', '') or (ft.get('free_time_date', '') if isinstance(ft, dict) else '')
                        if ftd:
                            _con_return = str(ftd)[:10]
                            break
                    _do_arr = getattr(do, 'arrival_date', None)
                    _do_arrival = (_do_arr.isoformat() if hasattr(_do_arr, 'isoformat') else str(_do_arr or '')) if _do_arr and str(_do_arr) != 'None' else ''
                    do_dict = {
                        'bl_no': str(getattr(do, 'bl_no', '') or ''),
                        'arrival_date': _do_arrival,
                        'free_time_date': _con_return,
                        'free_time': str(getattr(do, 'free_time', '') or ''),
                        'warehouse': str(getattr(do, 'warehouse', '') or ''),
                    }

                try:
                    result = self.engine.process_inbound(
                        packing_data=packing_dict, invoice_data=inv_dict,
                        bl_data=bl_dict, do_data=do_dict
                    )
                    if result.get('success'):
                        created_lots.append(getattr(lot, "lot_no", ""))
                    else:
                        err_msg = result.get('message', '') or ', '.join(result.get('errors', []))
                        errors.append(f"LOT {getattr(lot, 'lot_no', '')}: {err_msg}")
                        failed_rows.append({
                            'row': idx + 2, 'row_num': idx + 2,
                            'value': err_msg, 'column': 'LOT NO',
                            'missing_columns': [],
                        })
                except (ValueError, TypeError, AttributeError) as e:
                    errors.append(f"LOT {getattr(lot, 'lot_no', '')}: {e}")
                    failed_rows.append({
                        'row': idx + 2, 'row_num': idx + 2,
                        'value': str(e), 'column': 'LOT NO',
                        'missing_columns': [],
                    })

            if errors:
                self._log_safe(f"⚠️ 일부 오류: {len(errors)}건")
                for e in errors[:5]:
                    self._log_safe(f"  - {e}")
            if skipped_lots:
                self._log_safe(f"⏭ 중복 건너뜀: {len(skipped_lots)}건")

            if created_lots:
                self._log_safe(f"✅ 저장 완료: {len(created_lots)}건")
                return True, []
            elif skipped_lots and not errors:
                # 모든 LOT가 중복 — "필수 컬럼 누락" 대신 중복 안내
                self._log_safe(f"⏭ 모든 LOT 중복 ({len(skipped_lots)}건) — 신규 LOT 없음")
                return False, [{
                    'row': '?', 'row_num': '?', 'value': f'{len(skipped_lots)}건 모두 이미 DB에 존재',
                    'column': '', 'missing_columns': [], 'type': 'all_duplicate_lot'
                }]
            else:
                self._log_safe(f"❌ 저장된 LOT 없음 (오류 {len(errors)}건, 건너뜀 {len(skipped_lots)}건)")
                return False, failed_rows

        except (sqlite3.OperationalError, sqlite3.IntegrityError, OSError) as e:
            logger.error(f"DB 저장 실패: {e}", exc_info=True)
            self._log_safe(f"❌ DB 저장 실패: {e}")
            msg = str(e)
            err_type = 'db_schema' if 'no column named' in msg.lower() or 'no such column' in msg.lower() else 'db_error'
            try:
                row_num = _last_idx + 2 if _last_idx >= 0 else '?'
            except NameError:
                row_num = '?'
            return False, [{'row': row_num, 'row_num': row_num, 'value': msg, 'column': '', 'missing_columns': [], 'type': err_type}]

    def _export_to_excel(self) -> None:
        """미리보기 데이터 Excel 내보내기"""
        from .onestop_inbound import PREVIEW_COLUMNS
        if not self.preview_data:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            save_path = filedialog.asksaveasfilename(
                parent=self.dialog, title="Excel 내보내기",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"입고미리보기_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            if not save_path:
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "입고 미리보기"

            headers = [col[1] for col in PREVIEW_COLUMNS]
            hfill = PatternFill(start_color="2c6fbb", end_color="2c6fbb", fill_type="solid")
            hfont = Font(color="FFFFFF", bold=True, size=10)
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))

            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.fill = hfill
                cell.font = hfont
                cell.alignment = Alignment(horizontal='center')
                cell.border = border

            for ri, row_data in enumerate(self.preview_data, 2):
                for ci, (col_id, _, _, _) in enumerate(PREVIEW_COLUMNS, 1):
                    cell = ws.cell(row=ri, column=ci, value=row_data.get(col_id, ''))
                    cell.border = border

            for ci, (_, h, w, _) in enumerate(PREVIEW_COLUMNS, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(w / 7, len(h) + 2)

            wb.save(save_path)
            self._log_safe(f"📥 Excel 저장: {save_path}")

        except (ValueError, TypeError, AttributeError) as e:
            self._log_safe(f"❌ Excel 오류: {e}")
