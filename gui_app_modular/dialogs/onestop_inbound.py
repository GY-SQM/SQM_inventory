"""
from ..utils.custom_messagebox import CustomMessageBox
SQM v3.8.4 — 원스톱 입고 팝업
4종 서류(PL, Invoice, BL, DO)를 한 화면에서:
  파일 선택 → 체크 표시 → 파싱 → 미리보기 → DB 업로드

작성일: 2025-02-06
"""
import sqlite3
import os
import time
import tkinter as tk
from tkinter import ttk, filedialog, BOTH, YES, X, Y, LEFT, RIGHT, BOTTOM, END, VERTICAL, HORIZONTAL
import logging
import threading
from datetime import datetime, date as _date_type

# 비즈니스 기본값
from core.constants import DEFAULT_WAREHOUSE

from ..utils.ui_constants import ThemeColors, DialogSize, center_dialog
from core.types import safe_float

# v5.8.7: DatePicker 달력 UI — gui_bootstrap 통일 (ttkbootstrap.DateEntry, 없으면 텍스트 입력 폴백)
from ..utils.gui_bootstrap import DateEntry, HAS_DATEENTRY

logger = logging.getLogger(__name__)


# 미리보기 컬럼 정의 — 업로드3: 전 컬럼 가운데 정렬
PREVIEW_COLUMNS = [
    ("no",               "NO",               50,  "center"),
    ("sap_no",           "SAP NO",          110,  "center"),
    ("bl_no",            "BL NO",           150,  "center"),
    ("container_no",     "CONTAINER",       130,  "center"),
    ("product",          "PRODUCT",         180,  "center"),
    ("product_code",     "CODE",            100,  "center"),
    ("lot_no",           "LOT NO",          110,  "center"),
    ("lot_sqm",          "LOT SQM",          80,  "center"),
    ("mxbg_pallet",      "MXBG",             70,  "center"),
    ("net_weight",       "NET(Kg)",          90,  "center"),
    ("gross_weight",     "GROSS(kg)",         90,  "center"),
    ("salar_invoice_no", "INVOICE NO",      120,  "center"),
    ("ship_date",        "SHIP DATE",        90,  "center"),
    ("arrival_date",     "ARRIVAL",          90,  "center"),
    ("con_return",       "CON RETURN",       95,  "center"),
    ("free_time",        "FREE TIME",        80,  "center"),
    ("warehouse",        "WH",              100,  "center"),
    ("status",           "STATUS",           80,  "center"),
]

# 4종 서류 정의 (v3.8.7: 동그라미 번호 순서) — v5.7.5: Invoice/FA, Bill of Loading, Delivery Order
DOC_TYPES = [
    ('PACKING_LIST', '① Packing List (포장명세서)', True),
    ('INVOICE',      '② Invoice, FA (송장)',        True),
    ('BL',           '③ Bill of Loading (선하증권)', True),
    ('DO',           '④ Delivery Order (인도지시서) (선택사항)', False),
]


from .inbound_dialog_base import InboundDialogBase

# v5.7.5: 진행률 팝업 조정 — 업로드2: 창·폰트 더 키움
PROGRESS_POPUP_WIDTH = 880
PROGRESS_POPUP_HEIGHT = 380
PROGRESS_POPUP_CLOSE_DELAY_MS = 1600


class OneStopInboundDialog(InboundDialogBase):
    """v3.8.4 원스톱 입고 팝업
    
    하나의 팝업에서:
    1. 4종 파일 선택 (각각 [파일 선택] 버튼 + ✅ 체크)
    2. [파싱 시작] → 프로그레스 바
    3. 18열 미리보기 테이블
    4. [DB 업로드] 또는 [Excel 내보내기]
    """
    
    def __init__(self, parent, engine, log_fn=None, app=None):
        self.parent = parent
        self.engine = engine
        self.app = app  # v3.8.8: 메인 앱 참조 (새로고침용)
        self._log = log_fn or (lambda msg, **kw: logger.info(msg))
        
        # 파일 경로 저장
        self.file_paths = {}  # {doc_type: file_path}
        
        # 파싱 결과
        self.parsed_results = {}
        self.preview_data = []
        
        # 업로드 결과
        self.upload_success = False
        # v5.8.9: 컨테이너 번호 접미사(-숫자) 디폴트 숨김, 필요 시 표시
        self._show_container_suffix = False
        # 파싱 결과 팝업에서 DB 업로드 선택 시, 완료 후 엑셀 내보내기 여부 질의
        self._ask_excel_after_upload = False
        
        # UI 참조
        self.dialog = None
        self.file_labels = {}
        self.check_labels = {}
        self.tree = None
        self.btn_parse = None
        self.btn_upload = None
        self.btn_excel = None
    
    def show(self, initial_files: dict = None) -> None:
        """팝업 표시. initial_files: { 'DO': 경로 } 등 드래그앤드롭/캡처 이미지 사전 지정."""
        self._initial_files = initial_files or {}
        self._create_dialog()
    
    def _attach_doc_tooltip(self, widget, text: str):
        """v3.8.9: 문서 위젯에 툴팁 추가"""
        tip = None
        def enter(e):
            nonlocal tip
            tip = tk.Toplevel(widget)
            tip.wm_overrideredirect(True)
            tip.wm_geometry(f"+{e.x_root+15}+{e.y_root+10}")
            _od = ThemeColors.is_dark_theme(getattr(self.parent, 'current_theme', 'flatly'))
            lbl = tk.Label(tip, text=text, justify='left',
                          background=ThemeColors.get('bg_card', _od), foreground=ThemeColors.get('text_primary', _od),
                          relief='solid', borderwidth=1,
                          font=('맑은 고딕', 11), padx=8, pady=6)
            lbl.pack()
        def leave(e):
            nonlocal tip
            if tip:
                tip.destroy()
                tip = None
        widget.bind('<Enter>', enter)
        widget.bind('<Leave>', leave)
    
    def _create_dialog(self) -> None:
        """원스톱 입고 팝업 생성"""
        self.dialog = tk.Toplevel(self.parent)
        self.dialog.title("📥 입고 — SQM v3.9.4")
        self.dialog.geometry(DialogSize.get_geometry(self.parent, 'large'))
        self.dialog.minsize(720, 520)
        self.dialog.transient(self.parent)
        self.dialog.grab_set()
        center_dialog(self.dialog, self.parent)
        self.dialog.resizable(True, True)
        self.dialog.protocol("WM_DELETE_WINDOW", self._on_cancel)
        
        main = ttk.Frame(self.dialog, padding=6)
        main.pack(fill=BOTH, expand=YES)
        
        # ═══════════════════════════════════════════════════════════
        # 1. 상단: 4종 서류 + 파싱 버튼 (1줄 균등 배치)
        # ═══════════════════════════════════════════════════════════
        file_frame = ttk.Frame(main)
        file_frame.pack(fill=X, pady=(0, 4))
        
        # 6열 그리드: [①PL][②INV][③BL][④DO] [파싱][힌트]
        for i in range(4):
            file_frame.columnconfigure(i, weight=1, uniform='doc')
        file_frame.columnconfigure(4, weight=0)  # 파싱 버튼
        file_frame.columnconfigure(5, weight=0)  # 힌트
        
        short_names = {
            'PACKING_LIST': '① Packing List',
            'INVOICE':      '② Invoice, FA',
            'BL':           '③ Bill of Loading',
            'DO':           '④ Delivery Order',
        }
        
        # v3.8.9: 서류별 상세 툴팁 — v5.7.5: Invoice/FA, Bill of Loading, Delivery Order
        _tooltips = {
            'PACKING_LIST': '📦 Packing List (포장명세서)\n\n• LOT번호, 제품명, 수량, 중량 정보 추출\n• 필수 서류 — 없으면 입고 불가\n• PDF 또는 Excel 파일 지원',
            'INVOICE':      '📑 Invoice, FA (송장)\n\n• SAP번호, 단가, 총금액 정보 추출\n• 필수 서류 — 없으면 SAP번호 누락\n• PDF 파일 지원',
            'BL':           '🚢 Bill of Loading (선하증권)\n\n• BL번호, 선박명, 출항일, 도착일 추출\n• 필수 서류 — 없으면 선적 정보 누락\n• PDF 파일 지원',
            'DO':           '📋 Delivery Order (인도지시서)\n\n• 인도 장소, Free Time 정보 추출\n• 선택 서류 — 없어도 입고 가능\n• PDF 파일 지원',
        }
        
        for idx, (doc_type, doc_name, required) in enumerate(DOC_TYPES):
            cell = ttk.Frame(file_frame)
            cell.grid(row=0, column=idx, sticky='ew', padx=(0, 2))
            
            # 서류명
            lbl = ttk.Label(cell, text=short_names.get(doc_type, ''),
                      font=('맑은 고딕', 14, 'bold'))
            lbl.pack(side=LEFT, padx=(2, 2))
            self._attach_doc_tooltip(lbl, _tooltips.get(doc_type, ''))
            
            # 📂 폴더선택 버튼
            _os_dark = ThemeColors.is_dark_theme(getattr(self.parent, 'current_theme', 'flatly'))
            btn_sel = tk.Button(cell, text="📂",
                                command=lambda dt=doc_type: self._select_file(dt),
                                font=('', 13), bg=ThemeColors.get('btn_neutral', _os_dark), fg=ThemeColors.get('badge_text', _os_dark),
                                padx=4, pady=1, cursor='hand2', bd=0)
            btn_sel.pack(side=LEFT, padx=(0, 2))
            _req = '(필수)' if required else '(선택)'
            self._attach_doc_tooltip(btn_sel, f"클릭하여 {doc_name} 파일 선택 {_req}")
            
            # 체크 표시
            check_label = ttk.Label(cell, text="☐", font=('', 15))
            check_label.pack(side=LEFT, padx=(0, 2))
            self.check_labels[doc_type] = check_label
            
            # 파일명 (숨김 — 체크되면 표시)
            file_label = ttk.Label(cell, text="", foreground=ThemeColors.get('text_muted', _os_dark),
                                   font=('맑은 고딕', 12), anchor='w')
            file_label.pack(side=LEFT, fill=X, expand=True, padx=(0, 2))
            self.file_labels[doc_type] = file_label

        # 드래그앤드롭/캡처 이미지 등 초기 파일 지정
        for doc_type, path in getattr(self, '_initial_files', {}).items():
            if doc_type in self.file_labels and path:
                self.file_paths[doc_type] = path
                self.file_labels[doc_type].configure(text=os.path.basename(path))
                self.check_labels[doc_type].configure(text="☑")
        
        # [파싱 시작] 버튼
        self.btn_parse = ttk.Button(
            file_frame, text="▶ 파싱 시작",
            command=self._start_parsing,
            state='disabled', width=10
        )
        self.btn_parse.grid(row=0, column=4, padx=(6, 2))
        self._attach_doc_tooltip(self.btn_parse,
            "선택한 서류를 분석합니다\n\n• Packing List → LOT, 수량, 중량 추출\n• Invoice, FA → SAP번호, 금액 추출\n• Bill of Loading → BL번호, 선박, 일정 추출\n• Delivery Order → 인도장소, Free Time 추출")
        
        self.parse_hint = ttk.Label(
            file_frame, text="",
            foreground='white', font=('맑은 고딕', 12)
        )
        self.parse_hint.grid(row=0, column=5, padx=(2, 4), sticky='w')
        self._update_parse_hint()
        
        # v5.7.5: 프로그레스 (팝업 + 인라인)
        self.progress_var = tk.DoubleVar(value=0)
        self.status_var = tk.StringVar(value="")
        self._progress_popup = None
        self._progress_popup_label = None
        self._progress_popup_bar = None
        
        # ═══════════════════════════════════════════════════════════
        # 1.5 진행 상태 (미리보기 위에 고정 — 진행/데이터 혼동 방지)
        # ═══════════════════════════════════════════════════════════
        _pop_dark = ThemeColors.is_dark_theme(getattr(self.parent, 'current_theme', 'flatly'))
        progress_section = ttk.LabelFrame(main, text="⏱ 진행 상태", padding=8)
        progress_section.pack(fill=X, pady=(6, 4))
        self._progress_inline_placeholder = ttk.Label(
            progress_section, text="파싱을 시작하면 진행 상황이 여기에 표시됩니다.",
            font=('맑은 고딕', 11), foreground=ThemeColors.get('text_muted', _pop_dark))
        self._progress_inline_placeholder.pack(anchor='w')
        self._progress_inline_frame = ttk.Frame(progress_section)
        # 아래에서 pack하지 않음 — 파싱 시작 시 pack, 완료 후 forget
        self._progress_inline_msg = ttk.Label(self._progress_inline_frame, text="", font=('맑은 고딕', 12, 'bold'))
        self._progress_inline_msg.pack(anchor='w')
        _ps = ttk.Style()
        _ps.configure('Inline.Horizontal.TProgressbar', troughcolor=ThemeColors.get('bg_secondary', _pop_dark), thickness=12)
        self._progress_inline_bar = ttk.Progressbar(self._progress_inline_frame, maximum=100, mode='determinate', style='Inline.Horizontal.TProgressbar')
        self._progress_inline_bar.pack(fill=X, pady=(4, 2))
        _row2 = ttk.Frame(self._progress_inline_frame)
        _row2.pack(fill=X)
        self._progress_inline_pct_elapsed = ttk.Label(_row2, text="", font=('맑은 고딕', 10), foreground=ThemeColors.get('text_secondary', _pop_dark))
        self._progress_inline_pct_elapsed.pack(side=tk.RIGHT)
        self._progress_inline_busy = ttk.Label(_row2, text="", font=('맑은 고딕', 11), foreground=ThemeColors.get('statusbar_icon_warn', _pop_dark))
        self._progress_inline_busy.pack(side=tk.LEFT)
        
        # ═══════════════════════════════════════════════════════════
        # 2. 미리보기 테이블 (v3.8.7: 폰트 20% 확대)
        # ═══════════════════════════════════════════════════════════
        # v5.7.5: "업로드 2" 삭제 — "(확인 후 업로드)" 문구 제거
        tree_frame = ttk.LabelFrame(main, text="📊 미리보기 (스케일링·처리된 데이터)", padding=4)
        tree_frame.pack(fill=BOTH, expand=YES, pady=(0, 3))
        
        # ★ v3.8.7: 미리보기 Treeview 폰트 20% 확대 (기본 9pt → 11pt)
        import tkinter.font as tkfont
        preview_font = tkfont.Font(family='맑은 고딕', size=14)
        heading_font = tkfont.Font(family='맑은 고딕', size=13, weight='bold')
        row_height = preview_font.metrics('linespace') + 8
        
        _tree_dark = ThemeColors.is_dark_theme(getattr(self.parent, 'current_theme', 'flatly'))
        _tree_fg = ThemeColors.get('text_primary', _tree_dark)
        style = ttk.Style()
        style.configure('Preview.Treeview',
                        font=('맑은 고딕', 14),
                        rowheight=row_height,
                        foreground=_tree_fg,
                        fieldbackground=ThemeColors.get('bg_card', _tree_dark))
        style.configure('Preview.Treeview.Heading',
                        font=('맑은 고딕', 13, 'bold'))
        
        columns = tuple(col[0] for col in PREVIEW_COLUMNS)
        self.tree = ttk.Treeview(
            tree_frame, columns=columns, show="headings",
            height=18, selectmode='browse',
            style='Preview.Treeview'
        )
        self.tree.tag_configure('odd', background=ThemeColors.get('tree_stripe', _tree_dark), foreground=_tree_fg)
        self.tree.tag_configure('even', background=ThemeColors.get('bg_card', _tree_dark), foreground=_tree_fg)
        
        for col_id, header, width, anchor in PREVIEW_COLUMNS:
            self.tree.heading(col_id, text=header)
            self.tree.column(col_id, width=width, anchor=anchor, minwidth=35)
        
        scrollbar_y = ttk.Scrollbar(tree_frame, orient=VERTICAL, command=self.tree.yview)
        scrollbar_x = ttk.Scrollbar(tree_frame, orient=HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        scrollbar_x.pack(side=BOTTOM, fill=X)
        self.tree.pack(side=LEFT, fill=BOTH, expand=YES)
        scrollbar_y.pack(side=RIGHT, fill=Y)
        
        # v5.8.9: 컨테이너 번호 접미사(-숫자) 표시 옵션
        self._var_show_container_suffix = tk.BooleanVar(value=False)
        chk_container = ttk.Checkbutton(
            tree_frame, text="컨테이너 번호 접미사(-숫자) 표시",
            variable=self._var_show_container_suffix,
            command=self._on_toggle_container_suffix
        )
        chk_container.pack(anchor='w', padx=4, pady=(2, 0))
        
        # ═══════════════════════════════════════════════════════════
        # 4. 하단 한 줄 — 업로드5: 폰트 통일(15), 업로드6: 합계 가운데 배치
        # [엑셀][DB 업로드]  (합계: ... 가운데)  [취소]
        # ═══════════════════════════════════════════════════════════
        btn_frame = ttk.Frame(main)
        btn_frame.pack(fill=X, pady=(8, 0))
        
        _font = getattr(self, '_toolbar_font', '맑은 고딕') if hasattr(self, '_toolbar_font') else '맑은 고딕'
        _btn_font_size = 15
        _btn_fg = ThemeColors.get('badge_text', _tree_dark)
        _blue = ThemeColors.get('info', _tree_dark)
        _red = ThemeColors.get('statusbar_icon_err', _tree_dark)
        
        self.btn_excel = tk.Button(
            btn_frame, text="📥 Excel 내보내기",
            command=self._export_to_excel, state='disabled',
            font=(_font, _btn_font_size, 'bold'), bg=_blue, fg=_btn_fg,
            padx=15, pady=6, cursor='hand2', bd=0
        )
        self.btn_excel.pack(side=LEFT, padx=(0, 5))
        
        self.btn_upload = tk.Button(
            btn_frame, text="📤 DB 업로드",
            command=self._on_upload, state='disabled',
            font=(_font, _btn_font_size, 'bold'), bg=_blue, fg=_btn_fg,
            padx=20, pady=8, cursor='hand2', bd=0
        )
        self.btn_upload.pack(side=LEFT, padx=(5, 0))
        self._attach_doc_tooltip(self.btn_upload,
            "미리보기 데이터를 DB에 저장합니다\n\n• 저장 후 재고리스트에 자동 반영\n• 중복 LOT는 자동 스킵\n• 저장 완료 후 재고리스트 화면 표시")
        
        self.summary_var = tk.StringVar(value="")
        _summary_lbl = ttk.Label(btn_frame, textvariable=self.summary_var,
                                font=('맑은 고딕', 13, 'bold'),
                                foreground=ThemeColors.get('statusbar_progress', _tree_dark))
        _summary_lbl.pack(side=LEFT, fill=X, expand=True, padx=10)
        
        tk.Button(
            btn_frame, text="❌ 취소",
            command=self._on_cancel,
            font=(_font, _btn_font_size, 'bold'), bg=_red, fg=_btn_fg,
            padx=20, pady=8, cursor='hand2', bd=0
        ).pack(side=RIGHT, padx=(5, 0))
    
    # ═══════════════════════════════════════════════════════════
    # 파일 선택
    # ═══════════════════════════════════════════════════════════
    
    def _update_parse_hint(self) -> None:
        """파싱 시작 옆 업로드 상태 문구 갱신: 총 4개 중 N개 업로드되었습니다."""
        n = len(self.file_paths)
        if not getattr(self, 'parse_hint', None):
            return
        _hint_dark = ThemeColors.is_dark_theme(getattr(self.parent, 'current_theme', 'flatly'))
        if 'PACKING_LIST' not in self.file_paths:
            self.parse_hint.config(
                text="💡 최소 Packing List를 선택하세요",
                foreground=ThemeColors.get('text_muted', _hint_dark)
            )
            if self.btn_parse:
                self.btn_parse.config(state='disabled')
        else:
            if self.btn_parse:
                self.btn_parse.config(state='normal')
            self.parse_hint.config(
                text=f"총 4개 중 {n}개 업로드되었습니다.",
                foreground=ThemeColors.get('text_primary', _hint_dark)
            )
    
    def _select_file(self, doc_type: str):
        """서류별 파일 선택"""
        type_names = {
            'PACKING_LIST': 'Packing List',
            'INVOICE': 'Invoice, FA',
            'BL': 'Bill of Loading',
            'DO': 'Delivery Order',
        }
        
        file_path = filedialog.askopenfilename(
            parent=self.dialog,
            title=f"{type_names.get(doc_type, doc_type)} 파일 선택",
            filetypes=[
                ("PDF files", "*.pdf"),
                ("Image (D/O 캡처)", "*.png *.jpg *.jpeg"),
                ("All files", "*.*")
            ]
        )
        
        if not file_path:
            return
        
        self.file_paths[doc_type] = file_path
        fname = os.path.basename(file_path)
        
        # UI 업데이트
        self.file_labels[doc_type].config(text=fname, foreground=ThemeColors.get('text_primary', ThemeColors.is_dark_theme(getattr(self.parent, 'current_theme', 'flatly'))))
        self.check_labels[doc_type].config(text="✅")
        
        self._log(f"📂 {doc_type}: {fname}")
        
        # 파싱 버튼 활성화 조건: PL 필수
        self._update_parse_hint()
    
    # ═══════════════════════════════════════════════════════════
    # 파싱
    # ═══════════════════════════════════════════════════════════
    
    def _start_parsing(self) -> None:
        """v3.8.9: 파싱 시작 — 누락 서류 경고 후 진행"""
        # 서류 누락 검사
        missing_required = []
        missing_optional = []
        for doc_type, doc_name, required in DOC_TYPES:
            if doc_type not in self.file_paths:
                if required:
                    missing_required.append(doc_name)
                else:
                    missing_optional.append(doc_name)
        
        # 경고 메시지 구성
        if missing_required:
            warning_lines = ["⚠️ 다음 필수 서류가 선택되지 않았습니다:\n"]
            for name in missing_required:
                warning_lines.append(f"  • {name}")
            warning_lines.append("\n선택하지 않은 서류의 정보는 누락됩니다.")
            if missing_optional:
                warning_lines.append(f"\n📋 선택 서류 미선택: {', '.join(missing_optional)}")
            warning_lines.append("\n계속 진행하시겠습니까?")
            proceed = msgbox.askyesno(
                "서류 누락 확인",
                "\n".join(warning_lines),
                parent=self.dialog
            )
            if not proceed:
                return
        elif missing_optional:
            # 선택 서류만 누락 — 정보 안내만
            self._update_progress(0, f"ℹ️ {', '.join(missing_optional)} 미선택 — 해당 정보 생략")
        
        # D/O 미선택 시 경고: 선적일·도착일·반납일 직접 입력 안내
        if 'DO' not in self.file_paths:
            do_ok = msgbox.askyesno(
                "D/O 미첨부",
                "D/O 서류가 업로드되지 않았습니다. 계속할까요?\n\n"
                "이 경우 선적일, 도착일, 컨테이너 반납일을 사용자가 직접 입력해야 합니다.",
                parent=self.dialog
            )
            if not do_ok:
                return
        
        self.btn_parse.config(state='disabled')
        self._show_progress_inline()
        
        thread = threading.Thread(
            target=self._parse_thread,
            daemon=True
        )
        thread.start()
    
    def _show_progress_inline(self) -> None:
        """진행 상태를 미리보기 위 인라인 영역에만 표시 (팝업 없음, 움직임 표시 포함)"""
        ph = getattr(self, '_progress_inline_placeholder', None)
        fr = getattr(self, '_progress_inline_frame', None)
        if ph and ph.winfo_ismapped():
            ph.pack_forget()
        if fr:
            fr.pack(fill=X)
        self._progress_start_time = time.time()
        if getattr(self, '_progress_inline_bar', None):
            self._progress_inline_bar['value'] = 0
        if getattr(self, '_progress_inline_msg', None):
            self._progress_inline_msg.config(text="준비 중...")
        if getattr(self, '_progress_inline_pct_elapsed', None):
            self._progress_inline_pct_elapsed.config(text="0%  ·  경과: 0:00")
        if getattr(self, '_progress_inline_busy', None):
            self._progress_inline_busy.config(text="진행 중 ●")
        self._start_progress_elapsed_tick()
        self._start_progress_busy_animation()

    def _hide_progress_inline(self) -> None:
        """진행 완료 후 인라인 영역을 플레이스홀더로 복귀"""
        fr = getattr(self, '_progress_inline_frame', None)
        ph = getattr(self, '_progress_inline_placeholder', None)
        if fr and fr.winfo_ismapped():
            fr.pack_forget()
        if ph:
            ph.pack(anchor='w')

    def _show_progress_popup(self) -> None:
        """작업진행 전용 창 사용 안 함 — 기존 화면(인라인 진행 상태)만 사용"""
        pass

    def _progress_elapsed_tick(self) -> None:
        """경과 시간 표시 업데이트 (1초 간격) — 팝업·인라인 둘 다"""
        start = getattr(self, '_progress_start_time', None)
        if start is None:
            self._progress_elapsed_job = self.dialog.after(1000, self._progress_elapsed_tick) if self.dialog and self.dialog.winfo_exists() else None
            return
        secs = int(time.time() - start)
        if secs >= 3600:
            h, r = divmod(secs, 3600)
            m, s = divmod(r, 60)
            elapsed_text = f"경과: {h}:{m:02d}:{s:02d}"
        else:
            m, s = divmod(secs, 60)
            elapsed_text = f"경과: {m}:{s:02d}"
        # 인라인 경과 (현재 퍼센트 + 경과)
        pct_elapsed = getattr(self, '_progress_inline_pct_elapsed', None)
        if pct_elapsed and pct_elapsed.winfo_ismapped():
            pct = getattr(self, 'progress_var', None)
            pct_val = int(pct.get()) if pct else 0
            pct_elapsed.config(text=f"{pct_val}%  ·  {elapsed_text}")
        self._progress_elapsed_job = self.dialog.after(1000, self._progress_elapsed_tick) if self.dialog and self.dialog.winfo_exists() else None

    def _start_progress_elapsed_tick(self) -> None:
        """경과 시간 타이머 시작"""
        self._progress_elapsed_job = None
        if self.dialog and self.dialog.winfo_exists():
            self._progress_elapsed_job = self.dialog.after(1000, self._progress_elapsed_tick)

    def _stop_progress_elapsed_tick(self) -> None:
        """경과 시간 타이머 중지"""
        if getattr(self, '_progress_elapsed_job', None):
            try:
                if self.dialog and self.dialog.winfo_exists():
                    self.dialog.after_cancel(self._progress_elapsed_job)
            except (tk.TclError, ValueError):
                pass
        self._progress_elapsed_job = None

    def _progress_busy_tick(self) -> None:
        """진행 중 움직임 표시 — 기존 화면(인라인) 진행 상태 영역에만 표시"""
        phase = getattr(self, '_progress_busy_phase', 0) % 4
        self._progress_busy_phase = phase + 1
        texts = ['진행 중 ●  ', '진행 중 ●● ', '진행 중 ●●●', '진행 중 ●● ']
        inline_busy = getattr(self, '_progress_inline_busy', None)
        if inline_busy and inline_busy.winfo_ismapped():
            inline_busy.config(text=texts[phase])
        self._progress_busy_job = self.dialog.after(400, self._progress_busy_tick) if self.dialog and self.dialog.winfo_exists() else None

    def _start_progress_busy_animation(self) -> None:
        self._progress_busy_phase = 0
        if self.dialog and self.dialog.winfo_exists():
            self._progress_busy_job = self.dialog.after(400, self._progress_busy_tick)

    def _stop_progress_busy_animation(self) -> None:
        if getattr(self, '_progress_busy_job', None):
            try:
                if self.dialog and self.dialog.winfo_exists():
                    self.dialog.after_cancel(self._progress_busy_job)
            except (tk.TclError, ValueError):
                pass
        self._progress_busy_job = None

    def _hide_progress_popup(self) -> None:
        """진행률 팝업 닫기"""
        self._stop_progress_busy_animation()
        self._stop_progress_elapsed_tick()
        try:
            if getattr(self, '_progress_popup', None) and self._progress_popup.winfo_exists():
                self._progress_popup.destroy()
        except Exception as e:
            logger.debug(f"Suppressed: {e}")
        self._progress_popup = None
        self._progress_popup_label = None
        self._progress_popup_bar = None
        self._progress_popup_pct = None
        self._progress_popup_busy = None
        self._progress_popup_elapsed = None

    def _update_progress(self, pct: int, message: str):
        """프로그레스 바 업데이트 (스레드 안전) — 팝업 + 인라인 동기화, 완료 시 인라인 복귀"""
        def _update():
            self.progress_var.set(pct)
            self.status_var.set(message)
            # 팝업
            bar = getattr(self, '_progress_popup_bar', None)
            if bar and bar.winfo_exists():
                bar['value'] = max(0, min(100, pct))
                if self._progress_popup_label:
                    self._progress_popup_label.config(text=message)
                if getattr(self, '_progress_popup_pct', None):
                    self._progress_popup_pct.config(text=f"{pct}%" if pct >= 0 else "—")
            # 인라인 (미리보기 위) — 기존 화면 프로그레스 바만 사용
            inline_bar = getattr(self, '_progress_inline_bar', None)
            inline_msg = getattr(self, '_progress_inline_msg', None)
            inline_busy = getattr(self, '_progress_inline_busy', None)
            if inline_bar and inline_bar.winfo_ismapped():
                inline_bar['value'] = max(0, min(100, pct))
            if inline_msg and inline_msg.winfo_ismapped():
                inline_msg.config(text=message)
            if pct >= 100 or (pct == 0 and message.strip().startswith("❌")):
                self._stop_progress_busy_animation()
                if inline_busy and inline_busy.winfo_ismapped():
                    inline_busy.config(text="완료" if pct >= 100 else "오류")
                if self.dialog and self.dialog.winfo_exists():
                    self.dialog.after(PROGRESS_POPUP_CLOSE_DELAY_MS, self._hide_progress_popup)
                self.dialog.after(PROGRESS_POPUP_CLOSE_DELAY_MS + 100, self._hide_progress_inline)
        if self.dialog and self.dialog.winfo_exists():
            self.dialog.after(0, _update)
    
    def _parse_thread(self) -> None:
        """백그라운드 파싱"""
        try:
            from parsers.document_parser_v2 import DocumentParserV2
            
            gemini_key = os.environ.get('GEMINI_API_KEY', '')
            if not gemini_key:
                try:
                    from core.config import get_settings
                    settings = get_settings()
                    gemini_key = settings.get('gemini_api_key', '')
                except (ImportError, ModuleNotFoundError) as _e:
                    logger.debug(f"onestop_inbound: {_e}")
            
            # v5.5.1: 모든 파싱은 API(Gemini) 강제
            if not gemini_key or str(gemini_key).strip() == '' or str(gemini_key).startswith('your-'):
                raise RuntimeError("API-only 모드: Gemini API Key가 필요합니다. 설정에서 API Key를 입력하세요.")

            parser = DocumentParserV2(gemini_api_key=gemini_key)
            
            # 파싱 순서: PL → Invoice → BL → DO
            parse_order = ['PACKING_LIST', 'INVOICE', 'BL', 'DO']
            to_parse = [(dt, self.file_paths[dt]) for dt in parse_order if dt in self.file_paths]
            total = len(to_parse)
            if total == 0:
                self._update_progress(90, "파싱할 파일이 없습니다")
                return
            
            icons = {'PACKING_LIST': '📦', 'INVOICE': '📑', 'BL': '🚢', 'DO': '📋'}
            
            pl_result = None
            inv_result = None
            bl_result = None
            do_result = None
            
            # v5.7.5: 현재 파싱 중인 서류 이름 표시
            doc_type_display = {
                'PACKING_LIST': 'Packing List',
                'INVOICE': 'Invoice, FA',
                'BL': 'Bill of Loading',
                'DO': 'Delivery Order',
            }
            for idx, (doc_type, file_path) in enumerate(to_parse):
                fname = os.path.basename(file_path)
                icon = icons.get(doc_type, '📄')
                pct = int(10 + 70 * idx / total)
                doc_name = doc_type_display.get(doc_type, doc_type)
                self._update_progress(pct, f"현재 파싱 중: {doc_name} — {fname}")
                self._log_safe(f"{icon} {doc_type} 파싱: {fname}")
                
                try:
                    if doc_type == 'PACKING_LIST':
                        pl_result = parser.parse_packing_list(file_path)
                        self.parsed_results['packing_list'] = pl_result
                        _lots = getattr(pl_result, 'lots', []) if pl_result else []
                        if _lots:
                            _tnw = getattr(pl_result, 'total_net_weight_kg', 0) or 0
                            self._log_safe(f"  ✅ LOTs: {len(_lots)}, Net: {_tnw:,.0f}kg")
                    
                    elif doc_type == 'INVOICE':
                        inv_result = parser.parse_invoice(file_path)
                        self.parsed_results['invoice'] = inv_result
                        if inv_result:
                            self._log_safe(f"  ✅ SAP: {getattr(inv_result, 'sap_no', '')}, Invoice: {getattr(inv_result, 'salar_invoice_no', '')}")
                    
                    elif doc_type == 'BL':
                        bl_result = parser.parse_bl(file_path)
                        self.parsed_results['bl'] = bl_result
                        if bl_result:
                            self._log_safe(f"  ✅ B/L: {getattr(bl_result, 'bl_no', '')}, Containers: {getattr(bl_result, 'total_containers', 0)}")
                    
                    elif doc_type == 'DO':
                        do_result = parser.parse_do(file_path)
                        self.parsed_results['do'] = do_result
                        if do_result:
                            self._log_safe(f"  ✅ D/O: B/L={getattr(do_result, 'bl_no', '')}")
                
                except (ValueError, TypeError, AttributeError, RuntimeError) as e:
                    self._log_safe(f"  ❌ {doc_type} 파싱 오류: {e}")
                    logger.error(f"파싱 오류 [{doc_type}]: {e}", exc_info=True)
                    # RuntimeError: Gemini API-Only 실패(예: JSON 추출 실패) → 입고 미완료 → 재고/톤백 리스트에 데이터 없음
                    if isinstance(e, RuntimeError) and doc_type == 'PACKING_LIST':
                        self._log_safe("  💡 Packing List 실패 시 입고가 완료되지 않아 톤백 리스트에 표시되지 않습니다.")
                else:
                    # 서류 하나 파싱 직후마다 병합 후 미리보기 테이블·메인 화면에 실시간 반영
                    self._merge_results(inv_result, pl_result, bl_result, do_result)
                    if self.dialog and self.dialog.winfo_exists():
                        self.dialog.after(0, lambda: self._push_preview_to_main())
                        self.dialog.after(0, lambda: self._refresh_preview_tree_only())
            
            # 병합
            self._update_progress(85, "📊 데이터 병합 중...")
            self._merge_results(inv_result, pl_result, bl_result, do_result)
            
            # ═══════════════════════════════════════════════════════
            # ★★★ v5.8.7: D/O 없거나 arrival_date 누락 시 사용자 입력
            # ═══════════════════════════════════════════════════════
            self._do_deferred = False  # D/O 추후 첨부 플래그
            _need_date_input = False
            if not do_result:
                _need_date_input = True
                self._log_safe("📋 D/O 미첨부 — 날짜 정보 수동 입력 필요")
            elif self.preview_data and not (self.preview_data[0].get('arrival_date') or '').strip():
                _need_date_input = True
                self._log_safe("📋 D/O에서 입항일 추출 실패 — 수동 입력 필요")
            
            if _need_date_input and self.preview_data:
                prefilled_ship = ''
                if self.preview_data:
                    prefilled_ship = self.preview_data[0].get('ship_date', '') or ''
                
                import queue
                date_queue = queue.Queue()
                
                def _show_date_popup():
                    self._hide_progress_popup()
                    result = self._ask_missing_dates(prefilled_ship, do_result)
                    date_queue.put(result)
                
                if self.dialog and self.dialog.winfo_exists():
                    self.dialog.after(0, _show_date_popup)
                    try:
                        user_dates = date_queue.get(timeout=300)
                    except queue.Empty:
                        user_dates = None
                    
                    if user_dates:
                        if user_dates.get('deferred'):
                            # "D/O 추후 첨부" 선택
                            self._do_deferred = True
                            self._log_safe("  📋 D/O 추후 첨부 선택됨 — arrival_date 없이 진행")
                        else:
                            for row in self.preview_data:
                                if user_dates.get('ship_date') and not (row.get('ship_date') or '').strip():
                                    row['ship_date'] = user_dates['ship_date']
                                if user_dates.get('arrival_date'):
                                    row['arrival_date'] = user_dates['arrival_date']
                                if user_dates.get('free_time') and not (row.get('free_time') or '').strip():
                                    row['free_time'] = user_dates['free_time']
                            self._log_safe(f"  ✅ 수동 입력: arrival={user_dates.get('arrival_date')}, free_time={user_dates.get('free_time')}")
                    else:
                        self._log_safe("  ⚠️ 날짜 입력 취소 — arrival_date 없이 진행")
            
            # v3.8.9: 파싱 결과 경고 (누락된 정보)
            _warnings = []
            if not pl_result or not getattr(pl_result, 'lots', None):
                _warnings.append("⚠️ Packing List: LOT 정보 추출 실패")
            if not inv_result or not getattr(inv_result, 'sap_no', None):
                _warnings.append("⚠️ Invoice: SAP번호 추출 실패 — 수동 입력 필요")
            if not bl_result or not getattr(bl_result, 'bl_no', None):
                _warnings.append("⚠️ B/L: BL번호 추출 실패 — 수동 입력 필요")
            
            if _warnings:
                _warn_msg = "\n".join(_warnings)
                self._log_safe(f"\n{'='*40}\n{_warn_msg}\n{'='*40}")
                # GUI 경고
                def _show_warn():
                    CustomMessageBox.warning(None, "파싱 결과 확인", _warn_msg, parent=self.dialog)
                if self.dialog and self.dialog.winfo_exists():
                    self.dialog.after(500, _show_warn)
            
            # 병합 직후 메인 화면 재고 리스트에 실시간 반영
            if self.dialog and self.dialog.winfo_exists() and self.preview_data:
                self.dialog.after(0, lambda: self._push_preview_to_main())
            
            # 표시
            self._update_progress(95, "📋 미리보기 준비...")
            self._display_preview()
            # 파싱 완료 후 DB 업로드·Excel 버튼이 반드시 보이도록 폴백 (순차 삽입 완료 전에도 활성화)
            _preview_len = len(self.preview_data)
            def _ensure_buttons_visible():
                if not self.dialog or not self.dialog.winfo_exists():
                    return
                if _preview_len and getattr(self, 'preview_data', None) and len(self.preview_data) == _preview_len:
                    if getattr(self, 'btn_excel', None) and self.btn_excel.winfo_exists():
                        self.btn_excel.config(state='normal')
                    if getattr(self, 'btn_upload', None) and self.btn_upload.winfo_exists():
                        if self._has_required_docs():
                            self.btn_upload.config(state='normal')
                        else:
                            self.btn_upload.config(state='disabled')
            if self.dialog and self.dialog.winfo_exists():
                self.dialog.after(400, _ensure_buttons_visible)
            
            self._update_progress(100, f"✅ 파싱 완료 — {len(self.preview_data)}개 LOT")
            self._log_safe(f"✅ 파싱 완료: {len(self.preview_data)} LOT, {total}종 서류")
        
        except (RuntimeError, ValueError) as e:
            self._update_progress(0, f"❌ 오류: {e}")
            self._log_safe(f"❌ 파싱 오류: {e}")
            logger.error(f"원스톱 파싱 오류: {e}", exc_info=True)
            self._enable_parse_btn()
    
    # ═══════════════════════════════════════════════════════════
    # 데이터 병합 (4종 → 18열)
    # ═══════════════════════════════════════════════════════════
    
    def _merge_results(self, invoice, pl, bl, do) -> list:
        """4종 파싱 결과를 18열 미리보기 데이터로 병합"""
        self.preview_data = []
        
        if not pl or not getattr(pl, 'lots', None):
            if invoice and getattr(invoice, 'lot_numbers', None):
                for idx, lot_no in enumerate(getattr(invoice, 'lot_numbers', []), 1):
                    row = self._empty_row(idx)
                    row['sap_no'] = getattr(invoice, 'sap_no', '') or ''
                    row['lot_no'] = lot_no
                    row['product'] = getattr(invoice, 'product', '') or 'LITHIUM CARBONATE'
                    row['salar_invoice_no'] = getattr(invoice, 'salar_invoice_no', '') or ''
                    row['ship_date'] = str(getattr(invoice, 'invoice_date', '')) if getattr(invoice, 'invoice_date', None) else ''
                    if bl:
                        row['bl_no'] = self._format_bl(getattr(bl, 'bl_no', '') or '')
                    self._fill_do(row, do)
                    row['status'] = 'AVAILABLE'
                    self.preview_data.append(row)
            return
        
        for idx, lot in enumerate(getattr(pl, 'lots', []) or [], 1):
            row = self._empty_row(idx)
            row['sap_no'] = getattr(pl, 'sap_no', '') or (getattr(invoice, 'sap_no', '') if invoice else '') or ''
            row['container_no'] = getattr(lot, 'container_no', '') or ''
            row['product'] = getattr(pl, 'product', '') or 'LITHIUM CARBONATE'
            row['product_code'] = getattr(pl, 'code', '') or ''
            row['lot_no'] = getattr(lot, 'lot_no', '') or ''
            row['lot_sqm'] = getattr(lot, 'lot_sqm', '') or ''
            
            _mxbg = getattr(lot, 'mxbg_pallet', None)
            row['mxbg_pallet'] = str(_mxbg) if _mxbg else '10'
            
            _nw = getattr(lot, 'net_weight_kg', None)
            row['net_weight'] = f"{float(_nw):,.1f}" if _nw else ''
            
            _gw = getattr(lot, 'gross_weight_kg', None)
            row['gross_weight'] = f"{float(_gw):,.3f}" if _gw else ''
            
            # v3.8.8: B/L ship_date 우선, Invoice 폴백 — 업로드3/4: 파싱값으로 채움 (날짜는 YYYY-MM-DD)
            if bl:
                row['bl_no'] = self._format_bl(getattr(bl, 'bl_no', '') or '')
                _sd = getattr(bl, 'ship_date', None)
                if _sd:
                    row['ship_date'] = str(_sd)[:10] if len(str(_sd)) >= 10 else str(_sd)
            
            if invoice:
                row['salar_invoice_no'] = getattr(invoice, 'salar_invoice_no', '') or ''
                if not (row.get('ship_date') or '').strip():
                    _id = getattr(invoice, 'invoice_date', None)
                    if _id:
                        row['ship_date'] = str(_id)[:10] if len(str(_id)) >= 10 else str(_id)
                if not row['sap_no']:
                    row['sap_no'] = getattr(invoice, 'sap_no', '') or ''
            
            self._fill_do(row, do)
            if not (row.get('warehouse') or '').strip():
                row['warehouse'] = DEFAULT_WAREHOUSE
            row['status'] = 'AVAILABLE'
            self.preview_data.append(row)
    
    def _empty_row(self, no: int) -> dict:
        row = {col[0]: '' for col in PREVIEW_COLUMNS}
        row['no'] = str(no)
        return row
    
    def _date_str(self, val) -> str:
        """날짜를 YYYY-MM-DD 문자열로. None/'None'/비어있으면 '' 반환 (date.today() 사용 안 함)."""
        if val is None or (isinstance(val, str) and (not val.strip() or val.strip() in ('None', 'none'))):
            return ''
        if hasattr(val, 'isoformat'):
            return str(val.isoformat())[:10]
        s = str(val).strip()
        return s[:10] if len(s) >= 10 and s not in ('None', 'none') else (s if s and s not in ('None', 'none') else '')

    def _format_bl(self, bl_no) -> str:
        if not bl_no:
            return ''
        bl_no = str(bl_no).strip()
        if bl_no.isdigit() and len(bl_no) >= 9:
            return f"MAEU{bl_no}"
        return bl_no
    
    def _fill_do(self, row: dict, do) -> None:
        """v3.8.8: D/O 데이터로 미리보기 행 보완 (free_time 계산 포함)"""
        if not do:
            return
        if not row.get('bl_no') and getattr(do, 'bl_no', None):
            row['bl_no'] = str(getattr(do, 'bl_no', ''))
        
        # arrival_date (업로드3/4: D/O 파싱값으로 채움, YYYY-MM-DD)
        # v5.8.8: 날짜가 아닌 값(예: '광양')이면 넣지 않음 — ARRIVAL 컬럼 혼동 방지
        arr = getattr(do, 'arrival_date', None)
        if arr and str(arr) != 'None':
            _s = str(arr).strip()[:10]
            if len(_s) == 10 and _s.count('-') == 2 and _s.replace('-', '').isdigit():
                row['arrival_date'] = _s
        
        # warehouse
        wh = getattr(do, 'warehouse_name', '') or getattr(do, 'warehouse', '')
        if wh:
            row['warehouse'] = str(wh)
        
        # FREE TIME = con_return(컨테이너 반납일) - arrival_date (일수). D/O의 Free_Time 컬럼 = 반납일
        ft_infos = getattr(do, 'free_time_info', []) or []
        if ft_infos and arr and str(arr) != 'None':
            try:
                con_return_str = ''
                for ft in ft_infos:
                    ftd = (getattr(ft, 'free_time_date', '') or getattr(ft, 'free_time_until', '')) if not isinstance(ft, dict) else (ft.get('free_time_date') or ft.get('free_time_until') or '')
                    if ftd and str(ftd) != 'None':
                        con_return_str = str(ftd)[:10]
                        break
                if not con_return_str:
                    print(f"[원스톱 미리보기] D/O free_time_info 있으나 반납일 없음 — CON RETURN/FREE TIME 빈칸. 항목 수: {len(ft_infos)}")
                if con_return_str:
                    con_return_dt = datetime.strptime(con_return_str, '%Y-%m-%d').date()
                    arr_dt = datetime.strptime(str(arr)[:10], '%Y-%m-%d').date()
                    days = (con_return_dt - arr_dt).days
                    row['free_time'] = str(max(0, days))
                    row['con_return'] = str(con_return_str)[:10]
                    print(f"[원스톱 미리보기] D/O 반납일 적용: con_return={row['con_return']}, free_time(일수)={row['free_time']}")
            except (ValueError, TypeError) as e:
                logging.getLogger(__name__).debug(f"free_time 계산 실패: {e}")
        # 업로드4: free_time 일수만 있는 경우 (DO.free_time.storage_free_days)
        if not (row.get('free_time') or '').strip():
            ft_single = getattr(do, 'free_time', None)
            if ft_single is not None:
                days_val = getattr(ft_single, 'storage_free_days', None) or (ft_single.get('storage_free_days') if isinstance(ft_single, dict) else None)
                if days_val is not None:
                    row['free_time'] = str(int(days_val))
    
    # ═══════════════════════════════════════════════════════════
    # ★★★ v5.8.7: 날짜 입력 팝업 (DatePicker 달력 UI)
    # ═══════════════════════════════════════════════════════════
    
    def _ask_missing_dates(self, prefilled_ship: str = '', do_result=None) -> dict:
        """
        사용자에게 ship_date, arrival_date, free_time을 물어보는 DatePicker 팝업.
        
        호출 조건:
            1) D/O 자체가 없을 때
            2) D/O는 있는데 arrival_date 추출 실패 시
        
        UI:
            - gui_bootstrap HAS_DATEENTRY(ttkbootstrap.DateEntry)가 있으면 달력 위젯 사용
            - 없으면 텍스트 입력 폴백
            - "D/O 추후 첨부" 버튼으로 건너뛰기 가능
        
        Returns:
            dict: {'ship_date': str, 'arrival_date': str, 'free_time': str}
            또는 {'deferred': True} (D/O 추후 첨부)
            또는 None (취소)
        """
        result_holder = [None]
        
        def _build_popup():
            win = None
            try:
                win = tk.Toplevel(self.dialog)
                
                if not do_result:
                    win.title("📋 D/O 미첨부 — 날짜 정보 입력")
                    msg_text = "D/O가 없습니다. 입항일 등을 직접 입력하거나,\n나중에 D/O를 추가할 수 있습니다."
                else:
                    win.title("📋 D/O 파싱 실패 — 날짜 정보 입력")
                    msg_text = "D/O에서 날짜를 읽지 못했습니다.\n직접 입력하거나 나중에 D/O를 다시 첨부할 수 있습니다."
                
                win.geometry(DialogSize.get_geometry(self.dialog, 'medium'))
                win.resizable(False, False)
                win.transient(self.dialog)
                win.grab_set()
                center_dialog(win, self.dialog)
                
                frame = ttk.Frame(win, padding=20)
                frame.pack(fill=tk.BOTH, expand=True)
                
                # 안내 메시지
                ttk.Label(frame, text=msg_text,
                         font=('맑은 고딕', 11, 'bold'),
                         wraplength=460).pack(anchor='w', pady=(0, 12))
                
                # ── 헬퍼: DateEntry( gui_bootstrap ) 또는 텍스트 입력 생성 ──
                def _make_date_field(parent, label, hint, prefill='', required=False):
                    """HAS_DATEENTRY면 ttkbootstrap 달력, 없으면 텍스트 입력. 반환값은 .get()으로 문자열 조회."""
                    _cal_dark = ThemeColors.is_dark_theme(getattr(self.parent, 'current_theme', 'flatly'))
                    lf = ttk.LabelFrame(parent,
                        text=f"{'★ ' if required else ''}{label}{' — 필수' if required else ''}",
                        padding=8)
                    lf.pack(fill=tk.X, pady=(0, 8))
                    
                    var = tk.StringVar(value=prefill)
                    
                    if HAS_DATEENTRY and DateEntry is not None:
                        startdate = None
                        if prefill:
                            try:
                                parts = prefill.split('-')
                                startdate = _date_type(int(parts[0]), int(parts[1]), int(parts[2]))
                            except (ValueError, IndexError):
                                pass
                        de = DateEntry(lf, dateformat='%Y-%m-%d', startdate=startdate,
                                       bootstyle='info', width=16)
                        de.pack(side=tk.LEFT, padx=(0, 8))
                        ttk.Label(lf, text=hint,
                                 font=('맑은 고딕', 9), foreground=ThemeColors.get('text_muted', _cal_dark)).pack(side=tk.LEFT)
                        class _DateGetter:
                            def get(self):
                                return (de.entry.get() or '').strip() if de and de.winfo_exists() else ''
                        return _DateGetter()
                    else:
                        entry = ttk.Entry(lf, textvariable=var,
                                         font=('맑은 고딕', 11), width=16)
                        entry.pack(side=tk.LEFT, padx=(0, 8))
                        ttk.Label(lf, text=hint,
                                 font=('맑은 고딕', 9), foreground=ThemeColors.get('text_muted', _cal_dark)).pack(side=tk.LEFT)
                        return var
                
                # ── 3개 날짜 필드 ──
                ship_var = _make_date_field(frame,
                    "선적일 (Ship Date)",
                    "※ B/L에서 추출됨" if prefilled_ship else "YYYY-MM-DD",
                    prefill=prefilled_ship)
                
                arrival_var = _make_date_field(frame,
                    "입항일 (Arrival Date)",
                    "YYYY-MM-DD (예: 2025-10-17)",
                    required=True)
                
                ft_var = _make_date_field(frame,
                    "반납기한 (Free Time)",
                    "반납일 YYYY-MM-DD 또는 일수(14)")
                
                # 에러 표시
                err_var = tk.StringVar()
                _err_dark = ThemeColors.is_dark_theme(getattr(self.parent, 'current_theme', 'flatly'))
                ttk.Label(frame, textvariable=err_var,
                         font=('맑은 고딕', 10), foreground=ThemeColors.get('danger', _err_dark)).pack(anchor='w', pady=(4, 0))
                
                # ── 날짜 검증 함수 ──
                def _validate_date(s):
                    import re as _re
                    if not s:
                        return True
                    s = s.strip()
                    if _re.match(r'^\d{4}-\d{1,2}-\d{1,2}$', s):
                        try:
                            parts = s.split('-')
                            _date_type(int(parts[0]), int(parts[1]), int(parts[2]))
                            return True
                        except ValueError:
                            return False
                    return False
                
                # ── 확인 버튼 ──
                def _on_ok():
                    arr = arrival_var.get().strip()
                    
                    if not arr:
                        err_var.set("⚠️ 입항일은 필수입니다!")
                        return
                    if not _validate_date(arr):
                        err_var.set("⚠️ 입항일 형식 오류 (YYYY-MM-DD)")
                        return
                    
                    ship = ship_var.get().strip()
                    if ship and not _validate_date(ship):
                        err_var.set("⚠️ 선적일 형식 오류 (YYYY-MM-DD)")
                        return
                    
                    ft_raw = ft_var.get().strip()
                    free_time_str = ''
                    if ft_raw:
                        if ft_raw.isdigit():
                            free_time_str = ft_raw
                        elif _validate_date(ft_raw):
                            try:
                                ap = arr.split('-')
                                fp = ft_raw.split('-')
                                arr_d = _date_type(int(ap[0]), int(ap[1]), int(ap[2]))
                                ft_d = _date_type(int(fp[0]), int(fp[1]), int(fp[2]))
                                free_time_str = str(max(0, (ft_d - arr_d).days))
                            except (ValueError, IndexError):
                                free_time_str = ft_raw
                        else:
                            err_var.set("⚠️ 반납기한: YYYY-MM-DD 또는 일수")
                            return
                    else:
                        free_time_str = '14'
                    
                    result_holder[0] = {
                        'ship_date': ship,
                        'arrival_date': arr,
                        'free_time': free_time_str,
                    }
                    win.destroy()
                
                # ── D/O 추후 첨부 버튼 ──
                def _on_defer():
                    result_holder[0] = {'deferred': True}
                    win.destroy()
                
                # ── 취소 ──
                def _on_cancel():
                    result_holder[0] = None
                    win.destroy()
                
                # ── 버튼 배치 ──
                btn_frame = ttk.Frame(frame)
                btn_frame.pack(fill=tk.X, pady=(12, 0))
                
                ttk.Button(btn_frame, text="✅ 확인",
                          command=_on_ok, width=10).pack(side=tk.LEFT, padx=(0, 8))
                
                ttk.Button(btn_frame, text="📋 D/O 추후 첨부",
                          command=_on_defer, width=16).pack(side=tk.LEFT, padx=(0, 8))
                
                ttk.Button(btn_frame, text="❌ 취소",
                          command=_on_cancel, width=10).pack(side=tk.LEFT)
                
                win.protocol("WM_DELETE_WINDOW", _on_cancel)
                return win
                
            except Exception as e:
                logger.error(f"[_ask_missing_dates] 팝업 오류: {e}", exc_info=True)
                return None
        
        if not self.dialog or not self.dialog.winfo_exists():
            return None
        
        win = _build_popup()
        if win and win.winfo_exists():
            win.wait_window(win)
        return result_holder[0]
    
    # ═══════════════════════════════════════════════════════════
    # 미리보기 표시
    # ═══════════════════════════════════════════════════════════
    
    def _push_preview_to_main(self) -> None:
        """파싱된 미리보기 데이터를 메인 화면 재고 리스트에 실시간 반영"""
        if not getattr(self, 'app', None) or not hasattr(self.app, '_set_parsing_preview_data'):
            return
        if not self.preview_data:
            return
        try:
            self.app._set_parsing_preview_data(list(self.preview_data))
        except (RuntimeError, ValueError, TypeError) as e:
            logger.debug(f"푸시 미리보기 실패: {e}")

    def _clear_preview_from_main(self) -> None:
        """메인 화면 파싱 미리보기 해제 후 DB 기준으로 복원"""
        if not getattr(self, 'app', None) or not hasattr(self.app, '_set_parsing_preview_data'):
            return
        try:
            self.app._set_parsing_preview_data(None)
        except (RuntimeError, ValueError, TypeError) as e:
            logger.debug(f"미리보기 해제 실패: {e}")

    def _format_container_display(self, val) -> str:
        """컨테이너 번호: 디폴트로 접미사 -숫자 제거. 표시 옵션 켜면 원문 반환."""
        if not val or not isinstance(val, str):
            return val or ''
        if getattr(self, '_show_container_suffix', False):
            return val.strip()
        s = val.strip()
        if '-' in s:
            pre, _, suf = s.rpartition('-')
            if suf.isdigit():
                return pre
        return s
    
    def _on_toggle_container_suffix(self) -> None:
        """컨테이너 접미사 표시 체크 시 미리보기 테이블 갱신"""
        var = getattr(self, '_var_show_container_suffix', None)
        self._show_container_suffix = bool(var and var.get())
        if self.preview_data and getattr(self, 'tree', None) and self.tree.winfo_exists():
            self._refresh_preview_tree_only()
    
    def _row_display_values(self, row: dict) -> tuple:
        """한 행의 표시용 values (container_no는 접미사 옵션 적용)."""
        out = []
        for col in PREVIEW_COLUMNS:
            key = col[0]
            if key == 'container_no':
                out.append(self._format_container_display(row.get(key, '')))
            else:
                out.append(row.get(key, ''))
        return tuple(out)
    
    def _refresh_preview_tree_only(self) -> None:
        """미리보기 테이블만 현재 preview_data로 갱신 (요약/버튼/팝업 없음). 파싱 중 실시간 표시용."""
        if not getattr(self, 'tree', None) or not self.tree.winfo_exists():
            return
        for item in self.tree.get_children():
            self.tree.delete(item)
        if not self.preview_data:
            return
        for idx, row in enumerate(self.preview_data):
            values = self._row_display_values(row)
            tag = 'even' if idx % 2 == 0 else 'odd'
            self.tree.insert('', END, values=values, tags=(tag,))

    def _display_preview(self) -> None:
        """미리보기 테이블 표시 — 한 번에가 아니라 순차적으로 행 추가 (보기 편하게)"""
        def _update():
            if not self.tree:
                return
            self._push_preview_to_main()
            for item in self.tree.get_children():
                self.tree.delete(item)
            if not self.preview_data:
                self._update_summary()
                return
            # 행을 한꺼번에 넣지 않고 짧은 간격으로 순차 삽입
            delay_ms = 25
            data = list(self.preview_data)
            columns = PREVIEW_COLUMNS

            def _insert_row(idx: int) -> None:
                if idx >= len(data) or not self.tree.winfo_exists():
                    self._update_summary()
                    if self.preview_data and self._has_required_docs():
                        self.btn_upload.config(state='normal')
                    else:
                        self.btn_upload.config(state='disabled')
                    if self.preview_data:
                        self.btn_excel.config(state='normal')
                    return
                row = data[idx]
                values = self._row_display_values(row)
                tag = 'even' if idx % 2 == 0 else 'odd'
                self.tree.insert('', END, values=values, tags=(tag,))
                if self.dialog and self.dialog.winfo_exists():
                    self.dialog.after(delay_ms, lambda i=idx + 1: _insert_row(i))

            _insert_row(0)

        if self.dialog and self.dialog.winfo_exists():
            self.dialog.after(0, _update)
    
    def _has_required_docs(self) -> bool:
        """필수 서류 3종(Packing List, Invoice, B/L)이 모두 선택·파싱되었는지 확인"""
        for doc_type, _name, required in DOC_TYPES:
            if required and doc_type not in self.file_paths:
                return False
        return True
    
    def _update_summary(self) -> None:
        """합계행"""
        if not self.preview_data:
            self.summary_var.set("")
            return
        
        containers = set(r['container_no'] for r in self.preview_data if r['container_no'])
        total_tb = 0
        total_net = 0.0
        total_gross = 0.0
        
        for r in self.preview_data:
            try:
                total_tb += int(r.get('mxbg_pallet', '10')) if r.get('mxbg_pallet', '') else 0
            except (ValueError, TypeError) as _e:
                logger.debug(f"onestop_inbound: {_e}")
            try:
                total_net += safe_float(r['net_weight']) if r['net_weight'] else 0
            except (ValueError, TypeError) as _e:
                logger.debug(f"onestop_inbound: {_e}")
            try:
                total_gross += safe_float(r['gross_weight']) if r['gross_weight'] else 0
            except (ValueError, TypeError) as _e:
                logger.debug(f"onestop_inbound: {_e}")
        
        self.summary_var.set(
            f"합계: {len(self.preview_data)} LOT | "
            f"{len(containers)} 컨테이너 | "
            f"{total_tb} 톤백 | "
            f"Net {total_net:,.0f} kg | "
            f"Gross {total_gross:,.0f} kg"
        )
    
    def _check_parsing_duplicates(self) -> str:
        """파싱 결과에서 입고 중복 여부 검사 (lot_no 기준). 중복이 있으면 안내 문구 반환, 없으면 빈 문자열."""
        if not self.preview_data:
            return ""
        from collections import Counter
        lot_counts = Counter(str(r.get('lot_no', '')).strip() for r in self.preview_data if str(r.get('lot_no', '')).strip())
        dups = [(lot, cnt) for lot, cnt in lot_counts.items() if cnt > 1]
        if not dups:
            return ""
        parts = [f"LOT NO {lot} ({cnt}건)" for lot, cnt in dups[:10]]
        if len(dups) > 10:
            parts.append(f"외 {len(dups) - 10}건")
        return "중복: " + ", ".join(parts)

    # ═══════════════════════════════════════════════════════════
    # DB 업로드
    # ═══════════════════════════════════════════════════════════
    
    def _on_upload(self) -> None:
        """DB 업로드 (v3.8.8: 중복 LOT 사전 경고 + 위젯 안전 처리)"""
        if not self.preview_data:
            return
        # v5.7.0: 필수 3종(PL+FA+BL) 없으면 업로드 차단
        if not self._has_required_docs():
            missing = [name for (dt, name, req) in DOC_TYPES if req and dt not in self.file_paths]
            try:
                from ..utils.custom_messagebox import CustomMessageBox
                CustomMessageBox.showwarning(
                    self.dialog, "필수 서류 누락",
                    "DB 업로드를 하려면 다음 3종 서류가 모두 필요합니다:\n\n"
                    "  • ① Packing List (포장명세서)\n"
                    "  • ② Invoice, FA (송장)\n"
                    "  • ③ Bill of Loading (선하증권)\n\n"
                    f"누락: {', '.join(missing)}\n\n"
                    "Delivery Order(인도지시서)는 선택사항이며, 나중에 [📋 D/O 후속 연결] 메뉴로 보충할 수 있습니다."
                )
            except (ImportError, ModuleNotFoundError):
                from ..utils.ui_constants import CustomMessageBox
                CustomMessageBox.showwarning(
                    self.dialog,
                    "필수 서류 누락",
                    "Packing List, Invoice/FA, Bill of Loading 3종 모두 필요합니다."
                )
            return

        # v3.8.8: 중복 LOT 사전 체크
        dup_lots = []
        if hasattr(self.engine, '_check_lot_exists') or hasattr(self.engine, 'db'):
            try:
                db = getattr(self.engine, 'db', None)
                if db:
                    for row in self.preview_data:
                        lot_no = row.get('lot_no', '')
                        if lot_no:
                            existing = db.fetchone(
                                "SELECT 1 FROM inventory WHERE lot_no = ?", (lot_no,))
                            if existing:
                                dup_lots.append(lot_no)
            except (sqlite3.OperationalError, sqlite3.IntegrityError, OSError) as e:
                logger.debug(f"중복 체크 오류: {e}")
        
        if dup_lots:
            try:
                from ..utils.custom_messagebox import CustomMessageBox
                dup_msg = ', '.join(dup_lots[:5])
                if len(dup_lots) > 5:
                    dup_msg += f" 외 {len(dup_lots) - 5}건"
                ok = CustomMessageBox.askyesno(
                    self.dialog, "⚠️ 중복 LOT 경고",
                    f"다음 {len(dup_lots)}개 LOT가 이미 DB에 존재합니다:\n\n"
                    f"{dup_msg}\n\n"
                    f"중복 LOT는 건너뛰고 나머지만 입고합니다.\n계속하시겠습니까?"
                )
            except (ImportError, ModuleNotFoundError):
                ok = msgbox.askyesno("⚠️ 중복 LOT 경고",
                    f"{len(dup_lots)}개 LOT 중복! 건너뛰고 계속?")
            if not ok:
                return
        
        try:
            from ..utils.custom_messagebox import CustomMessageBox
            ok = CustomMessageBox.askyesno(
                self.dialog, "DB 업로드 확인",
                f"{len(self.preview_data)}개 LOT를 데이터베이스에 저장합니다.\n\n"
                f"이 작업은 되돌릴 수 없습니다.\n계속하시겠습니까?"
            )
        except (ImportError, ModuleNotFoundError):
            ok = msgbox.askyesno("DB 업로드 확인",
                f"{len(self.preview_data)}개 LOT 저장?")
        
        if not ok:
            return
        
        # v3.8.8: 위젯 존재 확인 후 비활성화
        try:
            if self.btn_upload and self.btn_upload.winfo_exists():
                self.btn_upload.config(state='disabled')
            if self.btn_excel and self.btn_excel.winfo_exists():
                self.btn_excel.config(state='disabled')
        except (RuntimeError, ValueError) as _e:
            logger.debug(f'Suppressed: {_e}')
        
        self._show_progress_inline()
        thread = threading.Thread(target=self._upload_thread, daemon=True)
        thread.start()
    
    def _upload_thread(self) -> None:
        """백그라운드 DB 업로드"""
        try:
            self._update_progress(0, "📤 DB 업로드 시작...")
            
            pl = self.parsed_results.get('packing_list')
            invoice = self.parsed_results.get('invoice')
            bl = self.parsed_results.get('bl')
            do = self.parsed_results.get('do')
            
            if not pl or not getattr(pl, 'lots', None):
                self._update_progress(0, "❌ Packing List 없음")
                self._enable_buttons()
                return
            # v5.7.0: 필수 3종(PL+FA+BL) 없으면 업로드 중단 — 중복/부분 업로드 방지
            if not invoice:
                self._update_progress(0, "❌ FA(송장) 필수 — 3종(PL+FA+BL) 모두 필요")
                self._enable_buttons()
                return
            if not bl:
                self._update_progress(0, "❌ B/L(선하증권) 필수 — 3종(PL+FA+BL) 모두 필요")
                self._enable_buttons()
                return

            success, failed_rows = self._save_to_db(pl, invoice, bl, do)
            
            if success:
                total = len(self.preview_data)
                self._update_progress(100, f"✅ 업로드 완료: {total} LOT")
                self._log_safe(f"✅ DB 업로드 완료: {total} LOT")
                self.upload_success = True
                self._show_success_and_close(total)
            else:
                self._update_progress(0, "❌ 업로드 실패")
                
                # v4.2.1: 상세 실패 팝업 표시 (실패한 행 번호 포함)
                try:
                    from ..utils.upload_error_dialog import show_upload_error_dialog
                    from ..utils.upload_error_template import UploadErrorTemplate
                    
                    rows_for_msg = failed_rows if failed_rows else [{'row': '?', 'value': '업로드 실패', 'column': ''}]
                    error_msg = UploadErrorTemplate.format_multiple_errors(
                        errors=[{
                            'type': 'missing_required',
                            'rows': rows_for_msg
                        }],
                        total_rows=len(self.preview_data)
                    )
                    
                    show_upload_error_dialog(
                        self.dialog,
                        "입고 업로드 실패",
                        error_msg
                    )
                except (ImportError, Exception) as _e:
                    # 팝업 실패 시 기존 방식으로 표시
                    from ..utils.ui_constants import CustomMessageBox
                    CustomMessageBox.showerror(
                        self.dialog,
                        "업로드 실패",
                        "입고 처리 중 오류가 발생했습니다.\n로그를 확인하세요."
                    )
                
                self._enable_buttons()
        
        except (ValueError, TypeError, AttributeError) as e:
            self._update_progress(0, f"❌ 오류: {e}")
            self._log_safe(f"❌ 업로드 오류: {e}")
            logger.error(f"업로드 오류: {e}", exc_info=True)
            
            # v4.2.1: 예외 발생 시에도 상세 팝업 표시
            try:
                from ..utils.upload_error_dialog import show_upload_error_dialog
                from ..utils.upload_error_template import UploadErrorTemplate
                
                error_msg = UploadErrorTemplate.format_multiple_errors(
                    errors=[{
                        'type': 'file_format',
                        'rows': [{'row': '?', 'value': str(e), 'column': ''}]
                    }],
                    total_rows=len(self.preview_data) if hasattr(self, 'preview_data') else 0
                )
                
                show_upload_error_dialog(
                    self.dialog,
                    "입고 처리 오류",
                    error_msg
                )
            except (ImportError, Exception):
                # 팝업 실패 시 기존 방식
                from ..utils.ui_constants import CustomMessageBox
                CustomMessageBox.showerror(
                    self.dialog,
                    "오류",
                    f"입고 처리 오류:\n{e}"
                )
            
            self._enable_buttons()
    
    def _save_to_db(self, pl, invoice, bl, do):
        """engine.process_inbound를 LOT별로 호출하여 DB 저장
        
        v3.8.8: 이중 트랜잭션 제거 — process_inbound 내부 트랜잭션에 위임
        각 LOT가 독립적으로 커밋됨. 중복 LOT는 자동 건너뜀.
        
        Returns:
            (success: bool, failed_rows: list) — 실패 시 실패한 행 정보 [{'row': N, 'value': msg, 'column': ''}, ...]
        """
        try:
            if not hasattr(self.engine, 'process_inbound'):
                self._log_safe("❌ engine.process_inbound 메서드 없음")
                return False, []

            _lots = getattr(pl, 'lots', []) or []
            total = len(_lots)
            if total == 0:
                return False, []

            created_lots = []
            skipped_lots = []
            errors = []
            failed_rows = []  # 업로드 실패 요약용: [{'row': Excel행, 'value': 오류메시지, 'column': ''}, ...]
            
            for idx, lot in enumerate(_lots):
                pct = 10 + int(80 * (idx + 1) / total)
                lot_no = getattr(lot, 'lot_no', '') or ''
                self._update_progress(pct, f"📦 LOT {idx+1}/{total}: {lot_no}")
                
                # v3.8.8: 중복 LOT 건너뛰기 (에러 대신 skip)
                if lot_no:
                    try:
                        existing = self.engine.db.fetchone(
                            "SELECT 1 FROM inventory WHERE lot_no = ?", (lot_no,))
                        if existing:
                            self._log_safe(f"  ⏭ LOT {lot_no}: 이미 존재 (건너뜀)")
                            skipped_lots.append(lot_no)
                            continue
                    except (sqlite3.OperationalError, sqlite3.IntegrityError, OSError) as _e:
                        logger.debug(f'Suppressed: {_e}')
                
                # PackingData 호환 dict
                # v5.7.5: TONBAG(tonbag_count) 빈 값 허용 — 미입력 시 mxbg_pallet 사용
                _tonbag = getattr(lot, 'tonbag_count', None)
                if _tonbag is None or (isinstance(_tonbag, str) and str(_tonbag).strip() == ''):
                    _tonbag = getattr(lot, 'mxbg_pallet', 10) or 10
                try:
                    _tonbag = int(float(_tonbag))
                except (TypeError, ValueError):
                    _tonbag = getattr(lot, 'mxbg_pallet', 10) or 10
                # v3.8.8: free_time 계산 (None 안전 처리)
                # ★ v5.8.7: D/O 없으면 preview_data에서 사용자 입력값 사용
                # ★ v5.8.8: arrival_date가 날짜가 아닌 값(예: '광양')이면 비움 — ARRIVAL 컬럼 혼동 방지
                _arrival_raw = getattr(do, 'arrival_date', None) if do else None
                _arrival = str(_arrival_raw) if _arrival_raw and str(_arrival_raw) != 'None' else ''
                if _arrival:
                    _a10 = (_arrival[:10] if len(_arrival) >= 10 else _arrival)
                    if not (len(_a10) == 10 and _a10.count('-') == 2 and _a10.replace('-', '').isdigit()):
                        _arrival = ''
                _free_time = 0
                _free_time_date = ''
                
                # D/O 없는 경우: preview_data에서 사용자 입력값 가져오기
                if not _arrival and self.preview_data and idx < len(self.preview_data):
                    _user_arr = self.preview_data[idx].get('arrival_date', '')
                    if _user_arr:
                        _ua = str(_user_arr)[:10]
                        if len(_ua) == 10 and _ua.count('-') == 2 and _ua.replace('-', '').isdigit():
                            _arrival = _ua
                    _user_ft = self.preview_data[idx].get('free_time', '')
                    if _user_ft:
                        try:
                            _free_time = int(_user_ft)
                        except (ValueError, TypeError):
                            pass
                    _user_ship = self.preview_data[idx].get('ship_date', '')
                
                if do:
                    _free_time_date = str(getattr(do, 'free_time_date', '') or '')
                    if not _free_time_date:
                        # free_time_info 리스트에서 추출
                        ft_infos = getattr(do, 'free_time_info', []) or []
                        for ft in ft_infos:
                            ftd = getattr(ft, 'free_time_date', '') or (ft.get('free_time_date', '') if isinstance(ft, dict) else '')
                            if ftd:
                                _free_time_date = str(ftd)
                                break
                    
                    # 계산: free_time_date - arrival_date
                    if _free_time_date and _arrival:
                        try:
                            _ft_dt = datetime.strptime(str(_free_time_date)[:10], '%Y-%m-%d').date()
                            _arr_dt = datetime.strptime(str(_arrival)[:10], '%Y-%m-%d').date()
                            _free_time = (_ft_dt - _arr_dt).days
                            if _free_time < 0:
                                _free_time = 0
                        except (ValueError, TypeError):
                            _free_time = 0
                
                packing_dict = {
                    'lot_no': getattr(lot, 'lot_no', '') or '',
                    'lot_sqm': getattr(lot, 'lot_sqm', '') or '',
                    'sap_no': getattr(pl, 'sap_no', '') or (getattr(invoice, 'sap_no', '') if invoice else '') or '',
                    'bl_no': self._format_bl(
                        (getattr(bl, 'bl_no', '') if bl else '') or 
                        (getattr(do, 'bl_no', '') if do else '') or ''
                    ),
                    'container_no': getattr(lot, 'container_no', '') or '',
                    'product': getattr(pl, 'product', '') or 'LITHIUM CARBONATE',
                    'product_code': getattr(pl, 'code', '') or '',
                    'net_weight': getattr(lot, 'net_weight_kg', 0) or 0,
                    'gross_weight': getattr(lot, 'gross_weight_kg', 0) or 0,
                    'mxbg_pallet': getattr(lot, 'mxbg_pallet', 10) or 10,
                    'tonbag_count': _tonbag,
                    'salar_invoice_no': getattr(invoice, 'salar_invoice_no', '') if invoice else '',
                    'ship_date': self._date_str(getattr(bl, 'ship_date', None) if bl else None) or self._date_str(getattr(invoice, 'invoice_date', None) if invoice else None) or '',
                    'arrival_date': _arrival,
                    'free_time': _free_time,
                    'free_time_date': _free_time_date,
                    'con_return': _free_time_date[:10] if _free_time_date else '',  # DB 컬럼명과 동일 키로 전달
                    'warehouse': str(getattr(do, 'warehouse', DEFAULT_WAREHOUSE)) if do else DEFAULT_WAREHOUSE,
                    'vessel': getattr(pl, 'vessel', '') or '',
                }
                
                # 필수 컬럼 검사 — 어떤 데이터가 빠졌는지 명확히 수집
                missing_display = []
                if not (str(packing_dict.get('lot_no', '') or '').strip()):
                    missing_display.append('LOT NO')
                if not (str(packing_dict.get('product', '') or '').strip()):
                    missing_display.append('PRODUCT')
                try:
                    nw = packing_dict.get('net_weight', 0)
                    if nw is None or (isinstance(nw, (int, float)) and float(nw) <= 0):
                        missing_display.append('NET(Kg)')
                except (TypeError, ValueError):
                    missing_display.append('NET(Kg)')
                try:
                    mx = packing_dict.get('mxbg_pallet', 0)
                    if mx is None or (isinstance(mx, (int, float)) and int(float(mx)) <= 0):
                        missing_display.append('MXBG')
                except (TypeError, ValueError):
                    missing_display.append('MXBG')
                if missing_display:
                    failed_rows.append({
                        'row': idx + 2,
                        'value': '비어 있음',
                        'column': ', '.join(missing_display),
                        'missing_columns': missing_display,
                    })
                    errors.append(f"행 {idx + 2}: {', '.join(missing_display)} 누락")
                    continue
                
                # invoice_data dict
                inv_dict = None
                if invoice:
                    inv_dict = {
                        'sap_no': getattr(invoice, 'sap_no', '') or '',
                        'salar_invoice_no': getattr(invoice, 'salar_invoice_no', '') or '',
                        'invoice_date': str(getattr(invoice, 'invoice_date', '')) if getattr(invoice, 'invoice_date', None) else '',
                    }
                
                # bl_data dict
                bl_dict = None
                if bl:
                    bl_dict = {
                        'bl_no': self._format_bl(getattr(bl, 'bl_no', '') or ''),
                        'ship_date': self._date_str(getattr(bl, 'ship_date', None)) or self._date_str(getattr(bl, 'shipped_date', None)) or '',
                        'vessel': getattr(bl, 'vessel', '') or '',
                    }
                
                # do_data dict (arrival_date=선박 입항일, free_time_date=con_return=컨테이너 반납일)
                do_dict = None
                if do:
                    _con_return = ''
                    ft_infos = getattr(do, 'free_time_info', []) or []
                    for ft in ft_infos:
                        ftd = getattr(ft, 'free_time_date', '') or (ft.get('free_time_date', '') if isinstance(ft, dict) else '')
                        if ftd:
                            _con_return = str(ftd)[:10]
                            break
                    if _con_return:
                        print(f"[원스톱 업로드] DO에서 con_return(반납일) 사용: {_con_return} — DB CON RETURN 컬럼에 저장")
                    else:
                        print(f"[원스톱 업로드] DO에 반납일 없음 — CON RETURN 빈 값으로 저장됨 (free_time_info 항목 수: {len(ft_infos)})")
                    _do_arr = getattr(do, 'arrival_date', None)
                    _do_arrival = (_do_arr.isoformat() if hasattr(_do_arr, 'isoformat') else str(_do_arr or '')) if _do_arr and str(_do_arr) != 'None' else ''
                    do_dict = {
                        'bl_no': str(getattr(do, 'bl_no', '') or ''),
                        'arrival_date': _do_arrival,
                        'free_time_date': _con_return,  # con_return: D/O Free_Time 컬럼(반납일)
                        'free_time': str(getattr(do, 'free_time', '') or ''),
                        'warehouse': str(getattr(do, 'warehouse', '') or ''),
                    }
                
                try:
                    result = self.engine.process_inbound(
                        packing_data=packing_dict,
                        invoice_data=inv_dict,
                        bl_data=bl_dict,
                        do_data=do_dict
                    )
                    
                    if result.get('success'):
                        created_lots.append(getattr(lot, "lot_no", ""))
                    else:
                        err_msg = result.get('message', '') or ', '.join(result.get('errors', []))
                        errors.append(f"LOT {getattr(lot, 'lot_no', '')}: {err_msg}")
                        # Excel 행 번호: 헤더 1행 + 데이터는 2부터 (idx 0 → 행 2)
                        failed_rows.append({'row': idx + 2, 'value': err_msg, 'column': 'LOT NO'})
                
                except (ValueError, TypeError, AttributeError) as e:
                    errors.append(f"LOT {getattr(lot, 'lot_no', '')}: {e}")
                    failed_rows.append({'row': idx + 2, 'value': str(e), 'column': 'LOT NO'})
            
            if errors:
                self._log_safe(f"⚠️ 일부 오류: {len(errors)}건")
                for e in errors[:5]:
                    self._log_safe(f"  - {e}")
            
            if skipped_lots:
                self._log_safe(f"⏭ 중복 건너뜀: {len(skipped_lots)}건")
            
            # v3.8.8: 각 LOT는 process_inbound 내부에서 개별 트랜잭션 처리됨
            # 외부 commit/rollback 불필요
            if created_lots:
                self._log_safe(f"✅ 저장 완료: {len(created_lots)}건")
                return True, []
            else:
                self._log_safe(f"❌ 저장된 LOT 없음 (오류 {len(errors)}건, 건너뜀 {len(skipped_lots)}건)")
                return False, failed_rows

        except (sqlite3.OperationalError, sqlite3.IntegrityError, OSError) as e:
            logger.error(f"DB 저장 실패: {e}", exc_info=True)
            self._log_safe(f"❌ DB 저장 실패: {e}")
            return False, []
    
    # ═══════════════════════════════════════════════════════════
    # Excel 내보내기
    # ═══════════════════════════════════════════════════════════
    
    def _export_to_excel(self) -> None:
        """미리보기 데이터 Excel 내보내기"""
        if not self.preview_data:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            save_path = filedialog.asksaveasfilename(
                parent=self.dialog, title="Excel 내보내기",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"입고미리보기_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            if not save_path:
                return
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "입고 미리보기"
            
            headers = [col[1] for col in PREVIEW_COLUMNS]
            hfill = PatternFill(start_color="2c6fbb", end_color="2c6fbb", fill_type="solid")
            hfont = Font(color="FFFFFF", bold=True, size=10)
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))
            
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.fill = hfill
                cell.font = hfont
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            
            for ri, row_data in enumerate(self.preview_data, 2):
                for ci, (col_id, _, _, _) in enumerate(PREVIEW_COLUMNS, 1):
                    cell = ws.cell(row=ri, column=ci, value=row_data.get(col_id, ''))
                    cell.border = border
            
            for ci, (_, h, w, _) in enumerate(PREVIEW_COLUMNS, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(w / 7, len(h) + 2)
            
            wb.save(save_path)
            self._log_safe(f"📥 Excel 저장: {save_path}")
        
        except (ValueError, TypeError, AttributeError) as e:
            self._log_safe(f"❌ Excel 오류: {e}")
    
    # ═══════════════════════════════════════════════════════════
    # 유틸리티
    # ═══════════════════════════════════════════════════════════
    
    def _show_success_and_close(self, count: int):
        def _close():
            if self.dialog and self.dialog.winfo_exists():
                _app = self.app if self.app else None
                # v5.8.9: 파싱 결과 팝업에서 DB 업로드 선택 시, 완료 후 엑셀 내보내기 여부 질의
                if getattr(self, '_ask_excel_after_upload', False):
                    self._ask_excel_after_upload = False
                    try:
                        from ..utils.custom_messagebox import CustomMessageBox
                        if CustomMessageBox.askyesno(self.dialog, "엑셀 내보내기",
                            "DB 업로드가 완료되었습니다.\n엑셀 내보내기도 하시겠습니까?\n(아니오를 누르면 여기서 종료합니다.)"):
                            self._export_to_excel()
                    except (ImportError, ModuleNotFoundError):
                        pass
                _deferred = getattr(self, '_do_deferred', False)
                _msg = f"✅ {count}개 LOT 저장 완료"
                if _deferred:
                    _msg += "\n\n📋 D/O가 미첨부된 상태입니다.\n나중에 D/O를 받으면 [📋 D/O 후속 연결] 메뉴에서\narrival_date와 Free Time을 업데이트할 수 있습니다."
                try:
                    from ..utils.custom_messagebox import CustomMessageBox
                    CustomMessageBox.showinfo(self.dialog, "업로드 완료", _msg)
                except (ImportError, ModuleNotFoundError):
                    CustomMessageBox.info(None, "완료", _msg)
                self.dialog.destroy()
                
                # v3.8.9: 업로드 후 재고리스트 탭 이동 + 자동 새로고침
                # dialog.destroy() 후이므로 app.root.after 사용
                if _app:
                    try:
                        _root = getattr(_app, 'root', None)
                        if _root:
                            if hasattr(_app, '_set_parsing_preview_data'):
                                _app._set_parsing_preview_data(None)
                            if hasattr(_app, 'notebook') and hasattr(_app, 'tab_inventory'):
                                _root.after(200, lambda: _app.notebook.select(_app.tab_inventory))
                            if hasattr(_app, '_refresh_inventory'):
                                _root.after(500, _app._refresh_inventory)
                                logger.info("[onestop] 재고 새로고침 예약 완료 (500ms)")
                    except (RuntimeError, ValueError) as e:
                        logger.debug(f"재고 새로고침 호출 실패: {e}")
        
        if self.dialog and self.dialog.winfo_exists():
            self.dialog.after(100, _close)
    
    def _enable_buttons(self) -> None:
        def _u():
            try:
                if self.btn_upload and self.btn_upload.winfo_exists():
                    self.btn_upload.config(state='normal')
                if self.btn_excel and self.btn_excel.winfo_exists():
                    self.btn_excel.config(state='normal')
            except (RuntimeError, ValueError) as _e:
                logger.debug(f'Suppressed: {_e}')
        if self.dialog and self.dialog.winfo_exists():
            self.dialog.after(0, _u)
    
    def _enable_parse_btn(self):
        def _u():
            if self.btn_parse:
                self.btn_parse.config(state='normal')
        if self.dialog and self.dialog.winfo_exists():
            self.dialog.after(0, _u)
    
    def _on_cancel(self):
        self._clear_preview_from_main()
        if self.dialog:
            self.dialog.destroy()
    
    def _log_safe(self, msg: str):
        try:
            if self._log:
                if self.dialog and self.dialog.winfo_exists():
                    self.dialog.after(0, lambda: self._log(msg))
                else:
                    self._log(msg)
        except (RuntimeError, ValueError):
            logger.info(msg)
