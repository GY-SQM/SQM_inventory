# -*- coding: utf-8 -*-
"""
Claude_sqm_allocation_testdata_generator_v712.py
================================================
SQM v7.1.2 — Allocation 7-Gate 테스트 데이터 자동 생성기

★ 실제 SQM allocation_parser.py alias_patterns 기준으로 작성
  헤더명: LOT_NO / QTY(MT) / SOLD TO / SALE_REF / OUTBOUND_DATE / CONTAINER_NO

생성 시나리오:
  - Normal       : 정상 통과
  - Gate1        : LOT 미존재
  - Gate2        : cargo 총량 초과 (G2-CARGO-EXCEED)
  - Gate4        : 샘플 포함량까지 Allocation 시도
  - Gate5        : 배치 내 동일 LOT 2행 합산 초과 (G5-BATCH-SUM)
  - Gate6        : 가용 톤백 부족 시나리오 설명용
  - Gate7        : random seed 고정 테스트용

출력: ./output_allocation_testdata/*.xlsx  (openpyxl)
      ./output_allocation_testdata/*.csv   (utf-8-sig)

작성: Ruby (Claude) / SQM v7.1.2
"""
from __future__ import annotations

import csv
import os
from dataclasses import dataclass, field, asdict
from datetime import datetime, timedelta
from typing import List, Optional

# openpyxl 있으면 xlsx도 생성
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

OUTPUT_DIR = "./output_allocation_testdata"

# ──────────────────────────────────────────────
# SQM 실제 헤더 (allocation_parser.py alias_patterns 완전 일치)
# ──────────────────────────────────────────────
HEADERS = [
    "LOT_NO",          # lot_no
    "QTY(MT)",         # qty_mt
    "SOLD TO",         # sold_to / customer
    "SALE_REF",        # sale_ref
    "OUTBOUND_DATE",   # outbound_date
    "CONTAINER_NO",    # container_no
    "REMARK",          # remark (자동 무시 컬럼 — SQM alias_patterns 미포함)
]

TODAY = datetime.today()
SHIP_DATE = (TODAY + timedelta(days=10)).strftime("%Y-%m-%d")


@dataclass
class AllocRow:
    lot_no: str
    qty_mt: float
    sold_to: str
    sale_ref: str
    outbound_date: str
    container_no: str
    remark: str = ""

    def to_row(self) -> list:
        return [
            self.lot_no,
            self.qty_mt,
            self.sold_to,
            self.sale_ref,
            self.outbound_date,
            self.container_no,
            self.remark,
        ]


# ──────────────────────────────────────────────
# 시나리오별 데이터
# ──────────────────────────────────────────────

def scenario_normal() -> List[AllocRow]:
    """Gate 통과: 정상 케이스 (cargo=10t, Allocation=8t)"""
    return [
        AllocRow(
            lot_no="1125072147",
            qty_mt=8.000,
            sold_to="CATL",
            sale_ref="SO-NORMAL-001",
            outbound_date=SHIP_DATE,
            container_no="TCKU1234567",
            remark="정상 케이스 — Gate1~7 전부 통과 예상"
        )
    ]


def scenario_gate1_missing_lot() -> List[AllocRow]:
    """Gate1: LOT 미존재 Hard Stop"""
    return [
        AllocRow(
            lot_no="9999999999",       # DB에 없는 LOT
            qty_mt=5.000,
            sold_to="BYD",
            sale_ref="SO-G1-MISS-001",
            outbound_date=SHIP_DATE,
            container_no="MSCU9876543",
            remark="[Gate1] LOT_NOT_FOUND Hard Stop 예상"
        )
    ]


def scenario_gate2_cargo_exceed() -> List[AllocRow]:
    """Gate2: cargo_weight_sum 초과 (G2-CARGO-EXCEED)
    cargo=10t, Allocation=10.001t → Hard Stop
    """
    return [
        AllocRow(
            lot_no="1125072148",
            qty_mt=10.001,             # cargo(10.000t) 초과
            sold_to="LG",
            sale_ref="SO-G2-OVER-001",
            outbound_date=SHIP_DATE,
            container_no="HLCU1111111",
            remark="[Gate2] G2-CARGO-EXCEED — cargo=10t, Alloc=10.001t"
        )
    ]


def scenario_gate4_sample_violation() -> List[AllocRow]:
    """Gate4: 샘플(1kg=0.001t) 포함량까지 Allocation 시도
    total=10.001t를 cargo인 줄 알고 입력하는 케이스
    """
    return [
        AllocRow(
            lot_no="1125072149",
            qty_mt=10.001,             # total(cargo+sample) 그대로 입력 오류
            sold_to="CATL",
            sale_ref="SO-G4-SMPL-001",
            outbound_date=SHIP_DATE,
            container_no="YMLU2222222",
            remark="[Gate4] 샘플 포함 총량 입력 오류 — cargo=10.000 초과"
        )
    ]


def scenario_gate5_duplicate_lot() -> List[AllocRow]:
    """Gate5: 배치 내 동일 LOT 2행 합산 초과 (G5-BATCH-SUM)
    LOT-A 6t + LOT-A 5t = 11t > cargo(10t) → Hard Stop
    """
    return [
        AllocRow(
            lot_no="1125072150",
            qty_mt=6.000,
            sold_to="CATL",
            sale_ref="SO-G5-DUP-001",
            outbound_date=SHIP_DATE,
            container_no="TCKU3333333",
            remark="[Gate5] 동일 LOT 1번째 행 — 6t"
        ),
        AllocRow(
            lot_no="1125072150",       # 같은 LOT
            qty_mt=5.000,              # 합계 11t > cargo(10t)
            sold_to="BYD",
            sale_ref="SO-G5-DUP-002",
            outbound_date=SHIP_DATE,
            container_no="TCKU3333334",
            remark="[Gate5] 동일 LOT 2번째 행 — 합계 11t G5-HARD-STOP 예상"
        ),
    ]


def scenario_gate6_selectable_shortage() -> List[AllocRow]:
    """Gate6: selectable pool 부족 안내용
    실제 DB tonbag 상태에 따라 결정되므로 데이터 설명용
    """
    return [
        AllocRow(
            lot_no="1125072151",
            qty_mt=8.000,              # 8개 필요
            sold_to="LG",
            sale_ref="SO-G6-SEL-001",
            outbound_date=SHIP_DATE,
            container_no="MSCU4444444",
            remark="[Gate6] DB AVAILABLE tonbag이 8개 미만이면 Hard Stop"
        )
    ]


def scenario_gate7_random_seed() -> List[AllocRow]:
    """Gate7: random seed 고정 테스트
    SQM reservation_mode='seeded' 로 실행 시 seed 저장 검증
    """
    return [
        AllocRow(
            lot_no="1125072152",
            qty_mt=5.000,
            sold_to="CATL",
            sale_ref="SO-G7-SEED-001",
            outbound_date=SHIP_DATE,
            container_no="HLCU5555555",
            remark="[Gate7] reservation_mode=seeded — audit_log ALLOC_RANDOM_LOG 확인"
        )
    ]


# ──────────────────────────────────────────────
# 저장 함수
# ──────────────────────────────────────────────

def ensure_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)


def save_csv(filename: str, rows: List[AllocRow]) -> str:
    ensure_dir(OUTPUT_DIR)
    path = os.path.join(OUTPUT_DIR, filename)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        for row in rows:
            writer.writerow(row.to_row())
    return path


def save_xlsx(filename: str, rows: List[AllocRow], gate_label: str = "") -> str:
    if not HAS_OPENPYXL:
        return ""
    ensure_dir(OUTPUT_DIR)
    path = os.path.join(OUTPUT_DIR, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Allocation"

    # 헤더 스타일
    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_widths = [18, 10, 14, 20, 14, 14, 40]
    for i, (h, w) in enumerate(zip(HEADERS, col_widths), start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = border
        ws.column_dimensions[chr(64 + i)].width = w
    ws.row_dimensions[1].height = 22

    # 데이터 행
    warn_fill = PatternFill("solid", fgColor="FFF2CC")  # 경고 노란색
    err_fill  = PatternFill("solid", fgColor="FFE0E0")  # 오류 빨간색

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row.to_row(), start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            # Gate2/4/5 오류 행 강조
            if "Gate5" in gate_label or "Gate2" in gate_label or "Gate4" in gate_label:
                if c_idx in (1, 2):  # lot_no, qty_mt 강조
                    cell.fill = err_fill
        ws.row_dimensions[r_idx].height = 18

    # 상단 설명 메모 (Gate 설명)
    if gate_label:
        ws.cell(row=1, column=len(HEADERS) + 1,
                value=f"테스트 목적: {gate_label}").font = Font(italic=True, color="888888")

    wb.save(path)
    return path


# ──────────────────────────────────────────────
# 메인
# ──────────────────────────────────────────────

SCENARIOS = [
    ("allocation_normal",              scenario_normal(),              "정상 통과 — Gate1~7 전부 PASS"),
    ("allocation_gate1_missing_lot",   scenario_gate1_missing_lot(),  "Gate1: LOT_NOT_FOUND Hard Stop"),
    ("allocation_gate2_cargo_exceed",  scenario_gate2_cargo_exceed(), "Gate2: G2-CARGO-EXCEED Hard Stop"),
    ("allocation_gate4_sample_viol",   scenario_gate4_sample_violation(), "Gate4: 샘플 포함량 초과 Hard Stop"),
    ("allocation_gate5_dup_lot",       scenario_gate5_duplicate_lot(), "Gate5: G5-BATCH-SUM Hard Stop"),
    ("allocation_gate6_sel_short",     scenario_gate6_selectable_shortage(), "Gate6: selectable 부족 (DB 상태 의존)"),
    ("allocation_gate7_seed",          scenario_gate7_random_seed(),  "Gate7: random seed 고정 audit_log 확인"),
]


def main() -> None:
    ensure_dir(OUTPUT_DIR)
    print(f"\nSQM v7.1.2 — Allocation 테스트 데이터 생성기")
    print(f"출력 폴더: {os.path.abspath(OUTPUT_DIR)}")
    print("=" * 60)

    for name, rows, label in SCENARIOS:
        csv_path  = save_csv(f"{name}.csv", rows)
        xlsx_path = save_xlsx(f"{name}.xlsx", rows, label) if HAS_OPENPYXL else "(openpyxl 없음)"
        row_cnt = len(rows)
        print(f"  ✅ {name}")
        print(f"     CSV : {csv_path} ({row_cnt}행)")
        if HAS_OPENPYXL:
            print(f"     XLSX: {xlsx_path}")
        print(f"     목적: {label}")

    print(f"\n총 {len(SCENARIOS)}개 시나리오 생성 완료")
    print("\n[SQM 업로드 방법]")
    print("  메뉴 → 출고 → Allocation 입력 → Excel 파일 선택")
    print("  헤더: LOT_NO / QTY(MT) / SOLD TO / SALE_REF / OUTBOUND_DATE")
    print("\n[Gate별 기대 결과]")
    for name, _, label in SCENARIOS:
        print(f"  {name:40s} → {label}")


if __name__ == "__main__":
    main()
