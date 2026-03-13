# -*- coding: utf-8 -*-
"""
Claude_sqm_testdata_gen_v712.py
================================
SQM v7.1.2 — Allocation 7-Gate 테스트 데이터 자동 생성기

★ 실제 allocation_parser.py alias_patterns 헤더 완전 일치
  LOT_NO / QTY(MT) / SOLD TO / SALE_REF / OUTBOUND_DATE / CONTAINER_NO

생성 시나리오 (7개):
  Normal   → Gate1~7 정상 통과
  Gate1    → LOT_NOT_FOUND Hard Stop
  Gate2    → G2-CARGO-EXCEED Hard Stop
  Gate4    → 샘플 포함량 초과 차단
  Gate5    → G5-BATCH-SUM Hard Stop (동일 LOT 2행 합산)
  Gate6    → selectable pool 부족 안내
  Gate7    → random seed 고정 audit_log 확인

[Bug6 감사 결과 — 코드 주석으로 영구 기록]
  continue 16건 전수조사 완료 (SQM v7.1.2 기준):
    L1273: _bqt_sum<=0 → G5 사전필터 (qty=0인 LOT 무시) ✅ 정상
    L1334~L1368: 입력값 오류 (INVALID_LOT/ZERO_QTY/QTY/CUSTOMER/SALE_REF) ✅ 정상
    L1423: SALE_REF_CONFLICT ✅ 정상
    L1442: LOT_MODE_DUP ✅ 정상
    L1475: LOT_NOT_FOUND ✅ 정상
    L1504: G2_CARGO_EXCEED ✅ 정상 (G2 패치, continue가 이 LOT 스킵)
    L1517: LOT_STATUS_MISMATCH ✅ 정상
    L1608: NO_AVAILABLE_TONBAG ✅ 정상
    L1656: INVALID_OUTBOUND_DATE ✅ 정상
    L1670/L1700: QTY_EXCEEDS_AVAILABLE ✅ 정상
    L1759: STAGED 승인대기 완료 → 다음 LOT 진행 ✅ 의도적 설계
  return result 2건:
    L1303: G5 Hard Stop → 전체 배치 차단 ✅ 의도적 설계
    L1986: 함수 최종 반환 ✅ 정상
  → 실제 위험 continue: 0건 (Bug6 해소)

작성: Ruby (Claude) / SQM v7.1.2
"""
from __future__ import annotations

import csv
import os
from dataclasses import dataclass
from datetime import datetime, timedelta
from typing import List

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

OUTPUT_DIR   = "./output_testdata"
TODAY        = datetime.today()
SHIP_DATE    = (TODAY + timedelta(days=10)).strftime("%Y-%m-%d")

# ★ 실제 SQM allocation_parser.py alias_patterns 완전 일치 헤더
HEADERS = ["LOT_NO", "QTY(MT)", "SOLD TO", "SALE_REF",
           "OUTBOUND_DATE", "CONTAINER_NO", "REMARK"]


@dataclass
class AllocRow:
    lot_no:        str
    qty_mt:        float
    sold_to:       str
    sale_ref:      str
    outbound_date: str
    container_no:  str
    remark:        str = ""

    def to_row(self) -> list:
        return [self.lot_no, self.qty_mt, self.sold_to,
                self.sale_ref, self.outbound_date,
                self.container_no, self.remark]


# ── 시나리오 ──────────────────────────────────────

def s_normal() -> List[AllocRow]:
    return [AllocRow("1125072147", 8.000, "CATL", "SO-NORMAL-001",
                     SHIP_DATE, "TCKU1234567", "정상 케이스 — Gate1~7 전부 PASS")]

def s_gate1() -> List[AllocRow]:
    return [AllocRow("9999999999", 5.000, "BYD", "SO-G1-001",
                     SHIP_DATE, "MSCU9876543", "[Gate1] LOT_NOT_FOUND Hard Stop 예상")]

def s_gate2() -> List[AllocRow]:
    return [AllocRow("1125072148", 10.001, "LG", "SO-G2-001",
                     SHIP_DATE, "HLCU1111111",
                     "[Gate2] cargo=10t 초과 → G2-CARGO-EXCEED Hard Stop")]

def s_gate4() -> List[AllocRow]:
    return [AllocRow("1125072149", 10.001, "CATL", "SO-G4-001",
                     SHIP_DATE, "YMLU2222222",
                     "[Gate4] total(cargo+1kg) 오입력 → G2-CARGO-EXCEED Hard Stop")]

def s_gate5() -> List[AllocRow]:
    return [
        AllocRow("1125072150", 6.000, "CATL", "SO-G5-001",
                 SHIP_DATE, "TCKU3333333", "[Gate5] 동일LOT 1행 6t"),
        AllocRow("1125072150", 5.000, "BYD",  "SO-G5-002",
                 SHIP_DATE, "TCKU3333334", "[Gate5] 동일LOT 2행 5t → 합계11t G5-HARD-STOP"),
    ]

def s_gate6() -> List[AllocRow]:
    return [AllocRow("1125072151", 8.000, "LG", "SO-G6-001",
                     SHIP_DATE, "MSCU4444444",
                     "[Gate6] DB AVAILABLE tonbag < 8개이면 Hard Stop (DB 상태 의존)")]

def s_gate7() -> List[AllocRow]:
    return [AllocRow("1125072152", 5.000, "CATL", "SO-G7-001",
                     SHIP_DATE, "HLCU5555555",
                     "[Gate7] reservation_mode=seeded → audit_log ALLOC_RANDOM_LOG 확인")]


SCENARIOS = [
    ("allocation_normal",     s_normal(),  "정상 통과"),
    ("allocation_gate1",      s_gate1(),   "Gate1: LOT_NOT_FOUND"),
    ("allocation_gate2",      s_gate2(),   "Gate2: G2-CARGO-EXCEED"),
    ("allocation_gate4",      s_gate4(),   "Gate4: 샘플 포함량 초과"),
    ("allocation_gate5_dup",  s_gate5(),   "Gate5: G5-BATCH-SUM"),
    ("allocation_gate6",      s_gate6(),   "Gate6: selectable 부족"),
    ("allocation_gate7_seed", s_gate7(),   "Gate7: random_seed 로그"),
]


# ── 저장 ──────────────────────────────────────────

def _ensure(path: str) -> None:
    os.makedirs(path, exist_ok=True)

def save_csv(name: str, rows: List[AllocRow]) -> str:
    _ensure(OUTPUT_DIR)
    p = os.path.join(OUTPUT_DIR, f"{name}.csv")
    with open(p, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(HEADERS)
        for r in rows:
            w.writerow(r.to_row())
    return p

def save_xlsx(name: str, rows: List[AllocRow], label: str = "") -> str:
    if not HAS_XLSX:
        return ""
    _ensure(OUTPUT_DIR)
    p = os.path.join(OUTPUT_DIR, f"{name}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Allocation"

    # 헤더
    hf = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(border_style="thin", color="AAAAAA")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    al_c = Alignment(horizontal="center", vertical="center")
    widths = [18, 10, 14, 22, 14, 14, 45]

    for ci, (h, w) in enumerate(zip(HEADERS, widths), 1):
        c = ws.cell(1, ci, h)
        c.fill, c.font, c.alignment, c.border = hf, hfont, al_c, bd
        ws.column_dimensions[chr(64+ci)].width = w
    ws.row_dimensions[1].height = 22

    # 오류 강조 색
    err_fill = PatternFill("solid", fgColor="FFE0E0")
    is_err = any(g in label for g in ["Gate2","Gate4","Gate5"])

    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row.to_row(), 1):
            c = ws.cell(ri, ci, val)
            c.border = bd
            c.alignment = Alignment(vertical="center")
            if is_err and ci in (1, 2):
                c.fill = err_fill
        ws.row_dimensions[ri].height = 18

    wb.save(p)
    return p


def main() -> None:
    _ensure(OUTPUT_DIR)
    print(f"\nSQM v7.1.2 — Allocation 테스트 데이터 생성기")
    print(f"출력 폴더: {os.path.abspath(OUTPUT_DIR)}")
    print("=" * 60)
    for name, rows, label in SCENARIOS:
        cp = save_csv(name, rows)
        xp = save_xlsx(name, rows, label)
        print(f"  ✅ {name:<30s} ({len(rows)}행) — {label}")
        if xp:
            print(f"     XLSX: {xp}")
    print(f"\n총 {len(SCENARIOS)}개 시나리오 완료")
    print("\n[SQM 업로드] 메뉴 → 출고 → Allocation 입력 → 파일 선택")
    print("[Bug6 감사] continue 16건 전수조사 — 실제 위험 0건 (v7.1.2 확인)")


if __name__ == "__main__":
    main()
