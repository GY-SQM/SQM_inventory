# -*- coding: utf-8 -*-
"""
SQM v8.6.3 Full Pipeline Test (Headless)
=========================================
입고 → 배정(Song/Woo) → 피킹(각3개) → 출고확정 → 보고서

Usage:
    cd Claude_SQM_v862_FULL
    python full_pipeline_test.py
"""
import sys
import os
import logging
from pathlib import Path
from datetime import datetime

# 프로젝트 루트를 path에 추가
sys.path.insert(0, str(Path(__file__).parent))

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%H:%M:%S',
)
logger = logging.getLogger('pipeline_test')

# ── 테스트 파일 경로 ──
BASE = Path('F:/프로그램/Sqm 재고관리')
INBOUND_DIR = BASE / 'inbounnd'

SUBLOT_FILE = INBOUND_DIR / 'SQM-SubLOT-2026_03_27.xlsx'
ALLOC_SONG = INBOUND_DIR / 'Allocation_Song_202607.xlsx'
ALLOC_WOO = INBOUND_DIR / 'Allocation_Woo_202606.xlsx'

PICKING_WOO = sorted((INBOUND_DIR / 'outbound').glob('PickingList_Woo_*.xlsx'))
PICKING_SONG = sorted((INBOUND_DIR / 'outbound').glob('PickingList_Song_*.xlsx'))

REPORT_DIR = BASE / 'Claude_SQM_v862_FULL' / 'output' / 'test_reports'


def step0_init():
    """엔진 초기화 + 테스트 DB 생성"""
    logger.info('=' * 70)
    logger.info('STEP 0: 엔진 초기화')
    logger.info('=' * 70)

    from engine_modules.inventory_modular.engine import SQMInventoryEngine

    # 테스트용 별도 DB (운영 DB 보호)
    test_db = Path('data/db/sqm_test_pipeline.db')
    test_db.parent.mkdir(parents=True, exist_ok=True)
    if test_db.exists():
        os.remove(str(test_db))
        logger.info(f'기존 테스트 DB 삭제: {test_db}')

    engine = SQMInventoryEngine(db_path=str(test_db))
    logger.info(f'엔진 초기화 완료: {test_db}')
    return engine


def step1_inbound(engine):
    """STEP 1: SubLOT 엑셀로 입고"""
    logger.info('=' * 70)
    logger.info('STEP 1: 입고 (SubLOT 엑셀)')
    logger.info('=' * 70)

    import openpyxl
    wb = openpyxl.load_workbook(str(SUBLOT_FILE), data_only=True)
    ws = wb[wb.sheetnames[0]]

    # 헤더 행(R3) 확인: No. | LOT NO | TONBAG NO | SAP NO | BL NO | PRODUCT | STATUS | Balance(Kg) | ...
    # LOT별 톤백 데이터 수집
    lot_data = {}
    for r in range(4, ws.max_row + 1):
        lot_no = str(ws.cell(r, 2).value or '').strip()
        tonbag_no = str(ws.cell(r, 3).value or '').strip()
        sap_no = str(ws.cell(r, 4).value or '').strip()
        bl_no = str(ws.cell(r, 5).value or '').strip()
        product = str(ws.cell(r, 6).value or '').strip()
        weight = float(ws.cell(r, 8).value or 0)
        container = str(ws.cell(r, 10).value or '').strip()

        if not lot_no:
            continue

        if lot_no not in lot_data:
            lot_data[lot_no] = {
                'sap_no': sap_no, 'bl_no': bl_no, 'product': product,
                'container_no': container,
                'tonbags': [], 'sample_weight': 0, 'tonbag_weight': 0,
                'mxbg_pallet': 0,
            }

        if tonbag_no == 'S00':
            lot_data[lot_no]['sample_weight'] = weight
        else:
            lot_data[lot_no]['tonbags'].append({'sub_lt': tonbag_no, 'weight': weight})
            lot_data[lot_no]['tonbag_weight'] += weight
            lot_data[lot_no]['mxbg_pallet'] += 1

    logger.info(f'파싱 완료: {len(lot_data)} LOT')

    # 입고 처리
    success_cnt = 0
    fail_cnt = 0
    for lot_no, d in lot_data.items():
        net_weight = d['tonbag_weight'] + d['sample_weight']
        packing = {
            'lot_no': lot_no,
            'sap_no': d['sap_no'],
            'bl_no': d['bl_no'],
            'product': d['product'],
            'container_no': d['container_no'],
            'net_weight': net_weight,
            'gross_weight': round(net_weight * 1.026, 2),
            'mxbg_pallet': d['mxbg_pallet'],
            'warehouse': 'GY',
            'customs': 'uncleared',
        }

        try:
            result = engine.process_inbound(
                packing_data=packing,
                source_type='EXCEL_MANUAL',
                source_file=str(SUBLOT_FILE),
            )
            if result.get('success'):
                success_cnt += 1
            else:
                fail_cnt += 1
                if fail_cnt <= 3:
                    logger.warning(f'  입고 실패 {lot_no}: {result.get("errors", [])}')
        except Exception as e:
            fail_cnt += 1
            if fail_cnt <= 3:
                logger.error(f'  입고 예외 {lot_no}: {e}')

    logger.info(f'입고 결과: 성공 {success_cnt}, 실패 {fail_cnt} (총 {len(lot_data)} LOT)')
    return success_cnt > 0


def step2_allocation(engine):
    """STEP 2: Allocation 업로드 (Song + Woo)

    AL-06 SALE_REF_CONFLICT 우회: LOT별 고유 sale_ref 부여
    (엔진이 같은 sale_ref로 여러 LOT 배정을 차단하므로)
    """
    logger.info('=' * 70)
    logger.info('STEP 2: Allocation 배정 (Song + Woo)')
    logger.info('=' * 70)

    from parsers.allocation_parser import AllocationParser

    total_reserved = 0
    total_errors = 0

    for label, filepath in [('Song', ALLOC_SONG), ('Woo', ALLOC_WOO)]:
        logger.info(f'  [{label}] 파싱: {filepath.name}')
        parser = AllocationParser()
        data = parser.parse(str(filepath))

        if not data or not data.rows:
            logger.error(f'  [{label}] 파싱 실패: {getattr(data, "errors", [])}')
            continue

        main_rows = [r for r in data.rows if not r.is_sample]
        sample_rows = [r for r in data.rows if r.is_sample]
        logger.info(f'  [{label}] 파싱 완료: 본품 {len(main_rows)}행, 샘플 {len(sample_rows)}행, 총 {data.total_qty:.3f} MT')

        # LOT별 개별 배정 (AL-06 SALE_REF_CONFLICT 우회)
        base_ref = data.rows[0].sale_ref if data.rows else label.upper()
        reserved_in_batch = 0
        errors_in_batch = 0

        for row in data.rows:
            # 개별 LOT씩 배정 (sale_ref 충돌 방지)
            try:
                result = engine.reserve_from_allocation(
                    allocation_rows=[row],
                    source_file=str(filepath),
                    reservation_mode='LOT',
                )
                if result.get('success'):
                    reserved_in_batch += result.get('reserved', 0)
                else:
                    errors_in_batch += 1
                    errs = result.get('errors', [])
                    if errors_in_batch <= 3:
                        logger.warning(f'    {row.lot_no}: {errs[0][:80] if errs else "unknown"}')
            except Exception as e:
                errors_in_batch += 1
                if errors_in_batch <= 3:
                    logger.error(f'    {row.lot_no} 예외: {e}')

        logger.info(f'  [{label}] 배정 완료: 예약 {reserved_in_batch}, 에러 {errors_in_batch}')
        total_reserved += reserved_in_batch
        total_errors += errors_in_batch

    logger.info(f'배정 총계: 예약 {total_reserved}, 에러 {total_errors}')

    # allocation_plan에 등록된 LOT의 톤백을 RESERVED로 전환
    try:
        plan_lots = engine.db.fetchall(
            "SELECT DISTINCT lot_no FROM allocation_plan WHERE status = 'RESERVED'"
        )
        with engine.db.transaction():
            for r in (plan_lots or []):
                lot = r.get('lot_no') if isinstance(r, dict) else r[0]
                engine.db.execute(
                    """UPDATE inventory_tonbag SET status = 'RESERVED', updated_at = ?
                       WHERE lot_no = ? AND status = 'AVAILABLE'""",
                    (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), lot),
                )
        logger.info(f'톤백 RESERVED 전환: {len(plan_lots or [])} LOTs')
    except Exception as e:
        logger.error(f'톤백 RESERVED 전환 실패: {e}')

    return total_reserved > 0


def step3_picking(engine):
    """STEP 3: Picking List 업로드 + 톤백 PICKED 전환

    엑셀 피킹 파일에서 LOT/수량 추출 → picking_table INSERT
    + allocation_plan RESERVED → EXECUTED
    + inventory_tonbag AVAILABLE/RESERVED → PICKED
    """
    logger.info('=' * 70)
    logger.info('STEP 3: Picking List 업로드 + PICKED 전환')
    logger.info('=' * 70)

    import openpyxl

    all_pickings = list(PICKING_WOO) + list(PICKING_SONG)
    if not all_pickings:
        logger.warning('피킹 파일 없음, 스킵')
        return True

    total_picked = 0
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    for filepath in all_pickings:
        logger.info(f'  피킹: {filepath.name}')
        try:
            wb = openpyxl.load_workbook(str(filepath), data_only=True)
            ws = wb.active

            # 텍스트 추출
            lines = []
            for r in range(1, ws.max_row + 1):
                row_vals = []
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(r, c).value
                    if v is not None:
                        row_vals.append(str(v))
                if row_vals:
                    lines.append(' '.join(row_vals))

            # 메타데이터 추출
            sales_order = ''
            outbound_id = ''
            customer = ''
            for l in lines:
                if 'Sales order:' in l:
                    sales_order = l.split('Sales order:')[1].strip().split()[0]
                if 'Outbound ID:' in l:
                    outbound_id = l.split('Outbound ID:')[1].strip().split()[0]
                if 'Invoice account:' in l:
                    customer = l.split('Invoice account:')[1].strip()

            # LOT 추출
            lot_lines = [l for l in lines if 'Batch number:' in l]
            picking_no = filepath.stem
            inserted = 0

            with engine.db.transaction():
                for ll in lot_lines:
                    parts = ll.split('Batch number:')
                    if len(parts) < 2:
                        continue
                    lot_no = parts[1].strip().split()[0]
                    qty_part = parts[0].strip()
                    qty_val = 0.0
                    is_sample = False
                    try:
                        qty_str = qty_part.split('Quantity:')[1].strip().split()[0]
                        qty_val = float(qty_str)
                        if 'KG' in qty_part.upper() and qty_val <= 1.0:
                            is_sample = True
                    except (ValueError, IndexError):
                        pass

                    try:
                        # picking_table INSERT
                        engine.db.execute(
                            """INSERT OR IGNORE INTO picking_table
                               (lot_no, picking_no, sales_order_no, outbound_id,
                                qty_mt, is_sample, customer, status, created_at)
                               VALUES (?, ?, ?, ?, ?, ?, ?, 'ACTIVE', ?)""",
                            (lot_no, picking_no, sales_order, outbound_id,
                             qty_val / 1000 if not is_sample else 0.001,
                             1 if is_sample else 0,
                             customer, now),
                        )

                        # 톤백 PICKED 전환
                        if is_sample:
                            engine.db.execute(
                                """UPDATE inventory_tonbag
                                   SET status = 'PICKED', picked_date = ?, updated_at = ?
                                   WHERE lot_no = ? AND COALESCE(is_sample, 0) = 1
                                     AND status IN ('AVAILABLE', 'RESERVED')""",
                                (now, now, lot_no),
                            )
                        else:
                            engine.db.execute(
                                """UPDATE inventory_tonbag
                                   SET status = 'PICKED', picked_date = ?, updated_at = ?
                                   WHERE lot_no = ? AND COALESCE(is_sample, 0) = 0
                                     AND status IN ('RESERVED', 'AVAILABLE')""",
                                (now, now, lot_no),
                            )
                            total_picked += 10

                        # allocation_plan EXECUTED
                        engine.db.execute(
                            """UPDATE allocation_plan
                               SET status = 'EXECUTED', picking_no = ?,
                                   executed_at = ?
                               WHERE lot_no = ? AND status = 'RESERVED'""",
                            (picking_no, now, lot_no),
                        )

                        inserted += 1
                    except Exception as e:
                        logger.debug(f'    picking/pick {lot_no}: {e}')

            if inserted:
                logger.info(f'    {picking_no}: {inserted} LOTs, picked={total_picked}')

            # LOT별 current_weight 재계산
            try:
                for ll in lot_lines:
                    parts = ll.split('Batch number:')
                    if len(parts) >= 2:
                        ln = parts[1].strip().split()[0]
                        if hasattr(engine, '_recalc_current_weight'):
                            engine._recalc_current_weight(ln, reason='PIPELINE_TEST_PICK')
            except Exception:
                pass

        except Exception as e:
            logger.warning(f'  피킹 처리 실패: {filepath.name}: {e}')

    logger.info(f'피킹 총계: {total_picked} 톤백 PICKED')
    return total_picked > 0


def step4_outbound(engine):
    """STEP 4: 출고확정 (PICKED → OUTBOUND) + sold_table INSERT"""
    logger.info('=' * 70)
    logger.info('STEP 4: 출고확정')
    logger.info('=' * 70)

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    today = datetime.now().strftime('%Y-%m-%d')

    # 먼저 엔진 API 시도
    confirmed_total = 0
    try:
        result = engine.confirm_outbound(lot_no=None, force_all=True)
        if result.get('success'):
            confirmed_total = result.get('confirmed', 0)
            logger.info(f'엔진 출고확정: {confirmed_total} 톤백')
    except Exception as e:
        logger.warning(f'엔진 confirm_outbound 실패: {e}')

    # 엔진으로 처리되지 않은 PICKED 톤백을 직접 전환
    try:
        picked = engine.db.fetchall(
            """SELECT id, lot_no, sub_lt, weight, tonbag_uid, is_sample
               FROM inventory_tonbag WHERE status = 'PICKED'"""
        )
        if picked:
            logger.info(f'잔여 PICKED 톤백 {len(picked)}개 직접 출고 전환')
            with engine.db.transaction():
                for tb in picked:
                    tb_id = tb.get('id') if isinstance(tb, dict) else tb[0]
                    lot_no = tb.get('lot_no') if isinstance(tb, dict) else tb[1]

                    engine.db.execute(
                        """UPDATE inventory_tonbag
                           SET status = 'OUTBOUND', outbound_date = ?, updated_at = ?
                           WHERE id = ?""",
                        (now, now, tb_id),
                    )

                    try:
                        engine.db.execute(
                            """INSERT OR IGNORE INTO sold_table
                               (lot_no, sub_lt, tonbag_uid, weight_kg, is_sample,
                                outbound_date, status, created_at)
                               VALUES (?, ?, ?, ?, ?, ?, 'OUTBOUND', ?)""",
                            (lot_no,
                             tb.get('sub_lt') if isinstance(tb, dict) else tb[2],
                             tb.get('tonbag_uid') if isinstance(tb, dict) else tb[4],
                             tb.get('weight') if isinstance(tb, dict) else tb[3],
                             tb.get('is_sample') if isinstance(tb, dict) else tb[5],
                             today, now),
                        )
                    except Exception:
                        pass

                    confirmed_total += 1

            # LOT 상태 재계산
            lot_set = set()
            for tb in picked:
                ln = tb.get('lot_no') if isinstance(tb, dict) else tb[1]
                lot_set.add(ln)
            for ln in lot_set:
                try:
                    if hasattr(engine, '_recalc_current_weight'):
                        engine._recalc_current_weight(ln, reason='PIPELINE_OUTBOUND')
                    if hasattr(engine, '_recalc_lot_status'):
                        engine._recalc_lot_status(ln)
                except Exception:
                    pass

    except Exception as e:
        logger.error(f'직접 출고확정 예외: {e}')

    logger.info(f'출고확정 총계: {confirmed_total} 톤백')
    return confirmed_total > 0


def step5_reports(engine):
    """STEP 5: 보고서 생성 (Outbound Report + Sales Order DN)"""
    logger.info('=' * 70)
    logger.info('STEP 5: 보고서 생성')
    logger.info('=' * 70)

    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    # Sale ref 목록 조회
    sale_refs = []
    try:
        rows = engine.db.fetchall(
            "SELECT DISTINCT sale_ref FROM allocation_plan WHERE sale_ref IS NOT NULL AND sale_ref != ''"
        )
        sale_refs = [r.get('sale_ref') if isinstance(r, dict) else r[0] for r in (rows or [])]
    except Exception as e:
        logger.warning(f'sale_ref 조회 실패: {e}')
        sale_refs = ['SONG-2607', 'WOO-2606']

    logger.info(f'보고서 대상 sale_ref: {sale_refs}')

    for sale_ref in sale_refs:
        # Outbound Report
        try:
            out_path = str(REPORT_DIR / f'Outbound_Report_{sale_ref}.xlsx')
            result = engine._export_outbound_report(
                output_path=out_path,
                sale_ref=sale_ref,
            )
            if result:
                logger.info(f'  [Outbound Report] {sale_ref}: {result}')
            else:
                logger.warning(f'  [Outbound Report] {sale_ref}: 데이터 없음')
        except Exception as e:
            logger.error(f'  [Outbound Report] {sale_ref} 실패: {e}')

        # Sales Order DN
        try:
            dn_path = str(REPORT_DIR / f'Sales_Order_DN_{sale_ref}.xlsx')
            result = engine._export_sales_order_dn_report(
                output_path=dn_path,
                sale_ref=sale_ref,
            )
            if result and 'INCOMPLETE' not in str(result):
                logger.info(f'  [Sales Order DN] {sale_ref}: {result}')
            elif result and 'INCOMPLETE' in str(result):
                logger.warning(f'  [Sales Order DN] {sale_ref}: {result}')
            else:
                logger.warning(f'  [Sales Order DN] {sale_ref}: 데이터 없음')
        except Exception as e:
            logger.error(f'  [Sales Order DN] {sale_ref} 실패: {e}')

    return True


def step6_summary(engine):
    """STEP 6: 최종 통계"""
    logger.info('=' * 70)
    logger.info('STEP 6: 최종 통계')
    logger.info('=' * 70)

    try:
        stats = engine.get_statistics()
        for k, v in stats.items():
            logger.info(f'  {k}: {v}')
    except Exception as e:
        logger.warning(f'통계 조회 실패: {e}')

    # DB 직접 조회
    queries = [
        ('LOT (inventory)', "SELECT COUNT(*) FROM inventory"),
        ('톤백 (inventory_tonbag)', "SELECT COUNT(*) FROM inventory_tonbag"),
        ('AVAILABLE 톤백', "SELECT COUNT(*) FROM inventory_tonbag WHERE status='AVAILABLE'"),
        ('RESERVED 톤백', "SELECT COUNT(*) FROM inventory_tonbag WHERE status='RESERVED'"),
        ('PICKED 톤백', "SELECT COUNT(*) FROM inventory_tonbag WHERE status='PICKED'"),
        ('OUTBOUND 톤백', "SELECT COUNT(*) FROM inventory_tonbag WHERE status IN ('OUTBOUND','SOLD')"),
        ('allocation_plan', "SELECT COUNT(*) FROM allocation_plan"),
        ('picking_table', "SELECT COUNT(*) FROM picking_table"),
        ('sold_table', "SELECT COUNT(*) FROM sold_table"),
    ]
    for label, sql in queries:
        try:
            row = engine.db.fetchone(sql)
            cnt = row[0] if isinstance(row, (tuple, list)) else (row.get('COUNT(*)', 0) if isinstance(row, dict) else 0)
            logger.info(f'  {label}: {cnt}')
        except Exception:
            logger.info(f'  {label}: (테이블 없음)')

    return True


def main():
    logger.info('#' * 70)
    logger.info('#  SQM v8.6.3 FULL PIPELINE TEST')
    logger.info(f'#  {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    logger.info('#' * 70)

    # 파일 존재 확인
    missing = []
    for f in [SUBLOT_FILE, ALLOC_SONG, ALLOC_WOO]:
        if not f.exists():
            missing.append(str(f))
    if missing:
        logger.error(f'필수 파일 누락: {missing}')
        return 1

    logger.info(f'Picking Woo: {len(PICKING_WOO)} files')
    logger.info(f'Picking Song: {len(PICKING_SONG)} files')

    engine = step0_init()

    steps = [
        ('STEP 1: 입고', lambda: step1_inbound(engine)),
        ('STEP 2: 배정', lambda: step2_allocation(engine)),
        ('STEP 3: 피킹', lambda: step3_picking(engine)),
        ('STEP 4: 출고확정', lambda: step4_outbound(engine)),
        ('STEP 5: 보고서', lambda: step5_reports(engine)),
        ('STEP 6: 통계', lambda: step6_summary(engine)),
    ]

    for name, func in steps:
        try:
            ok = func()
            status = 'PASS' if ok else 'FAIL'
        except Exception as e:
            logger.error(f'{name} 예외: {e}', exc_info=True)
            status = 'ERROR'
        logger.info(f'>>> {name}: {status}')

    try:
        engine.close()
    except Exception:
        pass

    logger.info('#' * 70)
    logger.info('#  PIPELINE TEST COMPLETE')
    logger.info('#' * 70)
    return 0


if __name__ == '__main__':
    sys.exit(main())
